from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_RE_SPLIT_TOKEN = re.compile(r"[;, \n]+")
_RE_ALLOWED_ROLE = re.compile(r"^[A-Z0-9_/:-]+$")


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().upper()


def _norm_col(value: Any) -> str:
    text = _norm_text(value)
    return re.sub(r"[^A-Z0-9_]", "", text)


def _split_tokens(raw: Any) -> list[str]:
    if raw is None:
        return []
    text = str(raw).replace("\r", "\n").replace("\t", " ").strip().upper()
    if not text:
        return []

    out: list[str] = []
    for token in _RE_SPLIT_TOKEN.split(text):
        token = token.strip()
        if not token:
            continue
        for prefix in ("TCODE=", "TCODE:", "T=", "T:", "ROLE=", "ROLE:", "/N", "/O"):
            if token.startswith(prefix):
                token = token[len(prefix) :].strip()
                break
        if token:
            out.append(token)
    return list(dict.fromkeys(out))


def _error_result(role_name: str, error_type: str, message: str, status: str = "ERRO") -> dict[str, Any]:
    return {
        "ok": False,
        "status": status,
        "role": role_name,
        "description": None,
        "language": None,
        "system": "DEV",
        "client": "100",
        "error_type": error_type,
        "message": message,
        "warnings": [],
        "errors": [message],
    }


def _analyze_create_sheet(ws: Any, sheet_name: str, role_name: str) -> dict[str, Any]:
    required_columns = {"AGR_NAME", "TEXT", "TCODE", "STATUS", "MSG", "TIMESTEMP"}
    header_row = None
    header_map: dict[str, int] = {}
    best_row = 0
    max_matches = 0

    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_values = [_norm_col(cell.value) for cell in ws[row_idx]]
        found = set(row_values).intersection(required_columns)
        if len(found) > max_matches:
            max_matches = len(found)
            best_row = row_idx
        if len(found) == len(required_columns):
            header_row = row_idx
            for col_idx, name in enumerate(row_values, start=1):
                if name:
                    header_map[name] = col_idx
            break

    if not header_row:
        missing = sorted(required_columns - set(header_map))
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": best_row or None,
                "required_columns": sorted(required_columns),
            },
            "warnings": [],
            "errors": [
                "Nao foi encontrada a linha de cabecalho completa.",
                *( [f"Colunas em falta: {', '.join(missing)}"] if missing else [] ),
            ],
        }

    col_agr = header_map.get("AGR_NAME")
    col_text = header_map.get("TEXT")
    col_tcode = header_map.get("TCODE")
    col_status = header_map.get("STATUS")
    col_msg = header_map.get("MSG")
    col_ts = header_map.get("TIMESTEMP")

    records: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        agr = str(ws.cell(row=row_idx, column=col_agr).value or "").strip() if col_agr else ""
        if not agr:
            continue
        record = {
            "_row": row_idx,
            "AGR_NAME": agr,
            "TEXT": str(ws.cell(row=row_idx, column=col_text).value or "").strip() if col_text else "",
            "TCODE": str(ws.cell(row=row_idx, column=col_tcode).value or "").strip() if col_tcode else "",
            "STATUS": str(ws.cell(row=row_idx, column=col_status).value or "").strip() if col_status else "",
            "MSG": str(ws.cell(row=row_idx, column=col_msg).value or "").strip() if col_msg else "",
            "TIMESTEMP": str(ws.cell(row=row_idx, column=col_ts).value or "").strip() if col_ts else "",
        }
        records.append(record)

    if not records:
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": header_row,
                "records": 0,
            },
            "warnings": [],
            "errors": ["Nao encontrei linhas para processar."],
        }

    roles_map: dict[str, dict[str, Any]] = {}
    ignored_completed = 0
    for record in records:
        if _norm_text(record["STATUS"]) == "CONCLUIDO":
            ignored_completed += 1
            continue

        agr = str(record["AGR_NAME"]).strip().upper()
        if agr not in roles_map:
            roles_map[agr] = {
                "AGR_NAME": agr,
                "TEXT": str(record["TEXT"]).strip(),
                "TCODE_LIST": [],
                "rows": [],
            }
        if not roles_map[agr]["TEXT"] and str(record["TEXT"]).strip():
            roles_map[agr]["TEXT"] = str(record["TEXT"]).strip()
        roles_map[agr]["TCODE_LIST"].extend(_split_tokens(record["TCODE"]))
        roles_map[agr]["rows"].append(record["_row"])

    if not roles_map:
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": header_row,
                "records": len(records),
                "ignored_completed": ignored_completed,
            },
            "warnings": [],
            "errors": ["Todas as linhas uteis estavam marcadas como CONCLUIDO."],
        }

    distinct_roles = sorted(roles_map)
    if role_name and (len(distinct_roles) != 1 or distinct_roles[0] != role_name):
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "role_in_excel": distinct_roles[0] if len(distinct_roles) == 1 else distinct_roles,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": header_row,
                "records": len(records),
                "roles_found": distinct_roles,
                "ignored_completed": ignored_completed,
            },
            "warnings": [],
            "errors": [
                "O Perfil informado no Excel nao corresponde ao Perfil analisado em PRD.",
                f"Perfil analisado: {role_name}",
                f"Perfil no Excel: {', '.join(distinct_roles)}",
            ],
        }

    target_role = role_name if role_name else distinct_roles[0]
    role_data = roles_map[target_role]
    tcode_list = list(dict.fromkeys(role_data["TCODE_LIST"]))
    description = str(role_data["TEXT"]).strip() or None

    warnings: list[str] = []
    errors: list[str] = []
    if not description:
        errors.append("Descricao da role nao encontrada no Excel.")
    if not tcode_list:
        errors.append("Nao foram encontrados TCODEs validos no Excel.")

    if errors:
        return {
            "ok": False,
            "status": "INVALID",
            "role": target_role,
            "description": description,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": header_row,
                "records": len(records),
                "rows": role_data["rows"],
                "tcode_count": len(tcode_list),
            },
            "warnings": warnings,
            "errors": errors,
        }

    return {
        "ok": True,
        "status": "VALID",
        "role": target_role,
        "description": description,
        "language": None,
        "system": "DEV",
        "client": "100",
        "sheet": sheet_name,
        "summary": {
            "header_row": header_row,
            "records": len(records),
            "rows": role_data["rows"],
            "tcode_count": len(tcode_list),
            "unique_tcodes": tcode_list,
        },
        "warnings": warnings,
        "errors": errors,
    }


def _analyze_composta_sheet(ws: Any, sheet_name: str, role_name: str) -> dict[str, Any]:
    min_required = {"TEXT", "STATUS", "MSG", "TIMESTEMP"}
    role_columns_opts = {"ROLES", "TCODE", "ROLES_FILHAS", "AGR_NAME_FILHA"}

    header_row = None
    header_map: dict[str, int] = {}
    best_row = 0
    max_matches = 0

    for row_idx in range(1, min(ws.max_row, 20) + 1):
        row_values = [_norm_col(cell.value) for cell in ws[row_idx]]
        row_set = set(row_values)
        if not min_required.issubset(row_set):
            continue

        if "AGR_NAME_COMPOSTA" in row_set and "AGR_NAME" in row_set:
            header_row = row_idx
            for idx, name in enumerate(row_values, start=1):
                if name:
                    header_map[name] = idx
            break

        if "AGR_NAME" in row_set and len(role_columns_opts.intersection(row_set)) >= 1:
            header_row = row_idx
            for idx, name in enumerate(row_values, start=1):
                if name:
                    header_map[name] = idx
            break

    if not header_row:
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": best_row or None,
                "required_columns": sorted(min_required),
            },
            "warnings": [],
            "errors": ["Nao foi encontrada a linha de cabecalho completa para a Funcao Composta."],
        }

    col_text = header_map.get("TEXT")
    col_status = header_map.get("STATUS")
    col_msg = header_map.get("MSG")
    col_ts = header_map.get("TIMESTEMP")

    if "AGR_NAME_COMPOSTA" in header_map:
        col_agr = header_map["AGR_NAME_COMPOSTA"]
        col_roles = header_map["AGR_NAME"]
    else:
        col_agr = header_map.get("AGR_NAME")
        col_roles = None
        for opt in ["ROLES", "TCODE", "ROLES_FILHAS", "AGR_NAME_FILHA"]:
            if opt in header_map:
                col_roles = header_map[opt]
                break

    records: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        agr = str(ws.cell(row=row_idx, column=col_agr).value or "").strip() if col_agr else ""
        if not agr:
            continue
        record = {
            "_row": row_idx,
            "AGR_NAME": agr,
            "TEXT": str(ws.cell(row=row_idx, column=col_text).value or "").strip() if col_text else "",
            "ROLES": str(ws.cell(row=row_idx, column=col_roles).value or "").strip() if col_roles else "",
            "STATUS": str(ws.cell(row=row_idx, column=col_status).value or "").strip() if col_status else "",
            "MSG": str(ws.cell(row=row_idx, column=col_msg).value or "").strip() if col_msg else "",
            "TIMESTEMP": str(ws.cell(row=row_idx, column=col_ts).value or "").strip() if col_ts else "",
        }
        records.append(record)

    if not records:
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {"header_row": header_row, "records": 0},
            "warnings": [],
            "errors": ["Nao encontrei linhas para processar na Funcao Composta."],
        }

    roles_map: dict[str, dict[str, Any]] = {}
    ignored_completed = 0
    for record in records:
        if _norm_text(record["STATUS"]) == "CONCLUIDO":
            ignored_completed += 1
            continue

        agr = str(record["AGR_NAME"]).strip().upper()
        if agr not in roles_map:
            roles_map[agr] = {
                "AGR_NAME": agr,
                "TEXT": str(record["TEXT"]).strip(),
                "ROLES_LIST": [],
                "rows": [],
            }
        if not roles_map[agr]["TEXT"] and str(record["TEXT"]).strip():
            roles_map[agr]["TEXT"] = str(record["TEXT"]).strip()
        roles_map[agr]["ROLES_LIST"].extend(_split_tokens(record["ROLES"]))
        roles_map[agr]["rows"].append(record["_row"])

    if not roles_map:
        return {
            "ok": False,
            "status": "INVALID",
            "role": role_name,
            "description": None,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {"header_row": header_row, "records": len(records), "ignored_completed": ignored_completed},
            "warnings": [],
            "errors": ["Todas as linhas uteis estavam marcadas como CONCLUIDO."],
        }

    distinct_roles = sorted(roles_map)
    target_role = role_name if role_name in roles_map else distinct_roles[0]
    role_data = roles_map[target_role]
    roles_list = list(dict.fromkeys(role_data["ROLES_LIST"]))
    description = str(role_data["TEXT"]).strip() or None

    warnings: list[str] = []
    errors: list[str] = []
    if not description:
        errors.append("Descricao da Funcao Composta nao encontrada no Excel.")
    if not roles_list:
        errors.append("Nao foram encontradas Funcoes Componentes (roles filhas) validas no Excel.")

    if errors:
        return {
            "ok": False,
            "status": "INVALID",
            "role": target_role,
            "description": description,
            "language": None,
            "system": "DEV",
            "client": "100",
            "sheet": sheet_name,
            "summary": {
                "header_row": header_row,
                "records": len(records),
                "rows": role_data["rows"],
                "tipo": "Função Composta",
                "roles_filhas_count": len(roles_list),
            },
            "warnings": warnings,
            "errors": errors,
        }

    return {
        "ok": True,
        "status": "VALID",
        "role": target_role,
        "description": description,
        "language": None,
        "system": "DEV",
        "client": "100",
        "sheet": sheet_name,
        "summary": {
            "header_row": header_row,
            "records": len(records),
            "rows": role_data["rows"],
            "tipo": "Função Composta",
            "roles_filhas_count": len(roles_list),
            "roles_filhas": roles_list,
        },
        "warnings": warnings,
        "errors": errors,
    }


def analyze_pfcg_create_excel(excel_path: str, expected_role_name: str) -> dict[str, Any]:
    role_name = _norm_text(expected_role_name)
    path = Path(str(excel_path or "").strip())
    if not path:
        return _error_result(role_name, "INVALID_INPUT", "Caminho do ficheiro vazio.")

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return _error_result(role_name, "UNSUPPORTED_FILE", f"Extensao nao suportada: {path.suffix or '<sem extensao>'}.")

    if not path.exists():
        return _error_result(role_name, "FILE_NOT_FOUND", f"Ficheiro nao encontrado: {path.name}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        return _error_result(role_name, "INVALID_WORKBOOK", f"Nao foi possivel abrir o Excel: {exc}")

    try:
        sheet_names = wb.sheetnames
        if "PFCG_COMPOSTA" in sheet_names:
            sheet_name = "PFCG_COMPOSTA"
            ws = wb[sheet_name]
            return _analyze_composta_sheet(ws, sheet_name, role_name)
        elif "PFCG_CREATE" in sheet_names:
            sheet_name = "PFCG_CREATE"
            ws = wb[sheet_name]
            return _analyze_create_sheet(ws, sheet_name, role_name)
        elif len(sheet_names) == 1:
            sheet_name = sheet_names[0]
            ws = wb[sheet_name]
            first_rows_text = " ".join(str(cell.value or "") for row in ws.iter_rows(max_row=5) for cell in row).upper()
            if "COMPOSTA" in first_rows_text or "ROLES_FILHAS" in first_rows_text or "AGR_NAME_COMPOSTA" in first_rows_text:
                return _analyze_composta_sheet(ws, sheet_name, role_name)
            return _analyze_create_sheet(ws, sheet_name, role_name)
        else:
            return _error_result(
                role_name,
                "MISSING_SHEET",
                "Sheet obrigatoria 'PFCG_CREATE' ou 'PFCG_COMPOSTA' nao encontrada.",
                status="INVALID",
            )
    finally:
        wb.close()
