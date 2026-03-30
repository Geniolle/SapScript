# -*- coding: utf-8 -*-

###################################################################################
# D. PFCG_AUTHORITY.py (NOVA VISÃO: PULL DINÂMICO + AUDITORIA LIMPA + RETORNO /N)
# PFCG - Inserção Massiva de Valores de Autorização via PFCGMASSVAL
###################################################################################

def executar(
    ambiente_cockpit,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True
):
    import sys
    import os
    import time
    import unicodedata
    import tkinter as tk

    import pandas as pd
    import win32com.client
    from tkinter import filedialog
    from openpyxl import load_workbook

    tempo_inicio_total = time.time()

    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual) 
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)

    NOME_SHEET = "PFCG_AUTHORITY"
    SEARCH_HEADER_IN_FIRST_ROWS = 20
    COLUNAS_MINIMAS = {"AGR_NAME", "STATUS", "MSG"} 

    MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P", "CUA": "SPA"}
    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'.")

    def formatar_tempo(segundos):
        h, resto = divmod(segundos, 3600)
        m, s = divmod(resto, 60)
        if h > 0: return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
        return f"{int(m):02d}m {int(s):02d}s"

    def norm_col(s):
        if s is None: s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def selecionar_ficheiro():
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
            filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*"))
        )
        root.destroy()
        return path

    ###################################################################################
    # LER EXCEL (MAPEAMENTO DINÂMICO DE TODAS AS COLUNAS)
    ###################################################################################
    if not caminho_ficheiro:
        print("📂 Selecione o ficheiro Excel…")
        caminho_ficheiro = selecionar_ficheiro()
        if not caminho_ficheiro: return

    try:
        wb = load_workbook(caminho_ficheiro)
        ws = wb[NOME_SHEET]
    except Exception as e:
        print(f"❌ Erro ao abrir Excel: {e}")
        return

    header_row = None
    header_map = {}
    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [norm_col(c.value) for c in ws[r]]
        colunas_encontradas = set(row_vals).intersection(COLUNAS_MINIMAS)
        if len(colunas_encontradas) >= len(COLUNAS_MINIMAS):
            header_row = r
            for idx, name in enumerate(row_vals, start=1):
                if name: header_map[name] = idx
            break

    if not header_row:
        print("\n❌ Cabeçalho não encontrado.")
        return wb.close()

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        agr = ws.cell(row=r, column=header_map.get("AGR_NAME", -1)).value
        if not agr: continue
        
        rec = {"_row": r}
        for col_name, col_idx in header_map.items():
            val = ws.cell(row=r, column=col_idx).value
            rec[col_name] = "" if val is None else str(val).strip()
            
        records.append(rec)

    df_original = pd.DataFrame(records)
    df_proc = df_original[~df_original["STATUS"].str.upper().str.contains("CONCLUIDO", na=False)].copy()
    if df_proc.empty:
        return print("⚠️ Tudo concluído.")

    ###################################################################################
    # CONEXÃO SAP
    ###################################################################################
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        session = next((sess for conn in application.Children for sess in conn.Children if sess.Info.SystemName.upper() == SISTEMA_ESPERADO), None)
    except: session = None

    if not session:
        return print(f"❌ Não encontrei sessão do ambiente '{ambiente_cockpit}'.")

    def object_exists(id_string):
        try:
            session.findById(id_string)
            return True
        except: return False

    ###################################################################################
    # LOGGER DE AUDITORIA MANUAL & EXECUTOR
    ###################################################################################
    class PFCG_AuthPage_Auditor:
        def __init__(self, sess):
            self.sess = sess

        def get_sbar(self):
            try:
                sbar = self.sess.findById("wnd[0]/sbar")
                return getattr(sbar, "MessageType", "").strip().upper(), (sbar.Text or "").strip()
            except: return "", ""

        def audit_step(self, descricao, path, acao="press", valor=None, vkey=None, silencioso=False):
            if not silencioso:
                print(f"\n  🔎 [AUDIT] {descricao}")
                log_detail = f"      ↳ ID: {path} | Ação: {acao}"
                if valor is not None: log_detail += f" | Valor: '{valor}'"
                if vkey is not None: log_detail += f" | VKey: {vkey}"
                print(log_detail)
                
                # Descomente ou remova a linha abaixo para automatizar 100%
                input("      ⏳ Prima [ENTER] para executar este passo no SAP...") 

            try:
                if path: elem = self.sess.findById(path)

                if acao == "text": elem.text = valor
                elif acao == "press": elem.press()
                elif acao == "select": 
                    if hasattr(elem, "selected"): elem.selected = True
                    else: elem.select()
                elif acao == "sendVKey":
                    if path: elem.sendVKey(vkey)
                    else: self.sess.findById("wnd[0]").sendVKey(vkey)

                if not silencioso:
                    print("      ✅ SUCESSO")
                    mtype, mtext = self.get_sbar()
                    if mtext:
                        icone = "🔴" if mtype in ["E", "A"] else ("🟡" if mtype == "W" else "🟢")
                        print(f"      {icone} SAP STATUS: [{mtype}] {mtext}")
                
            except Exception as e:
                if not silencioso:
                    print(f"      ❌ FALHA AQUI (Erro 619): O ID [{path}] falhou.")
                    raise Exception(f"FALHA NO PASSO: '{descricao}' -> ID: {path}")

        def ensure_role_exists(self, nome, desc):
            self.audit_step("Chamar transação /npfcg", "wnd[0]/tbar[0]/okcd", "text", "/npfcg")
            self.audit_step("Enter (Ir para PFCG)", "wnd[0]", "sendVKey", vkey=0)
            
            self.audit_step("Inserir Nome da Role", "wnd[0]/usr/ctxtAGR_NAME_NEU", "text", nome)
            self.audit_step("Clicar Criar Papel Simples", "wnd[0]/usr/btn%#AUTOTEXT003", "press")
            
            if object_exists("wnd[0]/usr/txtS_AGR_TEXTS-TEXT"):
                descricao_final = desc if desc else f"Criado via Script - {nome}"
                self.audit_step("Preencher Descrição da Role", "wnd[0]/usr/txtS_AGR_TEXTS-TEXT", "text", descricao_final)
                self.audit_step("Enter após descrição", "wnd[0]", "sendVKey", vkey=0)
                
                if self.sess.Children.Count > 1 and object_exists("wnd[1]/usr/btnBUTTON_1"):
                    self.audit_step("Confirmar popup criação (Sim)", "wnd[1]/usr/btnBUTTON_1", "press")
                
                if object_exists("wnd[0]/tbar[0]/btn[11]"):
                    self.audit_step("Guardar Role", "wnd[0]/tbar[0]/btn[11]", "press")
            else:
                print("  ├─ [SAP] A função já existe.")

        def update_mass_values_dynamic(self, nome, objeto, row_data):
            self.audit_step("Chamar transação /nPFCGMASSVAL", "wnd[0]/tbar[0]/okcd", "text", "/nPFCGMASSVAL")
            self.audit_step("Enter (Ir para MASSVAL)", "wnd[0]", "sendVKey", vkey=0)
            
            self.audit_step("Selecionar Execução Direta", "wnd[0]/usr/radMOD_EXE", "select")
            self.audit_step("Selecionar Inserção Manual", "wnd[0]/usr/radSEL_NAU", "select")
            self.audit_step("ENTER Crucial do VBS para atualizar tela", "wnd[0]", "sendVKey", vkey=0)
            
            self.audit_step("Preencher ROLE-LOW por segurança", "wnd[0]/usr/ctxtROLE-LOW", "text", nome)
            
            self.audit_step("Preencher OBJOBJ (Objeto de Autorização)", "wnd[0]/usr/ctxtOBJOBJ", "text", objeto)
            self.audit_step("Enter após OBJOBJ (A carregar campos...)", "wnd[0]", "sendVKey", vkey=0)
            
            time.sleep(0.5)

            # ---------------------------------------------------------
            # LEITURA DINÂMICA SILENCIOSA E EFICIENTE
            # ---------------------------------------------------------
            campos_encontrados_sap = 0
            
            for j in range(1, 15):
                btn_id = f"wnd[0]/usr/btnPOBJ{j}N"
                if not object_exists(btn_id):
                    break # Sem botão, acabaram os campos do SAP.
                
                campos_encontrados_sap += 1
                
                # Leitura silenciosa do nome do campo
                campo_sap_tecnico = ""
                ids_para_testar = [
                    f"wnd[0]/usr/txtOBJFLD{j}",      
                    f"wnd[0]/usr/ctxtOBJFLD{j}",     
                    f"wnd[0]/usr/ctxtS_AGR_DEFINE-FNAM{j}",
                    f"wnd[0]/usr/txtS_AGR_DEFINE-FNAM{j}",
                    f"wnd[0]/usr/ctxtFNAM{j}",
                    f"wnd[0]/usr/txtFNAM{j}",
                    f"wnd[0]/usr/ctxtFIELD{j}",
                    f"wnd[0]/usr/txtFIELD{j}"
                ]
                
                for cid in ids_para_testar:
                    if object_exists(cid):
                        campo_sap_tecnico = self.sess.findById(cid).text.strip().upper()
                        break

                # Fallback silencioso
                if not campo_sap_tecnico:
                    campo_sap_tecnico = str(row_data.get(f'CAMPO {j}', '')).strip().upper()

                if not campo_sap_tecnico:
                    continue # Linha vazia, salta sem imprimir nada

                # Cruzamento com o Excel
                valor_no_excel = str(row_data.get(campo_sap_tecnico, '')).strip()

                if valor_no_excel and valor_no_excel != 'NAN':
                    print(f"\n  ⭐ [CAMPO IDENTIFICADO] Nome: '{campo_sap_tecnico}' | Valor a inserir: '{valor_no_excel}'")
                    
                    self.audit_step(f"Clicar Botão VALS para '{campo_sap_tecnico}'", btn_id, "press")
                    time.sleep(0.5) 
                    
                    if valor_no_excel == "*":
                        self.audit_step(f"Full Auth (*) no Campo '{campo_sap_tecnico}'", "wnd[1]/usr/btnGES2", "press")
                    else:
                        valores_lista = [v.strip() for v in valor_no_excel.split(',')]
                        
                        if object_exists("wnd[1]/usr/tblSAPLSUPRNACT_TC"):
                            print(f"      ℹ️ Janela de Atividades Detetada! Valores: {valores_lista}")
                            for linha in range(15):
                                try:
                                    act_code = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNACT_TC/txtH_FVAL-LOW[1,{linha}]").text.strip()
                                    if act_code in valores_lista:
                                        self.audit_step(f"Marcar Checkbox '{act_code}'", f"wnd[1]/usr/tblSAPLSUPRNACT_TC/chkH_FVAL-MARK[0,{linha}]", "select", silencioso=True)
                                        print(f"      ✅ Checkbox '{act_code}' marcada com sucesso.")
                                except: break 
                        else:
                            for idx_val, val in enumerate(valores_lista):
                                try:
                                    self.audit_step(f"Preencher Valor '{val}' para '{campo_sap_tecnico}'", f"wnd[1]/usr/tblSAPLSUPRNVAL_TC/ctxtH_FVAL_LOW[0,{idx_val}]", "text", val, silencioso=True)
                                except Exception:
                                    self.audit_step(f"Preencher Valor '{val}' para '{campo_sap_tecnico}'", f"wnd[1]/usr/tblSAPLSUPRNVAL_TC/ctxtH_FVAL_LOW[1,{idx_val}]", "text", val, silencioso=True)
                                print(f"      ✅ Valor '{val}' inserido com sucesso.")

                    self.audit_step(f"Confirmar Popup de '{campo_sap_tecnico}'", "wnd[1]/tbar[0]/btn[0]", "press")
                
                # Opcional: Se quiser omitir também os ignorados, pode apagar este else:
                # else:
                #     print(f"  ⏭️ [IGNORADO] O SAP pediu '{campo_sap_tecnico}', mas a coluna no Excel está vazia.")

            if campos_encontrados_sap == 0:
                print("  ⚠️ AVISO: O Objeto inserido não gerou campos visíveis.")

            self.audit_step("Clicar Executar (Relógio)", "wnd[0]/tbar[1]/btn[8]", "press")
            self.audit_step("Clicar Guardar (Disquete)", "wnd[0]/tbar[1]/btn[20]", "press")
            
            if self.sess.Children.Count > 1 and object_exists("wnd[1]/tbar[0]/btn[0]"):
                self.audit_step("Confirmar popup Sucesso Gravação", "wnd[1]/tbar[0]/btn[0]", "press")

            self.audit_step("Voltar ao Ecrã Inicial (/N)", "wnd[0]/tbar[0]/okcd", "text", "/N")
            self.audit_step("Enter (/N)", "wnd[0]", "sendVKey", vkey=0)

    ###################################################################################
    # EXECUÇÃO
    ###################################################################################
    auditor = PFCG_AuthPage_Auditor(session)
    resultados = {}

    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
        session.findById("wnd[0]").sendVKey(0)
    except: pass

    for idx, rr in df_proc.iterrows():
        nome = str(rr["AGR_NAME"]).strip()
        desc = str(rr.get("TEXT", "")).strip()
        
        objeto = str(rr.get("OBJETO DE AUTORIZACAO", rr.get("OBJETO", ""))).strip()
        if not objeto: objeto = "F_KNA1_GRP"
        
        print("\n======================================================================")
        print(f"▶ [{idx+1}/{len(df_proc)}] LÓGICA DINÂMICA: {nome} | OBJ: {objeto}")
        print("======================================================================")
        
        try:
            print("\n[Etapa 1] Validação da Role")
            auditor.ensure_role_exists(nome, desc)
            
            print("\n[Etapa 2] Inserção MASSVAL (Modo Pull Dinâmico)")
            auditor.update_mass_values_dynamic(nome, objeto, rr)
            
            resultados[idx] = {"STATUS": "CONCLUIDO", "MSG": "Sucesso Auditado"}
            print("\n🟢 SUCESSO! Ecrã preenchido de forma dinâmica.")
            break # Remova ou comente este break quando quiser processar todas as linhas
            
        except Exception as e:
            resultados[idx] = {"STATUS": "ERRO", "MSG": str(e)}
            print(f"\n🔴 INTERRUPÇÃO DETETADA:\n{str(e)}")
            try: session.findById("wnd[0]/tbar[0]/okcd").text = "/N"; session.findById("wnd[0]").sendVKey(0)
            except: pass
            break 

    return True

if __name__ == "__main__":
    executar("DEV")