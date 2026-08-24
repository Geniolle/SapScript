from __future__ import annotations

import importlib.util
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


MODULE_PATH = Path("Processos") / "Códigos IVA" / "A. CRIAR_CODIGO_IVA.py"


def load_module():
    spec = importlib.util.spec_from_file_location("codigo_iva", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules.pop(spec.name, None)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def temp_dir() -> Path:
    return Path(tempfile.mkdtemp(prefix="codigo_iva_"))


def build_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "CRIAR_CODIGO_IVA"

    ws["B2"] = "texto solto"
    ws["A5"] = " ACTION "
    ws["B5"] = " COUNTRY "
    ws["C5"] = " VAT_CODE "
    ws["D5"] = " DESCRIPTION_PT "
    ws["E5"] = " RATE "
    ws["F5"] = " VAT_TYPE "
    ws["G5"] = " TAX_TYPE "
    ws["H5"] = " STATUS "
    ws["I5"] = " MSG "
    ws["J5"] = " TIMESTEMP "
    ws["K5"] = " OPERATION_DEBIT "
    ws["L5"] = " GL_ACCOUNT_DEBIT "
    ws["M5"] = " OPERATION_CREDIT "
    ws["N5"] = " GL_ACCOUNT_CREDIT "
    ws["O5"] = " REPORTING_COUNTRY "

    ws["A6"] = "CRIAR"
    ws["B6"] = "PT"
    ws["C6"] = "AC"
    ws["D6"] = "Descricao valida"
    ws["E6"] = "23,5"
    ws["F6"] = "I"
    ws["G6"] = "M"
    ws["H6"] = ""
    ws["I6"] = ""
    ws["J6"] = ""
    ws["K6"] = "VST"
    ws["L6"] = "000024320110"
    ws["M6"] = "MWS"
    ws["N6"] = "000024320120"
    ws["O6"] = "PT"

    ws["A7"] = "VALIDAR"
    ws["B7"] = "PT"
    ws["C7"] = "AC"
    ws["D7"] = "Descricao valida"
    ws["E7"] = 23.5
    ws["F7"] = "I"
    ws["G7"] = "M"
    ws["H7"] = "PENDENTE"
    ws["I7"] = ""
    ws["J7"] = ""
    ws["K7"] = "VST"
    ws["L7"] = "000024320110"
    ws["O7"] = "PT"

    ws["A8"] = "CRIAR"
    ws["B8"] = "ES"
    ws["C8"] = "ESE"
    ws["D8"] = "Descricao outra"
    ws["E8"] = 21
    ws["F8"] = "I"
    ws["G8"] = "M"
    ws["H8"] = ""
    ws["I8"] = ""
    ws["J8"] = ""
    ws["O8"] = "ES"

    wb.save(path)
    wb.close()
    return path


def build_exec_workbook(path: Path, rows: list[dict[str, object]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "CRIAR_CODIGO_IVA"
    headers = [
        "ACTION",
        "COUNTRY",
        "VAT_CODE",
        "DESCRIPTION_PT",
        "RATE",
        "VAT_TYPE",
        "TAX_TYPE",
        "STATUS",
        "MSG",
        "TIMESTEMP",
        "DESCRIPTION_EN",
        "DESCRIPTION_ES",
        "OPERATION_DEBIT",
        "GL_ACCOUNT_DEBIT",
        "OPERATION_CREDIT",
        "GL_ACCOUNT_CREDIT",
        "REPORTING_COUNTRY",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=5, column=idx).value = header
    for row_idx, row in enumerate(rows, start=6):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx).value = row.get(header, "")
    wb.save(path)
    wb.close()
    return path


def make_row(
    *,
    action: str,
    country: str,
    vat_code: str,
    description_pt: str = "Descricao valida",
    rate: object = "23,5",
    vat_type: str = "I",
    tax_type: str = "M",
    status: str = "",
    description_en: str = "",
    description_es: str = "",
    operation_debit: str = "",
    gl_account_debit: str = "",
    operation_credit: str = "",
    gl_account_credit: str = "",
    reporting_country: str = "",
) -> dict[str, object]:
    return {
        "ACTION": action,
        "COUNTRY": country,
        "VAT_CODE": vat_code,
        "DESCRIPTION_PT": description_pt,
        "RATE": rate,
        "VAT_TYPE": vat_type,
        "TAX_TYPE": tax_type,
        "STATUS": status,
        "MSG": "",
        "TIMESTEMP": "",
        "DESCRIPTION_EN": description_en,
        "DESCRIPTION_ES": description_es,
        "OPERATION_DEBIT": operation_debit,
        "GL_ACCOUNT_DEBIT": gl_account_debit,
        "OPERATION_CREDIT": operation_credit,
        "GL_ACCOUNT_CREDIT": gl_account_credit,
        "REPORTING_COUNTRY": reporting_country,
    }


def read_rows(path: Path):
    wb = load_workbook(path)
    ws = wb["CRIAR_CODIGO_IVA"]
    rows = {
        idx: (ws[f"H{idx}"].value, ws[f"I{idx}"].value, ws[f"J{idx}"].value)
        for idx in range(6, ws.max_row + 1)
    }
    wb.close()
    return rows


class FakeCodigoIvaSapGuiClient:
    def __init__(
        self,
        session=None,
        request_ctx=None,
        *,
        consultar_codigo_fn=None,
        criar_codigo_fn=None,
        consultar_conta_fn=None,
        configurar_conta_fn=None,
    ):
        self.session = session
        self.request_ctx = request_ctx or {}
        self.consultar_codigo_fn = consultar_codigo_fn
        self.criar_codigo_fn = criar_codigo_fn
        self.consultar_conta_fn = consultar_conta_fn
        self.configurar_conta_fn = configurar_conta_fn
        self.calls = []

    def consultar_codigo(self, payload):
        self.calls.append(("consultar_codigo", payload))
        if self.consultar_codigo_fn is not None:
            return self.consultar_codigo_fn(payload)
        return {"exists": False}

    def criar_codigo(self, payload):
        self.calls.append(("criar_codigo", payload))
        if self.criar_codigo_fn is not None:
            return self.criar_codigo_fn(payload)
        return {"success": True}

    def consultar_conta(self, payload):
        self.calls.append(("consultar_conta", payload))
        if self.consultar_conta_fn is not None:
            return self.consultar_conta_fn(payload)
        return {"exists": False, "equivalent": False, "gl_account": ""}

    def configurar_conta(self, payload):
        self.calls.append(("configurar_conta", payload))
        if self.configurar_conta_fn is not None:
            return self.configurar_conta_fn(payload)
        return {"success": True}


def install_fake_client(monkeypatch, mod, client, session_marker="SESSION", captured=None):
    if captured is None:
        captured = []

    def fake_obter_sessao_sap(ambiente):
        captured.append(ambiente)
        return session_marker

    monkeypatch.setattr(mod, "CodigoIvaSapGuiClient", lambda session, request_ctx=None, recordings_available=False: client)
    monkeypatch.setattr(mod, "obter_sessao_sap", fake_obter_sessao_sap)
    return captured


def test_normalizar_cabecalho():
    mod = load_module()
    assert mod.normalizar_cabecalho("  Descrição\nPT ") == "DESCRICAO PT"
    assert mod.normalizar_cabecalho("  vat_code  ") == "VAT_CODE"


def test_localizar_cabecalho_e_ler_registros():
    mod = load_module()
    path = build_workbook(temp_dir() / "codigo_iva_localizar.xlsx")
    wb = load_workbook(path)
    ws, header_row, header_map = mod.localizar_folha_cabecalho(wb)
    assert header_row == 5
    assert header_map["VAT_CODE"] == 3
    registros = mod.ler_registros(ws, header_row, header_map)
    assert len(registros) == 3
    wb.close()


def test_localizar_folha_por_nome_normalizado():
    mod = load_module()
    wb = Workbook()
    ws = wb.active
    ws.title = "Criar_Codigo_Iva"
    ws["A1"] = "ACTION"
    ws["B1"] = "COUNTRY"
    ws["C1"] = "VAT_CODE"
    ws["D1"] = "DESCRIPTION_PT"
    ws["E1"] = "RATE"
    ws["F1"] = "VAT_TYPE"
    ws["G1"] = "TAX_TYPE"
    ws["H1"] = "STATUS"
    ws["I1"] = "MSG"
    ws["J1"] = "TIMESTEMP"
    ws2, header_row, header_map = mod.localizar_folha_cabecalho(wb)
    assert ws2.title == "Criar_Codigo_Iva"
    assert header_row == 1
    assert header_map["VAT_CODE"] == 3
    wb.close()


def test_codigo_dois_caracteres_e_especial_valido():
    mod = load_module()
    grupo = {
        "COUNTRY": "PT",
        "VAT_CODE": "A1",
        "rows": [
            {
                "ACTION": "CRIAR",
                "COUNTRY": "PT",
                "VAT_CODE": "A1",
                "DESCRIPTION_PT": "Descricao valida",
                "RATE_RAW": "23,5",
                "VAT_TYPE": "I",
                "TAX_TYPE": "M",
            }
        ],
    }
    validacao = mod.validar_grupo(grupo)
    assert validacao.ok is True
    assert validacao.dados["vat_code"] == "A1"


def test_codigo_interrogacao_rejeitado():
    mod = load_module()
    grupo = {
        "COUNTRY": "PT",
        "VAT_CODE": "?1",
        "rows": [
            {
                "ACTION": "CRIAR",
                "COUNTRY": "PT",
                "VAT_CODE": "?1",
                "DESCRIPTION_PT": "Descricao valida",
                "RATE_RAW": "23",
                "VAT_TYPE": "I",
                "TAX_TYPE": "M",
            }
        ],
    }
    validacao = mod.validar_grupo(grupo)
    assert validacao.ok is False
    assert "?" in validacao.msg


def test_descricao_maior_que_50_chars():
    mod = load_module()
    grupo = {
        "COUNTRY": "PT",
        "VAT_CODE": "AC",
        "rows": [
            {
                "ACTION": "CRIAR",
                "COUNTRY": "PT",
                "VAT_CODE": "AC",
                "DESCRIPTION_PT": "x" * 51,
                "RATE_RAW": "23",
                "VAT_TYPE": "I",
                "TAX_TYPE": "M",
            }
        ],
    }
    validacao = mod.validar_grupo(grupo)
    assert validacao.ok is False
    assert "50 caracteres" in validacao.msg


@pytest.mark.parametrize("valor, esperado", [("23,5", 23.5), ("23.5", 23.5), (23, 23.0)])
def test_parse_rate(valor, esperado):
    mod = load_module()
    assert mod.parse_rate(valor) == esperado


def test_conta_com_zeros_preservada():
    mod = load_module()
    assert mod.valor_em_texto_preservado(24320110) == "24320110"
    assert mod.valor_em_texto_preservado("000024320110") == "000024320110"


def test_linhas_ja_concluidas_nao_entram_no_grupo():
    mod = load_module()
    registros = [
        {"_row": 1, "ACTION": "CRIAR", "COUNTRY": "PT", "VAT_CODE": "AC", "STATUS": "CONCLUIDO"},
        {"_row": 2, "ACTION": "CRIAR", "COUNTRY": "PT", "VAT_CODE": "AC", "STATUS": ""},
    ]
    grupos = mod.agrupar_registros(registros)
    assert len(grupos) == 1
    assert len(grupos[0]["rows"]) == 1


def test_agrupamento_por_pais_e_codigo():
    mod = load_module()
    registros = [
        {"_row": 1, "ACTION": "CRIAR", "COUNTRY": "PT", "VAT_CODE": "AC", "STATUS": ""},
        {"_row": 2, "ACTION": "VALIDAR", "COUNTRY": "PT", "VAT_CODE": "AC", "STATUS": ""},
        {"_row": 3, "ACTION": "CRIAR", "COUNTRY": "ES", "VAT_CODE": "E1", "STATUS": ""},
    ]
    grupos = mod.agrupar_registros(registros)
    assert [(g["COUNTRY"], g["VAT_CODE"]) for g in grupos] == [("ES", "E1"), ("PT", "AC")]


def test_operacao_debito_e_credito():
    mod = load_module()
    grupo = {
        "COUNTRY": "PT",
        "VAT_CODE": "AC",
        "rows": [
            {
                "ACTION": "CRIAR",
                "COUNTRY": "PT",
                "VAT_CODE": "AC",
                "DESCRIPTION_PT": "Descricao valida",
                "RATE_RAW": "23",
                "VAT_TYPE": "I",
                "TAX_TYPE": "M",
                "OPERATION_DEBIT": "VST",
                "GL_ACCOUNT_DEBIT": "000024320110",
                "OPERATION_CREDIT": "MWS",
                "GL_ACCOUNT_CREDIT": "000024320120",
            }
        ],
    }
    validacao = mod.validar_grupo(grupo)
    assert validacao.ok is True
    assert validacao.dados["ob40_pairs"] == [
        {"OPERATION": "VST", "GL_ACCOUNT": "000024320110", "DIRECTION": "DEBIT"},
        {"OPERATION": "MWS", "GL_ACCOUNT": "000024320120", "DIRECTION": "CREDIT"},
    ]


def test_validar_sem_criar():
    mod = load_module()
    grupo = {
        "COUNTRY": "PT",
        "VAT_CODE": "AC",
        "rows": [
            {
                "ACTION": "VALIDAR",
                "COUNTRY": "PT",
                "VAT_CODE": "AC",
                "DESCRIPTION_PT": "Descricao valida",
                "RATE_RAW": "23",
                "VAT_TYPE": "I",
                "TAX_TYPE": "M",
            }
        ],
    }
    validacao = mod.validar_grupo(grupo)
    assert validacao.ok is True
    assert validacao.dados["action"] == "VALIDAR"


def test_atualizacao_status_msg_timestamp():
    mod = load_module()
    path = build_workbook(temp_dir() / "codigo_iva_status.xlsx")
    wb = load_workbook(path)
    ws, header_row, header_map = mod.localizar_folha_cabecalho(wb)
    registros = mod.ler_registros(ws, header_row, header_map)
    resultados = {
        str(registros[0]["_row"]): {"STATUS": "CONCLUIDO", "MSG": "OK", "TIMESTEMP": "2026-07-23 10:11:12"},
    }
    assert mod.gravar_resultados_excel(path, "CRIAR_CODIGO_IVA", header_map, registros, resultados) is True
    wb2 = load_workbook(path)
    ws2 = wb2["CRIAR_CODIGO_IVA"]
    assert ws2["H6"].value == "CONCLUIDO"
    assert ws2["I6"].value == "OK"
    assert ws2["J6"].value == "2026-07-23 10:11:12"
    wb2.close()


def test_confirmacao_global(monkeypatch):
    mod = load_module()
    monkeypatch.setattr("builtins.input", lambda *_args, **_kwargs: "S")
    assert mod._confirmar_execucao("DEV", {"request_number": "S4DK900001"}, [{"COUNTRY": "PT"}], 0) is True


def test_validar_codigo_inexistente(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "validar_inexistente.xlsx", [make_row(action="VALIDAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient(consultar_codigo_fn=lambda _payload: {"exists": False})
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is True
    rows = read_rows(path)
    assert rows[6][0] == "IGNORADO"
    assert "não encontrado" in rows[6][1].lower()


def test_validar_codigo_equivalente(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "validar_equivalente.xlsx", [make_row(action="VALIDAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient(
        consultar_codigo_fn=lambda _payload: {
            "exists": True,
            "country": "PT",
            "vat_code": "AC",
            "description_pt": "Descricao valida",
            "rate": 23.5,
            "vat_type": "I",
            "tax_type": "M",
            "conditions": [{"condition_key": "I", "rate": 23.5, "direction": "PT"}],
            "accounts": [],
        },
    )
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is True
    rows = read_rows(path)
    assert rows[6][0] == "CONCLUIDO"
    assert "configuração equivalente" in rows[6][1].lower()


def test_validar_codigo_diferente(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "validar_diferente.xlsx", [make_row(action="VALIDAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient(
        consultar_codigo_fn=lambda _payload: {
            "exists": True,
            "country": "PT",
            "vat_code": "AC",
            "description_pt": "Outra descricao",
            "rate": 21,
            "vat_type": "I",
            "tax_type": "M",
            "conditions": [],
            "accounts": [],
        },
    )
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert "descrição diferente" in rows[6][1].lower() or "taxa diferente" in rows[6][1].lower()


def test_criar_sem_request_marca_erro(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "criar_sem_request.xlsx", [make_row(action="CRIAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient()
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert "obrigatória uma request" in rows[6][1].lower()
    assert client.calls == []


def test_criar_sem_transporte_rejeitado(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "criar_sem_transporte.xlsx", [make_row(action="CRIAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient()
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar(
        "DEV",
        request_ctx={"request_number": "S4DK900001", "request_option": "4"},
        caminho_ficheiro=str(path),
        modo_nao_interativo=True,
        pedir_confirmacao=False,
    )
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert "sem transporte" in rows[6][1].lower()


def test_criar_equivalente_nao_altera(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "criar_equivalente.xlsx", [make_row(action="CRIAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient(
        consultar_codigo_fn=lambda _payload: {
            "exists": True,
            "country": "PT",
            "vat_code": "AC",
            "description_pt": "Descricao valida",
            "rate": 23.5,
            "vat_type": "I",
            "tax_type": "M",
            "conditions": [{"condition_key": "I", "rate": 23.5, "direction": "PT"}],
            "accounts": [],
        },
    )
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar(
        "DEV",
        request_ctx={"request_number": "S4DK900001", "request_option": "1"},
        caminho_ficheiro=str(path),
        modo_nao_interativo=True,
        pedir_confirmacao=False,
    )
    assert ok is True
    rows = read_rows(path)
    assert rows[6][0] == "CONCLUIDO"
    assert "configuração equivalente" in rows[6][1].lower()
    assert [call[0] for call in client.calls] == ["consultar_codigo"]


def test_criar_com_ob40_sucesso(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(
        temp_dir() / "criar_ob40_sucesso.xlsx",
        [
            make_row(
                action="CRIAR",
                country="PT",
                vat_code="AC",
                operation_debit="VST",
                gl_account_debit="000024320110",
                operation_credit="MWS",
                gl_account_credit="000024320120",
            )
        ],
    )
    consult_state = {"calls": 0}

    def consultar_codigo(_payload):
        consult_state["calls"] += 1
        if consult_state["calls"] == 1:
            return {"exists": False}
        return {
            "exists": True,
            "country": "PT",
            "vat_code": "AC",
            "description_pt": "Descricao valida",
            "rate": 23.5,
            "vat_type": "I",
            "tax_type": "M",
            "conditions": [{"condition_key": "I", "rate": 23.5, "direction": "PT"}],
            "accounts": [],
        }

    client = FakeCodigoIvaSapGuiClient(
        consultar_codigo_fn=consultar_codigo,
        criar_codigo_fn=lambda _payload: {"success": True},
        consultar_conta_fn=lambda payload: {"exists": False, "equivalent": False, "gl_account": "", "payload": payload},
        configurar_conta_fn=lambda payload: {"success": True, "payload": payload},
    )
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar(
        "DEV",
        request_ctx={"request_number": "S4DK900001", "request_option": "1"},
        caminho_ficheiro=str(path),
        modo_nao_interativo=True,
        pedir_confirmacao=False,
    )
    assert ok is True
    rows = read_rows(path)
    assert rows[6][0] == "CONCLUIDO"
    assert "contas configuradas na ob40" in rows[6][1].lower()
    assert [call[0] for call in client.calls] == [
        "consultar_codigo",
        "criar_codigo",
        "consultar_codigo",
        "consultar_conta",
        "configurar_conta",
        "consultar_conta",
        "configurar_conta",
    ]


def test_criar_com_falha_ob40(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(
        temp_dir() / "criar_ob40_falha.xlsx",
        [make_row(action="CRIAR", country="PT", vat_code="AC", operation_debit="VST", gl_account_debit="000024320110")],
    )
    consult_state = {"calls": 0}

    def consultar_codigo(_payload):
        consult_state["calls"] += 1
        if consult_state["calls"] == 1:
            return {"exists": False}
        return {
            "exists": True,
            "country": "PT",
            "vat_code": "AC",
            "description_pt": "Descricao valida",
            "rate": 23.5,
            "vat_type": "I",
            "tax_type": "M",
            "conditions": [{"condition_key": "I", "rate": 23.5, "direction": "PT"}],
            "accounts": [],
        }

    client = FakeCodigoIvaSapGuiClient(
        consultar_codigo_fn=consultar_codigo,
        criar_codigo_fn=lambda _payload: {"success": True},
        consultar_conta_fn=lambda payload: {"exists": False, "equivalent": False, "gl_account": "", "payload": payload},
        configurar_conta_fn=lambda payload: {"success": False, "message": "OB40 indisponível", "payload": payload},
    )
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar(
        "DEV",
        request_ctx={"request_number": "S4DK900001", "request_option": "1"},
        caminho_ficheiro=str(path),
        modo_nao_interativo=True,
        pedir_confirmacao=False,
    )
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert "determinação de contas pendente" in rows[6][1].lower()


def test_condicao_fiscal_nao_encontrada(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "condicao_fiscal.xlsx", [make_row(action="VALIDAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient(
        consultar_codigo_fn=lambda _payload: {
            "exists": True,
            "country": "PT",
            "vat_code": "AC",
            "description_pt": "Descricao valida",
            "rate": 23.5,
            "vat_type": "Z",
            "tax_type": "M",
            "conditions": [{"condition_key": "VST", "rate": 23.5, "direction": "PT"}],
            "accounts": [],
        },
    )
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert "condição fiscal" in rows[6][1].lower()


def test_retorno_tipo_erro_e_recuperacao(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(
        temp_dir() / "recovery.xlsx",
        [
            make_row(action="VALIDAR", country="PT", vat_code="AC", rate=21),
            make_row(action="VALIDAR", country="ES", vat_code="E1", rate=21),
        ],
    )
    state = {"calls": 0}

    def consultar_codigo(payload):
        state["calls"] += 1
        if payload["vat_code"] == "AC":
            return {"exists": True, "country": "PT", "vat_code": "AC", "description_pt": "Outra", "rate": 21, "vat_type": "I", "tax_type": "M", "conditions": [], "accounts": []}
        return {"exists": True, "country": "ES", "vat_code": "E1", "description_pt": "Descricao valida", "rate": 21, "vat_type": "I", "tax_type": "M", "conditions": [], "accounts": []}

    client = FakeCodigoIvaSapGuiClient(consultar_codigo_fn=consultar_codigo)
    install_fake_client(monkeypatch, mod, client)
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert rows[7][0] == "CONCLUIDO"


def test_sessao_por_ambiente(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "ambiente.xlsx", [make_row(action="VALIDAR", country="PT", vat_code="AC")])
    client = FakeCodigoIvaSapGuiClient(consultar_codigo_fn=lambda _payload: {"exists": False})
    called = []
    install_fake_client(monkeypatch, mod, client, captured=called)
    mod.executar("QAD", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert called == ["QAD"]


def test_default_cliente_sap_gui_sem_gravacoes(monkeypatch):
    mod = load_module()
    path = build_exec_workbook(temp_dir() / "pendente_sap_gui.xlsx", [make_row(action="VALIDAR", country="PT", vat_code="AC")])
    monkeypatch.setattr(mod, "obter_sessao_sap", lambda _ambiente: object())
    ok = mod.executar("DEV", request_ctx={}, caminho_ficheiro=str(path), modo_nao_interativo=True, pedir_confirmacao=False)
    assert ok is False
    rows = read_rows(path)
    assert rows[6][0] == "ERRO"
    assert "gravações sap gui scripting não disponíveis" in rows[6][1].lower()
