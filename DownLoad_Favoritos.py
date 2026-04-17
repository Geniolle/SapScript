# -*- coding: utf-8 -*-
"""
Download dos favoritos SAP por utilizador
- Abre um ficheiro Excel por popup em primeiro plano
- Lê a sheet "Users Ativos"
- Procura dinamicamente a coluna "Usuário" pelo nome do cabeçalho
- Executa o download dos favoritos no SAP para cada utilizador
- O nome do ficheiro é sempre o próprio utilizador
- Preenche STATUS, MSG e TIMESTEMP no Excel
- Sempre lê o wnd[0]/sbar para preencher o STATUS

Requisitos:
    pip install openpyxl pywin32

Observações:
- Ajuste SAP_CONNECTION_INDEX e SAP_SESSION_INDEX se necessário
- O diretório de download padrão é C:\Favoritos
"""

import os
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog, messagebox

import pythoncom
import win32com.client
from openpyxl import load_workbook


# =========================================================
# CONFIGURAÇÕES
# =========================================================
SAP_CONNECTION_INDEX = 0
SAP_SESSION_INDEX = 0
SHEET_NAME = "Users Ativos"
DOWNLOAD_DIR = r"C:\Favoritos"

HEADER_SEARCH_MAX_ROWS = 20
HEADER_REQUIRED = ["Usuário"]
OPTIONAL_COLUMNS = ["STATUS", "MSG", "TIMESTEMP"]

SAP_PROGRAM = "MENU_FAVORITES_DOWNLOAD"
SAP_TCODE = "/nse37"

SLEEP_SHORT = 0.3
SLEEP_MEDIUM = 0.8


# =========================================================
# UTILITÁRIOS GERAIS
# =========================================================
def log(msg: str) -> None:
    print(msg, flush=True)


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_header(value) -> str:
    text = normalize_text(value).upper()
    return text


def safe_filename(name: str) -> str:
    name = normalize_text(name)
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    return name


def ensure_folder(path: str) -> None:
    Path(path).mkdir(parents=True, exist_ok=True)


# =========================================================
# POPUP DE SELEÇÃO DO EXCEL
# =========================================================
def select_excel_file() -> str:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update()

    file_path = filedialog.askopenfilename(
        parent=root,
        title="Selecionar ficheiro Excel",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("All files", "*.*"),
        ],
    )

    root.update()
    root.destroy()

    if not file_path:
        raise RuntimeError("Nenhum ficheiro foi selecionado.")

    return file_path


# =========================================================
# EXCEL
# =========================================================
def find_header_row_and_columns(ws):
    """
    Procura nas primeiras linhas a linha de cabeçalho e devolve:
    - header_row
    - dict com nome normalizado -> índice da coluna
    """
    max_row = min(ws.max_row, HEADER_SEARCH_MAX_ROWS)
    max_col = ws.max_column

    for row_idx in range(1, max_row + 1):
        headers_map = {}
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            header = normalize_header(cell_value)
            if header:
                headers_map[header] = col_idx

        if all(normalize_header(req) in headers_map for req in HEADER_REQUIRED):
            return row_idx, headers_map

    raise RuntimeError(
        f"Não foi possível localizar a linha de cabeçalho com a(s) coluna(s): {HEADER_REQUIRED}"
    )


def ensure_column(ws, header_row: int, headers_map: dict, column_name: str) -> int:
    key = normalize_header(column_name)
    if key in headers_map:
        return headers_map[key]

    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = column_name
    headers_map[key] = new_col
    return new_col


def read_users(ws, header_row: int, headers_map: dict):
    user_col = headers_map[normalize_header("Usuário")]

    status_col = ensure_column(ws, header_row, headers_map, "STATUS")
    msg_col = ensure_column(ws, header_row, headers_map, "MSG")
    ts_col = ensure_column(ws, header_row, headers_map, "TIMESTEMP")

    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        user_value = ws.cell(row=row_idx, column=user_col).value
        user_value = normalize_text(user_value)

        if not user_value:
            continue

        rows.append(
            {
                "row": row_idx,
                "user": user_value,
                "status_col": status_col,
                "msg_col": msg_col,
                "ts_col": ts_col,
            }
        )

    return rows


def write_result(ws, row_idx: int, status_col: int, msg_col: int, ts_col: int, status: str, msg: str):
    ws.cell(row=row_idx, column=status_col).value = status
    ws.cell(row=row_idx, column=msg_col).value = msg
    ws.cell(row=row_idx, column=ts_col).value = now_str()


# =========================================================
# SAP
# =========================================================
def get_sap_session():
    pythoncom.CoInitialize()

    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        application = sap_gui.GetScriptingEngine
        connection = application.Children(SAP_CONNECTION_INDEX)
        session = connection.Children(SAP_SESSION_INDEX)
        return application, connection, session
    except Exception as e:
        raise RuntimeError(
            "Não foi possível obter a sessão SAP GUI Scripting. "
            "Confirme que o SAP está aberto, logado e com scripting ativo."
        ) from e


def get_sbar_text(session) -> str:
    try:
        return normalize_text(session.findById("wnd[0]/sbar").Text)
    except Exception:
        return ""


def get_sbar_type(session) -> str:
    try:
        return normalize_text(session.findById("wnd[0]/sbar").MessageType)
    except Exception:
        return ""


def set_okcode(session, value: str) -> None:
    session.findById("wnd[0]/tbar[0]/okcd").text = value
    session.findById("wnd[0]").sendVKey(0)


def enter_program_se37(session, program_name: str) -> None:
    set_okcode(session, SAP_TCODE)
    time.sleep(SLEEP_SHORT)

    session.findById("wnd[0]/usr/ctxtRS38L-NAME").text = program_name
    session.findById("wnd[0]").sendVKey(8)
    time.sleep(SLEEP_MEDIUM)


def run_favorites_download(session, user_id: str, download_dir: str):
    """
    Equivalente ao script VB fornecido, adaptado para Python.
    Nome do ficheiro = utilizador.
    """
    user_id = normalize_text(user_id)
    file_name = safe_filename(user_id)

    if not user_id:
        raise RuntimeError("Utilizador vazio.")

    ensure_folder(download_dir)

    # Entra no programa
    enter_program_se37(session, SAP_PROGRAM)

    # Campo do utilizador conforme o mapeamento gravado
    # VBScript:
    # session.findById("wnd[0]/usr/txt[34,7]").text = "168801"
    session.findById("wnd[0]/usr/txt[34,7]").text = user_id
    session.findById("wnd[0]/usr/txt[34,7]").caretPosition = len(user_id)

    # Executar
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    time.sleep(SLEEP_MEDIUM)

    # Popup de gravação
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = download_dir
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = file_name
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len(file_name)
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(SLEEP_MEDIUM)

    # Ler statusbar após download
    sbar_type = get_sbar_type(session)
    sbar_text = get_sbar_text(session)

    # Voltar
    try:
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        time.sleep(SLEEP_SHORT)
    except Exception:
        pass

    try:
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        time.sleep(SLEEP_SHORT)
    except Exception:
        pass

    return sbar_type, sbar_text


def classify_status(sbar_type: str, sbar_text: str) -> str:
    """
    STATUS sempre baseado no wnd[0]/sbar.
    """
    sbar_type = normalize_text(sbar_type).upper()
    sbar_text = normalize_text(sbar_text)

    if sbar_type == "S":
        return "OK"
    if sbar_type == "W":
        return "WARNING"
    if sbar_type == "E":
        return "ERRO"
    if sbar_type == "A":
        return "ABORT"
    if sbar_type == "I":
        return "INFO"

    if sbar_text:
        return "INFO"

    return "SEM STATUS"


# =========================================================
# MAIN
# =========================================================
def main():
    log("###################################################################################")
    log("🚀 INÍCIO | DOWNLOAD FAVORITOS SAP")
    log("###################################################################################")

    excel_path = select_excel_file()
    log(f"[FILE] Excel selecionado: {excel_path}")

    ensure_folder(DOWNLOAD_DIR)
    log(f"[DIR] Pasta de download: {DOWNLOAD_DIR}")

    wb = load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f'A sheet "{SHEET_NAME}" não existe no ficheiro selecionado.')

    ws = wb[SHEET_NAME]

    header_row, headers_map = find_header_row_and_columns(ws)
    log(f"[EXCEL] Linha de cabeçalho localizada: {header_row}")

    rows = read_users(ws, header_row, headers_map)
    if not rows:
        raise RuntimeError("Nenhum utilizador encontrado na sheet 'Users Ativos'.")

    log(f"[EXCEL] Total de utilizadores encontrados: {len(rows)}")

    _, _, session = get_sap_session()
    log("[SAP] Sessão SAP obtida com sucesso")

    ok_count = 0
    err_count = 0

    for idx, item in enumerate(rows, start=1):
        row_idx = item["row"]
        user_id = item["user"]

        log("------------------------------------------------------------")
        log(f"[{idx}/{len(rows)}] PROCESSANDO | LINHA={row_idx} | USUÁRIO={user_id}")

        try:
            sbar_type, sbar_text = run_favorites_download(
                session=session,
                user_id=user_id,
                download_dir=DOWNLOAD_DIR,
            )

            status = classify_status(sbar_type, sbar_text)
            msg = sbar_text or "Sem mensagem no status bar"

            write_result(
                ws=ws,
                row_idx=row_idx,
                status_col=item["status_col"],
                msg_col=item["msg_col"],
                ts_col=item["ts_col"],
                status=status,
                msg=msg,
            )

            wb.save(excel_path)

            if status == "OK":
                ok_count += 1
            else:
                err_count += 1

            log(f"[RESULT] STATUS={status} | MSG={msg}")

        except Exception as e:
            try:
                sbar_text = get_sbar_text(session)
            except Exception:
                sbar_text = ""

            error_msg = normalize_text(str(e))
            final_msg = sbar_text if sbar_text else error_msg

            write_result(
                ws=ws,
                row_idx=row_idx,
                status_col=item["status_col"],
                msg_col=item["msg_col"],
                ts_col=item["ts_col"],
                status="ERRO",
                msg=final_msg,
            )

            wb.save(excel_path)
            err_count += 1

            log(f"[ERRO] LINHA={row_idx} | USUÁRIO={user_id} | MSG={final_msg}")

            # Tentativa de recuperação de ecrã/popups
            for _ in range(3):
                try:
                    if session.Children.Count > 1:
                        session.findById("wnd[1]/tbar[0]/btn[12]").press()
                        time.sleep(SLEEP_SHORT)
                    else:
                        break
                except Exception:
                    break

            for _ in range(3):
                try:
                    session.findById("wnd[0]/tbar[0]/btn[3]").press()
                    time.sleep(SLEEP_SHORT)
                except Exception:
                    break

    wb.save(excel_path)

    log("###################################################################################")
    log(f"[FIM] SUCESSO={ok_count} | ERROS={err_count} | FICHEIRO={excel_path}")
    log("###################################################################################")

    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showinfo(
        "Concluído",
        f"Processamento concluído.\n\nSucesso: {ok_count}\nErros: {err_count}\n\nExcel: {excel_path}",
        parent=root,
    )
    root.destroy()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        err = "".join(traceback.format_exception_only(type(exc), exc)).strip()
        print("\n[ERRO FATAL]")
        print(err)
        sys.exit(1)