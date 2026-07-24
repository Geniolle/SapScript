# -*- coding: utf-8 -*-

###################################################################################
# D. PFCG_COMPOSTA_RFC.py
# PFCG - Criar/Modificar Roles Compostas + Atribuir Roles Componentes via RFC
#
# Regras:
#  - 100% via RFC (sem GUI SAP)
#  - Lógica Create-or-Update automática
#  - Verificação de roles filhas via PRGN_CHECK_ROLE_EXISTS
#  - Leitura de componentes existentes via RFC_READ_TABLE (AGR_AGRS)
#  - Apenas roles em falta são adicionadas (sem duplicação)
#  - Gravação segura no Excel via COM Automation (fallback openpyxl)
####################################################################################

import sys
if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import functools
print = functools.partial(print, flush=True)


def executar(
    ambiente_cockpit,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    nome_pasta=None
):
    import os
    import time
    import re
    import unicodedata
    import tkinter as tk
    from tkinter import filedialog
    from datetime import datetime
    from openpyxl import load_workbook
    from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn
    from pyrfc import Connection
    from dotenv import load_dotenv

    tempo_inicio_total = time.time()

    # --- Correção da estrutura de pastas ---
    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual)
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)

    # Load .env
    env_path = os.path.join(os.getcwd(), ".env")
    load_dotenv(env_path)

    # Sheet base do processo; o sufixo _RFC do ficheiro não altera o nome da aba.
    NOME_SHEET = "PFCG_COMPOSTA"
    SEARCH_HEADER_IN_FIRST_ROWS = 20
    OPCOES_ROLES = {"ROLES", "TCODE", "ROLES_FILHAS", "AGR_NAME_FILHA"}

    MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    CLIENTES_POR_AMBIENTE = {"DEV": "100", "QAD": "100", "PRD": "100"}
    cliente_esperado = CLIENTES_POR_AMBIENTE.get(ambiente_cockpit, "100")

    def now_ts():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def norm_col(s):
        if s is None:
            s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def norm_txt(s):
        if s is None:
            s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def formatar_tempo(segundos):
        h, resto = divmod(segundos, 3600)
        m, s = divmod(resto, 60)
        if h > 0:
            return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
        return f"{int(m):02d}m {int(s):02d}s"

    def selecionar_ficheiro():
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
            filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*"))
        )
        root.destroy()
        return path

    def split_roles(raw):
        if not raw:
            return []
        s = str(raw).replace("\r", "\n").replace("\t", " ").strip().upper()
        parts = re.split(r"[;, \n]+", s)
        out = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if p.startswith("/N") or p.startswith("/O"):
                p = p[2:].strip()
            if p:
                out.append(p)
        return list(dict.fromkeys(out))

    def gravar_resultados_excel(caminho_ficheiro, sheet_name, header_map, records, resultados):
        col_st = header_map.get("STATUS")
        col_ms = header_map.get("MSG")
        col_tm = header_map.get("TIMESTEMP")

        # 1. Tentar via Excel COM primeiro (mais seguro no Windows)
        pythoncom = None
        excel_app = None
        wb_excel = None
        try:
            import win32com.client
            import pythoncom as _pythoncom

            pythoncom = _pythoncom
            pythoncom.CoInitialize()

            try:
                excel_app = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                excel_app = win32com.client.Dispatch("Excel.Application")

            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            abs_path = os.path.abspath(caminho_ficheiro)
            wb_excel = excel_app.Workbooks.Open(abs_path)
            ws_excel = wb_excel.Worksheets(sheet_name)

            for rec in records:
                chave_busca = str(rec["AGR_NAME"]).strip()
                res = resultados.get(chave_busca)
                if res:
                    if col_st:
                        ws_excel.Cells(rec["_row"], col_st).Value = res["STATUS"]
                    if col_ms:
                        ws_excel.Cells(rec["_row"], col_ms).Value = res["MSG"]
                    if col_tm:
                        ws_excel.Cells(rec["_row"], col_tm).Value = res["TIMESTEMP"]

            wb_excel.Save()
            return True
        except Exception as e_com:
            print(f"  [DEBUG] Falha ao gravar via Excel COM ({e_com}). Usando openpyxl como fallback...")
        finally:
            try:
                if wb_excel is not None:
                    wb_excel.Close(SaveChanges=True)
            except:
                pass
            try:
                if excel_app is not None:
                    excel_app.Quit()
            except:
                pass
            try:
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
            except:
                pass

        # 2. Fallback usando openpyxl
        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(caminho_ficheiro)
            ws = wb[sheet_name]
            for rec in records:
                chave_busca = str(rec["AGR_NAME"]).strip()
                res = resultados.get(chave_busca)
                if res:
                    if col_st:
                        ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
                    if col_ms:
                        ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
                    if col_tm:
                        ws.cell(row=rec["_row"], column=col_tm).value = res["TIMESTEMP"]
            wb.save(caminho_ficheiro)
            return True
        except Exception as e_openpyxl:
            print(f"  Falha critica ao gravar Excel com openpyxl: {e_openpyxl}")
            return False
        finally:
            try:
                if wb is not None:
                    wb.close()
            except:
                pass

    ###################################################################################
    # BLOCO 1: LER EXCEL
    ###################################################################################
    if not caminho_ficheiro:
        if os.path.exists("S4H_Perfis de autorização.xlsx"):
            caminho_ficheiro = "S4H_Perfis de autorização.xlsx"
            print("📂 Utilizando ficheiro Excel padrão encontrado na raiz: S4H_Perfis de autorização.xlsx")
        else:
            if modo_nao_interativo:
                raise ValueError("Faltou o parâmetro --xlsx em modo não-interativo.")
            print("📂 Selecione o ficheiro Excel…")
            caminho_ficheiro = selecionar_ficheiro()
            if not caminho_ficheiro:
                print("❌ Operação cancelada.")
                return

    if not os.path.exists(caminho_ficheiro):
        print(f"❌ Ficheiro não encontrado: {caminho_ficheiro}")
        return

    try:
        wb = load_workbook(caminho_ficheiro)
    except Exception as e:
        print(f"❌ Não consegui abrir o Excel: {e}")
        return

    if NOME_SHEET in wb.sheetnames:
        ws = wb[NOME_SHEET]
    else:
        if len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
        else:
            print(f"❌ Sheet '{NOME_SHEET}' não encontrada.")
            wb.close()
            return

    header_row = None
    header_map = {}
    COLUNAS_SAIDA = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [norm_col(c.value) for c in ws[r]]
        row_set = set(row_vals)

        if "TEXT" not in row_set or "AGR_NAME" not in row_set:
            continue

        is_format_1 = "AGR_NAME_COMPOSTA" in row_set
        is_format_2 = len(OPCOES_ROLES.intersection(row_set)) >= 1

        if is_format_1 or is_format_2:
            header_row = r
            for idx, name in enumerate(row_vals, start=1):
                if name:
                    header_map[name] = idx

            modificou = False
            last_col = 0
            for idx, val in enumerate(row_vals, start=1):
                if val:
                    last_col = idx

            for col_key, col_name in COLUNAS_SAIDA.items():
                if col_key not in header_map:
                    last_col += 1
                    ws.cell(row=r, column=last_col, value=col_name)
                    header_map[col_key] = last_col
                    print(f"➕ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                    modificou = True

            if modificou:
                try:
                    wb.save(caminho_ficheiro)
                    print("💾 Ficheiro Excel guardado com as novas colunas.")
                except Exception as e:
                    wb.close()
                    print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas: {e}")
                    return
            break

    if not header_row:
        wb.close()
        print("\n❌ Não encontrei a linha de cabeçalho completa com as colunas obrigatórias.")
        return

    # Detetar colunas
    col_text = header_map.get("TEXT")
    col_status = header_map.get("STATUS")
    col_msg = header_map.get("MSG")
    col_ts = header_map.get("TIMESTEMP")

    if "AGR_NAME_COMPOSTA" in header_map:
        col_agr = header_map["AGR_NAME_COMPOSTA"]
        col_roles = header_map.get("AGR_NAME")
    else:
        col_agr = header_map.get("AGR_NAME")
        col_roles = None
        for opt in ["ROLES", "TCODE", "ROLES_FILHAS", "AGR_NAME_FILHA"]:
            if opt in header_map:
                col_roles = header_map[opt]
                break

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        agr_val = ws.cell(row=r, column=col_agr).value if col_agr else None
        agr = "" if agr_val is None else str(agr_val).strip()
        if not agr:
            continue

        if " " in agr:
            orig = agr
            agr = agr.replace(" ", "")
            ws.cell(row=r, column=col_agr, value=agr)
            print(f"⚠️ Espaços detetados e removidos da Role: '{orig}' -> '{agr}' (atualizado no Excel)")
            try:
                wb.save(caminho_ficheiro)
            except Exception:
                pass

        text_val = ws.cell(row=r, column=col_text).value if col_text else None
        roles_val = ws.cell(row=r, column=col_roles).value if col_roles else None
        status_val = ws.cell(row=r, column=col_status).value if col_status else None
        msg_val = ws.cell(row=r, column=col_msg).value if col_msg else None
        ts_val = ws.cell(row=r, column=col_ts).value if col_ts else None

        records.append({
            "_row": r,
            "AGR_NAME": agr,
            "TEXT": "" if text_val is None else str(text_val).strip(),
            "ROLES": "" if roles_val is None else str(roles_val).strip(),
            "STATUS": "" if status_val is None else str(status_val).strip(),
            "MSG": "" if msg_val is None else str(msg_val).strip(),
            "TIMESTEMP": "" if ts_val is None else str(ts_val).strip(),
        })

    if not records:
        wb.close()
        print("⚠️ Não encontrei linhas para processar.")
        return

    # Agrupar por role composta (ignorar linhas CONCLUIDO)
    roles_map = {}
    for rec in records:
        status_norm = norm_txt(rec["STATUS"])
        if status_norm == "CONCLUIDO":
            continue

        agr = rec["AGR_NAME"].strip()
        if not agr:
            continue

        if agr not in roles_map:
            roles_map[agr] = {
                "AGR_NAME": agr,
                "TEXT": rec["TEXT"].strip(),
                "ROLES_LIST": []
            }

        if not roles_map[agr]["TEXT"] and rec["TEXT"].strip():
            roles_map[agr]["TEXT"] = rec["TEXT"].strip()

        roles_map[agr]["ROLES_LIST"].extend(split_roles(rec["ROLES"]))

    if not roles_map:
        wb.close()
        print("⚠️ Nada para processar (tudo CONCLUIDO).")
        return

    roles_agrupadas = []
    for item in roles_map.values():
        item["ROLES_LIST"] = list(dict.fromkeys(item["ROLES_LIST"]))
        roles_agrupadas.append(item)

    roles_agrupadas.sort(key=lambda x: x["AGR_NAME"])

    # --- Obter session GUI (opcional, para criação de request) ---
    session = None
    try:
        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        for i in range(application.Children.Count):
            conn_gui = application.Children(i)
            for j in range(conn_gui.Children.Count):
                sess = conn_gui.Children(j)
                if str(sess.Info.SystemName).upper() == SISTEMA_ESPERADO.upper():
                    session = sess
                    break
            if session:
                break
    except Exception:
        pass

    def _criar_nova_request_no_sap_local(sess):
        if not sess:
            print("❌ Sessão SAP GUI não disponível para criar request.")
            return None
        try:
            sess.findById("wnd[0]/tbar[0]/okcd").text = "/nSE10"
            sess.findById("wnd[0]").sendVKey(0)
            print("\nTipo da ordem:")
            print("1 - Ordem customizing")
            print("2 - Ordem workbench")
            while True:
                tipo = input("Digite a opção (1/2): ").strip()
                if tipo in ("1", "2"):
                    break
                print("❌ Opção inválida.")

            desc = input("Descrição da request (máx 60): ").strip()
            desc = desc[:60] if desc else "REQUEST CRIADA VIA SCRIPT"

            sess.findById("wnd[0]/tbar[1]/btn[6]").press()
            time.sleep(1)

            if tipo == "2":
                try:
                    sess.findById("wnd[1]/usr/radKO042-REQ_CONS_K").select()
                except:
                    pass

            sess.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(1)

            try:
                sess.findById("wnd[1]/usr/txtKO013-AS4TEXT").text = desc
            except:
                pass

            sess.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(1)

            trkorr = None
            for sap_id in ["wnd[0]/usr/lbl[20,9]", "wnd[0]/usr/lbl[1,1]"]:
                try:
                    txt = sess.findById(sap_id).Text
                    match = re.search(r"\b[A-Z0-9]{3,4}K\d{6,}\b", txt)
                    if match:
                        trkorr = match.group(0)
                except:
                    pass
                if trkorr:
                    break

            sess.findById("wnd[0]/tbar[0]/okcd").text = "/n"
            sess.findById("wnd[0]").sendVKey(0)

            tipo_txt = "Customizing" if tipo == "1" else "Workbench"
            print("\n✔️ Request criada.")
            print(f"Tipo: {tipo_txt} | Descrição: {desc}")
            if not trkorr:
                trkorr = input("Não consegui extrair a request automaticamente. Cole aqui: ").strip().upper()
            print(f"Request: {trkorr}")
            return trkorr
        except Exception as err:
            print(f"❌ Erro ao criar request no SAP GUI: {err}")
            return None

    # --- Menu de Transporte ---
    if not request_transporte and not modo_nao_interativo:
        print("\n============================================================")
        print("🚚 Opções de configuração de Transporte.\n")
        print("   1 - Escreva o número da Request")
        print("   2 - Criar nova ordem de transporte")
        print("   3 - Pesquisar suas request criadas.")
        print("   4 - Prima [Enter] vazio para NÃO transportar")
        print("============================================================")

        while True:
            req_input = input("\n👉 Opção: ").strip()
            if req_input in ("1", "2", "3", "4", ""):
                if req_input == "":
                    req_input = "4"
                break
            print("❌ Opção inválida. Use 1, 2, 3, 4 ou apenas pressione Enter.")

        if req_input == "1":
            request_transporte = input("🔢 Numero da Request (ex: S4QK900396): ").strip().upper()

        elif req_input == "2":
            request_transporte = _criar_nova_request_no_sap_local(session)

        elif req_input == "3":
            try:
                import pesquisar_request
                print("\n🔍 A abrir nova sessão em segundo plano para pesquisar (SE16H)...")
                resultados_pesquisa = pesquisar_request.listar_requests(
                    system_name=SISTEMA_ESPERADO,
                    include_requests=True,
                    use_new_mode=True,
                    minimize=False,
                    close_after=True,
                    session=session
                )
                if resultados_pesquisa:
                    escolha = input("\n👉 Digite o número (N) da Request que deseja utilizar (ou Enter para cancelar): ").strip()
                    if escolha.isdigit() and 1 <= int(escolha) <= len(resultados_pesquisa):
                        request_transporte = resultados_pesquisa[int(escolha) - 1][0]
                        print(f"✔️ Selecionou a Request: {request_transporte}")
                    else:
                        print("❌ Seleção cancelada. Não haverá transporte.")
                else:
                    print("⚠️ Não foram encontradas Requests abertas para o seu utilizador.")
            except Exception as e:
                print(f"❌ Erro ao pesquisar request: {e}")

        elif req_input == "4":
            print("⏭️  Nenhuma request selecionada (Transporte ignorado).")
            request_transporte = None
        print("============================================================")

    ambiente_up = str(ambiente_cockpit).upper().strip()
    # --- Credenciais RFC ---
    ashost = (
        os.getenv(f"SAP_ASHOST_{ambiente_up}")
        or os.getenv("SAP_ASHOST")
        or ""
    ).strip()
    sysnr = (
        os.getenv(f"SAP_SYSNR_{ambiente_up}")
        or os.getenv("SAP_SYSNR")
        or "00"
    ).strip() or "00"
    user_rfc = (
        os.getenv(f"SAP_USER_{ambiente_up}")
        or os.getenv("SAP_USER")
        or ""
    ).strip()
    client = (
        os.getenv(f"SAP_CLIENT_{ambiente_up}")
        or os.getenv("SAP_CLIENT")
        or cliente_esperado
    ).strip() or cliente_esperado
    lang = (
        os.getenv(f"SAP_LANGUAGE_{ambiente_up}")
        or os.getenv("SAP_CUA_LANGUAGE" if ambiente_up == "CUA" else "SAP_LANGUAGE")
        or "PT"
    ).strip() or "PT"
    env_pw_key = f"SAP_PASSWORD_{ambiente_up}"
    env_pw_key_legacy = f"SAP_PASSWORD_{SISTEMA_ESPERADO}CLNT{cliente_esperado}"
    passwd = (
        os.getenv(env_pw_key)
        or os.getenv(env_pw_key_legacy)
        or os.getenv("SAP_PASSWD")
        or os.getenv("SAP_PASSWORD")
        or ""
    ).strip()

    if not ashost or not user_rfc or not passwd:
        wb.close()
        print(
            "❌ Faltam credenciais SAP no ficheiro .env "
            f"(esperado: SAP_ASHOST_{ambiente_up} / SAP_USER_{ambiente_up} / SAP_PASSWORD_{ambiente_up} "
            "ou fallback antigo)"
        )
        return

    print(f"\n📋 Roles compostas a processar (agrupadas): {len(roles_agrupadas)}")
    for r in roles_agrupadas:
        print(f"  - {r['AGR_NAME']}: {r['TEXT']} (Componentes: {len(r['ROLES_LIST'])})")

    if pedir_confirmacao and not modo_nao_interativo:
        conf = input("\nDeseja lançar esses dados no SAP? [S/N]: ").strip().upper()
        if conf != "S":
            wb.close()
            print("❌ Execução cancelada pelo utilizador.")
            return

    ###################################################################################
    # BLOCO 2: PROCESSAR ROLES COMPOSTAS VIA RFC
    ###################################################################################
    resultados = {}
    role_metrics = {}

    def _rfc_conn():
        return Connection(
            ashost=ashost,
            sysnr=sysnr,
            client=client,
            user=user_rfc,
            passwd=passwd,
            lang=lang
        )

    def _role_exists_rfc(conn, role_name):
        """Verificar se uma role existe via RFC PRGN_CHECK_ROLE_EXISTS."""
        try:
            res = conn.call("PRGN_CHECK_ROLE_EXISTS", ROLE_NAME=role_name)
            ret = res.get("RETURN", [])
            for row in ret:
                if row.get("TYPE") == "E":
                    return False
            return True
        except Exception:
            return False

    def _ler_componentes_agr_agrs(conn, agr_name):
        """Ler os componentes já atribuídos à role composta via RFC_READ_TABLE."""
        try:
            res = conn.call(
                "RFC_READ_TABLE",
                QUERY_TABLE="AGR_AGRS",
                DELIMITER="|",
                FIELDS=[{"FIELDNAME": "AGR_NAME"}, {"FIELDNAME": "CHILD_AGR"}],
                OPTIONS=[{"TEXT": f"AGR_NAME = '{agr_name}'"}],
                ROWCOUNT=500
            )
            rows = res.get("DATA", [])
            componentes = set()
            for row in rows:
                wa = row.get("WA", "")
                parts = wa.split("|")
                if len(parts) >= 2:
                    child = parts[1].strip().upper()
                    if child:
                        componentes.add(child)
            return componentes
        except Exception as e:
            print(f"  ⚠️ Erro ao ler AGR_AGRS via RFC: {e}")
            return set()

    with Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total})"),
        TimeElapsedColumn(),
    ) as progress:
        task_roles = progress.add_task("Processando roles compostas via RFC...", total=len(roles_agrupadas))

        for idx_role, r_item in enumerate(roles_agrupadas, start=1):
            nome = r_item["AGR_NAME"]
            desc = r_item["TEXT"]
            componentes_excel = [c.strip().upper() for c in r_item["ROLES_LIST"] if c and str(c).strip()]
            tempo_inicio_role = time.time()

            print("======================================================================")
            print(f"▶ [{idx_role}/{len(roles_agrupadas)}] ROLE COMPOSTA: {nome} | Componentes: {len(componentes_excel)}")
            print("======================================================================")

            try:
                print(f"  ├─ Ligando via RFC ao sistema {SISTEMA_ESPERADO} (Host: {ashost}, Client: {cliente_esperado}, User: {user_rfc})...")
                conn = _rfc_conn()

                modo_operacao = None

                # Etapa 1: Criar role composta (ou detectar que já existe)
                print("  ├─ Etapa 1: Criar/verificar role composta...")
                try:
                    res_create = conn.call(
                        "PRGN_RFC_CREATE_AGR_MULTIPLE",
                        ACTIVITY_GROUP=nome,
                        ACTIVITY_GROUP_TEXT=desc,
                        COLLECTIVE_AGR="X",
                        NO_DIALOG="X",
                        REQUEST=request_transporte or ""
                    )
                    ret_create = res_create.get("RETURN", [])
                    err_msgs = [row.get("MESSAGE", "") for row in ret_create if row.get("TYPE") == "E"]
                    if err_msgs:
                        raise RuntimeError("Erro ao criar role composta: " + " | ".join(err_msgs))
                    modo_operacao = "CRIADA"
                    print("  ├─ Role composta criada com sucesso.")

                except Exception as e_create:
                    err_str = str(e_create)
                    if "ALREADY_EXISTS" in err_str or "ACTIVITY_GROUP_ALREADY_EXISTS" in err_str:
                        modo_operacao = "EXISTENTE"
                        print("  ├─ Role composta já existe — continuando com update...")
                    else:
                        raise

                # Etapa 1b: Verificar se a role existente é mesmo composta (tem COLL flag)
                if modo_operacao == "EXISTENTE":
                    try:
                        # Tentar um call de teste para PRGN_RFC_ADD_AGRS_TO_COLL_AGR com lista vazia
                        # Se retornar NO_COLLECTIVE_ACTIVITY_GROUP, é uma role simples
                        conn.call(
                            "PRGN_RFC_ADD_AGRS_TO_COLL_AGR",
                            ACTIVITY_GROUP=nome,
                            ACTIVITY_GROUPS=[],  # lista vazia - apenas verificar
                            NO_DIALOG="X",
                            REQUEST=""
                        )
                    except Exception as e_check:
                        if "NO_COLLECTIVE_ACTIVITY_GROUP" in str(e_check):
                            raise RuntimeError(
                                f"AVISO: A role '{nome}' existe no SAP mas é uma ROLE SIMPLES, não composta. "
                                f"Por favor, elimine-a manualmente em PFCG e execute novamente o script."
                            )

                # Etapa 2: Verificar se todos os componentes existem no SAP
                print(f"  ├─ Etapa 2: Verificar existência de {len(componentes_excel)} roles filhas via RFC...")
                roles_inexistentes = []
                for comp in componentes_excel:
                    existe = _role_exists_rfc(conn, comp)
                    if not existe:
                        roles_inexistentes.append(comp)

                if roles_inexistentes:
                    raise RuntimeError(f"Role filha inexistente: {roles_inexistentes[0]}. Composta NÃO atualizada.")

                print(f"  ├─ Todas as {len(componentes_excel)} roles filhas verificadas com sucesso.")

                # Etapa 3: Ler componentes já atribuídos via AGR_AGRS
                print("  ├─ Etapa 3: Ler componentes actuais via AGR_AGRS...")
                componentes_existentes = _ler_componentes_agr_agrs(conn, nome)
                print(f"  ├─ Componentes já atribuídos: {len(componentes_existentes)}")

                # Calcular o que falta inserir
                componentes_para_inserir = [c for c in componentes_excel if c not in componentes_existentes]
                componentes_ja_la = [c for c in componentes_excel if c in componentes_existentes]

                print(f"  ├─ Já existentes: {len(componentes_ja_la)} | A inserir: {len(componentes_para_inserir)}")

                if componentes_para_inserir:
                    for c in sorted(componentes_para_inserir):
                        print(f"  │  └─ [NOVO] {c}")
                else:
                    print("  │  └─ Nenhum componente novo a inserir.")

                # Etapa 4: Adicionar apenas os componentes em falta
                if componentes_para_inserir:
                    print(f"  ├─ Etapa 4: A adicionar {len(componentes_para_inserir)} roles filhas via RFC...")
                    agrs_param = [{"AGR_NAME": c, "TEXT": ""} for c in componentes_para_inserir]

                    res_add = conn.call(
                        "PRGN_RFC_ADD_AGRS_TO_COLL_AGR",
                        ACTIVITY_GROUP=nome,
                        ACTIVITY_GROUPS=agrs_param,
                        NO_DIALOG="X",
                        REQUEST=request_transporte or ""
                    )
                    ret_add = res_add.get("RETURN", [])
                    err_add_msgs = [row.get("MESSAGE", "") for row in ret_add if row.get("TYPE") == "E"]
                    if err_add_msgs:
                        raise RuntimeError("Erro ao adicionar componentes: " + " | ".join(err_add_msgs))

                    print(f"  ├─ {len(componentes_para_inserir)} roles filhas adicionadas com sucesso.")
                else:
                    print("  ├─ Etapa 4: Ignorada — sem roles filhas novas a adicionar.")

                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)
                msg_transporte = f" | Req {request_transporte}" if request_transporte else ""

                modo_label = modo_operacao
                if componentes_para_inserir:
                    modo_label += "+ATUALIZADA"
                elif modo_operacao == "EXISTENTE":
                    modo_label = "SEM ALTERACOES"

                resultados[nome] = {
                    "STATUS": "CONCLUIDO",
                    "MSG": f"{modo_label} (RFC) | {len(componentes_para_inserir)} novos / {len(componentes_ja_la)} existentes{msg_transporte}.",
                    "TIMESTEMP": now_ts()
                }

                icone = "🟢" if componentes_para_inserir or modo_operacao == "CRIADA" else "🔵"
                print(f"\n{icone} SUCESSO [{modo_label}]: Role composta processada via RFC! ⏱️ (Tempo: {str_tempo})")
                print("----------------------------------------------------------------------")

                role_metrics[nome] = {"total": int(tempo_decorrido_role)}

                conn.close()

            except Exception as e:
                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)
                err = str(e)
                resultados[nome] = {
                    "STATUS": "ERRO",
                    "MSG": err,
                    "TIMESTEMP": now_ts()
                }
                print(f"\n🔴 ERRO: {err} ⏱️ (Tempo: {str_tempo})")
                print("----------------------------------------------------------------------")
                role_metrics[nome] = {"total": int(tempo_decorrido_role)}

            finally:
                try:
                    gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados)
                except Exception as cp_exc:
                    print(f"  ⚠️ Erro ao salvar checkpoint do Excel: {cp_exc}")

                progress.advance(task_roles)

    ###################################################################################
    # BLOCO 3: GRAVAR EXCEL E TEMPO TOTAL
    ###################################################################################
    if gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados):
        print("\n💾 Resultados gravados com sucesso no Excel!")
    else:
        print("\n❌ Erro ao gravar resultados finais no Excel.")

    if role_metrics:
        try:
            fastest_role = min(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            slowest_role = max(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])

            print("\n=======================================================")
            print("📊 RESUMO COMPARATIVO DE PERFORMANCE DAS ROLES COMPOSTAS (RFC)")
            print("=======================================================")
            print(f" - Role mais rápida: {fastest_role} ({role_metrics[fastest_role]['total']}s)")
            print(f" - Role mais lenta: {slowest_role} ({role_metrics[slowest_role]['total']}s)")
            print("=======================================================")
        except Exception:
            pass

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
    print("🔁 Fim.")
    return True


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD"])
    parser.add_argument("--xlsx")
    parser.add_argument("--request", help="Número da Request de Transporte (Opcional)")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    args = parser.parse_args()

    env_cli = args.ambiente or (input("Ambiente (DEV/QAD/PRD): ").strip().upper() or "DEV")

    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        request_transporte=args.request,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm)
    )
