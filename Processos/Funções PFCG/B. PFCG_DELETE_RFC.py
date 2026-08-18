# -*- coding: utf-8 -*-

###################################################################################
# B. PFCG_DELETE_RFC.py
# PFCG - Eliminar Roles via RFC quando possivel, com fallback GUI
#
# Regras:
#  - RFC first com PRGN_RFC_DELETE_ACTIVITY_GROUP
#  - Fallback GUI via /NPFCGMASSDELETE quando necessario
#  - Logger visual estruturado por etapas
#  - Gravacao segura no Excel com checkpoint por role
###################################################################################

import sys

if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import functools

print = functools.partial(print, flush=True)


def executar(
    ambiente_cockpit,
    pfcg_object,
    caminho_ficheiro,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    **kwargs,
):
    import os
    import re
    import time
    import unicodedata
    from datetime import datetime

    import pyperclip
    import win32com.client
    from openpyxl import load_workbook
    from rich.progress import BarColumn, Progress, TextColumn, TimeElapsedColumn

    try:
        from dotenv import load_dotenv
    except Exception:
        load_dotenv = None

    try:
        from pyrfc import Connection
    except Exception:
        Connection = None

    tempo_inicio_total = time.time()

    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual)
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)

    if load_dotenv:
        load_dotenv(os.path.join(os.getcwd(), ".env"))

    # Sheet base do processo; não usar o nome do ficheiro nem o sufixo _RFC.
    NOME_SHEET = pfcg_object if pfcg_object else "PFCG_DELETE"
    SEARCH_HEADER_IN_FIRST_ROWS = 20
    TIMEOUT_SAP_BUSY = 180.0

    COL_ID = "ID"
    COL_AGR_NAME = "AGR_NAME"
    COL_TEXT = "TEXT"
    COL_STATUS = "STATUS"
    COL_MSG = "MSG"
    COL_TIMESTAMP = "TIMESTEMP"

    MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente invalido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    CLIENTES_POR_AMBIENTE = {"DEV": "100", "QAD": "100", "PRD": "100"}
    cliente_esperado = CLIENTES_POR_AMBIENTE.get(str(ambiente_cockpit).upper().strip(), "100")

    def agora_ts():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def log(msg):
        print(f"{agora_ts()} | {msg}", flush=True)

    def norm_col(s):
        if s is None:
            return ""
        return (
            unicodedata.normalize("NFKD", str(s))
            .encode("ASCII", "ignore")
            .decode("utf-8")
            .strip()
            .upper()
        )

    def norm_txt(s):
        if s is None:
            return ""
        return (
            unicodedata.normalize("NFKD", str(s))
            .encode("ASCII", "ignore")
            .decode("utf-8")
            .strip()
            .upper()
        )

    def formatar_tempo(segundos):
        h, resto = divmod(segundos, 3600)
        m, s = divmod(resto, 60)
        if h > 0:
            return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
        return f"{int(m):02d}m {int(s):02d}s"

    def existe(session, obj_id):
        try:
            session.findById(obj_id)
            return True
        except Exception:
            return False

    def esperar_sap_livre(session, timeout=120.0, pausa=0.2):
        limite = time.time() + timeout
        while time.time() < limite:
            try:
                busy = bool(getattr(session, "Busy", False))
            except Exception:
                busy = False
            if not busy:
                return True
            time.sleep(pausa)
        return False

    def mensagem_sem_resultado(msg):
        m = (msg or "").lower()
        return (
            ("nenhum" in m or "nenhuma" in m or "nenhumas" in m)
            and ("fun" in m or "regist" in m or "role" in m or "obj" in m)
            and ("encontrad" in m or "exist" in m)
        )

    def selecionar_ficheiro():
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
            filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*")),
        )
        root.destroy()
        return path

    def gravar_resultados_excel(caminho_ficheiro, sheet_name, header_map, records, resultados):
        col_st, col_ms, col_tm = (
            header_map.get(COL_STATUS),
            header_map.get(COL_MSG),
            header_map.get(COL_TIMESTAMP),
        )

        pythoncom = None
        excel_app = None
        wb_excel = None
        try:
            import pythoncom as _pythoncom

            pythoncom = _pythoncom

            pythoncom.CoInitialize()
            try:
                excel_app = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                excel_app = win32com.client.Dispatch("Excel.Application")

            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            abs_path = os.path.abspath(caminho_ficheiro)
            wb_excel = excel_app.Workbooks.Open(abs_path)
            ws_excel = wb_excel.Worksheets(sheet_name)

            for rec in records:
                chave = str(rec[COL_AGR_NAME]).strip()
                res = resultados.get(chave)
                if not res:
                    continue
                if col_st:
                    ws_excel.Cells(rec["_row"], col_st).Value = res["STATUS"]
                if col_ms:
                    ws_excel.Cells(rec["_row"], col_ms).Value = res["MSG"]
                if col_tm:
                    ws_excel.Cells(rec["_row"], col_tm).Value = res["TIMESTEMP"]

            wb_excel.Save()
            return True
        except Exception as e_com:
            print(f"  [DEBUG] Falha ao gravar via Excel COM ({e_com}). A usar openpyxl...")
        finally:
            try:
                if wb_excel is not None:
                    wb_excel.Close(SaveChanges=True)
            except Exception:
                pass
            try:
                if excel_app is not None:
                    excel_app.Quit()
            except Exception:
                pass
            try:
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
            except Exception:
                pass

        wb = None
        try:
            wb = load_workbook(caminho_ficheiro)
            ws = wb[sheet_name]
            for rec in records:
                chave = str(rec[COL_AGR_NAME]).strip()
                res = resultados.get(chave)
                if not res:
                    continue
                if col_st:
                    ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
                if col_ms:
                    ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
                if col_tm:
                    ws.cell(row=rec["_row"], column=col_tm).value = res["TIMESTEMP"]
            wb.save(caminho_ficheiro)
            return True
        except Exception as e_openpyxl:
            print(f"  [ERROR] Falha critica ao gravar Excel com openpyxl: {e_openpyxl}")
            return False
        finally:
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass

    def obter_session_gui():
        try:
            SapGuiAuto = win32com.client.GetObject("SAPGUI")
            app = SapGuiAuto.GetScriptingEngine
            for i in range(app.Children.Count):
                conn = app.Children(i)
                for j in range(conn.Children.Count):
                    sess = conn.Children(j)
                    try:
                        if str(sess.Info.SystemName).upper() == SISTEMA_ESPERADO.upper():
                            return sess
                    except Exception:
                        continue
        except Exception:
            pass
        return None

    def obter_credenciais_rfc():
        ashost = os.getenv("SAP_ASHOST")
        user = os.getenv("SAP_USER")
        sysnr = os.getenv("SAP_SYSNR", "00")
        env_pw_key = f"SAP_PASSWORD_{SISTEMA_ESPERADO}CLNT{cliente_esperado}"
        passwd = (
            os.getenv(env_pw_key)
            or os.getenv("SAP_PASSWD")
            or os.getenv("SAP_PASSWORD")
        )
        return ashost, user, sysnr, passwd

    def rfc_config():
        fm_name = (
            os.getenv("SAP_PFCG_DELETE_RFC_FM", "PRGN_RFC_DELETE_ACTIVITY_GROUP")
            .strip()
            or "PRGN_RFC_DELETE_ACTIVITY_GROUP"
        )
        allow_rfc_with_request = norm_txt(
            os.getenv("SAP_PFCG_DELETE_ALLOW_RFC_WITH_REQUEST", "")
        ) in {"1", "X", "YES", "TRUE", "SIM"}
        distribute = os.getenv("SAP_PFCG_DELETE_DISTRIBUTE", "").strip().upper()
        no_check = os.getenv("SAP_PFCG_DELETE_NO_CHECK_CUA", "").strip().upper()
        return fm_name, allow_rfc_with_request, distribute, no_check

    def classificar_erro_rfc(err_str):
        e = norm_txt(err_str)
        if any(
            token in e
            for token in [
                "CALL_FUNCTION_NOT_FOUND",
                "FUNCTION_NOT_FOUND",
                "COMMUNICATION_FAILURE",
                "SYSTEM_FAILURE",
                "LOGON_FAILURE",
                "NO RFC",
                "CONNECTION",
                "DESTINATION",
            ]
        ):
            return "fallback_gui"
        return "erro"

    def ler_roles_excel(wb):
        if NOME_SHEET in wb.sheetnames:
            ws = wb[NOME_SHEET]
        elif len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
        else:
            raise RuntimeError(f"Sheet '{NOME_SHEET}' nao encontrada.")

        header_row = None
        header_map = {}

        COLUNAS_ENTRADA = {COL_ID, COL_AGR_NAME}
        COLUNAS_SAIDA = {COL_STATUS: "STATUS", COL_MSG: "MSG", COL_TIMESTAMP: "TIMESTEMP"}

        for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
            row_vals = [norm_col(c.value) for c in ws[r]]
            if len(set(row_vals).intersection(COLUNAS_ENTRADA)) == len(COLUNAS_ENTRADA):
                header_row = r
                for idx, name in enumerate(row_vals, start=1):
                    if name:
                        header_map[name] = idx

                modificou = False
                last_col = 0
                for idx, val in enumerate(row_vals, start=1):
                    if val:
                        last_col = idx

                for col_key, col_name in COLUNAS_SAIDA.items():
                    if col_key not in header_map:
                        last_col += 1
                        ws.cell(row=r, column=last_col, value=col_name)
                        header_map[col_key] = last_col
                        log(f"+++ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                        modificou = True

                if modificou:
                    wb.save(caminho_ficheiro)
                break

        if not header_row:
            raise RuntimeError(
                f"Nao encontrei as colunas obrigatorias nas primeiras {SEARCH_HEADER_IN_FIRST_ROWS} linhas."
            )

        def get_cell(row_idx, col_name):
            if col_name not in header_map:
                return ""
            v = ws.cell(row=row_idx, column=header_map[col_name]).value
            return "" if v is None else str(v).strip()

        records = []
        for r in range(header_row + 1, ws.max_row + 1):
            agr = get_cell(r, COL_AGR_NAME)
            status = get_cell(r, COL_STATUS).upper()
            if not agr or status in {"CONCLUIDO", "CONCLUÍDO"}:
                continue

            if " " in agr:
                orig = agr
                agr = agr.replace(" ", "")
                ws.cell(row=r, column=header_map[COL_AGR_NAME], value=agr)
                log(f"!!! Espacos removidos da role: '{orig}' -> '{agr}'")
                try:
                    wb.save(caminho_ficheiro)
                except Exception:
                    pass

            records.append(
                {
                    "_row": r,
                    COL_AGR_NAME: agr,
                }
            )

        return ws, header_row, header_map, records

    def confirmar_execucao(roles_ordenadas):
        log(f"Ficheiro: {caminho_ficheiro}")
        log(f"Roles a processar ({len(roles_ordenadas)}):")
        for i, n in enumerate(roles_ordenadas, 1):
            print(f" {i:02d}. {n}", flush=True)

        if not modo_nao_interativo and pedir_confirmacao:
            if input("\nDeseja eliminar essas roles no SAP? [S/N]: ").strip().upper() != "S":
                log("Processo cancelado.")
                return False
        return True

    def executar_delete_gui(session, roles, request_transporte_local=None):
        if not roles:
            return "ERRO", "Nenhuma role para eliminar."

        if not esperar_sap_livre(session, timeout=TIMEOUT_SAP_BUSY):
            raise RuntimeError("SAP bloqueado antes de iniciar a eliminacao GUI.")

        log("A abrir transacao /NPFCGMASSDELETE...")
        session.findById("wnd[0]/tbar[0]/okcd").text = "/NPFCGMASSDELETE"
        session.findById("wnd[0]").sendVKey(0)

        try:
            sb = session.findById("wnd[0]/sbar").Text.strip()
            if sb:
                log(f"SAP: {sb}")
        except Exception:
            pass

        if existe(session, "wnd[0]/usr/radMOD_EXE"):
            session.findById("wnd[0]/usr/radMOD_EXE").select()

        log(f"A carregar lista de {len(roles)} role(s) no SAP GUI...")
        session.findById("wnd[0]/usr/btn%_ROLE_%_APP_%-VALU_PUSH").press()
        time.sleep(0.5)
        pyperclip.copy("\r\n".join(roles))
        session.findById("wnd[1]").sendVKey(24)
        time.sleep(0.3)
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        time.sleep(0.3)

        log("A executar a eliminacao em massa (GUI)...")
        session.findById("wnd[0]/tbar[1]/btn[8]").press()

        timeout = time.time() + 20.0
        ultimo_log = time.time()
        msg_final = "Execucao concluida via GUI."

        while time.time() < timeout:
            time.sleep(0.5)

            if time.time() - ultimo_log >= 3.0:
                log("SAP a processar...")
                ultimo_log = time.time()

            if existe(session, "wnd[1]/usr/ctxtKO008-TRKORR"):
                if request_transporte_local:
                    log(f"Popup de transporte detectado. A injetar request: {request_transporte_local}")
                    session.findById("wnd[1]/usr/ctxtKO008-TRKORR").text = request_transporte_local
                else:
                    log("Popup de transporte detectado. A ignorar (sem transporte).")
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                continue

            if existe(session, "wnd[1]/usr/btnSPOP-OPTION1"):
                log("Popup de confirmacao SAP. A aceitar...")
                session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
                continue

            if existe(session, "wnd[1]/tbar[0]/btn[0]"):
                log("Popup generico SAP. A fechar...")
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                continue

            if existe(session, "wnd[0]/usr/cntlGRID1/shellcont/shell"):
                break

        try:
            grid = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell")
            if int(getattr(grid, "RowCount", 0) or 0) > 0:
                for col in ["MESSAGE", "TEXT", "MSG"]:
                    try:
                        v = str(grid.GetCellValue(0, col)).strip()
                        if v:
                            msg_final = v
                            break
                    except Exception:
                        pass
        except Exception:
            pass

        try:
            msg_barra = session.findById("wnd[0]/sbar").Text.strip()
            if msg_barra:
                msg_final = msg_final or msg_barra
        except Exception:
            pass

        try:
            if existe(session, "wnd[0]/tbar[0]/btn[3]"):
                session.findById("wnd[0]/tbar[0]/btn[3]").press()
            session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
            session.findById("wnd[0]").sendVKey(0)
        except Exception:
            pass

        if mensagem_sem_resultado(msg_final):
            return "ERRO", f"{msg_final} - SAP nao encontrou as roles informadas."
        return "CONCLUIDO", msg_final or "Execucao concluida via GUI."

    def executar_delete_rfc(conn, role_name):
        fm_name, _, distribute, no_check = rfc_config()
        params = {"ACTIVITY_GROUP": role_name}
        if distribute:
            params["DISTRIBUTE"] = distribute
        if no_check:
            params["NO_CHECK_CENTRAL_USER_MANAGER"] = no_check

        resultado = conn.call(fm_name, **params)
        return resultado or {}

    print("\n[Etapa 1] Leitura do Excel")
    if not caminho_ficheiro:
        if os.path.exists("S4H_Perfis de autorização.xlsx"):
            caminho_ficheiro = "S4H_Perfis de autorização.xlsx"
            log("Usando ficheiro Excel padrao encontrado na raiz.")
        else:
            if modo_nao_interativo:
                raise ValueError("Faltou o parametro --xlsx em modo nao interativo.")
            log("Selecione o ficheiro Excel...")
            caminho_ficheiro = selecionar_ficheiro()
            if not caminho_ficheiro:
                log("Operacao cancelada.")
                return "voltar"

    if not os.path.exists(caminho_ficheiro):
        log(f"Ficheiro nao encontrado: {caminho_ficheiro}")
        return "voltar"

    try:
        wb = load_workbook(caminho_ficheiro)
    except Exception as e:
        log(f"Nao consegui abrir o Excel: {e}")
        return "voltar"

    try:
        ws, header_row, header_map, records = ler_roles_excel(wb)
    except Exception as e:
        wb.close()
        log(f"Erro a ler o Excel: {e}")
        return "voltar"

    if not records:
        wb.close()
        log("Nenhuma linha valida pendente encontrada na aba.")
        return "voltar"

    roles_ordenadas = [rec[COL_AGR_NAME] for rec in records]
    if not confirmar_execucao(roles_ordenadas):
        wb.close()
        return "voltar"

    fm_name, allow_rfc_with_request, distribute, no_check = rfc_config()
    ashost, user, sysnr, passwd = obter_credenciais_rfc()

    usar_rfc = bool(Connection) and bool(ashost and user and passwd)
    if request_transporte and not allow_rfc_with_request:
        usar_rfc = False
        log(
            "Request de transporte foi selecionada; a via RFC sera ignorada por defeito "
            "e o processo vai usar GUI para manter o comportamento de transporte."
        )

    conn = None
    if usar_rfc:
        try:
            log(
                f"A ligar via RFC ao sistema {SISTEMA_ESPERADO} "
                f"(Host: {ashost}, Client: {cliente_esperado}, User: {user})..."
            )
            conn = Connection(
                ashost=ashost,
                sysnr=sysnr,
                client=cliente_esperado,
                user=user,
                passwd=passwd,
                lang=os.getenv("SAP_LANGUAGE", "PT"),
            )
            log(f"RFC pronto. FM de delete configurado: {fm_name}")
        except Exception as e:
            conn = None
            usar_rfc = False
            log(f"Falha a ligar via RFC: {e}. A cair para GUI.")

    session_gui = None
    resultados = {}
    role_metrics = {}

    with Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total})"),
        TimeElapsedColumn(),
    ) as progress:
        task_roles = progress.add_task("Processando roles de eliminacao...", total=len(records))

        for idx_role, rec in enumerate(records, start=1):
            nome = rec[COL_AGR_NAME]
            tempo_inicio_role = time.time()

            print("======================================================================")
            print(f"-> [{idx_role}/{len(records)}] INICIANDO DELETE: {nome}")
            print("======================================================================")

            status_final = "ERRO"
            msg_final = "Erro desconhecido."
            modo_final = None

            try:
                if usar_rfc and conn is not None:
                    log(f"A eliminar via RFC ({fm_name}): {nome}")
                    try:
                        res_rfc = executar_delete_rfc(conn, nome)
                        retorno = res_rfc.get("RETURN", []) if isinstance(res_rfc, dict) else []
                        erros = [
                            row.get("MESSAGE", "")
                            for row in retorno
                            if str(row.get("TYPE", "")).upper() == "E"
                        ]
                        if erros:
                            raise RuntimeError(" | ".join(erros))
                        status_final = "CONCLUIDO"
                        msg_final = f"Eliminada via RFC ({fm_name})."
                        modo_final = "RFC"
                    except Exception as e_rfc:
                        err_str = str(e_rfc)
                        modo_erro = classificar_erro_rfc(err_str)
                        log(f"Erro na eliminacao via RFC: {err_str}")

                        if modo_erro == "fallback_gui":
                            log("A tentar fallback GUI para esta role...")
                            if session_gui is None:
                                session_gui = obter_session_gui()
                            if not session_gui:
                                raise RuntimeError(
                                    "Nao foi possivel localizar sessao SAP GUI para fallback."
                                )
                            status_final, msg_final = executar_delete_gui(
                                session_gui,
                                [nome],
                                request_transporte_local=request_transporte,
                            )
                            modo_final = "GUI"
                        else:
                            raise
                else:
                    if session_gui is None:
                        session_gui = obter_session_gui()
                    if not session_gui:
                        raise RuntimeError(
                            "Sessao SAP GUI nao encontrada e RFC nao disponivel."
                        )
                    status_final, msg_final = executar_delete_gui(
                        session_gui,
                        [nome],
                        request_transporte_local=request_transporte,
                    )
                    modo_final = "GUI"

                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)
                if status_final == "CONCLUIDO":
                    msg_final = f"{msg_final} [{modo_final}]"

                resultados[nome] = {
                    "STATUS": status_final,
                    "MSG": msg_final,
                    "TIMESTEMP": agora_ts(),
                }
                role_metrics[nome] = {"total": int(tempo_decorrido_role)}

                if status_final == "CONCLUIDO":
                    log(f"OK: {nome} ({modo_final}) | Tempo: {str_tempo}")
                else:
                    log(f"ERRO: {nome} | {msg_final} | Tempo: {str_tempo}")

            except Exception as e:
                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)
                err = str(e)
                resultados[nome] = {
                    "STATUS": "ERRO",
                    "MSG": err,
                    "TIMESTEMP": agora_ts(),
                }
                role_metrics[nome] = {"total": int(tempo_decorrido_role)}
                log(f"ERRO: {nome} | {err} | Tempo: {str_tempo}")

            finally:
                try:
                    gravar_resultados_excel(
                        caminho_ficheiro,
                        NOME_SHEET,
                        header_map,
                        records,
                        resultados,
                    )
                except Exception as cp_exc:
                    log(f"Falha ao salvar checkpoint do Excel: {cp_exc}")

                progress.advance(task_roles)

    if conn is not None:
        try:
            conn.close()
        except Exception:
            pass

    if gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados):
        print("\nFicheiro Excel atualizado com sucesso.")
    else:
        print("\nErro ao gravar os resultados finais no Excel.")

    if role_metrics:
        try:
            fastest_role = min(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            slowest_role = max(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            print("\n=======================================================")
            print("RESUMO COMPARATIVO DE PERFORMANCE DAS ROLES")
            print("=======================================================")
            print(f"Role mais rapida: {fastest_role} ({role_metrics[fastest_role]['total']}s)")
            print(f"Role mais lenta: {slowest_role} ({role_metrics[slowest_role]['total']}s)")
            print("=======================================================")
        except Exception:
            pass

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\nTempo total da operacao: {formatar_tempo(tempo_decorrido_total)}")
    print("Fim.")
    try:
        wb.close()
    except Exception:
        pass
    return True


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD"])
    parser.add_argument("--xlsx")
    parser.add_argument("--request", help="Numero da Request de Transporte (Opcional)")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    args = parser.parse_args()

    env_cli = args.ambiente or (input("Ambiente (DEV/QAD/PRD): ").strip().upper() or "DEV")

    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        request_transporte=args.request,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm),
    )
