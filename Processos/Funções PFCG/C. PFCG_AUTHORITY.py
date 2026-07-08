# -*- coding: utf-8 -*-

###################################################################################
# C. PFCG_AUTHORITY.py
# PFCG - Inserção Massiva de Valores de Autorização via PFCGMASSVAL & Funções Compostas
###################################################################################

def executar(
    ambiente_cockpit,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True
):
    import sys
    import os
    import time
    import re
    import unicodedata
    import tkinter as tk
    import win32com.client
    from tkinter import filedialog
    from datetime import datetime
    from openpyxl import load_workbook

    tempo_inicio_total = time.time()

    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual) 
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)

    NOME_SHEET = "PFCG_AUTHORITY"
    SEARCH_HEADER_IN_FIRST_ROWS = 20
    COLUNAS_MINIMAS = {"AGR_NAME", "STATUS", "MSG"} 

    MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P", "CUA": "SPA"}
    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'.")

    def formatar_tempo(segundos):
        h, resto = divmod(segundos, 3600)
        m, s = divmod(resto, 60)
        if h > 0: return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
        return f"{int(m):02d}m {int(s):02d}s"

    def norm_col(s):
        if s is None: s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def norm_txt(s):
        if s is None: s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def now_ts():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def selecionar_ficheiro():
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
            filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*"))
        )
        root.destroy()
        return path

    ###################################################################################
    # LER EXCEL (MAPEAMENTO DINÂMICO DE TODAS AS COLUNAS)
    ###################################################################################
    if not caminho_ficheiro:
        if modo_nao_interativo:
            raise ValueError("Faltou o parâmetro --xlsx em modo não-interativo.")
        print("📂 Selecione o ficheiro Excel…")
        caminho_ficheiro = selecionar_ficheiro()
        if not caminho_ficheiro: return

    try:
        wb = load_workbook(caminho_ficheiro)
        ws = wb[NOME_SHEET]
    except Exception as e:
        print(f"❌ Erro ao abrir Excel: {e}")
        return

    header_row = None
    col_agr_composta = None
    col_text_composta = None
    col_agr_simples = None
    col_text_simples = None
    col_objeto = None
    col_status = None
    col_msg = None
    col_timestamp = None
    dynamic_fields = {} # name -> col_idx

    COLUNAS_ENTRADA = {"AGR_NAME"}
    COLUNAS_SAIDA = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [norm_col(c.value) for c in ws[r]]
        colunas_entrada_encontradas = set(row_vals).intersection(COLUNAS_ENTRADA)
        if len(colunas_entrada_encontradas) == len(COLUNAS_ENTRADA):
            header_row = r
            
            seen_agr_name = False
            for idx, val in enumerate(row_vals, start=1):
                if not val:
                    continue
                if val == "AGR_NAME_COMPOSTA":
                    col_agr_composta = idx
                elif val == "AGR_NAME":
                    col_agr_simples = idx
                    seen_agr_name = True
                elif val == "TEXT":
                    if not seen_agr_name:
                        col_text_composta = idx
                    else:
                        col_text_simples = idx
                elif val in ("OBJETO DE AUTORIZACAO", "OBJETO DE AUTORIZAÇÃO", "OBJETO"):
                    col_objeto = idx
                elif val == "STATUS":
                    col_status = idx
                elif val == "MSG":
                    col_msg = idx
                elif val in ("TIMESTEMP", "TIMESTAMP"):
                    col_timestamp = idx
                elif val not in ("ID",):
                    dynamic_fields[val] = idx

            # Verificar se faltam colunas de saída e criá-las
            modificou = False
            last_col = 0
            for idx, val in enumerate(row_vals, start=1):
                if val:
                    last_col = idx

            if not col_status:
                last_col += 1
                ws.cell(row=r, column=last_col, value="STATUS")
                col_status = last_col
                print(f"➕ Coluna 'STATUS' criada automaticamente na coluna {last_col}.")
                modificou = True

            if not col_msg:
                last_col += 1
                ws.cell(row=r, column=last_col, value="MSG")
                col_msg = last_col
                print(f"➕ Coluna 'MSG' criada automaticamente na coluna {last_col}.")
                modificou = True

            if not col_timestamp:
                last_col += 1
                ws.cell(row=r, column=last_col, value="TIMESTEMP")
                col_timestamp = last_col
                print(f"➕ Coluna 'TIMESTEMP' criada automaticamente na coluna {last_col}.")
                modificou = True

            if modificou:
                try:
                    wb.save(caminho_ficheiro)
                    print("💾 Ficheiro Excel guardado com as novas colunas.")
                except Exception as e:
                    wb.close()
                    print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas: {e}")
                    return
            break

    if not header_row:
        print("\n❌ Cabeçalho não encontrado.")
        wb.close()
        return

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        agr_val = ws.cell(row=r, column=col_agr_simples).value if col_agr_simples else None
        agr = "" if agr_val is None else str(agr_val).strip()
        if not agr: continue
        
        # Validar e remover espaços do nome da Role simples
        if " " in agr:
            orig = agr
            agr = agr.replace(" ", "")
            ws.cell(row=r, column=col_agr_simples, value=agr)
            print(f"⚠️ Espaços detetados e removidos da Role Simples: '{orig}' -> '{agr}' (atualizado no Excel)")
            try:
                wb.save(caminho_ficheiro)
            except Exception:
                pass

        rec = {"_row": r}
        rec["AGR_NAME"] = agr
        
        if col_agr_composta:
            val = ws.cell(row=r, column=col_agr_composta).value
            agr_comp = "" if val is None else str(val).strip()
            if " " in agr_comp:
                orig_comp = agr_comp
                agr_comp = agr_comp.replace(" ", "")
                ws.cell(row=r, column=col_agr_composta, value=agr_comp)
                print(f"⚠️ Espaços detetados e removidos da Role Composta: '{orig_comp}' -> '{agr_comp}' (atualizado no Excel)")
                try:
                    wb.save(caminho_ficheiro)
                except Exception:
                    pass
            rec["AGR_NAME_COMPOSTA"] = agr_comp
        else:
            rec["AGR_NAME_COMPOSTA"] = ""
            
        if col_text_composta:
            val = ws.cell(row=r, column=col_text_composta).value
            rec["TEXT_COMPOSTA"] = "" if val is None else str(val).strip()
        else:
            rec["TEXT_COMPOSTA"] = ""
            
        if col_text_simples:
            val = ws.cell(row=r, column=col_text_simples).value
            rec["TEXT_AGR_NAME"] = "" if val is None else str(val).strip()
        else:
            rec["TEXT_AGR_NAME"] = ""
            
        if col_objeto:
            val = ws.cell(row=r, column=col_objeto).value
            rec["OBJETO DE AUTORIZACAO"] = "" if val is None else str(val).strip()
        else:
            rec["OBJETO DE AUTORIZACAO"] = ""
            
        if col_status:
            val = ws.cell(row=r, column=col_status).value
            rec["STATUS"] = "" if val is None else str(val).strip()
        else:
            rec["STATUS"] = ""
            
        if col_msg:
            val = ws.cell(row=r, column=col_msg).value
            rec["MSG"] = "" if val is None else str(val).strip()
        else:
            rec["MSG"] = ""
            
        if col_timestamp:
            val = ws.cell(row=r, column=col_timestamp).value
            rec["TIMESTEMP"] = "" if val is None else str(val).strip()
        else:
            rec["TIMESTEMP"] = ""
            
        for f_name, f_idx in dynamic_fields.items():
            val = ws.cell(row=r, column=f_idx).value
            rec[f_name] = "" if val is None else str(val).strip()
            
        records.append(rec)

    # Identificar linhas que não estão com status "CONCLUIDO"
    pending_records = []
    for idx, rec in enumerate(records):
        status_norm = norm_txt(rec["STATUS"])
        if "CONCLUIDO" not in status_norm:
            pending_records.append((idx, rec))

    if not pending_records:
        print("⚠️ Tudo concluído.")
        wb.close()
        return

    ###################################################################################
    # CONEXÃO SAP E MENU PREMIUM DE TRANSPORTE
    ###################################################################################
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        session = next((sess for conn in application.Children for sess in conn.Children if sess.Info.SystemName.upper() == SISTEMA_ESPERADO), None)
    except: session = None

    if not session:
        wb.close()
        print(f"❌ Não encontrei sessão do ambiente '{ambiente_cockpit}'.")
        return

    def object_exists(id_string):
        try:
            session.findById(id_string)
            return True
        except: return False

    def _safe_find(sap_id):
        try: return session.findById(sap_id)
        except: return None

    def _sap_busy():
        try: return bool(getattr(session, "Busy", False))
        except: return False

    def _esperar_sap_livre(timeout=8.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            if not _sap_busy(): return True
            time.sleep(pausa)
        return False

    def _esperar_objeto(sap_id, timeout=5.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            obj = _safe_find(sap_id)
            if obj: return obj
            time.sleep(pausa)
        return None

    def _esperar_sumir(sap_id, timeout=5.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            if not _safe_find(sap_id): return True
            time.sleep(pausa)
        return False

    def _send_vkey(vkey, wait_after=True):
        session.findById("wnd[0]").sendVKey(vkey)
        if wait_after: _esperar_sap_livre()

    def _press_if_exists(sap_id, timeout=2.0):
        obj = _esperar_objeto(sap_id, timeout=timeout)
        if not obj: return False
        try:
            obj.press()
            _esperar_sap_livre()
            return True
        except: return False

    sap_id_cache = {}
    def _resolver_id(cache_key, candidatos):
        sap_id = sap_id_cache.get(cache_key)
        if sap_id:
            obj = _safe_find(sap_id)
            if obj: return sap_id, obj
            sap_id_cache.pop(cache_key, None)
        for sap_id in candidatos:
            obj = _safe_find(sap_id)
            if obj:
                sap_id_cache[cache_key] = sap_id
                return sap_id, obj
        return None, None

    def _resolver_id_esperando(cache_key, candidatos, timeout=3.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            sap_id, obj = _resolver_id(cache_key, candidatos)
            if obj: return sap_id, obj
            time.sleep(pausa)
        return None, None

    def get_statusbar():
        try:
            sbar = session.findById("wnd[0]/sbar")
            tipo = getattr(sbar, "MessageType", "").strip().upper()
            texto = (sbar.Text or "").strip()
            if texto: print(f"[SAP_SBAR] {texto}")
            return (tipo, texto)
        except: return ("", "")

    def try_actions(actions):
        for a in actions:
            try:
                ctrl = session.findById(a["path"])
                if a["op"] == "text":
                    ctrl.setFocus()
                    ctrl.text = a["val"]
                    time.sleep(0.1)
                    return True
                elif a["op"] == "press":
                    if hasattr(ctrl, "Enabled") and not ctrl.Enabled: continue
                    ctrl.press()
                    time.sleep(0.15)
                    return True
                elif a["op"] == "select":
                    ctrl.select()
                    time.sleep(0.1)
                    return True
            except: continue
        return False

    def tratar_popup_modal(max_loops=6):
        for _ in range(max_loops):
            try:
                time.sleep(0.1)
                if session.ActiveWindow.Type != "GuiModalWindow": return False
                candidatos = [
                    "wnd[1]/usr/btnBUTTON_1",
                    "wnd[1]/usr/btnSPOP-OPTION1",
                    "wnd[1]/tbar[0]/btn[0]",
                    "wnd[1]/tbar[0]/btn[19]",
                    "wnd[1]/tbar[0]/btn[11]"
                ]
                for p in candidatos:
                    if try_actions([{"path": p, "op": "press"}]): return True
                return True
            except: return False
        return True

    def _criar_nova_request_no_sap_local(sess):
        okcd = _safe_find("wnd[0]/tbar[0]/okcd")
        if okcd:
            okcd.text = "/nSE10"
            sess.findById("wnd[0]").sendVKey(0)
            time.sleep(0.8)

        print("\nTipo da ordem:")
        print('1 - Ordem customizing')
        print('2 - Ordem workbench')

        while True:
            tipo = input("Digite a opção (1/2): ").strip()
            if tipo in ("1", "2"): break
            print("❌ Opção inválida. Use apenas 1 ou 2.")

        desc = input("Descrição da request (máx 60): ").strip()
        desc = desc[:60] if desc else "REQUEST CRIADA VIA SCRIPT"

        sess.findById("wnd[0]/tbar[1]/btn[6]").press()
        time.sleep(0.4)

        if tipo == "2":
            try: sess.findById("wnd[1]/usr/radKO042-REQ_CONS_K").select()
            except: pass

        sess.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(0.4)

        try: sess.findById("wnd[1]/usr/txtKO013-AS4TEXT").text = desc
        except: pass
        
        sess.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(0.6)

        trkorr = None
        for sap_id in ["wnd[0]/usr/lbl[20,9]", "wnd[0]/usr/lbl[1,1]"]:
            try:
                txt = sess.findById(sap_id).Text
                match = re.search(r"\b[A-Z0-9]{3,4}K\d{6,}\b", txt)
                if match: trkorr = match.group(0)
            except: pass
            if trkorr: break

        if okcd:
            okcd.text = "/n"
            sess.findById("wnd[0]").sendVKey(0)

        tipo_txt = "Customizing" if tipo == "1" else "Workbench"
        print("\n✔️ Request criada.")
        print(f"Tipo: {tipo_txt} | Descrição: {desc}")

        if not trkorr:
            trkorr = input("Não consegui extrair a request automaticamente. Cole aqui: ").strip().upper()
        
        print(f"Request: {trkorr}")
        return trkorr

    print("\n📋 Resumo das Funções a processar:")
    for idx_item, rec_item in pending_records:
        r_nome = rec_item["AGR_NAME"]
        r_obj = rec_item["OBJETO DE AUTORIZACAO"] or "F_KNA1_GRP"
        comp_info = f" -> Composta: {rec_item['AGR_NAME_COMPOSTA']}" if rec_item['AGR_NAME_COMPOSTA'] else ""
        print(f"   - {r_nome} (Obj: {r_obj}){comp_info}")
        
    print(f"\n🔢 Total de linhas para processamento: {len(pending_records)}")

    if not request_transporte and not modo_nao_interativo:
        print("\n============================================================")
        print("🚚 Opções de configuração de Transporte.\n")
        print("   1 - Escreva o número da Request")
        print("   2 - Criar nova ordem de transporte")
        print("   3 - Pesquisar suas request criadas.")
        print("   4 - Prima [Enter] vazio para NÃO transportar")
        print("============================================================")
        
        while True:
            req_input = input("\n👉 Opção: ").strip()
            if req_input in ("1", "2", "3", "4", ""):
                if req_input == "": req_input = "4"
                break
            print("❌ Opção inválida. Use 1, 2, 3, 4 ou apenas pressione Enter.")
        
        if req_input == "1":
            request_transporte = input("🔢 Numero da Request (ex: S4QK900396): ").strip().upper()
        
        elif req_input == "2":
            request_transporte = _criar_nova_request_no_sap_local(session)

        elif req_input == "3":
            try:
                import pesquisar_request
                print("\n🔍 A abrir nova sessão em segundo plano para pesquisar (SE16H)...")
                
                resultados_pesquisa = pesquisar_request.listar_requests(
                    system_name=SISTEMA_ESPERADO, 
                    include_requests=True, 
                    use_new_mode=True,     
                    minimize=True,         
                    close_after=True,
                    session=session
                )
                
                if resultados_pesquisa:
                    escolha = input("\n👉 Digite o número (N) da Request que deseja utilizar (ou Enter para cancelar): ").strip()
                    if escolha.isdigit() and 1 <= int(escolha) <= len(resultados_pesquisa):
                        request_transporte = resultados_pesquisa[int(escolha) - 1][0]
                        print(f"✔️ Selecionou a Request: {request_transporte}")
                    else:
                        print("❌ Seleção cancelada. Não haverá transporte.")
                else:
                    print("⚠️ Não foram encontradas Requests abertas para o seu utilizador.")
            except ImportError as e:
                print(f"❌ Erro de Importação: Não consegui encontrar o módulo pesquisar_request.py. Detalhe: {e}")
        
        elif req_input == "4":
            print("⏭️  Nenhuma request selecionada (Transporte ignorado).")
            request_transporte = None
        print("============================================================")

    if pedir_confirmacao and not modo_nao_interativo:
        if input("\nDeseja iniciar o processamento no SAP? [S/N]: ").strip().upper() != "S":
            wb.close()
            print("❌ Operação cancelada pelo utilizador.")
            return

    ###################################################################################
    # LOGGER DE AUDITORIA & EXECUTOR
    ###################################################################################
    class PFCG_AuthPage_Auditor:
        def __init__(self, sess):
            self.sess = sess

        def get_sbar(self):
            try:
                sbar = self.sess.findById("wnd[0]/sbar")
                return getattr(sbar, "MessageType", "").strip().upper(), (sbar.Text or "").strip()
            except: return "", ""

        def audit_step(self, descricao, path, acao="press", valor=None, vkey=None, silencioso=False):
            if not silencioso:
                print(f"\n  🔎 [AUDIT] {descricao}")
                log_detail = f"      ↳ ID: {path} | Ação: {acao}"
                if valor is not None: log_detail += f" | Valor: '{valor}'"
                if vkey is not None: log_detail += f" | VKey: {vkey}"
                print(log_detail)

            try:
                if path: elem = self.sess.findById(path)

                if acao == "text": elem.text = valor
                elif acao == "press": elem.press()
                elif acao == "select": 
                    if hasattr(elem, "selected"): elem.selected = True
                    else: elem.select()
                elif acao == "sendVKey":
                    if path: elem.sendVKey(vkey)
                    else: self.sess.findById("wnd[0]").sendVKey(vkey)

                if not silencioso:
                    print("      ✅ SUCESSO")
                    mtype, mtext = self.get_sbar()
                    if mtext:
                        icone = "🔴" if mtype in ["E", "A"] else ("🟡" if mtype == "W" else "🟢")
                        print(f"      {icone} SAP STATUS: [{mtype}] {mtext}")
                
            except Exception as e:
                if not silencioso:
                    print(f"      ❌ FALHA AQUI (Erro 619): O ID [{path}] falhou.")
                    raise Exception(f"FALHA NO PASSO: '{descricao}' -> ID: {path}")

        def ensure_role_exists(self, nome, desc):
            print("├─ Abrindo /NPFCG...")
            self.audit_step("Chamar transação /npfcg", "wnd[0]/tbar[0]/okcd", "text", "/npfcg", silencioso=True)
            self.audit_step("Enter (Ir para PFCG)", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
            
            print(f"├─ Validando role: {nome}")
            self.audit_step("Inserir Nome da Role", "wnd[0]/usr/ctxtAGR_NAME_NEU", "text", nome, silencioso=True)
            
            try_actions([
                {"path": "wnd[0]/usr/btn%#AUTOTEXT003", "op": "press"},
                {"path": "wnd[0]/tbar[1]/btn[5]", "op": "press"}
            ])
            tratar_popup_modal()
            
            mt, sb = get_statusbar()
            if "EXISTE" in norm_txt(sb) or "EXISTS" in norm_txt(sb):
                print(f"├─ Role já existe. Abrindo em alteração...")
                try_actions([
                    {"path": "wnd[0]/usr/btn%#AUTOTEXT001", "op": "press"},
                    {"path": "wnd[0]/tbar[1]/btn[2]", "op": "press"}
                ])
                tratar_popup_modal()
                
                desc_atual = ""
                sap_id, obj = _resolver_id(
                    "role_desc_field",
                    ["wnd[0]/usr/txtS_AGR_TEXTS-TEXT", "wnd[0]/usr/txtS_AGR_TEXTS-TEXT1", "wnd[0]/usr/txtAGR_TEXTS-TEXT"]
                )
                if obj: desc_atual = obj.text
                print(f"├─ Descrição: {desc_atual}")
                return "CHANGE"
            else:
                print(f"├─ Role não existe. Criando role simples...")
                sap_id, obj = _resolver_id(
                    "role_desc_field",
                    ["wnd[0]/usr/txtS_AGR_TEXTS-TEXT", "wnd[0]/usr/txtS_AGR_TEXTS-TEXT1", "wnd[0]/usr/txtAGR_TEXTS-TEXT"]
                )
                if obj:
                    descricao_final = desc if desc else f"Criada via Script - {nome}"
                    obj.text = descricao_final
                    _send_vkey(0)
                    tratar_popup_modal()
                    
                try:
                    self.sess.findById("wnd[0]").sendVKey(11)
                except:
                    try_actions([{"path": "wnd[0]/tbar[0]/btn[11]", "op": "press"}])
                _esperar_sap_livre()
                tratar_popup_modal()
                return "CREATE"

        def update_mass_values_dynamic(self, nome, objeto, row_data):
            self.audit_step("Chamar transação /nPFCGMASSVAL", "wnd[0]/tbar[0]/okcd", "text", "/nPFCGMASSVAL", silencioso=True)
            self.audit_step("Enter (Ir para MASSVAL)", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
            
            self.audit_step("Selecionar Execução Direta", "wnd[0]/usr/radMOD_EXE", "select", silencioso=True)
            self.audit_step("Selecionar Inserção Manual", "wnd[0]/usr/radSEL_NAU", "select", silencioso=True)
            self.audit_step("ENTER Crucial do VBS para atualizar tela", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
            
            self.audit_step("Preencher ROLE-LOW por segurança", "wnd[0]/usr/ctxtROLE-LOW", "text", nome, silencioso=True)
            
            self.audit_step("Preencher OBJOBJ (Objeto de Autorização)", "wnd[0]/usr/ctxtOBJOBJ", "text", objeto, silencioso=True)
            self.audit_step("Enter após OBJOBJ (A carregar campos...)", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
            
            time.sleep(0.5)

            campos_encontrados_sap = 0
            alterou_qualquer_campo = False
            
            detalhe_existentes = []
            detalhe_novos = []
            
            for j in range(1, 15):
                btn_id = f"wnd[0]/usr/btnPOBJ{j}N"
                if not object_exists(btn_id):
                    break 
                
                campos_encontrados_sap += 1
                campo_sap_tecnico = ""
                ids_para_testar = [
                    f"wnd[0]/usr/txtOBJFLD{j}",      
                    f"wnd[0]/usr/ctxtOBJFLD{j}",     
                    f"wnd[0]/usr/ctxtS_AGR_DEFINE-FNAM{j}",
                    f"wnd[0]/usr/txtS_AGR_DEFINE-FNAM{j}",
                    f"wnd[0]/usr/ctxtFNAM{j}",
                    f"wnd[0]/txtFNAM{j}",
                    f"wnd[0]/usr/ctxtFIELD{j}",
                    f"wnd[0]/usr/txtFIELD{j}"
                ]
                
                for cid in ids_para_testar:
                    if object_exists(cid):
                        campo_sap_tecnico = self.sess.findById(cid).text.strip().upper()
                        break

                if not campo_sap_tecnico:
                    campo_sap_tecnico = str(row_data.get(f'CAMPO {j}', '')).strip().upper()

                if not campo_sap_tecnico:
                    continue 

                valor_no_excel = str(row_data.get(campo_sap_tecnico, '')).strip()
                if not valor_no_excel or valor_no_excel == 'NAN':
                    continue

                # Open the values popup to read/edit
                self.audit_step(f"Clicar Botão VALS para '{campo_sap_tecnico}'", btn_id, "press", silencioso=True)
                time.sleep(0.5) 

                existing_vals = []
                missing_vals = []
                alterou_este_campo = False
                
                # Check which table control is open in popup wnd[1]
                if object_exists("wnd[1]/usr/tblSAPLSUPRNACT_TC"):
                    # Checkbox table
                    for linha in range(40):
                        try:
                            act_code = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNACT_TC/txtH_FVAL-LOW[1,{linha}]").text.strip()
                            if not act_code:
                                break
                            chk = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNACT_TC/chkH_FVAL-MARK[0,{linha}]")
                            if chk.selected:
                                existing_vals.append(act_code)
                        except:
                            break
                            
                    valores_lista = [v.strip() for v in valor_no_excel.split(',')]
                    missing_vals = [v for v in valores_lista if v not in existing_vals]
                    
                    if missing_vals:
                        alterou_este_campo = True
                        for linha in range(40):
                            try:
                                act_code = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNACT_TC/txtH_FVAL-LOW[1,{linha}]").text.strip()
                                if not act_code:
                                    break
                                if act_code in missing_vals:
                                    chk = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNACT_TC/chkH_FVAL-MARK[0,{linha}]")
                                    if not chk.selected:
                                        self.audit_step(f"Marcar Checkbox '{act_code}'", f"wnd[1]/usr/tblSAPLSUPRNACT_TC/chkH_FVAL-MARK[0,{linha}]", "select", silencioso=True)
                            except:
                                break
                
                elif object_exists("wnd[1]/usr/tblSAPLSUPRNVAL_TC"):
                    # Values table
                    empty_row_indices = []
                    for linha in range(40):
                        val_text = ""
                        try:
                            val_text = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNVAL_TC/ctxtH_FVAL_LOW[0,{linha}]").text.strip()
                        except:
                            try:
                                val_text = self.sess.findById(f"wnd[1]/usr/tblSAPLSUPRNVAL_TC/ctxtH_FVAL_LOW[1,{linha}]").text.strip()
                            except:
                                break
                        if val_text:
                            existing_vals.append(val_text)
                        else:
                            empty_row_indices.append(linha)
                            
                    if valor_no_excel == "*":
                        if "*" not in existing_vals:
                            self.audit_step(f"Full Auth (*) no Campo '{campo_sap_tecnico}'", "wnd[1]/usr/btnGES2", "press", silencioso=True)
                            alterou_este_campo = True
                            missing_vals = ["*"]
                        else:
                            missing_vals = []
                    else:
                        valores_lista = [v.strip() for v in valor_no_excel.split(',')]
                        missing_vals = [v for v in valores_lista if v not in existing_vals]
                        if missing_vals:
                            alterou_este_campo = True
                            for i, val in enumerate(missing_vals):
                                if i < len(empty_row_indices):
                                    empty_r = empty_row_indices[i]
                                    try:
                                        self.audit_step(f"Preencher Valor '{val}'", f"wnd[1]/usr/tblSAPLSUPRNVAL_TC/ctxtH_FVAL_LOW[0,{empty_r}]", "text", val, silencioso=True)
                                    except Exception:
                                        self.audit_step(f"Preencher Valor '{val}'", f"wnd[1]/usr/tblSAPLSUPRNVAL_TC/ctxtH_FVAL_LOW[1,{empty_r}]", "text", val, silencioso=True)
                                else:
                                    raise Exception(f"Tabela de valores cheia, não foi possível inserir '{val}' no campo '{campo_sap_tecnico}'")
                else:
                    raise Exception(f"Não foi possível validar o estado da autorização para o campo '{campo_sap_tecnico}'. Interface desconhecida.")

                if existing_vals:
                    detalhe_existentes.append(f"{campo_sap_tecnico} ({','.join(existing_vals)})")
                else:
                    detalhe_existentes.append(f"{campo_sap_tecnico} (nenhuma)")
                    
                if missing_vals:
                    detalhe_novos.append(f"{campo_sap_tecnico} ({','.join(missing_vals)})")
                    alterou_qualquer_campo = True

                self.audit_step(f"Confirmar Popup de '{campo_sap_tecnico}'", "wnd[1]/tbar[0]/btn[0]", "press", silencioso=True)
                time.sleep(0.2)

            str_existentes = " | ".join(detalhe_existentes) if detalhe_existentes else "nenhum"
            str_novos = " | ".join(detalhe_novos) if detalhe_novos else "nenhuma"
            print(f"├─ Valores já existentes: {str_existentes}")
            print(f"├─ Valores novos a inserir: {str_novos}")

            if not alterou_qualquer_campo:
                # Regresso sem salvar para evitar transportes/gerações desnecessárias
                self.audit_step("Voltar ao Ecrã Inicial (/N) por falta de deltas", "wnd[0]/tbar[0]/okcd", "text", "/N", silencioso=True)
                self.audit_step("Enter (/N)", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
                return False

            if campos_encontrados_sap == 0:
                print("  ⚠️ AVISO: O Objeto inserido não gerou campos visíveis.")

            self.audit_step("Clicar Executar (Relógio)", "wnd[0]/tbar[1]/btn[8]", "press", silencioso=True)
            self.audit_step("Clicar Guardar (Disquete)", "wnd[0]/tbar[1]/btn[20]", "press", silencioso=True)
            
            if self.sess.Children.Count > 1 and object_exists("wnd[1]/tbar[0]/btn[0]"):
                self.audit_step("Confirmar popup Sucesso Gravação", "wnd[1]/tbar[0]/btn[0]", "press", silencioso=True)

            mt, sb = self.get_sbar()
            if mt in ("E", "A"):
                raise Exception(f"Erro ao salvar autorizações: {sb}")
                
            mudou = True
            if sb:
                sb_norm = norm_txt(sb)
                if "NAO FORAM MODIFICADOS DADOS" in sb_norm or "SEM ALTERACOES" in sb_norm or "NO CHANGES" in sb_norm:
                    mudou = False
            return mudou

        def execute_transport(self, req_num, nome):
            self.audit_step("Chamar transação /nPFCG para Transporte", "wnd[0]/tbar[0]/okcd", "text", "/nPFCG", silencioso=True)
            self.audit_step("Enter", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
            
            self.audit_step("Preencher Função", "wnd[0]/usr/ctxtAGR_NAME_NEU", "text", nome, silencioso=True)
            self.audit_step("Selecionar Menu Transporte", "wnd[0]/mbar/menu[0]/menu[9]", "select", silencioso=True)
            self.audit_step("Executar Transporte", "wnd[0]/tbar[1]/btn[8]", "press", silencioso=True)
            time.sleep(0.3)
            
            if self.sess.Children.Count > 1 and object_exists("wnd[1]/usr/ctxtKO008-TRKORR"):
                self.audit_step("Inserir Request", "wnd[1]/usr/ctxtKO008-TRKORR", "text", req_num, silencioso=True)
                self.audit_step("Confirmar Request", "wnd[1]/tbar[0]/btn[0]", "press", silencioso=True)

    auditor = PFCG_AuthPage_Auditor(session)
    resultados_simples = {}
    resultados = {}

    # Contadores de Resumo
    linhas_processadas = len(pending_records)
    agr_criadas = 0
    agr_existentes = 0
    agr_com_erro = 0
    objetos_alterados = 0
    objetos_sem_alteracao = 0
    
    compostas_processadas = 0
    compostas_criadas = 0
    compostas_existentes = 0
    compostas_sem_alteracao = 0
    compostas_com_erro = 0
    componentes_inseridos_total = 0
    componentes_existentes_total = 0

    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
        session.findById("wnd[0]").sendVKey(0)
    except: pass

    # =====================================================================
    # FASE 1: Funções Simples (Agrupadas por AGR_NAME para otimização)
    # =====================================================================
    # Encontrar a lista distinta de AGR_NAME que estão pendentes
    agrs_pendentes = list(dict.fromkeys([rec["AGR_NAME"] for _, rec in pending_records if rec["AGR_NAME"]]))
    total_agrs = len(agrs_pendentes)
    
    for idx_agr, agr_name in enumerate(agrs_pendentes, start=1):
        # Achar todas as linhas pendentes deste AGR_NAME
        linhas_grupo = [(idx, rec) for idx, rec in pending_records if rec["AGR_NAME"] == agr_name]
        
        print("\n======================================================================")
        print(f"▶ [{idx_agr}/{total_agrs}] INICIANDO AGR_NAME: {agr_name}")
        print("======================================================================")
        
        # Obter descrição da primeira linha que tiver
        desc_role = ""
        for _, rec in linhas_grupo:
            if rec.get("TEXT_AGR_NAME"):
                desc_role = rec["TEXT_AGR_NAME"]
                break
                
        try:
            print("[Etapa 1] Preparação da Role Simples")
            modo = auditor.ensure_role_exists(agr_name, desc_role)
            if modo == "CREATE":
                agr_criadas += 1
            else:
                agr_existentes += 1
                
            for idx_row, row_data in linhas_grupo:
                objeto = row_data["OBJETO DE AUTORIZACAO"] or "F_KNA1_GRP"
                print("\n[Etapa 2] Autorizações")
                print(f"├─ Objeto de autorização: {objeto}")
                
                # Logar campos
                for campo in ["KTOKD", "ACTVT", "RLTYP"]:
                    val_c = str(row_data.get(campo, "")).strip()
                    if val_c and val_c != "NAN":
                        print(f"├─ Campo {campo}: {val_c}")
                        
                # Executar MASSVAL
                mudou = auditor.update_mass_values_dynamic(agr_name, objeto, row_data)
                if mudou:
                    objetos_alterados += 1
                else:
                    objetos_sem_alteracao += 1
                    
                resultados_simples[idx_row] = {
                    "STATUS": "CONCLUIDO",
                    "MSG": "AGR_NAME OK"
                }
                print(f"└─ Resultado: CONCLUIDO")
                
            if request_transporte:
                print("\n[Etapa 3] Inserir na Request de Transporte")
                auditor.execute_transport(request_transporte, agr_name)
                
            auditor.audit_step("Voltar ao Ecrã Inicial (/N)", "wnd[0]/tbar[0]/okcd", "text", "/N", silencioso=True)
            auditor.audit_step("Enter (/N)", "wnd[0]", "sendVKey", vkey=0, silencioso=True)
            
        except Exception as e:
            err_msg = str(e)
            print(f"🔴 ERRO no processamento da role simples {agr_name}: {err_msg}")
            agr_com_erro += 1
            
            for idx_row, _ in linhas_grupo:
                resultados_simples[idx_row] = {
                    "STATUS": "ERRO",
                    "MSG": f"AGR_NAME ERRO: {err_msg}"
                }
                
            try:
                session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
                session.findById("wnd[0]").sendVKey(0)
            except: pass
            
        # Checkpoint incremental a cada AGR_NAME processado
        try:
            col_st, col_ms, col_tm = col_status, col_msg, col_timestamp
            for idx_row, _ in linhas_grupo:
                res = resultados_simples.get(idx_row)
                if res:
                    ws.cell(row=records[idx_row]["_row"], column=col_st).value = res["STATUS"]
                    ws.cell(row=records[idx_row]["_row"], column=col_ms).value = res["MSG"]
                    ws.cell(row=records[idx_row]["_row"], column=col_tm).value = now_ts()
            wb.save(caminho_ficheiro)
        except Exception as checkpoint_exc:
            print(f"  ⚠️ Erro ao salvar checkpoint do Excel na Fase 1: {checkpoint_exc}")

    # =====================================================================
    # FASE 2: Funções Compostas
    # =====================================================================
    if col_agr_composta:
        # Achar todas as compostas pendentes na execução
        compostas_a_processar = list(dict.fromkeys([rec["AGR_NAME_COMPOSTA"] for _, rec in pending_records if rec["AGR_NAME_COMPOSTA"].strip()]))
        
        if len(compostas_a_processar) > 0:
            print("\n======================================================================")
            print(f"▶ [FASE 2] INICIANDO FASE COMPOSTA: {len(compostas_a_processar)} Funções Compostas a processar")
            print("======================================================================")
            
            grupos_para_processar = []
            
            for nome_comp in compostas_a_processar:
                nome_comp = str(nome_comp).strip()
                
                # Obter todas as linhas desta composta no Excel completo (records)
                linhas_composta_original = [rec for rec in records if rec["AGR_NAME_COMPOSTA"] == nome_comp]
                
                # Obter a lista distinta de AGR_NAME (componentes) para esta composta
                componentes_todos = list(dict.fromkeys([r["AGR_NAME"] for r in linhas_composta_original if r["AGR_NAME"]]))
                
                # Obter a descrição da composta
                desc_comp = ""
                for rec in linhas_composta_original:
                    if rec.get("TEXT_COMPOSTA"):
                        desc_comp = rec["TEXT_COMPOSTA"]
                        break
                
                # Validar componentes
                componentes_validos = []
                componentes_erros = []
                
                for comp in componentes_todos:
                    # Encontrar todas as linhas correspondentes a este comp no Excel
                    # Para saber se ele foi processado nesta execução ou se já estava concluído
                    rows_comp_processados = [idx for idx, rec in pending_records if rec["AGR_NAME"] == comp]
                    
                    if len(rows_comp_processados) > 0:
                        teve_erro = False
                        for r_idx in rows_comp_processados:
                            res_simp = resultados_simples.get(r_idx, {})
                            if res_simp.get("STATUS") == "ERRO":
                                teve_erro = True
                                break
                        if teve_erro:
                            componentes_erros.append(comp)
                        else:
                            componentes_validos.append(comp)
                    else:
                        componentes_validos.append(comp)
                        
                # Obter os índices das linhas pendentes desta composta
                linhas_pendentes_comp = [idx for idx, rec in pending_records if rec["AGR_NAME_COMPOSTA"] == nome_comp]
                
                if componentes_erros:
                    err_msg = f"Composta não processada: existem componentes AGR_NAME com erro na fase anterior. Falhas: {', '.join(componentes_erros)}."
                    print(f"🔴 {err_msg}")
                    compostas_com_erro += 1
                    for idx_row in linhas_pendentes_comp:
                        res_simp = resultados_simples.get(idx_row, {"STATUS": "ERRO", "MSG": "Não processado"})
                        if res_simp["STATUS"] == "CONCLUIDO":
                            resultados[idx_row] = {
                                "STATUS": "ERRO",
                                "MSG": f"AGR_NAME OK | {err_msg}"
                            }
                        else:
                            resultados[idx_row] = {
                                "STATUS": "ERRO",
                                "MSG": f"{res_simp['MSG']} | Composta não processada."
                            }
                else:
                    grupos_para_processar.append({
                        "agr_name_composta": nome_comp,
                        "text_composta": desc_comp,
                        "componentes": componentes_validos,
                        "linhas_excel": linhas_pendentes_comp
                    })
                    
            if grupos_para_processar:
                print(f"├─ Carregando módulo D. PFCG_COMPOSTA.py dinamicamente...")
                try:
                    import importlib.util
                    spec = importlib.util.spec_from_file_location(
                        "pfcg_composta",
                        os.path.join(dir_atual, "D. PFCG_COMPOSTA.py")
                    )
                    pfcg_composta = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(pfcg_composta)
                    
                    # Pre-populate cache with successful components from Phase 1
                    pfcg_composta.roles_existentes_cache = {}
                    for idx_row, res_s in resultados_simples.items():
                        if res_s.get("STATUS") == "CONCLUIDO":
                            rec_item = records[idx_row]
                            pfcg_composta.roles_existentes_cache[rec_item["AGR_NAME"].upper().strip()] = True
                    
                    res_compostas = pfcg_composta.executar_grupos_compostos(
                        ambiente_cockpit=ambiente_cockpit,
                        grupos_compostos=grupos_para_processar,
                        caminho_ficheiro=caminho_ficheiro,
                        request_transporte=request_transporte,
                        modo_nao_interativo=True,
                        pedir_confirmacao=False,
                        origem="PFCG_AUTHORITY",
                        sess_externa=session
                    )
                    
                    for gp in grupos_para_processar:
                        nome_comp = gp["agr_name_composta"]
                        res_comp = res_compostas.get(nome_comp, {"STATUS": "ERRO", "MSG": "Erro desconhecido na composta.", "modo": "CHANGE", "inseridos": 0, "existentes": 0})
                        
                        compostas_processadas += 1
                        if res_comp["STATUS"] == "CONCLUIDO":
                            if res_comp.get("modo") == "CREATE":
                                compostas_criadas += 1
                            else:
                                if res_comp.get("inseridos", 0) > 0:
                                    compostas_existentes += 1
                                else:
                                    compostas_sem_alteracao += 1
                                    
                            componentes_inseridos_total += res_comp.get("inseridos", 0)
                            componentes_existentes_total += res_comp.get("existentes", 0)
                        else:
                            compostas_com_erro += 1
                            
                        for idx_row in gp["linhas_excel"]:
                            res_simp = resultados_simples.get(idx_row, {"STATUS": "ERRO", "MSG": "Não processado"})
                            if res_simp["STATUS"] == "CONCLUIDO":
                                if res_comp["STATUS"] == "CONCLUIDO":
                                    resultados[idx_row] = {
                                        "STATUS": "CONCLUIDO",
                                        "MSG": f"AGR_NAME OK | COMPOSTA OK: {res_comp['MSG']}"
                                    }
                                else:
                                    resultados[idx_row] = {
                                        "STATUS": "ERRO",
                                        "MSG": f"AGR_NAME OK | COMPOSTA ERRO: {res_comp['MSG']}"
                                    }
                            else:
                                resultados[idx_row] = {
                                    "STATUS": "ERRO",
                                    "MSG": f"{res_simp['MSG']} | COMPOSTA não processada."
                                }
                except Exception as e_import:
                    err_msg = f"Erro técnico ao invocar D. PFCG_COMPOSTA.py: {e_import}"
                    print(f"🔴 {err_msg}")
                    for gp in grupos_para_processar:
                        compostas_com_erro += 1
                        for idx_row in gp["linhas_excel"]:
                            res_simp = resultados_simples.get(idx_row, {"STATUS": "ERRO", "MSG": "Não processado"})
                            resultados[idx_row] = {
                                "STATUS": "ERRO",
                                "MSG": f"{res_simp.get('MSG', 'AGR_NAME OK')} | COMPOSTA ERRO: {err_msg}"
                            }
                            
                # Checkpoint incremental para as compostas
                try:
                    col_st, col_ms, col_tm = col_status, col_msg, col_timestamp
                    for gp in grupos_para_processar:
                        for idx_row in gp["linhas_excel"]:
                            res = resultados.get(idx_row)
                            if res:
                                ws.cell(row=records[idx_row]["_row"], column=col_st).value = res["STATUS"]
                                ws.cell(row=records[idx_row]["_row"], column=col_ms).value = res["MSG"]
                                ws.cell(row=records[idx_row]["_row"], column=col_tm).value = now_ts()
                    wb.save(caminho_ficheiro)
                except Exception as checkpoint_exc:
                    print(f"  ⚠️ Erro ao salvar checkpoint do Excel na Fase 2: {checkpoint_exc}")

    # Resolver linhas sem composta associada
    for idx_row, rec in pending_records:
        if idx_row not in resultados:
            res_simp = resultados_simples.get(idx_row, {"STATUS": "ERRO", "MSG": "Não processado"})
            if res_simp["STATUS"] == "CONCLUIDO":
                resultados[idx_row] = {"STATUS": "CONCLUIDO", "MSG": "AGR_NAME OK | Sem composta associada."}
            else:
                resultados[idx_row] = res_simp

    ###################################################################################
    # GRAVAR EXCEL FINAL E EXIBIR RESUMOS
    ###################################################################################
    try:
        col_st = col_status
        col_ms = col_msg
        col_tm = col_timestamp
        for rec in records:
            # Encontrar no pending_records
            match_idx = [idx for idx, r in pending_records if r['_row'] == rec['_row']]
            if not match_idx: continue
            res = resultados.get(match_idx[0])
            if res:
                if col_st: ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
                if col_ms: ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
                if col_tm: ws.cell(row=rec["_row"], column=col_tm).value = now_ts()
        wb.save(caminho_ficheiro)
        wb.close()
        print("\n💾 Resultados gravados no Excel.")
    except Exception as e:
        print(f"❌ Erro ao gravar no Excel: {e}")

    # Exibir logs de resumo finais
    print("\n📊 RESUMO PFCG_AUTHORITY")
    print(f"- Linhas processadas: {linhas_processadas}")
    print(f"- AGR_NAME criadas: {agr_criadas}")
    print(f"- AGR_NAME já existentes: {agr_existentes}")
    print(f"- AGR_NAME com erro: {agr_com_erro}")
    print(f"- Objetos de autorização alterados: {objetos_alterados}")
    print(f"- Objetos de autorização sem alteração: {objetos_sem_alteracao}")
    
    print("\n📊 RESUMO AGR_NAME_COMPOSTA")
    print(f"- Compostas processadas: {compostas_processadas}")
    print(f"- Compostas criadas: {compostas_criadas}")
    print(f"- Compostas já existentes: {compostas_existentes}")
    print(f"- Componentes inseridos: {componentes_inseridos_total}")
    print(f"- Componentes já existentes: {componentes_existentes_total}")
    print(f"- Compostas sem alteração: {compostas_sem_alteracao}")
    print(f"- Compostas com erro: {compostas_com_erro}")

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")

    return True

if __name__ == "__main__":
    executar("DEV")