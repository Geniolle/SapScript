"""
Processador de perfis SAP com controlo por departamento.

O script segue a lógica validada:
1. Lê a sheet "Controlo" e processa apenas os departamentos com STATUS vazio.
2. Lê a sheet "Proposta" como base de funções/transações.
3. Para cada departamento pendente no Controlo:
   - lê a sheet do departamento;
   - identifica as transações marcadas com X por utilizador;
   - procura na sheet "Proposta" quais funções SAP contêm essas transações;
   - complementa com regras da sheet "DEFINIÇÕES";
   - escreve as funções finais na linha do utilizador em "Proposta Ativa", a partir da coluna "Descrição".
4. Gera/atualiza "PFCG_CREATE" apenas com as transações efetivamente marcadas com X para os utilizadores dos departamentos pendentes.
5. Atualiza "PFCG_COMPOSTA" apenas para as Composite Roles dos utilizadores dos departamentos pendentes, preservando os restantes dados.
6. Atualiza "CUA_ADICIONAR" apenas para os utilizadores dos departamentos pendentes, preservando os restantes dados.
7. Atualiza a sheet "Controlo" com STATUS e TIMESTAMP.
8. Guarda uma cópia do ficheiro com sufixo "_processado_YYYYMMDD_HHMMSS.xlsx".

Requisitos:
    pip install openpyxl

Execução:
    python processar_perfis_sap_controlo_v3.py

Também pode executar passando o caminho do ficheiro:
    python processar_perfis_sap_controlo_v3.py "C:\\caminho\\S4H_Perfis de autorização.xlsx"
"""

from __future__ import annotations

import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:  # pragma: no cover - ambientes sem UI
    tk = None
    filedialog = None
    messagebox = None


# =========================
# Configurações principais
# =========================

SISTEMA_FIXO = "S4DCLNT100"

SHEET_CONTROLO = "Controlo"
SHEET_PROPOSTA = "Proposta"
SHEET_PROPOSTA_ATIVA = "Proposta Ativa"
SHEET_DEFINICOES = "DEFINIÇÕES"
SHEET_PFCG_CREATE = "PFCG_CREATE"
SHEET_PFCG_COMPOSTA = "PFCG_COMPOSTA"
SHEET_CUA_ADICIONAR = "CUA_ADICIONAR"

PFCG_CREATE_HEADERS = ["ID", "AGR_NAME", "TEXT", "TCODE"]
PFCG_COMPOSTA_HEADERS = ["ID", "AGR_NAME_COMPOSTA", "TEXT", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]
CUA_ADICIONAR_HEADERS = ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]

STATUS_PROCESSADO = "PROCESSADO"
STATUS_ERRO = "ERRO"

# Por padrão, as roles organizacionais/BP da sheet DEFINIÇÕES são atribuídas ao utilizador
# na CUA_ADICIONAR, mas não entram na composição funcional da Composite Role.
EXCLUIR_REGRAS_ORG_DA_PFCG_COMPOSTA = True
PFCG_COMPOSTA_EXCLUDE_PREFIXES = ("ZORG_",)
PFCG_COMPOSTA_EXCLUDE_EXACT = {
    "Z_BR_TYPE_BP_GERAL",
    "Z_BR_BUSINESS_PARTNER",
    "Z_BR_BUSINESS_PARTNER_GL",
}

MAIN_AND_OUTPUT_SHEETS = {
    SHEET_CONTROLO,
    SHEET_PROPOSTA,
    SHEET_PROPOSTA_ATIVA,
    SHEET_DEFINICOES,
    SHEET_PFCG_CREATE,
    SHEET_PFCG_COMPOSTA,
    SHEET_CUA_ADICIONAR,
    "Users Ativos",
}


@dataclass
class DepartamentoPendente:
    nome: str
    row_idx: int


# =========================
# Utilitários genéricos
# =========================

def normalizar_texto(value: Any) -> str:
    """Normaliza texto para comparação de cabeçalhos/chaves."""
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    text = " ".join(text.split())
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.upper()


def texto_limpo(value: Any) -> str:
    """Converte célula em texto limpo."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").strip().split())


def valor_marcado_com_x(value: Any) -> bool:
    """Valida marcação X nas matrizes dos departamentos."""
    return normalizar_texto(value) == "X"


def split_valores(value: Any) -> list[str]:
    """Divide valores separados por vírgula, ponto e vírgula ou quebra de linha."""
    if value is None:
        return []
    raw = str(value).replace("\r", "\n")
    parts = re.split(r"[,;\n]+", raw)
    return [texto_limpo(part) for part in parts if texto_limpo(part)]


def unique_keep_order(values: Iterable[str]) -> list[str]:
    """Remove duplicados mantendo a ordem."""
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        clean = texto_limpo(value)
        if not clean:
            continue
        key = normalizar_texto(clean)
        if key not in seen:
            seen.add(key)
            result.append(clean)
    return result


def user_id_from_header(value: Any) -> str:
    """Extrai o ID do utilizador de cabeçalhos como 'S425 - José Penouço'."""
    text = texto_limpo(value)
    if not text:
        return ""
    if "-" in text:
        return texto_limpo(text.split("-", 1)[0])
    return texto_limpo(text.split()[0])


def output_path_for(input_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return input_path.with_name(f"{input_path.stem}_processado_{timestamp}{input_path.suffix}")


def escolher_ficheiro_excel() -> Path:
    """Abre pop-up em primeiro plano para seleção do ficheiro Excel."""
    if tk is None or filedialog is None:
        raise RuntimeError("Tkinter não está disponível neste ambiente. Passe o caminho do ficheiro por argumento.")

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.lift()
    root.focus_force()

    file_path = filedialog.askopenfilename(
        parent=root,
        title="Selecione o ficheiro Excel de perfis SAP",
        filetypes=[("Ficheiros Excel", "*.xlsx *.xlsm"), ("Todos os ficheiros", "*.*")],
    )

    root.destroy()

    if not file_path:
        raise SystemExit("Nenhum ficheiro selecionado. Processo cancelado.")

    return Path(file_path)


def mostrar_mensagem(titulo: str, mensagem: str) -> None:
    """Mostra mensagem final quando houver UI disponível."""
    if tk is None or messagebox is None:
        return
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showinfo(titulo, mensagem, parent=root)
        root.destroy()
    except Exception:
        pass


# =========================
# Utilitários Excel
# =========================

def resolver_nome_sheet(wb, name: str) -> str | None:
    """
    Resolve o nome real da sheet no workbook usando comparação normalizada.

    Isto evita erro quando o ficheiro tem, por exemplo:
    - controlo em vez de Controlo
    - DEFINICOES em vez de DEFINIÇÕES
    - espaços acidentais antes/depois do nome
    """
    wanted = normalizar_texto(name)
    for sheet_name in wb.sheetnames:
        if normalizar_texto(sheet_name) == wanted:
            return sheet_name
    return None


def get_sheet_or_raise(wb, name: str) -> Worksheet:
    real_name = resolver_nome_sheet(wb, name)
    if real_name is None:
        disponiveis = ", ".join(wb.sheetnames)
        raise ValueError(
            f"A sheet obrigatória '{name}' não existe no ficheiro. "
            f"Sheets disponíveis: {disponiveis}"
        )
    return wb[real_name]


def get_or_create_sheet(wb, name: str, headers: list[str]) -> Worksheet:
    real_name = resolver_nome_sheet(wb, name)
    if real_name is not None:
        ws = wb[real_name]
    else:
        ws = wb.create_sheet(name)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = header
    return ws


def clear_below_header(ws: Worksheet) -> None:
    """Limpa apenas linhas abaixo do cabeçalho, preservando a linha 1."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def encontrar_linha_cabecalho(ws: Worksheet, required_headers: Iterable[str], max_scan_rows: int = 15) -> int:
    required = {normalizar_texto(h) for h in required_headers}
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_values = {normalizar_texto(ws.cell(row=row_idx, column=col).value) for col in range(1, ws.max_column + 1)}
        if required.issubset(row_values):
            return row_idx
    raise ValueError(f"Não encontrei cabeçalho {list(required_headers)} na sheet '{ws.title}'.")


def mapa_cabecalhos(ws: Worksheet, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalizar_texto(ws.cell(row=header_row, column=col).value)
        if key and key not in result:
            result[key] = col
    return result


def get_col(header_map: dict[str, int], header_name: str, sheet_name: str) -> int:
    key = normalizar_texto(header_name)
    if key not in header_map:
        raise ValueError(f"Cabeçalho '{header_name}' não encontrado na sheet '{sheet_name}'.")
    return header_map[key]


def get_col_any(header_map: dict[str, int], header_names: list[str], sheet_name: str) -> int:
    for header_name in header_names:
        key = normalizar_texto(header_name)
        if key in header_map:
            return header_map[key]
    raise ValueError(f"Nenhum dos cabeçalhos {header_names} foi encontrado na sheet '{sheet_name}'.")


def escrever_linhas(ws: Worksheet, rows: list[list[Any]]) -> None:
    """Escreve linhas em bloco simples, começando na linha 2."""
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value


# =========================
# 0) Sheet Controlo
# =========================

def ler_departamentos_pendentes(ws_controlo: Worksheet) -> tuple[list[DepartamentoPendente], int, int, int]:
    """
    Lê a sheet Controlo.
    Só retorna departamentos onde STATUS está vazio.
    Cabeçalhos esperados: DEPARTAMENTO, STATUS, TIMESTAMP ou TIMESTEMP.
    """
    header_row = encontrar_linha_cabecalho(ws_controlo, ["DEPARTAMENTO", "STATUS"], max_scan_rows=10)
    headers = mapa_cabecalhos(ws_controlo, header_row)
    col_departamento = get_col(headers, "DEPARTAMENTO", ws_controlo.title)
    col_status = get_col(headers, "STATUS", ws_controlo.title)
    col_timestamp = get_col_any(headers, ["TIMESTAMP", "TIMESTEMP"], ws_controlo.title)

    pendentes: list[DepartamentoPendente] = []
    for row_idx in range(header_row + 1, ws_controlo.max_row + 1):
        departamento = texto_limpo(ws_controlo.cell(row=row_idx, column=col_departamento).value)
        status = texto_limpo(ws_controlo.cell(row=row_idx, column=col_status).value)
        if departamento and not status:
            pendentes.append(DepartamentoPendente(nome=departamento, row_idx=row_idx))

    return pendentes, col_status, col_timestamp, header_row


def atualizar_linha_controlo(ws_controlo: Worksheet, row_idx: int, col_status: int, col_timestamp: int, status: str) -> None:
    ws_controlo.cell(row=row_idx, column=col_status).value = status
    ws_controlo.cell(row=row_idx, column=col_timestamp).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# =========================
# 1) Sheet Proposta -> PFCG_CREATE
# =========================

def ler_blocos_da_proposta(ws_proposta: Worksheet) -> tuple[list[list[Any]], dict[str, list[tuple[str, str]]]]:
    """
    Lê a sheet Proposta em blocos:
    - Linha com FUNÇÃO + DESCRIÇÃO = função/role AGR_NAME + TEXT.
    - Linhas seguintes com FUNÇÃO preenchida e DESCRIÇÃO vazia = TCODE dessa função.
    """
    header_row = encontrar_linha_cabecalho(ws_proposta, ["FUNÇÃO", "DESCRIÇÃO"])
    headers = mapa_cabecalhos(ws_proposta, header_row)
    col_funcao = get_col(headers, "FUNÇÃO", ws_proposta.title)
    col_descricao = get_col(headers, "DESCRIÇÃO", ws_proposta.title)

    current_agr_name = ""
    current_text = ""
    pfcg_rows: list[list[Any]] = []
    tcode_to_roles: dict[str, list[tuple[str, str]]] = defaultdict(list)
    seen_create: set[tuple[str, str]] = set()
    seq = 1

    for row_idx in range(header_row + 1, ws_proposta.max_row + 1):
        funcao = texto_limpo(ws_proposta.cell(row=row_idx, column=col_funcao).value)
        descricao = texto_limpo(ws_proposta.cell(row=row_idx, column=col_descricao).value)

        if not funcao:
            continue

        if funcao and descricao:
            current_agr_name = funcao
            current_text = descricao
            continue

        if funcao and not descricao and current_agr_name:
            tcode = funcao
            create_key = (normalizar_texto(current_agr_name), normalizar_texto(tcode))
            if create_key not in seen_create:
                seen_create.add(create_key)
                pfcg_rows.append([seq, current_agr_name, current_text, tcode])
                seq += 1

            tcode_key = normalizar_texto(tcode)
            pair = (current_agr_name, current_text)
            if pair not in tcode_to_roles[tcode_key]:
                tcode_to_roles[tcode_key].append(pair)

    return pfcg_rows, tcode_to_roles


def atualizar_pfcg_create(wb, pfcg_rows: list[list[Any]]) -> None:
    ws = get_or_create_sheet(wb, SHEET_PFCG_CREATE, PFCG_CREATE_HEADERS)
    clear_below_header(ws)
    escrever_linhas(ws, pfcg_rows)


def coletar_tcodes_usadas(dept_sheet_map: dict[str, dict[str, list[str]]]) -> list[str]:
    """
    Recolhe todas as transações marcadas com X para os utilizadores
    dos departamentos pendentes no Controlo.
    """
    tcodes: list[str] = []
    for users_map in dept_sheet_map.values():
        for user_tcodes in users_map.values():
            tcodes.extend(user_tcodes)
    return unique_keep_order(tcodes)


def filtrar_pfcg_create_por_tcodes_usadas(
    pfcg_rows_base: list[list[Any]],
    tcodes_usadas: list[str],
) -> list[list[Any]]:
    """
    A PFCG_CREATE deve conter apenas as transações que os utilizadores
    dos departamentos pendentes vão utilizar.

    A sheet Proposta continua a ser lida por completo para permitir descobrir
    a relação TCODE -> AGR_NAME, mas a escrita na PFCG_CREATE é filtrada pelas
    transações realmente marcadas com X nas sheets de departamento.
    """
    tcodes_validas = {normalizar_texto(tcode) for tcode in tcodes_usadas if texto_limpo(tcode)}
    rows_filtradas: list[list[Any]] = []
    seen: set[tuple[str, str]] = set()
    seq = 1

    for row in pfcg_rows_base:
        if len(row) < 4:
            continue

        _old_id, agr_name, text, tcode = row[:4]
        if normalizar_texto(tcode) not in tcodes_validas:
            continue

        key = (normalizar_texto(agr_name), normalizar_texto(tcode))
        if key in seen:
            continue

        seen.add(key)
        rows_filtradas.append([seq, agr_name, text, tcode])
        seq += 1

    return rows_filtradas


# =========================
# 2) Sheet de departamento -> transações por utilizador
# =========================

def extrair_transacoes_por_utilizador_de_sheet(ws: Worksheet) -> dict[str, list[str]]:
    """
    Lê uma sheet de departamento com estrutura:
    Transação | Descrição | Sxxx - Nome | Syyy - Nome | ...
    Cada célula com X na coluna do utilizador atribui a transação da linha.
    """
    header_row = encontrar_linha_cabecalho(ws, ["Transação", "Descrição"], max_scan_rows=10)
    headers = mapa_cabecalhos(ws, header_row)
    col_transacao = get_col(headers, "Transação", ws.title)
    col_descricao = get_col(headers, "Descrição", ws.title)

    user_cols: dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        if col_idx in (col_transacao, col_descricao):
            continue
        user_id = user_id_from_header(ws.cell(row=header_row, column=col_idx).value)
        if user_id and normalizar_texto(user_id).startswith("S"):
            user_cols[col_idx] = user_id

    result: dict[str, list[str]] = {user: [] for user in user_cols.values()}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        tcode = texto_limpo(ws.cell(row=row_idx, column=col_transacao).value)
        if not tcode:
            continue

        if normalizar_texto(tcode) in {"TOTAL", "TOTAIS", "TOTAL GERAL"}:
            continue

        for col_idx, user_id in user_cols.items():
            if valor_marcado_com_x(ws.cell(row=row_idx, column=col_idx).value):
                result[user_id].append(tcode)

    return {user: unique_keep_order(tcodes) for user, tcodes in result.items() if tcodes}


def construir_mapa_departamentos_pendentes(wb, pendentes: list[DepartamentoPendente]) -> tuple[dict[str, dict[str, list[str]]], list[DepartamentoPendente]]:
    """Cria mapa apenas para os departamentos pendentes do Controlo."""
    dept_map: dict[str, dict[str, list[str]]] = {}
    departamentos_com_erro: list[DepartamentoPendente] = []

    for item in pendentes:
        real_sheet_name = resolver_nome_sheet(wb, item.nome)
        if real_sheet_name is None:
            departamentos_com_erro.append(item)
            continue
        ws = wb[real_sheet_name]
        try:
            dept_map[item.nome] = extrair_transacoes_por_utilizador_de_sheet(ws)
        except Exception:
            departamentos_com_erro.append(item)

    return dept_map, departamentos_com_erro


# =========================
# 3) Sheet DEFINIÇÕES -> regras por departamento
# =========================

def ler_regras_definicoes(ws_definicoes: Worksheet) -> dict[str, list[str]]:
    """Lê a sheet DEFINIÇÕES e retorna: departamento -> roles/regras adicionais."""
    header_row = encontrar_linha_cabecalho(ws_definicoes, ["DEPARTAMENTO"])
    headers = mapa_cabecalhos(ws_definicoes, header_row)
    col_departamento = get_col(headers, "DEPARTAMENTO", ws_definicoes.title)

    regras: dict[str, list[str]] = {}
    for row_idx in range(header_row + 1, ws_definicoes.max_row + 1):
        departamento = texto_limpo(ws_definicoes.cell(row=row_idx, column=col_departamento).value)
        if not departamento:
            continue

        valores: list[str] = []
        for col_idx in range(1, ws_definicoes.max_column + 1):
            if col_idx == col_departamento:
                continue
            valores.extend(split_valores(ws_definicoes.cell(row=row_idx, column=col_idx).value))

        regras[normalizar_texto(departamento)] = unique_keep_order(valores)

    return regras


# =========================
# 4) Preencher Proposta Ativa apenas para departamentos pendentes
# =========================

def roles_por_transacoes(tcodes: list[str], tcode_to_roles: dict[str, list[tuple[str, str]]]) -> list[str]:
    roles: list[str] = []
    for tcode in tcodes:
        for agr_name, _text in tcode_to_roles.get(normalizar_texto(tcode), []):
            roles.append(agr_name)
    return unique_keep_order(roles)


def limpar_roles_existentes(ws: Worksheet, row_idx: int, start_col: int, clear_until_col: int) -> None:
    for col_idx in range(start_col, clear_until_col + 1):
        ws.cell(row=row_idx, column=col_idx).value = None


def escrever_roles_na_linha(ws: Worksheet, row_idx: int, start_col: int, roles: list[str]) -> None:
    for offset, role in enumerate(roles):
        ws.cell(row=row_idx, column=start_col + offset).value = role


def construir_indice_usuarios_pendentes(dept_sheet_map: dict[str, dict[str, list[str]]]) -> dict[str, dict[str, Any]]:
    """
    Constrói índice:
    usuario -> {departamentos: [...], tcodes: [...]}
    """
    result: dict[str, dict[str, Any]] = {}
    for dept_name, users_map in dept_sheet_map.items():
        for user_id, tcodes in users_map.items():
            key = normalizar_texto(user_id)
            if key not in result:
                result[key] = {"usuario": user_id, "departamentos": [], "tcodes": []}
            result[key]["departamentos"].append(dept_name)
            result[key]["tcodes"].extend(tcodes)

    for data in result.values():
        data["departamentos"] = unique_keep_order(data["departamentos"])
        data["tcodes"] = unique_keep_order(data["tcodes"])

    return result


def preencher_proposta_ativa_para_pendentes(
    ws_ativa: Worksheet,
    dept_sheet_map: dict[str, dict[str, list[str]]],
    tcode_to_roles: dict[str, list[tuple[str, str]]],
    definicoes: dict[str, list[str]],
) -> dict[str, Any]:
    """Preenche roles na Proposta Ativa apenas para utilizadores dos departamentos pendentes."""
    header_row = encontrar_linha_cabecalho(
        ws_ativa,
        ["Usuário", "Departamento/Direção", "Departamento", "Composite Role", "Descrição"],
    )
    headers = mapa_cabecalhos(ws_ativa, header_row)

    col_usuario = get_col(headers, "Usuário", ws_ativa.title)
    col_dept_direcao = get_col(headers, "Departamento/Direção", ws_ativa.title)
    col_departamento = get_col(headers, "Departamento", ws_ativa.title)
    col_composite = get_col(headers, "Composite Role", ws_ativa.title)
    col_descricao = get_col(headers, "Descrição", ws_ativa.title)

    clear_until_col = max(ws_ativa.max_column, col_descricao + 160)
    user_index = construir_indice_usuarios_pendentes(dept_sheet_map)

    linhas_processadas = 0
    usuarios_processados: set[str] = set()
    usuarios_processados_ids: list[str] = []
    composite_roles_processadas: list[str] = []
    usuarios_nao_encontrados = set(user_index.keys())
    departamentos_sem_definicao: list[str] = []
    tcodes_sem_role: set[str] = set()

    for row_idx in range(header_row + 1, ws_ativa.max_row + 1):
        usuario = texto_limpo(ws_ativa.cell(row=row_idx, column=col_usuario).value)
        if not usuario:
            continue

        user_key = normalizar_texto(usuario)
        if user_key not in user_index:
            continue

        usuarios_nao_encontrados.discard(user_key)
        usuarios_processados.add(user_key)
        usuarios_processados_ids.append(usuario)

        dept_direcao = texto_limpo(ws_ativa.cell(row=row_idx, column=col_dept_direcao).value)
        departamento = texto_limpo(ws_ativa.cell(row=row_idx, column=col_departamento).value)
        composite_role = texto_limpo(ws_ativa.cell(row=row_idx, column=col_composite).value)
        if composite_role:
            composite_roles_processadas.append(composite_role)

        tcodes = user_index[user_key]["tcodes"]

        for tcode in tcodes:
            if normalizar_texto(tcode) not in tcode_to_roles:
                tcodes_sem_role.add(tcode)

        roles_funcionais = roles_por_transacoes(tcodes, tcode_to_roles)
        roles_departamento = definicoes.get(normalizar_texto(departamento), [])

        if departamento and normalizar_texto(departamento) not in definicoes:
            departamentos_sem_definicao.append(f"Linha {row_idx}: {usuario} -> {departamento}")

        roles_finais = unique_keep_order(roles_funcionais + roles_departamento)

        limpar_roles_existentes(ws_ativa, row_idx, col_descricao, clear_until_col)
        escrever_roles_na_linha(ws_ativa, row_idx, col_descricao, roles_finais)

        linhas_processadas += 1

    return {
        "linhas_processadas": linhas_processadas,
        "usuarios_processados": len(usuarios_processados),
        "usuarios_processados_ids": unique_keep_order(usuarios_processados_ids),
        "composite_roles_processadas": unique_keep_order(composite_roles_processadas),
        "usuarios_nao_encontrados": [user_index[k]["usuario"] for k in sorted(usuarios_nao_encontrados)],
        "departamentos_sem_definicao": unique_keep_order(departamentos_sem_definicao),
        "tcodes_sem_role": sorted(tcodes_sem_role),
    }


# =========================
# 5) Proposta Ativa -> PFCG_COMPOSTA
# =========================

def role_pode_entrar_na_pfcg_composta(role: str) -> bool:
    if not EXCLUIR_REGRAS_ORG_DA_PFCG_COMPOSTA:
        return True
    role_norm = normalizar_texto(role)
    if role_norm in {normalizar_texto(v) for v in PFCG_COMPOSTA_EXCLUDE_EXACT}:
        return False
    return not any(role_norm.startswith(normalizar_texto(prefix)) for prefix in PFCG_COMPOSTA_EXCLUDE_PREFIXES)


def texto_composite_role(composite_role: str, existing_texts: dict[str, str]) -> str:
    key = normalizar_texto(composite_role)
    if key in existing_texts and existing_texts[key]:
        return existing_texts[key]
    text = composite_role
    text = re.sub(r"^Z_BR_?", "", text, flags=re.IGNORECASE)
    text = text.replace("_", " ")
    return " ".join(text.split()).upper()


def ler_textos_existentes_pfcg_composta(ws: Worksheet | None) -> dict[str, str]:
    if ws is None or ws.max_row < 2:
        return {}
    try:
        header_row = encontrar_linha_cabecalho(ws, ["AGR_NAME_COMPOSTA", "TEXT"], max_scan_rows=5)
        headers = mapa_cabecalhos(ws, header_row)
        col_comp = get_col(headers, "AGR_NAME_COMPOSTA", ws.title)
        col_text = get_col(headers, "TEXT", ws.title)
    except ValueError:
        return {}

    result: dict[str, str] = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        comp = texto_limpo(ws.cell(row=row_idx, column=col_comp).value)
        text = texto_limpo(ws.cell(row=row_idx, column=col_text).value)
        if comp and text:
            result[normalizar_texto(comp)] = text
    return result


def ler_roles_da_proposta_ativa(ws_ativa: Worksheet, row_idx: int, start_col: int) -> list[str]:
    values: list[str] = []
    for col_idx in range(start_col, ws_ativa.max_column + 1):
        values.extend(split_valores(ws_ativa.cell(row=row_idx, column=col_idx).value))
    return unique_keep_order(values)


def gerar_pfcg_composta(
    wb,
    ws_ativa: Worksheet,
    composite_roles_alvo: Iterable[str] | None = None,
) -> list[list[Any]]:
    """
    Gera linhas da PFCG_COMPOSTA a partir da Proposta Ativa.

    Quando composite_roles_alvo é informado, gera apenas as linhas dessas
    Composite Roles. Isto permite atualizar somente o departamento em processamento.
    """
    pfcg_composta_real_name = resolver_nome_sheet(wb, SHEET_PFCG_COMPOSTA)
    existing_ws = wb[pfcg_composta_real_name] if pfcg_composta_real_name else None
    existing_texts = ler_textos_existentes_pfcg_composta(existing_ws)
    alvo_norm = {normalizar_texto(v) for v in composite_roles_alvo or [] if texto_limpo(v)}

    header_row = encontrar_linha_cabecalho(ws_ativa, ["Composite Role", "Descrição"])
    headers = mapa_cabecalhos(ws_ativa, header_row)
    col_composite = get_col(headers, "Composite Role", ws_ativa.title)
    col_descricao = get_col(headers, "Descrição", ws_ativa.title)

    rows: list[list[Any]] = []
    seen: set[tuple[str, str]] = set()
    seq = 1

    for row_idx in range(header_row + 1, ws_ativa.max_row + 1):
        composite = texto_limpo(ws_ativa.cell(row=row_idx, column=col_composite).value)
        if not composite:
            continue
        if alvo_norm and normalizar_texto(composite) not in alvo_norm:
            continue

        roles = ler_roles_da_proposta_ativa(ws_ativa, row_idx, col_descricao)
        for role in roles:
            if not role_pode_entrar_na_pfcg_composta(role):
                continue
            key = (normalizar_texto(composite), normalizar_texto(role))
            if key in seen:
                continue
            seen.add(key)
            rows.append([seq, composite, texto_composite_role(composite, existing_texts), role, "", "", ""])
            seq += 1

    return rows


def ler_linhas_existentes_pfcg_composta(ws: Worksheet) -> list[list[Any]]:
    """Lê as linhas atuais da PFCG_COMPOSTA no formato padrão."""
    if ws.max_row < 2:
        return []
    header_row = encontrar_linha_cabecalho(ws, ["AGR_NAME_COMPOSTA", "TEXT", "AGR_NAME"], max_scan_rows=5)
    headers = mapa_cabecalhos(ws, header_row)
    col_comp = get_col(headers, "AGR_NAME_COMPOSTA", ws.title)
    col_text = get_col(headers, "TEXT", ws.title)
    col_agr = get_col(headers, "AGR_NAME", ws.title)
    col_status = headers.get(normalizar_texto("STATUS"), 0)
    col_msg = headers.get(normalizar_texto("MSG"), 0)
    col_timestamp = headers.get(normalizar_texto("TIMESTEMP"), headers.get(normalizar_texto("TIMESTAMP"), 0))

    rows: list[list[Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        comp = texto_limpo(ws.cell(row=row_idx, column=col_comp).value)
        agr = texto_limpo(ws.cell(row=row_idx, column=col_agr).value)
        if not comp or not agr:
            continue
        text = texto_limpo(ws.cell(row=row_idx, column=col_text).value)
        status = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        msg = ws.cell(row=row_idx, column=col_msg).value if col_msg else ""
        timestamp = ws.cell(row=row_idx, column=col_timestamp).value if col_timestamp else ""
        rows.append([0, comp, text, agr, status or "", msg or "", timestamp or ""])
    return rows


def resequenciar_e_deduplicar_pfcg_composta(rows: list[list[Any]]) -> list[list[Any]]:
    result: list[list[Any]] = []
    seen: set[tuple[str, str]] = set()
    seq = 1
    for row in rows:
        if len(row) < 4:
            continue
        _id, comp, text, agr = row[:4]
        if not texto_limpo(comp) or not texto_limpo(agr):
            continue
        key = (normalizar_texto(comp), normalizar_texto(agr))
        if key in seen:
            continue
        seen.add(key)
        status = row[4] if len(row) > 4 else ""
        msg = row[5] if len(row) > 5 else ""
        timestamp = row[6] if len(row) > 6 else ""
        result.append([seq, comp, text, agr, status, msg, timestamp])
        seq += 1
    return result


def atualizar_pfcg_composta(
    wb,
    rows_novas: list[list[Any]],
    composite_roles_alvo: Iterable[str] | None = None,
) -> list[list[Any]]:
    """
    Atualiza a PFCG_COMPOSTA de forma incremental.

    - Remove apenas as linhas das Composite Roles processadas agora.
    - Mantém as Composite Roles dos outros departamentos.
    - Acrescenta as linhas novas e reordena o ID.
    """
    ws = get_or_create_sheet(wb, SHEET_PFCG_COMPOSTA, PFCG_COMPOSTA_HEADERS)
    alvo_norm = {normalizar_texto(v) for v in composite_roles_alvo or [] if texto_limpo(v)}

    if not alvo_norm:
        rows_finais = resequenciar_e_deduplicar_pfcg_composta(ler_linhas_existentes_pfcg_composta(ws))
    else:
        rows_existentes = ler_linhas_existentes_pfcg_composta(ws)
        rows_preservadas = [
            row for row in rows_existentes
            if normalizar_texto(row[1]) not in alvo_norm
        ]
        rows_finais = resequenciar_e_deduplicar_pfcg_composta(rows_preservadas + rows_novas)

    clear_below_header(ws)
    escrever_linhas(ws, rows_finais)
    return rows_finais


# =========================
# 6) Proposta Ativa -> CUA_ADICIONAR
# =========================

def gerar_cua_adicionar(
    ws_ativa: Worksheet,
    usuarios_alvo: Iterable[str] | None = None,
    definicoes: dict[str, list[str]] | None = None,
) -> list[list[Any]]:
    """
    Gera linhas da CUA_ADICIONAR a partir da Proposta Ativa.

    Quando usuarios_alvo é informado, gera apenas as linhas desses utilizadores.
    """
    if definicoes is None:
        definicoes = {}

    if usuarios_alvo is not None:
        usuarios_list = list(usuarios_alvo)
        if not usuarios_list:
            return []
        alvo_norm = {normalizar_texto(v) for v in usuarios_list if texto_limpo(v)}
        if not alvo_norm:
            return []
    else:
        alvo_norm = set()

    header_row = encontrar_linha_cabecalho(ws_ativa, ["Usuário", "Departamento", "Composite Role"])
    headers = mapa_cabecalhos(ws_ativa, header_row)
    col_usuario = get_col(headers, "Usuário", ws_ativa.title)
    col_departamento = get_col(headers, "Departamento", ws_ativa.title)
    col_composite = get_col(headers, "Composite Role", ws_ativa.title)

    rows: list[list[Any]] = []
    seen: set[tuple[str, str, str]] = set()
    seq = 1

    for row_idx in range(header_row + 1, ws_ativa.max_row + 1):
        usuario = texto_limpo(ws_ativa.cell(row=row_idx, column=col_usuario).value)
        departamento = texto_limpo(ws_ativa.cell(row=row_idx, column=col_departamento).value)
        composite = texto_limpo(ws_ativa.cell(row=row_idx, column=col_composite).value)
        if not usuario or not composite:
            continue
        if alvo_norm and normalizar_texto(usuario) not in alvo_norm:
            continue

        roles_definicoes = definicoes.get(normalizar_texto(departamento), []) if definicoes else []
        roles_cua = unique_keep_order([composite] + roles_definicoes)
        for role in roles_cua:
            key = (normalizar_texto(usuario), normalizar_texto(SISTEMA_FIXO), normalizar_texto(role))
            if key in seen:
                continue
            seen.add(key)
            rows.append([seq, usuario, SISTEMA_FIXO, role, "", "", ""])
            seq += 1

    return rows


def ler_linhas_existentes_cua_adicionar(ws: Worksheet) -> list[list[Any]]:
    """Lê as linhas atuais da CUA_ADICIONAR no formato padrão."""
    if ws.max_row < 2:
        return []
    header_row = encontrar_linha_cabecalho(ws, ["UTILIZADOR", "SISTEMA", "AGR_NAME"], max_scan_rows=5)
    headers = mapa_cabecalhos(ws, header_row)
    col_user = get_col(headers, "UTILIZADOR", ws.title)
    col_sistema = get_col(headers, "SISTEMA", ws.title)
    col_agr = get_col(headers, "AGR_NAME", ws.title)
    col_status = headers.get(normalizar_texto("STATUS"), 0)
    col_msg = headers.get(normalizar_texto("MSG"), 0)
    col_timestamp = headers.get(normalizar_texto("TIMESTEMP"), headers.get(normalizar_texto("TIMESTAMP"), 0))

    rows: list[list[Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        usuario = texto_limpo(ws.cell(row=row_idx, column=col_user).value)
        sistema = texto_limpo(ws.cell(row=row_idx, column=col_sistema).value)
        agr = texto_limpo(ws.cell(row=row_idx, column=col_agr).value)
        if not usuario or not sistema or not agr:
            continue
        status = ws.cell(row=row_idx, column=col_status).value if col_status else ""
        msg = ws.cell(row=row_idx, column=col_msg).value if col_msg else ""
        timestamp = ws.cell(row=row_idx, column=col_timestamp).value if col_timestamp else ""
        rows.append([0, usuario, sistema, agr, status or "", msg or "", timestamp or ""])
    return rows


def resequenciar_e_deduplicar_cua(rows: list[list[Any]]) -> list[list[Any]]:
    result: list[list[Any]] = []
    seen: set[tuple[str, str, str]] = set()
    seq = 1
    for row in rows:
        if len(row) < 4:
            continue
        _id, usuario, sistema, agr = row[:4]
        if not texto_limpo(usuario) or not texto_limpo(sistema) or not texto_limpo(agr):
            continue
        key = (normalizar_texto(usuario), normalizar_texto(sistema), normalizar_texto(agr))
        if key in seen:
            continue
        seen.add(key)
        status = row[4] if len(row) > 4 else ""
        msg = row[5] if len(row) > 5 else ""
        timestamp = row[6] if len(row) > 6 else ""
        result.append([seq, usuario, sistema, agr, status, msg, timestamp])
        seq += 1
    return result


def atualizar_cua_adicionar(
    wb,
    rows_novas: list[list[Any]],
    usuarios_alvo: Iterable[str] | None = None,
) -> list[list[Any]]:
    """
    Atualiza a CUA_ADICIONAR de forma incremental.

    - Remove apenas as linhas dos utilizadores processados agora.
    - Mantém os utilizadores dos outros departamentos.
    - Acrescenta as linhas novas e reordena o ID.
    """
    ws = get_or_create_sheet(wb, SHEET_CUA_ADICIONAR, CUA_ADICIONAR_HEADERS)
    alvo_norm = {normalizar_texto(v) for v in usuarios_alvo or [] if texto_limpo(v)}

    if not alvo_norm:
        rows_finais = resequenciar_e_deduplicar_cua(ler_linhas_existentes_cua_adicionar(ws))
    else:
        rows_existentes = ler_linhas_existentes_cua_adicionar(ws)
        rows_preservadas = [
            row for row in rows_existentes
            if normalizar_texto(row[1]) not in alvo_norm
        ]
        rows_finais = resequenciar_e_deduplicar_cua(rows_preservadas + rows_novas)

    clear_below_header(ws)
    escrever_linhas(ws, rows_finais)
    return rows_finais


def atualizar_cua_adicionar_rebuild_total(
    wb,
    rows_novas: list[list[Any]],
) -> list[list[Any]]:
    """
    Reconstrói totalmente a CUA_ADICIONAR.

    - Obtém ou cria a sheet CUA_ADICIONAR.
    - Deduplica as rows_novas.
    - Limpa toda a sheet abaixo do cabeçalho.
    - Escreve as novas linhas sem preservar as antigas.
    """
    ws = get_or_create_sheet(wb, SHEET_CUA_ADICIONAR, CUA_ADICIONAR_HEADERS)
    rows_finais = resequenciar_e_deduplicar_cua(rows_novas)
    clear_below_header(ws)
    escrever_linhas(ws, rows_finais)
    return rows_finais


# =========================
# Processo principal
# =========================

def processar_ficheiro(input_path: Path) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"Ficheiro não encontrado: {input_path}")

    try:
        wb = load_workbook(input_path)
    except PermissionError as exc:
        raise PermissionError(
            "Sem permissão para abrir o ficheiro. Feche o Excel, confirme que o ficheiro não está bloqueado "
            "pelo OneDrive e tente usar uma cópia local, por exemplo C:\\Temp.\n\n"
            f"Detalhe: {exc}"
        ) from exc

    ws_controlo = get_sheet_or_raise(wb, SHEET_CONTROLO)
    ws_proposta = get_sheet_or_raise(wb, SHEET_PROPOSTA)
    ws_ativa = get_sheet_or_raise(wb, SHEET_PROPOSTA_ATIVA)
    ws_definicoes = get_sheet_or_raise(wb, SHEET_DEFINICOES)

    print("0/7 A ler sheet Controlo e departamentos pendentes...")
    pendentes, col_status, col_timestamp, _header_row = ler_departamentos_pendentes(ws_controlo)
    if not pendentes:
        raise ValueError("Não existem departamentos pendentes na sheet Controlo. Preencha DEPARTAMENTO e deixe STATUS vazio para processar.")

    print("Departamentos pendentes:")
    for item in pendentes:
        print(f"- {item.nome}")

    print("\n1/8 A ler Proposta como base de funções e transações...")
    pfcg_create_rows_base, tcode_to_roles = ler_blocos_da_proposta(ws_proposta)

    print("2/8 A ler apenas sheets dos departamentos pendentes...")
    dept_sheet_map, departamentos_com_erro = construir_mapa_departamentos_pendentes(wb, pendentes)

    if not dept_sheet_map:
        for item in departamentos_com_erro:
            atualizar_linha_controlo(ws_controlo, item.row_idx, col_status, col_timestamp, STATUS_ERRO)
        raise ValueError("Nenhum departamento pendente conseguiu ser lido. Verifique se o nome em Controlo é igual ao nome da sheet.")

    print("3/8 A gerar PFCG_CREATE apenas com transações usadas pelos departamentos pendentes...")
    tcodes_usadas = coletar_tcodes_usadas(dept_sheet_map)
    pfcg_create_rows = filtrar_pfcg_create_por_tcodes_usadas(pfcg_create_rows_base, tcodes_usadas)
    atualizar_pfcg_create(wb, pfcg_create_rows)

    print("4/8 A ler regras da sheet DEFINIÇÕES...")
    definicoes = ler_regras_definicoes(ws_definicoes)

    print("5/8 A preencher Proposta Ativa apenas para utilizadores dos departamentos pendentes...")
    stats_ativa = preencher_proposta_ativa_para_pendentes(ws_ativa, dept_sheet_map, tcode_to_roles, definicoes)

    print("6/8 A atualizar PFCG_COMPOSTA apenas para as Composite Roles dos departamentos pendentes...")
    composite_roles_alvo = stats_ativa["composite_roles_processadas"]
    pfcg_composta_rows_novas = gerar_pfcg_composta(wb, ws_ativa, composite_roles_alvo)
    pfcg_composta_rows = atualizar_pfcg_composta(wb, pfcg_composta_rows_novas, composite_roles_alvo)

    print("7/8 A atualizar CUA_ADICIONAR (reconstrução total)...")
    usuarios_alvo = stats_ativa["usuarios_processados_ids"]
    cua_rows_novas = gerar_cua_adicionar(ws_ativa, usuarios_alvo=usuarios_alvo, definicoes=definicoes)
    cua_rows = atualizar_cua_adicionar_rebuild_total(wb, cua_rows_novas)

    print("8/8 A atualizar Controlo...")
    departamentos_ok = {normalizar_texto(nome) for nome in dept_sheet_map.keys()}
    departamentos_erro = {normalizar_texto(item.nome) for item in departamentos_com_erro}
    for item in pendentes:
        item_key = normalizar_texto(item.nome)
        if item_key in departamentos_erro:
            atualizar_linha_controlo(ws_controlo, item.row_idx, col_status, col_timestamp, STATUS_ERRO)
        elif item_key in departamentos_ok:
            atualizar_linha_controlo(ws_controlo, item.row_idx, col_status, col_timestamp, STATUS_PROCESSADO)

    output_path = output_path_for(input_path)
    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise PermissionError(
            "Sem permissão para guardar o ficheiro processado. Feche o Excel/OneDrive ou execute numa pasta local, "
            "por exemplo C:\\Temp.\n\n"
            f"Tentativa de gravação: {output_path}\n"
            f"Detalhe: {exc}"
        ) from exc

    print("\nProcessamento concluído.")
    print(f"Ficheiro de saída: {output_path}")
    print("\nResumo:")
    print(f"- Departamentos pendentes no Controlo: {len(pendentes)}")
    print(f"- Departamentos processados: {len(dept_sheet_map)}")
    print(f"- Departamentos com erro: {len(departamentos_com_erro)}")
    print(f"- Transações únicas usadas pelos departamentos pendentes: {len(tcodes_usadas)}")
    print(f"- PFCG_CREATE base lida da Proposta: {len(pfcg_create_rows_base)} linhas")
    print(f"- PFCG_CREATE gerada/filtrada: {len(pfcg_create_rows)} linhas")
    print(f"- Departamentos com regras em DEFINIÇÕES: {len(definicoes)}")
    print(f"- Linhas atualizadas na Proposta Ativa: {stats_ativa['linhas_processadas']}")
    print(f"- Utilizadores processados: {stats_ativa['usuarios_processados']}")
    print(f"- Composite Roles atualizadas na PFCG_COMPOSTA: {len(composite_roles_alvo)}")
    print(f"- PFCG_COMPOSTA novas linhas para o processamento atual: {len(pfcg_composta_rows_novas)}")
    print(f"- PFCG_COMPOSTA total após atualização incremental: {len(pfcg_composta_rows)} linhas")
    print(f"- Utilizadores atualizados na CUA_ADICIONAR: {len(usuarios_alvo)}")
    print(f"- CUA_ADICIONAR novas linhas para o processamento atual: {len(cua_rows_novas)}")
    print(f"- CUA_ADICIONAR total após atualização incremental: {len(cua_rows)} linhas")

    if departamentos_com_erro:
        print("\nAtenção - Departamentos do Controlo com erro de leitura ou sheet inexistente:")
        for item in departamentos_com_erro:
            print(f"  - {item.nome}")

    if stats_ativa["usuarios_nao_encontrados"]:
        print("\nAtenção - Utilizadores encontrados na sheet do departamento, mas não encontrados na Proposta Ativa:")
        for user_id in stats_ativa["usuarios_nao_encontrados"][:50]:
            print(f"  - {user_id}")
        if len(stats_ativa["usuarios_nao_encontrados"]) > 50:
            print(f"  ... +{len(stats_ativa['usuarios_nao_encontrados']) - 50} utilizadores")

    if stats_ativa["departamentos_sem_definicao"]:
        print("\nAtenção - Departamento sem regra na sheet DEFINIÇÕES:")
        for item in stats_ativa["departamentos_sem_definicao"][:30]:
            print(f"  - {item}")
        if len(stats_ativa["departamentos_sem_definicao"]) > 30:
            print(f"  ... +{len(stats_ativa['departamentos_sem_definicao']) - 30} ocorrências")

    if stats_ativa["tcodes_sem_role"]:
        print("\nAtenção - Transações sem função encontrada na sheet Proposta:")
        for tcode in stats_ativa["tcodes_sem_role"]:
            print(f"  - {tcode}")

    return output_path


def main() -> None:
    if len(sys.argv) > 1:
        input_path = Path(sys.argv[1])
    else:
        input_path = escolher_ficheiro_excel()

    try:
        output_path = processar_ficheiro(input_path)
        mostrar_mensagem("Processamento concluído", f"Ficheiro criado:\n{output_path}")
    except Exception as exc:
        mostrar_mensagem("Erro no processamento", str(exc))
        raise


if __name__ == "__main__":
    main()
