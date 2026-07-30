# -*- coding: utf-8 -*-

###################################################################################
# PROCESSO: Criar Utilizador CUA / SU01 (por cópia ou criação direta)
# Ex.: "L. CUA_CRIAR_USER.py"  →  Sheet "CUA_CRIAR_USER"
#
# ESTRUTURA ESPERADA DA SHEET:
# ID | UTILIZADOR | REF_USER | FIRST_NAME | LAST_NAME | EMAIL | FUNCTION | DEPARTMENT | MOB_NUMBER | STATUS | MSG | TIMESTEMP
#
# PADRÃO APLICADO:
# - Prefixo 'S' automático no ID do utilizador (ex: 80002000 -> S80002000)
# - Criação por cópia (SU01 btn[17] Shift+F6) se REF_USER preenchido
# - Criação direta (SU01 btn[8]) se REF_USER vazio
# - Preenchimento dos 6 campos CUA no tabpADDR (Nome, Sobrenome, Email, Função, Depto, Telefone)
# - Leitura e gravação com preservação de formatação Excel via openpyxl
###################################################################################

import sys
import os
import time
import unicodedata
from datetime import datetime

import pandas as pd
import win32com.client
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# Carregar variáveis de ambiente .env
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# Configurar stdout/stderr para UTF-8 no Windows
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# Tentar importar consulta de dados RH via RFC
try:
    from sap_script_web_cockpit_v2.worker.hr_data_analysis import search_hr_user_data_rfc
except ImportError:
    try:
        sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
        from sap_script_web_cockpit_v2.worker.hr_data_analysis import search_hr_user_data_rfc
    except Exception:
        search_hr_user_data_rfc = None

###################################################################################
# BLOCO 2: NOME DO SCRIPT / SHEET / MAPA DE SISTEMAS
###################################################################################

try:
    NOME_SCRIPT = os.path.splitext(os.path.basename(__file__))[0]
except NameError:
    NOME_SCRIPT = "L. CUA_CRIAR_USER"

NOME_SHEET = NOME_SCRIPT.split(".", 1)[-1].strip() if "." in NOME_SCRIPT else NOME_SCRIPT

MAPA_SISTEMA = {
    "DEV": "S4D",
    "QAD": "S4Q",
    "PRD": "S4P",
    "CUA": "SPA",
}

WEB_CONFIG = {
    "manages_own_session": True,
}

###################################################################################
# BLOCO 3: UTILITÁRIOS & FORMATAÇÃO S8000XXXX
###################################################################################

def formatar_tempo(segundos):
    m = int(segundos // 60)
    s = int(segundos % 60)
    return f"{m:02d}m {s:02d}s"


def agora_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalizar_coluna(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("utf-8")
        .strip()
        .upper()
    )


def normalizar_valor(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("utf-8")
        .strip()
        .upper()
    )


def texto_limpo(valor):
    if pd.isna(valor):
        return ""
    txt = str(valor).strip()
    if txt.lower() in ("nan", "none", "<na>"):
        return ""
    return txt


def format_sap_user_id(user_val: str) -> str:
    """
    Garante o prefixo 'S' no número de colaborador / utilizador SAP CUA (ex: 80002000 -> S80002000).
    Utilizadores alfanuméricos reais (ex: CSILVA) não recebem o prefixo S.
    """
    if not user_val:
        return ""
    clean = str(user_val).strip().upper()
    if not clean:
        return ""
    if clean.startswith("S"):
        return clean
    if clean.isdigit():
        clean = clean.lstrip("0")
        return f"S{clean}"
    return clean


def buscar_dados_rh_rfc(target_user: str) -> tuple[str, str, str]:
    """
    Busca automaticamente Nome (PA0002-VORNA), Sobrenome (PA0002-NACHN) e Email (PA0105)
    via RFC no sistema produtivo S4P a partir do ID do utilizador / colaborador.
    """
    first_name = ""ç
    last_name = ""
    email = ""

    if not search_hr_user_data_rfc or not target_user:
        return first_name, last_name, email

    try:
        query_clean = str(target_user).strip()
        print(f"\n🔍 A pesquisar dados do colaborador {query_clean} no RH (S4P via RFC)...")
        res = search_hr_user_data_rfc(query=query_clean, target_system_key="S4PCLNT100")
        if res.get("success") and res.get("data"):
            item = res["data"][0]
            first_name = item.get("first_name", "")
            last_name = item.get("last_name", "")
            email = item.get("email", "")
            print(f"✅ Dados obtidos no RH (S4P):")
            print(f"   • CUA-NAME_FIRST: {first_name} (PA0002-VORNA)")
            print(f"   • CUA-NAME_LAST: {last_name} (PA0002-NACHN)")
            print(f"   • CUA-SMTP_ADDR: {email} (PA0002/PA0105-USRID_LONG)")
        else:
            print("⚠️ Nenhum registo encontrado no RH (S4P). Pode preencher manualmente.")
    except Exception as e:
        print(f"⚠️ Não foi possível consultar o RH via RFC: {e}")

    return first_name, last_name, email


###################################################################################
# BLOCO 4: CONEXÃO E POSICIONAMENTO SAP GUI
###################################################################################

def conectar_sap(sistema_desejado, session_override=None):
    """
    Obtém uma sessão ativa do SAP GUI para o sistema desejado (ex: SPA / S4P / S4D).
    1. Se uma session válida for fornecida via session_override, utiliza-a.
    2. Procura nas conexões ativas do SAP GUI.
    3. Se não encontrar, tenta auto-login via obter_sessao_sap().
    """
    if session_override:
        try:
            sys_name = texto_limpo(getattr(session_override.Info, "SystemName", "")).upper()
            if sys_name == sistema_desejado.upper() or not sistema_desejado:
                print(f"✅ Utilizando sessão SAP fornecida: {sys_name}")
                return session_override
        except Exception:
            pass

    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application = sap_gui_auto.GetScriptingEngine

        for conn in application.Children:
            for sess in conn.Children:
                try:
                    if texto_limpo(sess.Info.SystemName).upper() == sistema_desejado.upper():
                        print(
                            f"✅ Conectado ao SAP: {sess.Info.SystemName} "
                            f"| Utilizador: {sess.Info.User} "
                            f"| Mandante: {sess.Info.Client}"
                        )
                        return sess
                except Exception:
                    continue
    except Exception:
        pass

    # 3. Fallback: tentar auto-iniciar sessão no SAP GUI se o SAP Logon não estiver aberto
    try:
        print(f"🔑 A iniciar ligação automática no SAP GUI ({sistema_desejado})...")
        try:
            from sap_script_web_cockpit_v2.sap_cockpit_web_ready import obter_sessao_sap
        except ImportError:
            try:
                sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
                from sap_script_web_cockpit_v2.sap_cockpit_web_ready import obter_sessao_sap
            except Exception:
                obter_sessao_sap = None

        if obter_sessao_sap:
            amb_key = "CUA" if sistema_desejado.upper() in ("SPA", "CUA") else sistema_desejado.upper()
            sess, _ = obter_sessao_sap(amb_key, interactive=False)
            if sess:
                print(f"✅ Sessão SAP iniciada com sucesso via auto-login ({amb_key}).")
                return sess
    except Exception as exc_auto:
        print(f"⚠️ Não foi possível auto-iniciar a sessão no SAP GUI: {exc_auto}")

    print(f"❌ Sessão SAP GUI não encontrada para o sistema {sistema_desejado}.")
    return None


def posicionar_sap_meia_tela_direita(session) -> bool:
    try:
        import win32api  # type: ignore
        import win32con  # type: ignore
        import win32gui  # type: ignore
    except Exception:
        return False

    try:
        wnd0 = session.findById("wnd[0]")
        hwnd = int(getattr(wnd0, "Handle"))
        if not hwnd:
            return False

        screen_w = int(win32api.GetSystemMetrics(0))
        screen_h = int(win32api.GetSystemMetrics(1))
        half_w = screen_w // 2
        target_w = max(900, half_w)
        target_h = max(700, screen_h)
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.MoveWindow(hwnd, half_w, 0, target_w, target_h, True)
        return True
    except Exception:
        return False


def ir_para_transacao(session, tcode):
    session.findById("wnd[0]/tbar[0]/okcd").text = f"/N{tcode}"
    session.findById("wnd[0]").sendVKey(0)


def capturar_status_bar(session, tentativas=3, espera=0.3):
    """
    Captura a mensagem do StatusBar (wnd[0]/sbar) e verifica pop-ups de erro genuínos em wnd[1].
    """
    for _ in range(tentativas):
        # 1. Verificar StatusBar principal (wnd[0]/sbar)
        try:
            sbar = session.findById("wnd[0]/sbar")
            msg_type = texto_limpo(getattr(sbar, "MessageType", ""))
            msg_text = texto_limpo(getattr(sbar, "Text", ""))
            if msg_type or msg_text:
                print(f"📢 [SAP StatusBar] ({msg_type}) {msg_text}")
                return msg_type, msg_text
        except Exception:
            pass

        # 2. Verificar Janela Pop-up de erro/aviso genuína em wnd[1]
        try:
            wnd1 = session.findById("wnd[1]")
            if wnd1:
                w_title = texto_limpo(getattr(wnd1, "Text", "")).lower()
                # Ignorar a janela de diálogo normal do Copy Users
                if "copy users" not in w_title and "copiar utilizador" not in w_title and "copia de usuario" not in w_title:
                    msg_inner = ""
                    try:
                        for child in wnd1.usr.Children:
                            t = texto_limpo(getattr(child, "Text", ""))
                            if t and t not in ("OK", "Cancelar", "Continuar", "Sim", "Não"):
                                msg_inner += " " + t
                    except Exception:
                        pass
                    full_msg = f"{getattr(wnd1, 'Text', '')}: {msg_inner.strip()}" if msg_inner else getattr(wnd1, 'Text', '')
                    print(f"📢 [SAP Pop-up] (E) {full_msg}")
                    return "E", full_msg
        except Exception:
            pass

        time.sleep(espera)

    return "", ""


def tratar_popups_pos_save(session, max_popups=3):
    for _ in range(max_popups):
        try:
            if not session.findById("wnd[1]"):
                break
            if session.findById("wnd[1]/tbar[0]/btn[0]"):
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
            elif session.findById("wnd[1]/tbar[0]/btn[11]"):
                session.findById("wnd[1]/tbar[0]/btn[11]").press()
            else:
                session.findById("wnd[1]").sendVKey(0)
            time.sleep(0.3)
        except Exception:
            break


###################################################################################
# BLOCO 5: ROTINA DE CRIAÇÃO SU01 (DIRETA E POR CÓPIA)
###################################################################################

def criar_utilizador_su01(
    session,
    target_user: str,
    reference_user: str = "",
    first_name: str = "",
    last_name: str = "",
    email: str = "",
    function: str = "",
    department: str = "",
    mob_number: str = "",
    language: str = "PT",
    initial_password: str = ""
) -> tuple[bool, str]:
    """
    Cria um utilizador CUA na transação SU01 via SAP GUI Scripting.
    Suporta criação por cópia (se reference_user fornecido) ou criação direta.
    """
    target_user_fmt = format_sap_user_id(target_user)
    ref_user_fmt = str(reference_user).strip().upper() if reference_user else ""

    if not target_user_fmt:
        return False, "Utilizador alvo inválido/vazio."

    try:
        # 1. Navegar para SU01
        ir_para_transacao(session, "SU01")
        time.sleep(0.6)

        if ref_user_fmt:
            # 2A. Criação Por Cópia (btn[17] / Shift+F6)
            print(f"📋 A executar criação por cópia na SU01: {target_user_fmt} (Referência: {ref_user_fmt})...")
            
            try:
                session.findById("wnd[0]/usr/txtSUID_ST_BNAME-BNAME").text = ref_user_fmt
            except Exception:
                pass

            try:
                session.findById("wnd[0]/tbar[1]/btn[17]").press()
            except Exception as e_btn:
                return False, f"Falha ao premir botão Copiar (btn[17]) na SU01: {e_btn}"
            time.sleep(0.5)

            # Preencher os campos de Origem e Destino no Pop-up de Cópia (wnd[1])
            try:
                session.findById("wnd[1]/usr/txtGV_COPY_UNAME_SRC").text = ref_user_fmt
                session.findById("wnd[1]/usr/txtGV_COPY_UNAME_DST").text = target_user_fmt
                time.sleep(0.3)
            except Exception as e_pop:
                return False, f"Falha ao preencher os campos do pop-up de cópia (wnd[1]): {e_pop}"

            # Premir botão de confirmação no Pop-up (btn[5], btn[0], btn[8] ou Enter)
            btn_confirmed = False
            for btn_path in ("wnd[1]/tbar[0]/btn[5]", "wnd[1]/tbar[0]/btn[0]", "wnd[1]/tbar[0]/btn[8]"):
                try:
                    session.findById(btn_path).press()
                    btn_confirmed = True
                    print(f"✅ Confirmada cópia via botão: {btn_path}")
                    break
                except Exception:
                    continue

            if not btn_confirmed:
                try:
                    session.findById("wnd[1]").sendVKey(0)
                    btn_confirmed = True
                    print("✅ Confirmada cópia via sendVKey(0)")
                except Exception as e_vkey:
                    return False, f"Não foi possível premir o botão de confirmação no pop-up: {e_vkey}"

            time.sleep(0.8)

            # Verificar se ocorreu algum erro real no SAP após confirmar a cópia
            m_type, m_txt = capturar_status_bar(session, tentativas=3)
            if m_type in ("E", "A"):
                return False, f"Erro SAP ao confirmar cópia ({ref_user_fmt} -> {target_user_fmt}): {m_txt}"

        else:
            # 2B. Criação Direta (btn[8] / Criar)
            print(f"👤 A executar criação direta na SU01: {target_user_fmt}...")
            try:
                session.findById("wnd[0]/usr/txtSUID_ST_BNAME-BNAME").text = target_user_fmt
                session.findById("wnd[0]/tbar[1]/btn[8]").press()  # Criar (F8)
            except Exception as e_create:
                return False, f"Falha ao iniciar criação direta do utilizador {target_user_fmt} na SU01: {e_create}"
            time.sleep(0.8)

        # 3. Preencher Aba Endereço (tabpADDR)
        try:
            session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR").select()
        except Exception:
            pass

        if last_name:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/txtSUID_ST_NODE_PERSON_NAME-NAME_LAST").text = str(last_name).strip()
            except Exception:
                pass

        if first_name:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/txtSUID_ST_NODE_PERSON_NAME-NAME_FIRST").text = str(first_name).strip()
            except Exception:
                pass

        if language:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/cmbSUID_ST_NODE_PERSON_NAME-LANGU").key = str(language).strip()
            except Exception:
                pass

        if function:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/txtSUID_ST_NODE_WORKPLACE-FUNCTION").text = str(function).strip()
            except Exception:
                pass

        if department:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/txtSUID_ST_NODE_WORKPLACE-DEPARTMENT").text = str(department).strip()
            except Exception:
                pass

        if mob_number:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/txtSUID_ST_NODE_COMM_DATA-MOB_NUMBER").text = str(mob_number).strip()
            except Exception:
                pass

        if email:
            try:
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpADDR/ssubMAINAREA:SAPLSUID_MAINTENANCE:1900/txtSUID_ST_NODE_COMM_DATA-SMTP_ADDR").text = str(email).strip()
            except Exception:
                pass

        try:
            session.findById("wnd[0]").sendVKey(0)
        except Exception:
            pass
        time.sleep(0.4)

        # 4. Preencher Aba Dados de Logon (tabpLOGO) - Definir Senha Inicial Obrigatória
        pwd_final = initial_password or os.environ.get("SAP_PASSWORD_RESET", "Acesso.2026")
        if pwd_final:
            try:
                print(f"🔑 A definir senha inicial no separador Logon Data (tabpLOGO)...")
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpLOGO").select()
                time.sleep(0.3)
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpLOGO/ssubMAINAREA:SAPLSUID_MAINTENANCE:1101/pwdSUID_ST_NODE_PASSWORD_EXT-PASSWORD").text = str(pwd_final).strip()
                session.findById("wnd[0]/usr/tabsTABSTRIP1/tabpLOGO/ssubMAINAREA:SAPLSUID_MAINTENANCE:1101/pwdSUID_ST_NODE_PASSWORD_EXT-PASSWORD2").text = str(pwd_final).strip()
            except Exception as e_pwd:
                print(f"[AVISO] Não foi possível definir a senha inicial no tabpLOGO: {e_pwd}")

        # 5. Gravar / Salvar (btn[11] / Ctrl+S)
        try:
            session.findById("wnd[0]/tbar[0]/btn[11]").press()
        except Exception as e_save:
            return False, f"Falha ao premir botão Gravar (btn[11]): {e_save}"
        time.sleep(0.8)

        # 6. Capturar mensagens da StatusBar / Popups / Estado "Not saved"
        tratar_popups_pos_save(session)
        msg_type, msg_text = capturar_status_bar(session, tentativas=5)

        # Verificar se o SAP recusou a gravação ("Not saved")
        try:
            for obj_id in ("wnd[0]/usr/txtSUID_ST_BNAME-STATUS", "wnd[0]/usr/txtSUID_ST_BNAME-STATUS_TEXT", "wnd[0]/sbar"):
                try:
                    st_val = texto_limpo(session.findById(obj_id).Text).lower()
                    if "not saved" in st_val or "não gravado" in st_val:
                        return False, f"Erro SAP: Utilizador não foi gravado ({msg_text or 'Aviso de validação no SAP'})."
                except Exception:
                    pass
        except Exception:
            pass

        modo_desc = f"por cópia de {ref_user_fmt}" if ref_user_fmt else "direta"
        if msg_type in ("E", "A"):
            return False, f"Erro SAP ao criar utilizador {target_user_fmt} ({modo_desc}): {msg_text}"

        return True, f"Utilizador {target_user_fmt} criado com sucesso na SU01 ({modo_desc}). Senha inicial definida. SAP: {msg_text or 'Gravado'}"

    except Exception as exc:
        return False, f"Exceção na rotina GUI de criação de utilizador (SU01): {exc}"


###################################################################################
# BLOCO 6: LEITURA E GRAVAÇÃO EXCEL
###################################################################################

def selecionar_ficheiro_excel():
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        caminho = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel para o processo {NOME_SCRIPT}",
            filetypes=[("Ficheiros Excel", "*.xlsx;*.xlsm"), ("Todos os ficheiros", "*.*")],
        )
        root.destroy()
        return caminho
    except Exception:
        return None


def ler_ficheiro(caminho_ficheiro, nome_sheet):
    if not caminho_ficheiro or not os.path.exists(caminho_ficheiro):
        print("❌ Caminho inválido ou ficheiro inexistente.")
        return None

    try:
        df = pd.read_excel(caminho_ficheiro, sheet_name=nome_sheet, dtype=object)
        df.columns = [normalizar_coluna(c) for c in df.columns]

        if "ID" not in df.columns:
            df["ID"] = [str(i + 1) for i in range(len(df))]
        if "UTILIZADOR" not in df.columns and "USER" in df.columns:
            df["UTILIZADOR"] = df["USER"]
        if "STATUS" not in df.columns:
            df["STATUS"] = ""
        if "MSG" not in df.columns:
            df["MSG"] = ""
        if "TIMESTEMP" not in df.columns:
            df["TIMESTEMP"] = ""

        return df
    except Exception as e:
        print(f"❌ Erro ao ler a sheet '{nome_sheet}': {e}")
        return None


def gravar_preservando_formatacao(caminho_ficheiro, nome_sheet, df_resultados):
    if df_resultados is None or df_resultados.empty:
        return True

    try:
        ext = os.path.splitext(caminho_ficheiro)[1].lower()
        wb = load_workbook(caminho_ficheiro, keep_vba=(ext == ".xlsm"))
        if nome_sheet not in wb.sheetnames:
            wb.close()
            return False

        ws = wb[nome_sheet]
        headers = [c.value for c in ws[1]]
        headers_norm = [normalizar_coluna(h) if h is not None else "" for h in headers]

        col_id = headers_norm.index("ID") + 1 if "ID" in headers_norm else 1
        col_status = headers_norm.index("STATUS") + 1 if "STATUS" in headers_norm else None
        col_msg = headers_norm.index("MSG") + 1 if "MSG" in headers_norm else None
        col_ts = headers_norm.index("TIMESTEMP") + 1 if "TIMESTEMP" in headers_norm else None

        mapa_res = {}
        for _, row in df_resultados.iterrows():
            cid = str(row.get("ID", "")).strip()
            if cid:
                mapa_res[cid] = (str(row.get("STATUS", "")), str(row.get("MSG", "")), str(row.get("TIMESTEMP", "")))

        atualizados = 0
        for r in range(2, ws.max_row + 1):
            val_id = ws.cell(row=r, column=col_id).value
            if val_id is not None:
                cid = str(val_id).strip()
                if cid in mapa_res:
                    st, msg, ts = mapa_res[cid]
                    if col_status:
                        ws.cell(row=r, column=col_status, value=st)
                    if col_msg:
                        ws.cell(row=r, column=col_msg, value=msg)
                    if col_ts:
                        ws.cell(row=r, column=col_ts, value=ts)
                    atualizados += 1

        wb.save(caminho_ficheiro)
        wb.close()
        print(f"💾 Ficheiro atualizado com sucesso ({atualizados} linha(s) gravada(s)).")
        return True
    except Exception as e:
        print(f"❌ Erro ao gravar resultados no Excel: {e}")
        return False


###################################################################################
# BLOCO 7: PONTO DE ENTRADA WEB / COCKPIT (executar)
###################################################################################

def executar(
    ambiente_cockpit="CUA",
    pfcg_object=None,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    target_user=None,
    reference_user=None,
    first_name=None,
    last_name=None,
    email=None,
    function=None,
    department=None,
    mob_number=None,
    **kwargs
):
    """
    Função de entrada unificada para o SAP Cockpit Web API.
    Suporta tanto a chamada individual direta (passando parâmetros no kwargs)
    como o processamento em lote via ficheiro Excel.
    """
    tempo_inicio_total = time.time()
    print(f"✅ Processo selecionado: {NOME_SCRIPT}")

    target_user = target_user or kwargs.get("user") or kwargs.get("utilizador")
    reference_user = reference_user or kwargs.get("target_user_ref") or kwargs.get("ref_user")
    if first_name and "first_name" not in kwargs: kwargs["first_name"] = first_name
    if last_name and "last_name" not in kwargs: kwargs["last_name"] = last_name
    if email and "email" not in kwargs: kwargs["email"] = email
    if function and "function" not in kwargs: kwargs["function"] = function
    if department and "department" not in kwargs: kwargs["department"] = department
    if mob_number and "mob_number" not in kwargs: kwargs["mob_number"] = mob_number

    # 0. Modo Interativo via Terminal: Perguntar ao utilizador se pretende Massivo ou Individual
    if not modo_nao_interativo and not target_user and not caminho_ficheiro:
        print("\n==================================================")
        print("📌 PROCESSO: Criar Utilizador CUA (SU01)")
        print("==================================================")
        print("Escolha o modo de execução:")
        print("  [1] 📊 Alteração Massiva (via Ficheiro Excel)")
        print("  [2] 👤 Alteração Individual (via Terminal)")
        opcao_modo = input("Selecione a opção [1/2] (Predefinição: 1): ").strip()

        if opcao_modo == "2":
            print("\n--- 👤 Modo de Alteração Individual ---")
            ref_user_in = input("Utilizador de Referência (REF_USER) [Opcional/Enter para omitir]: ").strip()
            target_user_in = input("Utilizador Destino / Colaborador (TARGET_USER) [ex: 80001996]: ").strip()
            while not target_user_in:
                print("⚠️ O utilizador destino é obrigatório!")
                target_user_in = input("Utilizador Destino / Colaborador (TARGET_USER): ").strip()

            # Consultar automaticamente no RH (S4P via RFC)
            rh_first, rh_last, rh_email = buscar_dados_rh_rfc(target_user_in)

            first_name_in = rh_first if rh_first else input("Nome (NAME_FIRST): ").strip()
            last_name_in = rh_last if rh_last else input("Sobrenome (NAME_LAST): ").strip()
            email_in = rh_email if rh_email else input("Email (SMTP_ADDR): ").strip()

            print("\nPor favor, introduza os 3 campos obrigatórios não existentes na tabela RH:")
            func_in = input("1• Função (FUNCTION) * (Solicitar): ").strip()
            dept_in = input("2• Departamento (DEPARTMENT) * (Solicitar): ").strip()
            mob_in = input("3• Telefone (MOB_NUMBER) * (Solicitar): ").strip()

            target_user = target_user_in
            reference_user = ref_user_in
            kwargs["first_name"] = first_name_in
            kwargs["last_name"] = last_name_in
            kwargs["email"] = email_in
            kwargs["function"] = func_in
            kwargs["department"] = dept_in
            kwargs["mob_number"] = mob_in
        else:
            caminho_ficheiro = selecionar_ficheiro_excel()

    # 1. Modo de Execução Individual via Chat / Web API / Terminal
    if target_user:
        target_user_fmt = format_sap_user_id(target_user)
        ref_user_fmt = str(reference_user).strip().upper() if reference_user else ""

        # Auto-completar dados de RH via RFC se Nome, Sobrenome ou Email não tiverem sido informados
        fn = kwargs.get("first_name", "")
        ln = kwargs.get("last_name", "")
        em = kwargs.get("email", "")
        if not fn or not ln or not em:
            rh_fn, rh_ln, rh_em = buscar_dados_rh_rfc(target_user_fmt)
            if not fn: kwargs["first_name"] = rh_fn
            if not ln: kwargs["last_name"] = rh_ln
            if not em: kwargs["email"] = rh_em

        sistema_desejado = MAPA_SISTEMA.get(ambiente_cockpit, "SPA")
        sess_override = kwargs.get("session")
        session = conectar_sap(sistema_desejado, session_override=sess_override)
        if not session:
            return {
                "success": False,
                "message": f"Não foi possível conectar ao SAP GUI no sistema {sistema_desejado}."
            }

        posicionar_sap_meia_tela_direita(session)

        ok, msg = criar_utilizador_su01(
            session=session,
            target_user=target_user_fmt,
            reference_user=ref_user_fmt,
            first_name=kwargs.get("first_name", ""),
            last_name=kwargs.get("last_name", ""),
            email=kwargs.get("email", ""),
            function=kwargs.get("function", ""),
            department=kwargs.get("department", ""),
            mob_number=kwargs.get("mob_number", ""),
            language=kwargs.get("language", "PT"),
            initial_password=kwargs.get("initial_password", kwargs.get("password", ""))
        )

        tempo_decorrido = time.time() - tempo_inicio_total
        if ok:
            print(f"\n==================================================")
            print(f"✅ SUCESSO: {msg}")
            print(f"==================================================")
        else:
            print(f"\n==================================================")
            print(f"❌ ERRO SAP: {msg}")
            print(f"==================================================")

        print(f"⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido)}")
        return {"success": ok, "message": msg}

    # 2. Modo de Execução Massiva via Excel
    if not modo_nao_interativo and not caminho_ficheiro:
        caminho_ficheiro = selecionar_ficheiro_excel()

    if not caminho_ficheiro:
        return {"success": False, "message": "Nenhum ficheiro Excel fornecido ou selecionado."}

    df = ler_ficheiro(caminho_ficheiro, sheet_alvo)
    if df is None or df.empty:
        return {"success": False, "message": f"A sheet '{sheet_alvo}' está vazia ou não foi encontrada."}

    sistema_desejado = MAPA_SISTEMA.get(ambiente_cockpit, "SPA")
    sess_override = kwargs.get("session")
    session = conectar_sap(sistema_desejado, session_override=sess_override)
    if not session:
        return {"success": False, "message": f"Sessão SAP não encontrada para {sistema_desejado}."}

    posicionar_sap_meia_tela_direita(session)

    for idx, row in df.iterrows():
        status_norm = normalizar_valor(row.get("STATUS", ""))
        if status_norm == "CONCLUIDO":
            continue

        u_target = format_sap_user_id(row.get("UTILIZADOR", ""))
        u_ref = format_sap_user_id(row.get("REF_USER", ""))

        if not u_target:
            df.at[idx, "STATUS"] = "ERRO"
            df.at[idx, "MSG"] = "Utilizador alvo não informado."
            df.at[idx, "TIMESTEMP"] = agora_str()
            continue

        ok, msg = criar_utilizador_su01(
            session=session,
            target_user=u_target,
            reference_user=u_ref,
            first_name=texto_limpo(row.get("FIRST_NAME", "")),
            last_name=texto_limpo(row.get("LAST_NAME", "")),
            email=texto_limpo(row.get("EMAIL", "")),
            function=texto_limpo(row.get("FUNCTION", "")),
            department=texto_limpo(row.get("DEPARTMENT", "")),
            mob_number=texto_limpo(row.get("MOB_NUMBER", "")),
        )

        df.at[idx, "STATUS"] = "CONCLUIDO" if ok else "ERRO"
        df.at[idx, "MSG"] = msg
        df.at[idx, "TIMESTEMP"] = agora_str()

    gravar_preservando_formatacao(caminho_ficheiro, sheet_alvo, df)

    tempo_decorrido = time.time() - tempo_inicio_total
    print(f"⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido)}")
    return {"success": True, "message": "Processamento massivo concluído."}


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD", "CUA"], default="CUA")
    parser.add_argument("--xlsx")
    parser.add_argument("--user")
    parser.add_argument("--ref-user")
    args = parser.parse_args()

    executar(
        ambiente_cockpit=args.ambiente,
        caminho_ficheiro=args.xlsx,
        target_user=args.user,
        reference_user=args.ref_user
    )
