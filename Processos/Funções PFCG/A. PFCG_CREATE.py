# -*- coding: utf-8 -*-

###################################################################################
# A. PFCG_CREATE.py
# PFCG - Criar/Modificar Roles + Atribuir TCODEs + Perfil + Transporte
#
# Regras:
#  - Logger visual estruturado por Etapas
#  - Integração com 'pesquisar_request.py'
#  - Inserção direta e rápida de TCODEs e de Ordem de Transporte
#  - Menu de Request Unificado
#  - Barra de progresso por Role
#  - Etapa 1 de performance: esperas inteligentes
#  - Etapa 2 de performance: sem pandas
#  - Etapa 3 de performance: cache de IDs SAP
###################################################################################
import sys
if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import functools
print = functools.partial(print, flush=True)


def executar(
    ambiente_cockpit,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    nome_pasta=None
):
    import os
    import time
    import re
    import unicodedata
    import tkinter as tk

    import win32com.client
    from tkinter import filedialog
    from datetime import datetime
    from math import ceil
    from openpyxl import load_workbook
    from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn

    tempo_inicio_total = time.time()

    # --- CORREÇÃO DA ESTRUTURA DE PASTAS ---
    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual)
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)
    # ---------------------------------------

    NOME_SHEET = "PFCG_CREATE"
    SEARCH_HEADER_IN_FIRST_ROWS = 20

    COLUNAS_OBRIGATORIAS = {"AGR_NAME", "TEXT", "TCODE", "STATUS", "MSG", "TIMESTEMP"}

    MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    SLEEP_UI = 0.08
    SLEEP_ACTION = 0.15
    TCODE_BLOCK_SIZE = 20

    def now_ts():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _sleep(t=SLEEP_UI):
        time.sleep(t)

    def norm_col(s):
        if s is None:
            s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def norm_txt(s):
        if s is None:
            s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def formatar_tempo(segundos):
        h, resto = divmod(segundos, 3600)
        m, s = divmod(resto, 60)
        if h > 0:
            return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
        return f"{int(m):02d}m {int(s):02d}s"

    def selecionar_ficheiro():
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
            filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*"))
        )
        root.destroy()
        return path

    def gravar_resultados_excel(caminho_ficheiro, sheet_name, header_map, records, resultados):
        col_st, col_ms, col_tm = header_map.get("STATUS"), header_map.get("MSG"), header_map.get("TIMESTEMP")
        
        # 1. Tentar via Excel COM primeiro (mais seguro no Windows para não corromper abas complexas)
        try:
            import win32com.client
            import pythoncom
            pythoncom.CoInitialize()
            
            try:
                excel_app = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                excel_app = win32com.client.Dispatch("Excel.Application")
            
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            abs_path = os.path.abspath(caminho_ficheiro)
            wb_excel = excel_app.Workbooks.Open(abs_path)
            ws_excel = wb_excel.Worksheets(sheet_name)
            
            for rec in records:
                chave_busca = str(rec["AGR_NAME"]).strip()
                res = resultados.get(chave_busca)
                if res:
                    if col_st:
                        ws_excel.Cells(rec["_row"], col_st).Value = res["STATUS"]
                    if col_ms:
                        ws_excel.Cells(rec["_row"], col_ms).Value = res["MSG"]
                    if col_tm:
                        ws_excel.Cells(rec["_row"], col_tm).Value = res["TIMESTEMP"]
            
            wb_excel.Save()
            wb_excel.Close(SaveChanges=True)
            try:
                excel_app.Quit()
            except:
                pass
            return True
        except Exception as e_com:
            print(f"  [DEBUG] Falha ao gravar via Excel COM ({e_com}). Usando openpyxl como fallback...")
            
        # 2. Fallback usando openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(caminho_ficheiro)
            ws = wb[sheet_name]
            for rec in records:
                chave_busca = str(rec["AGR_NAME"]).strip()
                res = resultados.get(chave_busca)
                if res:
                    if col_st:
                        ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
                    if col_ms:
                        ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
                    if col_tm:
                        ws.cell(row=rec["_row"], column=col_tm).value = res["TIMESTEMP"]
            wb.save(caminho_ficheiro)
            wb.close()
            return True
        except Exception as e_openpyxl:
            print(f"  ❌ Falha crítica ao gravar Excel com openpyxl: {e_openpyxl}")
            return False

    def split_tcodes(raw):
        if not raw:
            return []
        s = str(raw).replace("\r", "\n").replace("\t", " ").strip().upper()
        parts = re.split(r"[;, \n]+", s)
        out = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if p.startswith("/N") or p.startswith("/O"):
                p = p[2:].strip()
            if p:
                out.append(p)
        return list(dict.fromkeys(out))

    ###################################################################################
    # BLOCO 1: LER EXCEL
    ###################################################################################
    if not caminho_ficheiro:
        if os.path.exists("S4H_Perfis de autorização.xlsx"):
            caminho_ficheiro = "S4H_Perfis de autorização.xlsx"
            print("📂 Utilizando ficheiro Excel padrão encontrado na raiz: S4H_Perfis de autorização.xlsx")
        else:
            if modo_nao_interativo:
                raise ValueError("Faltou o parâmetro --xlsx em modo não-interativo.")
            print("📂 Selecione o ficheiro Excel…")
            caminho_ficheiro = selecionar_ficheiro()
            if not caminho_ficheiro:
                print("❌ Operação cancelada.")
                return

    if not os.path.exists(caminho_ficheiro):
        print(f"❌ Ficheiro não encontrado: {caminho_ficheiro}")
        return

    try:
        wb = load_workbook(caminho_ficheiro)
    except Exception as e:
        print(f"❌ Não consegui abrir o Excel: {e}")
        return

    if NOME_SHEET in wb.sheetnames:
        ws = wb[NOME_SHEET]
    else:
        if len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
        else:
            print(f"❌ Sheet '{NOME_SHEET}' não encontrada.")
            wb.close()
            return

    header_row = None
    header_map = {}
    
    COLUNAS_ENTRADA = {"AGR_NAME", "TEXT", "TCODE"}
    COLUNAS_SAIDA = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [norm_col(c.value) for c in ws[r]]
        colunas_entrada_encontradas = set(row_vals).intersection(COLUNAS_ENTRADA)

        if len(colunas_entrada_encontradas) == len(COLUNAS_ENTRADA):
            header_row = r
            for idx, name in enumerate(row_vals, start=1):
                if name:
                    header_map[name] = idx
            
            modificou = False
            last_col = 0
            for idx, val in enumerate(row_vals, start=1):
                if val:
                    last_col = idx

            for col_key, col_name in COLUNAS_SAIDA.items():
                if col_key not in header_map:
                    last_col += 1
                    ws.cell(row=r, column=last_col, value=col_name)
                    header_map[col_key] = last_col
                    print(f"➕ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                    modificou = True
            
            if modificou:
                try:
                    wb.save(caminho_ficheiro)
                    print("💾 Ficheiro Excel guardado com as novas colunas.")
                except Exception as e:
                    wb.close()
                    print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas: {e}")
                    print("💡 Certifique-se de que o ficheiro Excel não está aberto no Excel e tente novamente.")
                    return
            break

    if not header_row:
        wb.close()
        print("\n❌ Não encontrei a linha de cabeçalho completa.")
        return

    col_agr = header_map.get("AGR_NAME")
    col_text = header_map.get("TEXT")
    col_tcode = header_map.get("TCODE")
    col_status = header_map.get("STATUS")
    col_msg = header_map.get("MSG")
    col_ts = header_map.get("TIMESTEMP")

    records = []

    for r in range(header_row + 1, ws.max_row + 1):
        agr_val = ws.cell(row=r, column=col_agr).value if col_agr else None
        agr = "" if agr_val is None else str(agr_val).strip()
        if not agr:
            continue

        # Validar e remover espaços do nome da Role
        if " " in agr:
            orig = agr
            agr = agr.replace(" ", "")
            ws.cell(row=r, column=col_agr, value=agr)
            print(f"⚠️ Espaços detetados e removidos da Role: '{orig}' -> '{agr}' (atualizado no Excel)")
            try:
                wb.save(caminho_ficheiro)
            except Exception:
                pass

        text_val = ws.cell(row=r, column=col_text).value if col_text else None
        tcode_val = ws.cell(row=r, column=col_tcode).value if col_tcode else None
        status_val = ws.cell(row=r, column=col_status).value if col_status else None
        msg_val = ws.cell(row=r, column=col_msg).value if col_msg else None
        ts_val = ws.cell(row=r, column=col_ts).value if col_ts else None

        records.append({
            "_row": r,
            "AGR_NAME": agr,
            "TEXT": "" if text_val is None else str(text_val).strip(),
            "TCODE": "" if tcode_val is None else str(tcode_val).strip(),
            "STATUS": "" if status_val is None else str(status_val).strip(),
            "MSG": "" if msg_val is None else str(msg_val).strip(),
            "TIMESTEMP": "" if ts_val is None else str(ts_val).strip(),
        })

    if not records:
        wb.close()
        print("⚠️ Não encontrei linhas para processar.")
        return

    roles_map = {}

    for rec in records:
        status_norm = norm_txt(rec["STATUS"])
        if status_norm == "CONCLUIDO":
            continue

        agr = rec["AGR_NAME"].strip()
        if not agr:
            continue

        if agr not in roles_map:
            roles_map[agr] = {
                "AGR_NAME": agr,
                "TEXT": rec["TEXT"].strip(),
                "TCODE_LIST": []
            }

        if not roles_map[agr]["TEXT"] and rec["TEXT"].strip():
            roles_map[agr]["TEXT"] = rec["TEXT"].strip()

        roles_map[agr]["TCODE_LIST"].extend(split_tcodes(rec["TCODE"]))

    if not roles_map:
        wb.close()
        print("⚠️ Nada para processar (tudo CONCLUIDO).")
        return

    roles_agrupadas = []
    for item in roles_map.values():
        item["TCODE_LIST"] = list(dict.fromkeys(item["TCODE_LIST"]))
        roles_agrupadas.append(item)

    roles_agrupadas.sort(key=lambda x: x["AGR_NAME"])

    # =================================================================================
    # CAPTURA SESSÃO SAP
    # =================================================================================
    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        session = next((sess for conn in application.Children for sess in conn.Children if sess.Info.SystemName.upper() == SISTEMA_ESPERADO), None)
    except Exception:
        session = None

    if not session:
        wb.close()
        print(f"❌ Não encontrei sessão do ambiente '{ambiente_cockpit}'.")
        return

    ###################################################################################
    # HELPERS SAP - PERFORMANCE + CACHE DE IDs
    ###################################################################################
    sap_id_cache = {}

    def _safe_find(sap_id):
        try:
            return session.findById(sap_id)
        except:
            return None

    def _sap_busy():
        try:
            return bool(getattr(session, "Busy", False))
        except:
            return False

    def _esperar_sap_livre(timeout=8.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            if not _sap_busy():
                return True
            time.sleep(pausa)
        return False

    def _esperar_objeto(sap_id, timeout=5.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            obj = _safe_find(sap_id)
            if obj:
                return obj
            time.sleep(pausa)
        return None

    def _esperar_sumir(sap_id, timeout=5.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            if not _safe_find(sap_id):
                return True
            time.sleep(pausa)
        return False

    def _send_vkey(vkey, wait_after=True):
        session.findById("wnd[0]").sendVKey(vkey)
        if wait_after:
            _esperar_sap_livre()

    def _press_if_exists(sap_id, timeout=2.0):
        obj = _esperar_objeto(sap_id, timeout=timeout)
        if not obj:
            return False
        try:
            obj.press()
            _esperar_sap_livre()
            return True
        except:
            return False

    def _resolver_id(cache_key, candidatos):
        sap_id = sap_id_cache.get(cache_key)
        if sap_id:
            obj = _safe_find(sap_id)
            if obj:
                return sap_id, obj
            sap_id_cache.pop(cache_key, None)

        for sap_id in candidatos:
            obj = _safe_find(sap_id)
            if obj:
                sap_id_cache[cache_key] = sap_id
                return sap_id, obj
        return None, None

    def _resolver_id_esperando(cache_key, candidatos, timeout=3.0, pausa=0.05):
        limite = time.time() + timeout
        while time.time() < limite:
            sap_id, obj = _resolver_id(cache_key, candidatos)
            if obj:
                return sap_id, obj
            time.sleep(pausa)
        return None, None

    def _criar_nova_request_no_sap_local(sess):
        okcd = _safe_find("wnd[0]/tbar[0]/okcd")
        if okcd:
            okcd.text = "/nSE10"
            _send_vkey(0)

        print("\nTipo da ordem:")
        print("1 - Ordem customizing")
        print("2 - Ordem workbench")

        while True:
            tipo = input("Digite a opção (1/2): ").strip()
            if tipo in ("1", "2"):
                break
            print("❌ Opção inválida. Use apenas 1 ou 2.")

        desc = input("Descrição da request (máx 60): ").strip()
        desc = desc[:60] if desc else "REQUEST CRIADA VIA SCRIPT"

        sess.findById("wnd[0]/tbar[1]/btn[6]").press()
        _esperar_objeto("wnd[1]", timeout=3.0)

        if tipo == "2":
            try:
                radio = _safe_find("wnd[1]/usr/radKO042-REQ_CONS_K")
                if radio:
                    radio.select()
            except:
                pass

        _press_if_exists("wnd[1]/tbar[0]/btn[0]", timeout=3.0)

        try:
            campo_desc = _esperar_objeto("wnd[1]/usr/txtKO013-AS4TEXT", timeout=3.0)
            if campo_desc:
                campo_desc.text = desc
        except:
            pass

        _press_if_exists("wnd[1]/tbar[0]/btn[0]", timeout=3.0)

        trkorr = None
        for sap_id in ["wnd[0]/usr/lbl[20,9]", "wnd[0]/usr/lbl[1,1]"]:
            try:
                obj = _esperar_objeto(sap_id, timeout=1.0)
                if obj:
                    txt = obj.Text
                    match = re.search(r"\b[A-Z0-9]{3,4}K\d{6,}\b", txt)
                    if match:
                        trkorr = match.group(0)
            except:
                pass
            if trkorr:
                break

        if okcd:
            okcd.text = "/n"
            _send_vkey(0)

        tipo_txt = "Customizing" if tipo == "1" else "Workbench"
        print("\n✔️ Request criada.")
        print(f"Tipo: {tipo_txt} | Descrição: {desc}")

        if not trkorr:
            trkorr = input("Não consegui extrair a request automaticamente. Cole aqui: ").strip().upper()

        print(f"Request: {trkorr}")
        return trkorr

    if not request_transporte and not modo_nao_interativo:
        print("\n============================================================")
        print("🚚 Opções de configuração de Transporte.\n")
        print("   1 - Escreva o número da Request")
        print("   2 - Criar nova ordem de transporte")
        print("   3 - Pesquisar suas request criadas.")
        print("   4 - Prima [Enter] vazio para NÃO transportar")
        print("============================================================")

        while True:
            req_input = input("\n👉 Opção: ").strip()
            if req_input in ("1", "2", "3", "4", ""):
                if req_input == "":
                    req_input = "4"
                break
            print("❌ Opção inválida. Use 1, 2, 3, 4 ou apenas pressione Enter.")

        if req_input == "1":
            request_transporte = input("🔢 Numero da Request (ex: S4QK900396): ").strip().upper()

        elif req_input == "2":
            request_transporte = _criar_nova_request_no_sap_local(session)

        elif req_input == "3":
            try:
                import pesquisar_request
                print("\n🔍 A abrir nova sessão em segundo plano para pesquisar (SE16H)...")

                resultados_pesquisa = pesquisar_request.listar_requests(
                    system_name=SISTEMA_ESPERADO,
                    include_requests=True,
                    use_new_mode=True,
                    minimize=False,
                    close_after=True,
                    session=session
                )

                if resultados_pesquisa:
                    escolha = input("\n👉 Digite o número (N) da Request que deseja utilizar (ou Enter para cancelar): ").strip()
                    if escolha.isdigit() and 1 <= int(escolha) <= len(resultados_pesquisa):
                        request_transporte = resultados_pesquisa[int(escolha) - 1][0]
                        print(f"✔️ Selecionou a Request: {request_transporte}")
                    else:
                        print("❌ Seleção cancelada. Não haverá transporte.")
                else:
                    print("⚠️ Não foram encontradas Requests abertas para o seu utilizador.")
            except ImportError as e:
                print(f"❌ Erro de Importação: Não consegui encontrar o módulo pesquisar_request.py. Detalhe: {e}")

        elif req_input == "4":
            print("⏭️  Nenhuma request selecionada (Transporte ignorado).")
            request_transporte = None
        print("============================================================")

    print(f"\n📋 Roles a processar (agrupadas): {len(roles_agrupadas)}")
    for rr in roles_agrupadas:
        print(f" - {rr['AGR_NAME']}: {rr['TEXT']} (TCODEs: {len(rr['TCODE_LIST'])})")

    if pedir_confirmacao and not modo_nao_interativo:
        if input("\nDeseja lançar esses dados no SAP? [S/N]: ").strip().upper() != "S":
            wb.close()
            return

    # Inicializar documentação se nome_pasta estiver preenchido
    doc_session = None
    if nome_pasta and str(nome_pasta).strip():
        try:
            from sap_script_web_cockpit_v2.documentation.functional_doc import FunctionalDocSession
            
            PROCESS_DOC_CONFIG = {
                "processo": "PFCG_CREATE",
                "transacao": "PFCG",
                "titulo": "Criação/Atualização de Roles e Perfis de Autorização",
                "modulos_afetados": ["SAP Basis", "Segurança", "Autorizações"],
                "processos_afetados": [
                    "Gestão de roles e perfis de autorização",
                    "Atribuição de transações a roles",
                    "Geração de perfis de autorização"
                ],
                "objeto_principal": "Role",
                "solucao_proposta": "A solução consiste em processar automaticamente as roles informadas no ficheiro de entrada, validar as transações associadas, aceder à transação PFCG, criar ou atualizar a role, atribuir as transações na aba Menu, gravar as alterações e gerar o perfil de autorização."
            }

            doc_session = FunctionalDocSession(
                nome_pasta=nome_pasta,
                processo=PROCESS_DOC_CONFIG["processo"],
                transacao=PROCESS_DOC_CONFIG["transacao"],
                config=PROCESS_DOC_CONFIG
            )
            sap_user = ""
            sap_client = ""
            try:
                sap_user = session.Info.User
                sap_client = session.Info.Client
            except:
                pass
                
            metadata = {
                "ambiente": ambiente_cockpit,
                "sistema": SISTEMA_ESPERADO,
                "cliente": sap_client,
                "utilizador_sap": sap_user,
                "total_roles": len(roles_agrupadas),
                "excel_utilizado": caminho_ficheiro,
                "request_transporte": request_transporte,
            }
            doc_session.start_execution(metadata)
            print(f"[DOC] Documentação funcional iniciada na pasta: {doc_session.output_dir}")
        except Exception as doc_exc:
            print(f"[DOC_WARN] Não foi possível inicializar documentação funcional: {doc_exc}")

    # ---------------------------------------------------------------------------
    # Constantes de posicionamento da janela SAP para evidências de documentação
    # ---------------------------------------------------------------------------
    DOC_SAP_WINDOW_WIDTH_RATIO   = 0.60   # 60% da largura do monitor
    DOC_SAP_WINDOW_HEIGHT_RATIO  = 1.00   # 100% da altura útil
    DOC_SAP_WINDOW_X_RATIO       = 0.00   # Começa no lado esquerdo
    DOC_SAP_WINDOW_Y_RATIO       = 0.00
    DOC_SCREENSHOT_DELAY_SECONDS = 0.75   # Aguardar antes do print

    def prepare_sap_window_for_evidence():
        """
        Traz a janela SAP GUI para primeiro plano, restaura se minimizada,
        posiciona-a no lado esquerdo com 60% da largura e 100% da altura útil.
        Usa o handle real da janela SAP GUI Scripting sempre que disponível
        para evitar mover qualquer outra janela (browser, Word, VS Code, etc.).
        Não interrompe o SAP se falhar.
        """
        t_start = time.time()
        try:
            import win32gui
            import win32con
            import ctypes

            # ------------------------------------------------------------------
            # 1. Obter o HWND da janela SAP GUI diretamente pela sessão de scripting.
            #    Isto garante que apenas a janela SAP é movida — nunca o browser.
            # ------------------------------------------------------------------
            hwnd = None
            try:
                hwnd = int(session.findById("wnd[0]").Handle)
            except Exception:
                hwnd = None

            # Fallback: localizar por título com fragmentos específicos do SAP GUI.
            # "SAP" genérico é excluído intencionalmente para evitar match com o browser.
            if not hwnd:
                SAP_TITLE_FRAGMENTS = (
                    "SAP Easy Access",
                    "PFCG",
                    "DESENVOLVIMENTO",
                    "PRODUÇÃO",
                    "QUALIDADE",
                    "S4H",
                    "SAP Logon",
                )
                for title_fragment in SAP_TITLE_FRAGMENTS:
                    def _enum_cb(h, _):
                        nonlocal hwnd
                        if hwnd is None and win32gui.IsWindowVisible(h):
                            if title_fragment in win32gui.GetWindowText(h):
                                hwnd = h
                    win32gui.EnumWindows(_enum_cb, None)
                    if hwnd:
                        break

            if not hwnd:
                print("[DOC_WARN] Não foi possível localizar a janela SAP GUI para captura de evidência.")
                return

            # ------------------------------------------------------------------
            # 2. Obter área útil do monitor principal (sem barra de tarefas)
            # ------------------------------------------------------------------
            try:
                work_area = ctypes.wintypes.RECT()
                ctypes.windll.user32.SystemParametersInfoW(0x0030, 0, ctypes.byref(work_area), 0)
                screen_w = work_area.right  - work_area.left
                screen_h = work_area.bottom - work_area.top
                origin_x = work_area.left
                origin_y = work_area.top
            except Exception:
                screen_w = win32gui.GetSystemMetrics(0)
                screen_h = win32gui.GetSystemMetrics(1)
                origin_x, origin_y = 0, 0

            win_w = int(screen_w * DOC_SAP_WINDOW_WIDTH_RATIO)
            win_h = int(screen_h * DOC_SAP_WINDOW_HEIGHT_RATIO)
            win_x = origin_x + int(screen_w * DOC_SAP_WINDOW_X_RATIO)
            win_y = origin_y + int(screen_h * DOC_SAP_WINDOW_Y_RATIO)

            # ------------------------------------------------------------------
            # 3. Forçar minimização e depois restauro garante que a janela ganha foco/primeiro plano no Windows
            # ------------------------------------------------------------------
            win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
            time.sleep(0.15)
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            time.sleep(0.2)

            win32gui.SetWindowPos(
                hwnd,
                win32con.HWND_TOP,
                win_x, win_y, win_w, win_h,
                win32con.SWP_SHOWWINDOW,
            )

            # Trazer para primeiro plano (com tratamento de erro silencioso se o Windows bloquear)
            try:
                win32gui.SetForegroundWindow(hwnd)
            except Exception:
                pass
            time.sleep(DOC_SCREENSHOT_DELAY_SECONDS)

        except Exception as _pos_exc:
            t_spent = time.time() - t_start
            os.environ["CURRENT_ROLE_DOC_WARN"] = str(int(os.environ.get("CURRENT_ROLE_DOC_WARN", 0)) + 1)
            print(f"[DOC_WARN] Não foi possível posicionar a janela SAP para captura de evidência: {_pos_exc} | Tempo gasto antes do aviso: {t_spent:.2f}s")


    def capture_screenshot(role_name, momento, idx):
        if not doc_session or not doc_session.enabled:
            return None
        t_start = time.time()
        try:
            safe_role = re.sub(r"[^a-zA-Z0-9_-]", "_", role_name)
            filename = f"{idx:02d}_{safe_role}_{momento}.png"
            filepath = doc_session.image_dir / filename

            # Preparar janela SAP: primeiro plano + posicionamento controlado
            prepare_sap_window_for_evidence()

            wnd = session.findById("wnd[0]")
            try:
                wnd.hardCopy(str(filepath), 2)
            except Exception:
                try:
                    wnd.HardCopy(str(filepath), 2)
                except Exception:
                    try:
                        wnd.hardCopy(str(filepath))
                    except Exception:
                        wnd.HardCopy(str(filepath))
            if filepath.exists():
                return str(filepath)
            else:
                t_spent = time.time() - t_start
                os.environ["CURRENT_ROLE_DOC_WARN"] = str(int(os.environ.get("CURRENT_ROLE_DOC_WARN", 0)) + 1)
                print(f"[DOC_WARN] Não foi possível capturar evidência da role {role_name} no momento {momento} | Tempo gasto antes do aviso: {t_spent:.2f}s")
                return None
        except Exception as exc:
            t_spent = time.time() - t_start
            os.environ["CURRENT_ROLE_DOC_WARN"] = str(int(os.environ.get("CURRENT_ROLE_DOC_WARN", 0)) + 1)
            print(f"[DOC_WARN] Não foi possível capturar evidência: {exc} | Tempo gasto antes do aviso: {t_spent:.2f}s")
            return None

    ###################################################################################
    # BLOCO 2: SAP GUI helpers
    ###################################################################################
    def encontrar_grids_alv(root):
        """
        Percorre recursivamente Children de root e retorna objetos que parecem ALV/Grid.
        Deve ignorar Toolbar.
        Deve logar Id, Type, SubType e Name dos candidatos encontrados.
        """
        grids = []
        try:
            stack = [root]
            while stack:
                obj = stack.pop()
                try:
                    obj_id = obj.Id
                    obj_type = obj.Type
                    obj_subtype = getattr(obj, "SubType", "")
                    obj_name = obj.Name
                except:
                    continue
                    
                is_toolbar = "toolbar" in str(obj_type).lower() or "toolbar" in str(obj_subtype).lower() or "toolbar" in str(obj_name).lower()
                
                has_row_count = False
                try:
                    rc = int(obj.RowCount)
                    if rc >= 0:
                        has_row_count = True
                except:
                    pass
                    
                if has_row_count and not is_toolbar:
                    print(f"├─ Candidato ALV/Grid encontrado: Id={obj_id} | Type={obj_type} | SubType={obj_subtype} | Name={obj_name}")
                    grids.append(obj)
                    
                try:
                    for idx in range(obj.Children.Count):
                        stack.append(obj.Children(idx))
                except:
                    pass
        except Exception as e_find:
            print(f"  ⚠️ Erro ao procurar ALV: {e_find}")
        return grids

    def consultar_tcodes_agr_tcodes(sess_principal, agr_name):
        import time
        res = {
            "ok": False,
            "fonte": "AGR_TCODES",
            "tcodes": set(),
            "qtd": 0,
            "mensagem": "Inicializado",
            "erro_tecnico": False,
            "debug": {}
        }
        
        # --- TENTAR VIA RFC PRIMEIRO ---
        try:
            system_name = str(sess_principal.Info.SystemName).upper()
            client = str(sess_principal.Info.Client)
            
            ashost = os.getenv("SAP_ASHOST")
            user = os.getenv("SAP_USER")
            
            env_passwd_key = f"SAP_PASSWORD_{system_name}CLNT{client}"
            passwd = os.getenv(env_passwd_key) or os.getenv("SAP_PASSWD") or os.getenv("SAP_PASSWORD")
            
            sysnr = os.getenv("SAP_SYSNR") or "00"
            lang = os.getenv("SAP_LANG") or os.getenv("SAP_LANGUAGE") or "PT"
            
            if ashost and user and passwd:
                from pyrfc import Connection
                conn_params = {
                    "ashost": ashost,
                    "sysnr": sysnr,
                    "client": client,
                    "user": user,
                    "passwd": passwd,
                    "lang": lang
                }
                print(f"├─ Tentando ligar via RFC (Host: {ashost}, Client: {client}, User: {user})...")
                rfc_conn = Connection(**conn_params)
                
                options = [{"TEXT": f"AGR_NAME = '{agr_name}'"}]
                fields = [{"FIELDNAME": "TCODE"}]
                rfc_res = rfc_conn.call(
                    "RFC_READ_TABLE",
                    QUERY_TABLE="AGR_TCODES",
                    DELIMITER="|",
                    FIELDS=fields,
                    OPTIONS=options,
                    ROWCOUNT=9999
                )
                tcodes_set = set()
                for row in rfc_res.get("DATA", []):
                    wa = row.get("WA", "").strip()
                    if wa:
                        tc = wa.split("|")[0].strip().upper()
                        tc_norm = normalizar_tcode(tc)
                        if tc_norm:
                            tcodes_set.add(tc_norm)
                rfc_conn.close()
                
                res["tcodes"] = tcodes_set
                res["qtd"] = len(tcodes_set)
                res["ok"] = True
                res["mensagem"] = f"Consulta RFC concluída com sucesso. {len(tcodes_set)} tcodes lidas."
                print(f"├─ Consulta RFC concluída com sucesso. {len(tcodes_set)} tcodes encontradas.")
                
                tcodes_log = sorted(list(tcodes_set))
                print("├─ TCODEs encontradas na AGR_TCODES:")
                if tcodes_log:
                    for tc_item in tcodes_log[:10]:
                        print(f"│  └─ {tc_item}")
                    if len(tcodes_log) > 10:
                        print(f"│  └─ ... (+ {len(tcodes_log) - 10} adicionais)")
                else:
                    print("│  └─ nenhuma")
                    
                return res
            else:
                print("  ⚠️ Parâmetros RFC incompletos no .env. Ignorando RFC.")
        except Exception as rfc_exc:
            print(f"  ⚠️ Não foi possível consultar via RFC: {rfc_exc}")
            print("  ├─ Revertendo para consulta via GUI (SE16H)...")
            
        # --- FALLBACK SE16H GUI ---
        new_session = None
        try:
            print(f"├─ Consultando AGR_TCODES para a role {agr_name}...")
            # 1. Obter a conexão e a lista de IDs de sessão antes
            connection = sess_principal.Parent
            before_ids = set()
            for i in range(connection.Children.Count):
                before_ids.add(connection.Children(i).Id)
                
            # 2. Iniciar nova sessão via comando /ose16h na sessão principal
            sess_principal.findById("wnd[0]/tbar[0]/okcd").text = "/ose16h"
            sess_principal.findById("wnd[0]").sendVKey(0)
            
            # 3. Aguardar a criação da nova sessão
            t0 = time.time()
            while time.time() - t0 <= 10:
                for i in range(connection.Children.Count):
                    c = connection.Children(i)
                    if c.Id not in before_ids:
                        new_session = c
                        break
                if new_session:
                    break
                time.sleep(0.2)
                
            if not new_session:
                res["mensagem"] = "Não foi possível abrir uma nova sessão SAP (SE16H)."
                res["erro_tecnico"] = True
                print("  ⚠️ Não foi possível abrir uma nova sessão SAP para consultar a AGR_TCODES.")
                return res

            print("├─ SE16H aberta em novo modo.")
            res["debug"]["session_id"] = new_session.Id

            # Aguardar que a nova sessão carregue e não esteja Busy
            t_wait = time.time()
            while time.time() - t_wait <= 8:
                if not getattr(new_session, "Busy", False):
                    break
                time.sleep(0.1)
                
            # 4. Configurar a tabela AGR_TCODES na nova sessão
            tab_field_candidates = [
                "wnd[0]/usr/ctxtGD-TAB",
                "wnd[0]/usr/ctxtDATABROWSE-TABLENAME",
                "wnd[0]/usr/ctxtTABNAME",
            ]
            tab_field_id = None
            for cid in tab_field_candidates:
                try:
                    if new_session.findById(cid):
                        tab_field_id = cid
                        break
                except:
                    pass
                    
            if not tab_field_id:
                res["mensagem"] = "Não encontrou o campo de nome da tabela na SE16H."
                res["erro_tecnico"] = True
                print("  ⚠️ Não encontrou o campo de nome da tabela na SE16H.")
                return res
                
            # Escrever AGR_TCODES e dar Enter
            new_session.findById(tab_field_id).text = "AGR_TCODES"
            new_session.findById("wnd[0]").sendVKey(0)
            print("├─ Tabela informada: AGR_TCODES")
            
            # Aguardar carregar campos
            t_wait = time.time()
            tbl_control = None
            while time.time() - t_wait <= 5:
                try:
                    root = new_session.findById("wnd[0]/usr")
                    stack = [root]
                    while stack:
                        obj = stack.pop()
                        if obj.Name == "SAPLSE16NSELFIELDS_TC" or obj.Id.endswith("tblSAPLSE16NSELFIELDS_TC"):
                            tbl_control = obj
                            break
                        for idx in range(obj.Children.Count):
                            stack.append(obj.Children(idx))
                    if tbl_control:
                        break
                except:
                    pass
                time.sleep(0.1)
                
            if not tbl_control:
                res["mensagem"] = "Tabela de critérios da SE16H não carregou."
                res["erro_tecnico"] = True
                print("  ⚠️ Tabela de critérios da SE16H não carregou.")
                return res
                
            # Definir Max Ocorrências
            max_cids = ["wnd[0]/usr/txtMAX_SEL", "wnd[0]/usr/txtGD-MAXROWS", "wnd[0]/usr/txtMAX_HITS"]
            for mcid in max_cids:
                try:
                    new_session.findById(mcid).text = "9999"
                    break
                except:
                    pass
                    
            # Procurar o campo AGR_NAME nos critérios
            row_count = int(tbl_control.RowCount)
            visible_rows = int(tbl_control.VisibleRowCount)
            
            col_fieldname_orig = 13
            col_fieldname_prefix = "txtGS_SELFIELDS-FIELDNAME"
            col_low = 2
            col_low_prefix = "ctxtGS_SELFIELDS-LOW"
            
            try:
                for idx in range(tbl_control.Children.Count):
                    child = tbl_control.Children(idx)
                    id_str = child.Id
                    if "[" in id_str and "]" in id_str:
                        bracket_part = id_str.rsplit("[", 1)[-1].split("]")[0]
                        parts = bracket_part.split(",")
                        if len(parts) == 2 and int(parts[1]) == 0:
                            col_idx = int(parts[0])
                            prefix_path = id_str.rsplit("[", 1)[0]
                            prefix = prefix_path.split("/")[-1]
                            name = child.Name.upper()
                            if name.endswith("-LOW") or name == "GS_SELFIELDS-LOW":
                                col_low = col_idx
                                col_low_prefix = prefix
                            elif "FIELDNAME" in name:
                                if col_idx > 10:
                                    col_fieldname_orig = col_idx
                                    col_fieldname_prefix = prefix
            except:
                pass

            # Buscar a linha correspondente a AGR_NAME
            row_agr_name = None
            for r in range(min(row_count, visible_rows)):
                try:
                    fname_id = f"{tbl_control.Id}/{col_fieldname_prefix}[{col_fieldname_orig},{r}]"
                    fieldname = new_session.findById(fname_id).text.strip().upper()
                    if fieldname == "AGR_NAME":
                        row_agr_name = r
                        break
                except:
                    continue
                    
            if row_agr_name is None:
                res["mensagem"] = "Não encontrou o critério 'AGR_NAME' na tabela SE16H."
                res["erro_tecnico"] = True
                print("  ⚠️ Não encontrou o critério 'AGR_NAME' na tabela SE16H.")
                return res
                
            # Inserir o nome da role no filtro
            low_id = f"{tbl_control.Id}/{col_low_prefix}[{col_low},{row_agr_name}]"
            new_session.findById(low_id).text = agr_name
            print(f"├─ Filtro aplicado: AGR_NAME = {agr_name}")
            
            # 5. Executar a consulta (F8)
            new_session.findById("wnd[0]/tbar[1]/btn[8]").press()
            
            # Aguardar execução
            t_wait = time.time()
            while time.time() - t_wait <= 10:
                if not getattr(new_session, "Busy", False):
                    break
                time.sleep(0.15)
                
            # Capturar e logar statusbar se existir mensagem
            try:
                sbar = new_session.findById("wnd[0]/sbar")
                sbar_text = str(sbar.Text).strip()
                sbar_type = str(sbar.MessageType).strip().upper()
                if sbar_text:
                    print(f"├─ [SE16H_SBAR] Tipo: {sbar_type} | Texto: {sbar_text}")
                    res["debug"]["sbar"] = {"type": sbar_type, "text": sbar_text}
                    # Se for erro real na consulta
                    if sbar_type in ("E", "A"):
                        res["mensagem"] = f"Erro no statusbar da SE16H: {sbar_text}"
                        return res
            except:
                pass

            print("├─ Consulta executada.")

            # 6. Localizar o ALV Grid na tela de resultados usando a nova função encontrar_grids_alv
            grid = None
            try:
                root_wnd = new_session.findById("wnd[0]")
                grids_found = encontrar_grids_alv(root_wnd)
                if grids_found:
                    grid = grids_found[0]
            except Exception as e_grid:
                print(f"  ⚠️ Exceção ao varrer tela por ALV/Grid: {e_grid}")
                
            if not grid:
                # Se não encontrou ALV, mas o statusbar disse que "Nenhum dado selecionado",
                # então a consulta funcionou, mas não há linhas!
                try:
                    sbar = new_session.findById("wnd[0]/sbar")
                    sbar_text = str(sbar.Text).strip().upper()
                    if "NENHUM" in sbar_text or "NO DATA" in sbar_text or "ZERO" in sbar_text:
                        res["ok"] = True
                        res["qtd"] = 0
                        res["mensagem"] = "Consulta com sucesso, 0 registos encontrados."
                        print("├─ Linhas retornadas: 0 (Nenhum dado selecionado)")
                        return res
                except:
                    pass
                
                res["mensagem"] = "Não consegui localizar o ALV/Grid de resultados da SE16H."
                res["erro_tecnico"] = True
                print("  ⚠️ Não consegui localizar o ALV/Grid de resultados da SE16H.")
                return res

            print(f"├─ ALV/Grid encontrado: {grid.Id}")
            res["debug"]["grid_id"] = grid.Id

            # 7. Ler as transações
            r_count = int(grid.RowCount)
            print(f"├─ Linhas retornadas: {r_count}")
            res["debug"]["row_count"] = r_count

            # Descobrir a coluna de TCODE
            tcode_col = None
            colunas_candidatas = ["TCODE", "AGR_TCODES-TCODE", "S_TCODE"]
            
            # Tentar obter lista de colunas do grid
            alv_cols = []
            try:
                for col in grid.ColumnOrder:
                    alv_cols.append(str(col))
            except:
                pass
                
            print(f"├─ Colunas encontradas: {', '.join(alv_cols) if alv_cols else 'não listadas'}")
            res["debug"]["colunas"] = alv_cols
            
            # 1. Procurar nas colunas encontradas se alguma bate com as candidatas
            for c_cand in colunas_candidatas:
                if c_cand in alv_cols or c_cand.upper() in [c.upper() for c in alv_cols]:
                    tcode_col = c_cand
                    break
                    
            # 2. Se não encontrou, testar se GetCellValue funciona diretamente com alguma das candidatas no primeiro registo
            if not tcode_col and r_count > 0:
                for c_cand in colunas_candidatas:
                    try:
                        val = grid.GetCellValue(0, c_cand)
                        if val is not None:
                            tcode_col = c_cand
                            break
                    except:
                        pass
                        
            # 3. Se ainda não encontrou, testar todas as colunas lidas do ALV no primeiro registo
            if not tcode_col and r_count > 0 and alv_cols:
                for col in alv_cols:
                    try:
                        val = str(grid.GetCellValue(0, col)).strip()
                        # Uma TCODE válida tem formato alfanumérico não muito longo e não vazio
                        if val and len(val) <= 20 and re.match(r"^[A-Z0-9_/]+$", val, re.IGNORECASE):
                            tcode_col = col
                            break
                    except:
                        pass

            if r_count > 0 and not tcode_col:
                res["mensagem"] = "Não foi possível identificar a coluna TCODE no ALV/Grid."
                res["erro_tecnico"] = True
                print("  ⚠️ Não foi possível identificar a coluna TCODE no ALV/Grid.")
                return res

            if tcode_col:
                print(f"├─ Coluna TCODE identificada: {tcode_col}")

            tcodes_set = set()
            for r in range(r_count):
                try:
                    tc = str(grid.GetCellValue(r, tcode_col)).strip().upper()
                    if tc:
                        # Normalizar TCODE
                        tc_norm = normalizar_tcode(tc)
                        if tc_norm:
                            tcodes_set.add(tc_norm)
                except:
                    pass

            res["tcodes"] = tcodes_set
            res["qtd"] = len(tcodes_set)
            res["ok"] = True
            res["mensagem"] = f"Consulta concluída com sucesso. {len(tcodes_set)} tcodes lidas."

            # Logar primeiras TCODEs, máx 10
            tcodes_log = sorted(list(tcodes_set))
            print("├─ TCODEs encontradas na AGR_TCODES:")
            if tcodes_log:
                for tc_item in tcodes_log[:10]:
                    print(f"│  └─ {tc_item}")
                if len(tcodes_log) > 10:
                    print(f"│  └─ ... (+ {len(tcodes_log) - 10} adicionais)")
            else:
                print("│  └─ nenhuma")

        except Exception as e_tcode:
            res["mensagem"] = f"Exceção técnica: {e_tcode}"
            res["erro_tecnico"] = True
            print(f"  ⚠️ Erro técnico ao consultar a tabela AGR_TCODES: {e_tcode}")
            
        finally:
            # 8. Fechar a sessão temporária
            if new_session:
                try:
                    new_session.findById("wnd[0]").close()
                    time.sleep(0.3)
                    try:
                        if new_session.ActiveWindow.Type == "GuiModalWindow":
                            candidatos_btn = [
                                "wnd[1]/usr/btnSPOP-OPTION1",
                                "wnd[1]/usr/btnBUTTON_1",
                                "wnd[1]/tbar[0]/btn[0]",
                            ]
                            for c_btn in candidatos_btn:
                                try:
                                    new_session.findById(c_btn).press()
                                    break
                                except:
                                    pass
                    except:
                        pass
                except:
                    pass
                    
        return res

    def ler_tcodes_existentes_menu_pfcg(sess):
        res = {
            "ok": False,
            "fonte": "PFCG_MENU",
            "tcodes": set(),
            "qtd": 0,
            "mensagem": "Inicializado"
        }
        try:
            print("├─ Validando aba Menu da PFCG...")
            pfcg.goto_menu_tab()
            
            menu_shell_id, menu_shell_obj = _resolver_id(
                "menu_shell",
                [
                    "wnd[0]/usr/tabsTABSTRIP1/tabpTAB9/ssubSUB1:SAPLPRGN_TREE:0321/cntlTOOL_CONTROL/shellcont/shell",
                    "wnd[0]/usr/tabsTABSTRIP1/tabpTAB9/ssubSUB1:SAPLPRGN_TREE:0320/cntlTOOL_CONTROL/shellcont/shell"
                ]
            )
            if not menu_shell_obj:
                res["mensagem"] = "Não localizou o controlo do Menu na aba TAB9."
                print("  ⚠️ Controlo do Menu não encontrado na aba TAB9.")
                return res
                
            obj_type = getattr(menu_shell_obj, "Type", "desconhecido")
            obj_subtype = getattr(menu_shell_obj, "SubType", "desconhecido")
            obj_id = getattr(menu_shell_obj, "Id", "desconhecido")
            obj_name = getattr(menu_shell_obj, "Name", "desconhecido")
            print(f"├─ Controlo do Menu: Id={obj_id} | Type={obj_type} | SubType={obj_subtype} | Name={obj_name}")
            
            keys = None
            metodo_usado = None
            
            tentativas_metodos = [
                ("GetAllNodeKeys", lambda o: o.GetAllNodeKeys()),
                ("GetNodeKeys", lambda o: o.GetNodeKeys()),
                ("nodeKeys", lambda o: o.nodeKeys),
            ]
            
            for name_met, getter in tentativas_metodos:
                try:
                    keys = getter(menu_shell_obj)
                    if keys is not None:
                        metodo_usado = name_met
                        break
                except:
                    pass
                    
            if keys is None:
                res["mensagem"] = f"Nenhum método compatível para ler TreeView encontrado no controlo (Type={obj_type})."
                print("  ⚠️ Nenhum método compatível (GetAllNodeKeys/GetNodeKeys) funcionou no TreeView.")
                return res
                
            tcodes_tree = set()
            try:
                for k in keys:
                    text = str(menu_shell_obj.GetNodeTextByKey(k)).strip().upper()
                    match = re.search(r"\(([^)]+)\)", text)
                    if match:
                        tcodes_tree.add(match.group(1).strip())
                    else:
                        parts = text.split()
                        for p in parts:
                            p_clean = re.sub(r"[^A-Z0-9_]", "", p)
                            if p_clean:
                                tcodes_tree.add(p_clean)
            except Exception as e_node:
                res["mensagem"] = f"Erro ao ler nós da árvore usando método {metodo_usado}: {e_node}"
                print(f"  ⚠️ Erro técnico ao ler nós da árvore: {e_node}")
                return res

            normalized_tcodes = {normalizar_tcode(tc) for tc in tcodes_tree if tc}
            res["tcodes"] = normalized_tcodes
            res["qtd"] = len(normalized_tcodes)
            res["ok"] = True
            res["mensagem"] = f"TreeView lido com sucesso usando {metodo_usado}."
            
            tcodes_log = sorted(list(normalized_tcodes))
            print("├─ TCODEs encontradas na aba Menu:")
            if tcodes_log:
                for tc_item in tcodes_log[:10]:
                    print(f"│  └─ {tc_item}")
                if len(tcodes_log) > 10:
                    print(f"│  └─ ... (+ {len(tcodes_log) - 10} adicionais)")
            else:
                print("│  └─ nenhuma")
                
        except Exception as e_menu:
            res["mensagem"] = f"Exceção técnica no Menu: {e_menu}"
            print(f"  ⚠️ Erro ao ler aba Menu: {e_menu}")
            
        return res

    def normalizar_tcode(tc):
        if not tc:
            return ""
        tc = str(tc).strip().upper().replace(" ", "")
        if tc.startswith("/N") or tc.startswith("/O"):
            tc = tc[2:]
        return tc

    def get_statusbar():
        try:
            sbar = session.findById("wnd[0]/sbar")
            tipo = getattr(sbar, "MessageType", "").strip().upper()
            texto = (sbar.Text or "").strip()
            if texto:
                print(f"[SAP_SBAR] {texto}")
            return (tipo, texto)
        except:
            return ("", "")

    def try_actions(actions):
        for a in actions:
            try:
                ctrl = session.findById(a["path"])
                if a["op"] == "text":
                    ctrl.setFocus()
                    ctrl.text = a["val"]
                    _esperar_sap_livre()
                    return True
                elif a["op"] == "press":
                    if hasattr(ctrl, "Enabled") and not ctrl.Enabled:
                        continue
                    ctrl.press()
                    _esperar_sap_livre()
                    return True
                elif a["op"] == "select":
                    ctrl.select()
                    _esperar_sap_livre()
                    return True
            except:
                continue
        return False

    def tratar_popup_modal(max_loops=6):
        for _ in range(max_loops):
            try:
                _esperar_sap_livre(timeout=2.0, pausa=0.05)

                if session.ActiveWindow.Type != "GuiModalWindow":
                    return False

                try:
                    if session.findById("wnd[1]/usr/tblSAPLPRGN_WIZARDCTRL_TCODE", False) or \
                       session.findById("wnd[1]/usr/tblSAPLPRGN_WIZARDCTRL_TCODE1", False):
                        return False
                except:
                    pass

                candidatos = [
                    "wnd[1]/usr/btnBUTTON_1",
                    "wnd[1]/usr/btnSPOP-OPTION1",
                    "wnd[1]/tbar[0]/btn[0]",
                    "wnd[1]/tbar[0]/btn[19]",
                    "wnd[1]/tbar[0]/btn[11]"
                ]
                for p in candidatos:
                    if try_actions([{"path": p, "op": "press"}]):
                        return True
                return True
            except:
                return False
        return True

    def wait_wnd1_close(timeout=3.0):
        return _esperar_sumir("wnd[1]", timeout=timeout, pausa=0.05)

    ###################################################################################
    # BLOCO 3: Page Object PFCG
    ###################################################################################
    class PFCGPage:
        def __init__(self, sess):
            self.sess = sess

        def open(self):
            print("  ├─ Abrindo a transação /NPFCG...")
            # try:
            #     self.sess.findById("wnd[0]").maximize()
            # except Exception:
            #     pass

            self.sess.findById("wnd[0]/tbar[0]/okcd").text = "/NPFCG"
            _send_vkey(0)
            tratar_popup_modal()

        def set_role_name(self, nome):
            print(f"  ├─ Inserindo o nome da Role: {nome}")
            sap_id, obj = _resolver_id(
                "role_name_field",
                ["wnd[0]/usr/ctxtAGR_NAME_NEU", "wnd[0]/usr/ctxtAGR_NAME"]
            )
            if not obj:
                return False
            try:
                obj.setFocus()
                obj.text = nome
                _esperar_sap_livre()
                return True
            except:
                return False

        def open_for_edit(self):
            print("  ├─ Tentando abrir em modo de 'Criação'...")
            if not try_actions([
                {"path": "wnd[0]/usr/btn%#AUTOTEXT003", "op": "press"},
                {"path": "wnd[0]/tbar[1]/btn[5]", "op": "press"}
            ]):
                raise Exception("Não consegui clicar em Criar.")

            tratar_popup_modal()

            mt, sb = get_statusbar()
            if "EXISTE" in norm_txt(sb) or "EXISTS" in norm_txt(sb):
                print("  ├─ A Role já existe. Alterando para modo de 'Alteração'...")
                if not try_actions([
                    {"path": "wnd[0]/usr/btn%#AUTOTEXT001", "op": "press"},
                    {"path": "wnd[0]/tbar[1]/btn[2]", "op": "press"}
                ]):
                    raise Exception("Role já existe, mas não consegui abrir Alterar.")
                tratar_popup_modal()
                return "CHANGE"
            return "CREATE"

        def set_description(self, desc):
            print("  ├─ Preenchendo a descrição da Role...")
            sap_id, obj = _resolver_id(
                "role_desc_field",
                ["wnd[0]/usr/txtS_AGR_TEXTS-TEXT", "wnd[0]/usr/txtS_AGR_TEXTS-TEXT1", "wnd[0]/usr/txtAGR_TEXTS-TEXT"]
            )
            if not obj:
                return False
            try:
                obj.text = desc
                _send_vkey(0)
                tratar_popup_modal()
                return True
            except:
                return False

        def save(self, log_msg="  └─ Guardando alterações..."):
            print(log_msg)
            try:
                self.sess.findById("wnd[0]").sendVKey(11)
            except:
                try_actions([{"path": "wnd[0]/tbar[0]/btn[11]", "op": "press"}])

            _esperar_sap_livre()
            tratar_popup_modal()

            mt, sb = get_statusbar()
            if sb:
                if mt in ("E", "A"):
                    print(f"     ❌ SAP Erro: {sb}")
                    raise Exception(f"Falha ao guardar: {sb}")
                else:
                    print(f"     ✔️ SAP: {sb}")
            else:
                print("     ✔️ SAP: Operação concluída sem mensagem do sistema.")

        def goto_menu_tab(self):
            print("  ├─ Acedendo à aba 'Menu' (TAB9)...")
            sap_id, obj = _resolver_id("menu_tab", ["wnd[0]/usr/tabsTABSTRIP1/tabpTAB9"])
            if not obj:
                raise Exception("Não consegui abrir a aba Menu (TAB9).")
            try:
                obj.select()
                _esperar_sap_livre()
            except:
                raise Exception("Não consegui abrir a aba Menu (TAB9).")
            tratar_popup_modal()

        def goto_auth_tab(self):
            print("  ├─ Acedendo à aba 'Autorizações' (TAB5)...")
            sap_id, obj = _resolver_id("auth_tab", ["wnd[0]/usr/tabsTABSTRIP1/tabpTAB5"])
            if not obj:
                return False
            try:
                obj.select()
                _esperar_sap_livre()
                return True
            except:
                return False

        def _open_tcode_wizard(self):
            sap_id, obj = _resolver_id(
                "menu_shell",
                [
                    "wnd[0]/usr/tabsTABSTRIP1/tabpTAB9/ssubSUB1:SAPLPRGN_TREE:0321/cntlTOOL_CONTROL/shellcont/shell",
                    "wnd[0]/usr/tabsTABSTRIP1/tabpTAB9/ssubSUB1:SAPLPRGN_TREE:0320/cntlTOOL_CONTROL/shellcont/shell"
                ]
            )
            if obj:
                print("  ├─ Abrindo o Wizard de inserção de Transações (TB03)...")
                obj.pressButton("TB03")
                _esperar_objeto("wnd[1]", timeout=3.0)
                _esperar_sap_livre()
                return True
            raise Exception("Não encontrei o botão TB03.")

        def _fill_tcodes_fast(self, table_base, tcodes):
            if not tcodes:
                return 0
            inserted = 0
            for i, t in enumerate(tcodes):
                if i >= TCODE_BLOCK_SIZE:
                    break
                try:
                    cell_id = f"{table_base}/ctxtS_TCODES-TCODE[0,{i}]"
                    cell = self.sess.findById(cell_id)
                    cell.text = t
                    inserted += 1
                except Exception:
                    continue
            _esperar_sap_livre()
            return inserted

        def add_tcodes(self, tcodes):
            if not tcodes:
                return 0

            print(f"  ├─ Preparando a inserção rápida de {len(tcodes)} TCODE(s)...")
            inserted_total = 0
            total_blocos = max(1, ceil(len(tcodes) / TCODE_BLOCK_SIZE))

            for bloco in range(total_blocos):
                sub = tcodes[bloco * TCODE_BLOCK_SIZE: bloco * TCODE_BLOCK_SIZE + TCODE_BLOCK_SIZE]
                self._open_tcode_wizard()

                table_id, table_obj = _resolver_id(
                    "tcode_wizard_table",
                    ["wnd[1]/usr/tblSAPLPRGN_WIZARDCTRL_TCODE", "wnd[1]/usr/tblSAPLPRGN_WIZARDCTRL_TCODE1"]
                )
                if not table_id:
                    raise Exception("Não encontrei a tabela do Wizard de TCODE.")

                print(f"  ├─ Injetando bloco {bloco+1} na tabela...")
                qtd = self._fill_tcodes_fast(table_id, sub)
                inserted_total += qtd

                try:
                    self.sess.findById("wnd[1]").sendVKey(0)
                except:
                    pass

                _esperar_sap_livre()

                print("  ├─ Confirmando transações (Transferir)...")
                if not try_actions([
                    {"path": "wnd[1]/tbar[0]/btn[19]", "op": "press"},
                    {"path": "wnd[1]/tbar[0]/btn[0]", "op": "press"}
                ]):
                    pass

                wait_wnd1_close(timeout=2.0)
                tratar_popup_modal()

            return inserted_total

        def generate_authorization_profile(self):
            stats = {}
            
            t0 = time.time()
            if not self.goto_auth_tab():
                return False, stats
            tratar_popup_modal()
            stats["aba_autorizacoes"] = time.time() - t0
 
            t0 = time.time()
            print("  ├─ Clicando em 'Sugerir nome de perfil'...")
            try_actions([{
                "path": "wnd[0]/usr/tabsTABSTRIP1/tabpTAB5/ssubSUB1:SAPLPRGN_TREE:0350/btnPROFIL1",
                "op": "press"
            }])
 
            if _safe_find("wnd[1]"):
                print("  ├─ Confirmando a sugestão de nome de perfil no popup...")
                try_actions([{"path": "wnd[1]/tbar[0]/btn[11]", "op": "press"}])
            stats["sugerir_nome_perfil"] = time.time() - t0
 
            t0 = time.time()
            self.save("  ├─ Guardando a Role antes de gerar as autorizações...")
            stats["guardar_alteracoes"] = time.time() - t0
 
            t0 = time.time()
            print("  ├─ Acionando a Geração de Perfil... a aguardar...")
            try_actions([{"path": "wnd[0]/tbar[1]/btn[17]", "op": "press"}])
 
            if _safe_find("wnd[1]"):
                print("  ├─ Confirmando a geração de perfil na janela intermédia...")
                try_actions([{"path": "wnd[1]/usr/btnBUTTON_1", "op": "press"}])
            stats["gerar_perfil"] = time.time() - t0
 
            t0 = time.time()
            if _safe_find("wnd[1]"):
                print("  └─ Fechando popup de logs de autorização...")
                try:
                    self.sess.findById("wnd[1]").sendVKey(0)
                except:
                    pass
                _esperar_sap_livre()
            stats["fechar_popups"] = time.time() - t0
 
            try:
                self.sess.findById("wnd[0]").sendVKey(0)
            except:
                pass
            _esperar_sap_livre()
            tratar_popup_modal()
 
            mt, sb = get_statusbar()
            if sb and mt not in ("E", "A"):
                print(f"     ✔️ SAP: {sb}")
            else:
                print("     ✔️ SAP: Perfil gerado e confirmado.")
            return True, stats

        def execute_transport_and_exit(self, req_num):
            if req_num:
                print("  ├─ Recuando para a base da PFCG para pedir Transporte (F3 x2)...")
                for _ in range(2):
                    try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
                    tratar_popup_modal()

                print("  ├─ Acedendo ao Menu Função -> Transporte...")
                try_actions([{"path": "wnd[0]/mbar/menu[0]/menu[9]", "op": "select"}])
                tratar_popup_modal()

                print("  ├─ Clicando em Executar transporte...")
                try_actions([{"path": "wnd[0]/tbar[1]/btn[8]", "op": "press"}])

                field_id, req_field = _resolver_id_esperando(
                    "transport_req_field",
                    ["wnd[1]/usr/ctxtKO008-TRKORR"],
                    timeout=3.0
                )

                print(f"  ├─ Injetando a Request ({req_num}) diretamente no popup...")
                if req_field:
                    req_field.text = str(req_num)

                try_actions([{"path": "wnd[1]/tbar[0]/btn[0]", "op": "press"}])
                tratar_popup_modal()

                mt, sb = get_statusbar()
                if sb and mt not in ("E", "A"):
                    print(f"     ✔️ SAP: {sb}")
                else:
                    print("     ✔️ SAP: Transporte associado com sucesso!")

            print("  └─ Regressando em segurança ao ecrã principal SAP Easy Access (F3)...")
            for _ in range(3):
                try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
                tratar_popup_modal()

    ###################################################################################
    # BLOCO 4: EXECUÇÃO ESTRUTURADA
    ###################################################################################
    pfcg = PFCGPage(session)
    resultados = {}
    role_metrics = {}

    total_roles = len(roles_agrupadas)

    with Progress(
        TextColumn("[bold cyan]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total})"),
        TimeElapsedColumn(),
        transient=False,
    ) as progress:
        task_roles = progress.add_task("A processar roles...", total=total_roles)

        for idx, rr in enumerate(roles_agrupadas, start=1):
            nome, desc = (rr["AGR_NAME"] or "").strip(), (rr["TEXT"] or "").strip()
            tcodes = rr["TCODE_LIST"]

            progress.update(task_roles, description=f"A processar role: {nome}")

            # Reset counts in environment variables for the current role
            os.environ["CURRENT_ROLE_DOC_WARN"] = "0"
            os.environ["CURRENT_ROLE_POLLER_TIMEOUT"] = "0"

            if doc_session and doc_session.enabled:
                doc_session.start_role_section(nome, desc, len(tcodes))

            print("\n======================================================================")
            print(f"▶ [{idx}/{len(roles_agrupadas)}] INICIANDO ROLE: {nome} | TCODEs: {len(tcodes)}")
            print("======================================================================")

            tempo_inicio_role = time.time()

            # Subaction timing variables
            t_abrir_pfcg = 0.0
            t_nome_role = 0.0
            t_modo_edicao = 0.0
            t_descricao = 0.0
            t_guardar_inicial = 0.0
            
            t_aba_menu = 0.0
            t_tcodes = 0.0
            t_guardar_tcodes = 0.0
            t_evidencia_menu = 0.0
            
            t_aba_autorizacoes = 0.0
            t_sugerir_nome_perfil = 0.0
            t_guardar_alteracoes_auth = 0.0
            t_gerar_perfil = 0.0
            t_fechar_popups = 0.0
            t_evidencia_auth = 0.0
            
            t_transporte = 0.0

            t_etapa1 = 0.0
            t_etapa2 = 0.0
            t_etapa3 = 0.0
            t_etapa4 = 0.0

            try:
                # [Etapa 1] Preparação e Dados Básicos
                t_inicio_etapa1 = time.time()
                print("\n[Etapa 1] Preparação e Dados Básicos")
                
                t0 = time.time()
                pfcg.open()
                t_abrir_pfcg = time.time() - t0
                
                t0 = time.time()
                set_role_ok = pfcg.set_role_name(nome)
                t_nome_role = time.time() - t0
                if not set_role_ok:
                    raise Exception("Falha ao escrever AGR_NAME.")
                
                t0 = time.time()
                modo = pfcg.open_for_edit()
                t_modo_edicao = time.time() - t0
                
                t0 = time.time()
                pfcg.set_description(desc)
                t_descricao = time.time() - t0
                
                t0 = time.time()
                pfcg.save("  └─ Guardando alterações iniciais...")
                t_guardar_inicial = time.time() - t0
                t_etapa1 = time.time() - t_inicio_etapa1

                t_inicio_etapa2 = time.time()
                t_inicio_etapa3 = time.time()
                t_inicio_etapa4 = time.time()
                
                skip_rest = False
                tcodes_para_inserir = tcodes
                tcodes_existentes = set()
                fonte_final = None
                qtd_ins = 0

                if modo == "CHANGE":
                    print("\n[Etapa 2] Validação de Transações")
                    
                    tcodes_input = {normalizar_tcode(tc) for tc in tcodes if str(tc).strip()}
                    tcodes_sap = set()
                    
                    # 1. Consultar AGR_TCODES
                    res_agr = consultar_tcodes_agr_tcodes(session, nome)
                    
                    if res_agr["ok"]:
                        tcodes_sap = res_agr["tcodes"]
                        fonte_final = "AGR_TCODES"
                    else:
                        print(f"├─ ERRO técnico na consulta AGR_TCODES: {res_agr['mensagem']}")
                        raise Exception("Validação inconclusiva: não foi possível ler AGR_TCODES. Inserção bloqueada para evitar duplicação.")
                        
                    # 3. Calcular conjuntos delta
                    tcodes_ja_existentes = tcodes_input.intersection(tcodes_sap)
                    tcodes_para_inserir_set = tcodes_input.difference(tcodes_sap)
                    tcodes_extra_sap = tcodes_sap.difference(tcodes_input)
                    
                    tcodes_existentes = list(tcodes_sap)
                    
                    # 4. Logs detalhados das transações analisadas (Parte 5)
                    print(f"├─ TCODEs encontradas na {fonte_final}:")
                    if tcodes_sap:
                        for tc_item in sorted(list(tcodes_sap))[:10]:
                            print(f"│  └─ {tc_item}")
                        if len(tcodes_sap) > 10:
                            print(f"│  └─ ... (+ {len(tcodes_sap)-10} adicionais)")
                    else:
                        print("│  └─ nenhuma")
                        
                    print("├─ TCODEs no ficheiro:")
                    for tc_item in sorted(list(tcodes_input))[:10]:
                        print(f"│  └─ {tc_item}")
                    if len(tcodes_input) > 10:
                        print(f"│  └─ ... (+ {len(tcodes_input)-10} adicionais)")
                        
                    print("├─ Já existentes:")
                    if tcodes_ja_existentes:
                        for tc_item in sorted(list(tcodes_ja_existentes))[:10]:
                            print(f"│  └─ {tc_item}")
                        if len(tcodes_ja_existentes) > 10:
                            print(f"│  └─ ... (+ {len(tcodes_ja_existentes)-10} adicionais)")
                    else:
                        print("│  └─ nenhuma")
                        
                    print("├─ Novas a inserir:")
                    if tcodes_para_inserir_set:
                        for tc_item in sorted(list(tcodes_para_inserir_set))[:10]:
                            print(f"│  └─ {tc_item}")
                        if len(tcodes_para_inserir_set) > 10:
                            print(f"│  └─ ... (+ {len(tcodes_para_inserir_set)-10} adicionais)")
                    else:
                        print("│  └─ nenhuma")
                        
                    print(f"└─ Fonte final da validação: {fonte_final}")
                    
                    t_etapa2_validacao = time.time() - t_inicio_etapa2
                    
                    if not tcodes_para_inserir_set:
                        print(f"\n🟢 SUCESSO: Sem alterações necessárias. Todas as transações já estavam atribuídas na AGR_TCODES.")
                        resultados[nome] = {
                            "STATUS": "CONCLUIDO",
                            "MSG": "Sem alterações: todas as transações já estavam atribuídas na AGR_TCODES.",
                            "TIMESTEMP": now_ts()
                        }
                        skip_rest = True
                        if doc_session and doc_session.enabled:
                            doc_session.add_role_summary(nome, desc, len(tcodes), "Concluída (Sem alterações)", "00m 00s")
                        
                        print("  └─ Regressando ao ecrã principal SAP Easy Access (F3)...")
                        for _ in range(2):
                            try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
                            tratar_popup_modal()
                        
                        t_etapa2 = t_etapa2_validacao
                        t_etapa3 = 0.0
                        t_etapa4 = 0.0
                    else:
                        tcodes_para_inserir = list(tcodes_para_inserir_set)

                if not skip_rest:
                    # [Etapa 2] Atribuição de Transações (Aba Menu)
                    print("\n[Etapa 2] Atribuição de Transações (Aba Menu)")
                    if modo == "CHANGE":
                        print("├─ Inserindo apenas TCODEs novas...")
                    
                    t_inicio_etapa2_insercao = time.time()
                    
                    t0 = time.time()
                    pfcg.goto_menu_tab()
                    t_aba_menu = time.time() - t0
                    
                    t0 = time.time()
                    qtd_ins = pfcg.add_tcodes(tcodes_para_inserir)
                    t_tcodes = time.time() - t0
                    
                    t0 = time.time()
                    pfcg.save("  └─ Guardando Transações inseridas...")
                    t_guardar_tcodes = time.time() - t0

                    t0 = time.time()
                    if doc_session and doc_session.enabled:
                        shot = capture_screenshot(nome, "menu_transacoes_gravadas", idx)
                        if shot:
                            doc_session.add_evidence(
                                nome,
                                "Evidência 1 — Transações atribuídas e gravadas",
                                f"Aba Menu da role {nome} após atribuição das transações e confirmação de gravação no SAP.",
                                shot
                            )
                    t_evidencia_menu = time.time() - t0
                    
                    # Consolidar tempo de etapa2
                    if modo == "CHANGE":
                        t_etapa2 = t_etapa2_validacao + (time.time() - t_inicio_etapa2_insercao)
                    else:
                        t_etapa2 = time.time() - t_inicio_etapa2

                    # [Etapa 3] Geração do Perfil de Autorizações
                    print("\n[Etapa 3] Geração do Perfil de Autorizações")
                    success_auth, t_stats_auth = pfcg.generate_authorization_profile()
                    if not success_auth:
                        raise Exception("Falha na geração do perfil de autorizações.")
                    
                    t_aba_autorizacoes = t_stats_auth.get("aba_autorizacoes", 0.0)
                    t_sugerir_nome_perfil = t_stats_auth.get("sugerir_nome_perfil", 0.0)
                    t_guardar_alteracoes_auth = t_stats_auth.get("guardar_alteracoes", 0.0)
                    t_gerar_perfil = t_stats_auth.get("gerar_perfil", 0.0)
                    t_fechar_popups = t_stats_auth.get("fechar_popups", 0.0)

                    t0 = time.time()
                    if doc_session and doc_session.enabled:
                        shot = capture_screenshot(nome, "perfil_gerado", idx)
                        if shot:
                            doc_session.add_evidence(
                                nome,
                                "Evidência 2 — Perfil de autorizações gerado",
                                f"Perfil de autorizações da role {nome} gerado e confirmado no SAP.",
                                shot
                            )
                    t_evidencia_auth = time.time() - t0
                    t_etapa3 = time.time() - t_inicio_etapa3

                    # [Etapa 4] Ordem de Transporte e Encerramento
                    print("\n[Etapa 4] Ordem de Transporte e Encerramento")
                    
                    t0 = time.time()
                    pfcg.execute_transport_and_exit(request_transporte)
                    t_transporte = time.time() - t0
                    t_etapa4 = time.time() - t_inicio_etapa4

                    tempo_decorrido_role = time.time() - tempo_inicio_role
                    str_tempo = formatar_tempo(tempo_decorrido_role)

                    msg_transporte = f" | Add Req {request_transporte}" if request_transporte else ""
                    if modo == "CHANGE":
                        msg_status = f"Sucesso (CHANGE) | Inseridas {len(tcodes_para_inserir)}/{len(tcodes)} TCODEs | Já existentes {len(tcodes_ja_existentes)}/{len(tcodes)} | Fonte validação: {fonte_final}{msg_transporte}."
                    else:
                        msg_status = f"Sucesso (CREATE) | {qtd_ins}/{len(tcodes)} TCODEs | Perfil Gerado{msg_transporte}."

                    resultados[nome] = {
                        "STATUS": "CONCLUIDO",
                        "MSG": msg_status,
                        "TIMESTEMP": now_ts()
                    }
                    print(f"\n🟢 SUCESSO: Role tratada por completo! Inseridas: {len(tcodes_para_inserir)} | Já existentes: {len(tcodes_ja_existentes)} ⏱️ (Tempo: {str_tempo})")
                    print("----------------------------------------------------------------------")

                    if doc_session and doc_session.enabled:
                        doc_session.add_role_summary(nome, desc, len(tcodes), "Concluída", str_tempo)
                else:
                    print("  └─ Regressando ao ecrã principal SAP Easy Access (F3)...")
                    for _ in range(2):
                        try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
                        tratar_popup_modal()
                    t_etapa2 = 0.0
                    t_etapa3 = 0.0
                    t_etapa4 = 0.0

            except Exception as e:
                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)

                err = str(e)
                if "Validação inconclusiva" in err:
                    err = "Validação inconclusiva: não foi possível confirmar TCODEs existentes. Inserção bloqueada para evitar duplicação."
                else:
                    mt, sb = get_statusbar()
                    if mt in ("E", "A"):
                        err = sb
                resultados[nome] = {"STATUS": "ERRO", "MSG": err, "TIMESTEMP": now_ts()}

                print(f"\n🔴 ERRO: {err} ⏱️ (Tempo: {str_tempo})")
                print("----------------------------------------------------------------------")

                if doc_session and doc_session.enabled:
                    doc_session.add_role_summary(nome, desc, len(tcodes), "Não concluída", str_tempo)

                try:
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
                    _send_vkey(0)
                except:
                    pass

            finally:
                t_total_role_s = int(time.time() - tempo_inicio_role)
                doc_warn_count = int(os.environ.get("CURRENT_ROLE_DOC_WARN", 0))
                poller_timeout_count = int(os.environ.get("CURRENT_ROLE_POLLER_TIMEOUT", 0))
                
                role_metrics[nome] = {
                    "total": t_total_role_s,
                    "etapa1": int(t_etapa1),
                    "etapa2": int(t_etapa2),
                    "etapa3": int(t_etapa3),
                    "etapa4": int(t_etapa4),
                    "doc_warn": doc_warn_count,
                    "poller_timeout": poller_timeout_count
                }
                
                print(f"[METRIC] Role {nome} | total={t_total_role_s}s | etapa1={int(t_etapa1)}s | etapa2={int(t_etapa2)}s | etapa3={int(t_etapa3)}s | etapa4={int(t_etapa4)}s | doc_warn={doc_warn_count} | poller_timeout={poller_timeout_count}")
                print(f"[METRIC_DETAIL] Subactions: abrir_pfcg={t_abrir_pfcg:.2f}s, nome_role={t_nome_role:.2f}s, modo_edicao={t_modo_edicao:.2f}s, descricao={t_descricao:.2f}s, guardar_inicial={t_guardar_inicial:.2f}s, aba_menu={t_aba_menu:.2f}s, tcodes={t_tcodes:.2f}s, guardar_tcodes={t_guardar_tcodes:.2f}s, evidencia_menu={t_evidencia_menu:.2f}s, aba_autorizacoes={t_aba_autorizacoes:.2f}s, sugerir_nome_perfil={t_sugerir_nome_perfil:.2f}s, guardar_alteracoes_auth={t_guardar_alteracoes_auth:.2f}s, gerar_perfil={t_gerar_perfil:.2f}s, fechar_popups={t_fechar_popups:.2f}s, evidencia_auth={t_evidencia_auth:.2f}s, transporte={t_transporte:.2f}s")
                
                # Checkpoint Excel save
                try:
                    gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados)
                except Exception as cp_exc:
                    print(f"  ⚠️ Erro ao salvar checkpoint do Excel: {cp_exc}")

                progress.advance(task_roles)

    ###################################################################################
    # BLOCO 5: GRAVAR EXCEL E TEMPO TOTAL
    ###################################################################################
    if gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados):
        print("\n💾 Resultados gravados com sucesso no Excel!")
    else:
        print("\n❌ Erro ao gravar resultados finais no Excel.")

    # Imprimir Resumo Comparativo das Roles
    if role_metrics:
        try:
            fastest_role = min(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            slowest_role = max(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            
            print("\n=======================================================")
            print("📊 RESUMO COMPARATIVO DE PERFORMANCE DAS ROLES")
            print("=======================================================")
            print(f" - Role mais rápida: {fastest_role} ({role_metrics[fastest_role]['total']}s)")
            print(f" - Role mais lenta: {slowest_role} ({role_metrics[slowest_role]['total']}s)")
            
            print("\n⏱️ Etapa mais lenta por Role:")
            for r_name, m in role_metrics.items():
                stages = {
                    "Etapa 1": m["etapa1"],
                    "Etapa 2": m["etapa2"],
                    "Etapa 3": m["etapa3"],
                    "Etapa 4": m["etapa4"]
                }
                slowest_stage = max(stages.keys(), key=lambda k: stages[k])
                print(f"   • {r_name}: {slowest_stage} ({stages[slowest_stage]}s)")
                
            print("\n⚠️ Alertas técnicos acumulados por Role:")
            for r_name, m in role_metrics.items():
                print(f"   • {r_name}: DOC_WARN={m['doc_warn']} | timeouts={m['poller_timeout']}")
            print("=======================================================\n")
        except Exception as summary_exc:
            print(f"\n[DEBUG] Falha ao gerar resumo comparativo: {summary_exc}")

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")

    # Finalizar documentação funcional se ativada
    if doc_session and doc_session.enabled:
        try:
            final_path = doc_session.finalize()
            if final_path:
                print(f"[DOC] Documento funcional gerado: {final_path}")
        except Exception as doc_exc:
            print(f"[DOC_WARN] Falha ao finalizar documento funcional: {doc_exc}")

    print("🔁 Fim.")
    return True


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD"])
    parser.add_argument("--xlsx")
    parser.add_argument("--request", help="Número da Request de Transporte (Opcional)")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    parser.add_argument("--nome_pasta", help="Nome da pasta para documentação funcional")
    args = parser.parse_args()

    env_cli = args.ambiente or (input("Ambiente (DEV/QAD/PRD): ").strip().upper() or "DEV")

    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        request_transporte=args.request,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm),
        nome_pasta=args.nome_pasta
    )