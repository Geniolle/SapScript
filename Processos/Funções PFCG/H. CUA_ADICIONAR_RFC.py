# -*- coding: utf-8 -*-

###################################################################################
# PROCESSO: Adicionar Função CUA via RFC  (sheet = nome do .py SEM o prefixo)
# Ex.: "H. CUA_ADICIONAR.py"  →  Sheet "CUA_ADICIONAR"
#
# ESTRUTURA ESPERADA DA SHEET:
# ID | UTILIZADOR | SISTEMA | AGR_NAME | STATUS | MSG | TIMESTEMP
#
# PADRÃO APLICADO:
# - processamento totalmente via RFC
# - guarda a última mensagem relevante do SAP
# - grava no Excel atualizando APENAS STATUS / MSG / TIMESTEMP
# - preserva formatação, fórmulas, filtros e restantes colunas
# - popup sem diretório fixo (abre no último local usado)
###################################################################################

###################################################################################
# BLOCO 1: IMPORTAÇÕES
###################################################################################

import os
import time
import unicodedata
from datetime import date, datetime

import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from dotenv import load_dotenv

try:
    from pyrfc import Connection
except Exception:
    Connection = None

load_dotenv(os.path.join(os.getcwd(), ".env"))

###################################################################################
# BLOCO 2: NOME DO SCRIPT / SHEET / MAPA DE SISTEMAS
###################################################################################

try:
    NOME_SCRIPT = os.path.splitext(os.path.basename(__file__))[0]
except NameError:
    NOME_SCRIPT = "H. CUA_ADICIONAR"

NOME_SHEET = NOME_SCRIPT.split(".", 1)[-1].strip() if "." in NOME_SCRIPT else NOME_SCRIPT

MAPA_SISTEMA = {
    "DEV": "S4D",
    "QAD": "S4Q",
    "PRD": "S4P",
    "CUA": "SPA",
}

###################################################################################
# BLOCO 3: UTILITÁRIOS
###################################################################################

DEBUG_SAP_CONTROLS = os.environ.get("DEBUG_SAP_CONTROLS", "FALSE").upper() == "TRUE"

def formatar_tempo(segundos):
    m = int(segundos // 60)
    s = int(segundos % 60)
    return f"{m:02d}m {s:02d}s"


def agora_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalizar_coluna(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("utf-8")
        .strip()
        .upper()
    )


def normalizar_valor(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("utf-8")
        .strip()
        .upper()
    )


def texto_limpo(valor):
    if pd.isna(valor):
        return ""
    txt = str(valor).strip()
    if txt.lower() in ("nan", "none", "<na>"):
        return ""
    return txt


def _flag_presente(valor):
    return "sim" if texto_limpo(valor) else "não"


def _resumo_secreto(valor):
    txt = texto_limpo(valor)
    return f"sim (len={len(txt)})" if txt else "não"


def _first_env_value(pairs):
    for chave, fallback in pairs:
        valor = texto_limpo(os.getenv(chave, fallback or ""))
        if valor:
            return valor, chave
    return "", ""


def valor_vazio(valor):
    return texto_limpo(valor) == ""


def chave_id(valor):
    if pd.isna(valor):
        return ""

    if isinstance(valor, int):
        return str(valor)

    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else str(valor).strip()

    txt = str(valor).strip()

    if txt.endswith(".0"):
        base = txt[:-2]
        if base.isdigit():
            return base

    return txt


def mapear_cabecalhos_openpyxl(ws):
    mapa = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val is None:
            continue
        mapa[normalizar_coluna(val)] = c
    return mapa


###################################################################################
# BLOCO 4: SELEÇÃO DO FICHEIRO
###################################################################################

def selecionar_ficheiro_excel():
    """
    Popup sem diretório fixo.
    O Windows tende a abrir no último local utilizado.
    """
    try:
        root = tk.Tk()
        root.withdraw()
        root.update_idletasks()
        root.attributes("-topmost", True)

        caminho = filedialog.askopenfilename(
            title="Selecione o ficheiro Excel",
            filetypes=[
                ("Ficheiros Excel", "*.xlsx *.xlsm"),
                ("Todos os ficheiros", "*.*"),
            ],
        )

        root.destroy()

        if not caminho:
            print("⚠️ Seleção cancelada pelo utilizador.")
            return None

        ext = os.path.splitext(caminho)[1].lower()
        if ext not in (".xlsx", ".xlsm"):
            print("❌ Apenas ficheiros .xlsx e .xlsm são suportados neste processo.")
            return None

        print(f"✅ Ficheiro a processar: {caminho}")
        return caminho

    except Exception as e:
        print(f"❌ Erro ao abrir o popup: {e}")
        return None


###################################################################################
# BLOCO 5: LEITURA DO EXCEL
###################################################################################

def ler_ficheiro(caminho_ficheiro, nome_sheet):
    """
    Lê a sheet alvo, normaliza cabeçalhos e valida estrutura obrigatória.
    Harmoniza variantes para:
    ID | UTILIZADOR | SISTEMA | AGR_NAME | STATUS | MSG | TIMESTEMP
    """
    if not caminho_ficheiro or not os.path.exists(caminho_ficheiro):
        print("❌ Caminho inválido ou ficheiro inexistente.")
        return None

    try:
        ext = os.path.splitext(caminho_ficheiro)[1].lower()
        keep_vba = ext == ".xlsm"

        # 1. Carregar em openpyxl para verificar e criar cabeçalhos em falta
        wb = load_workbook(
            caminho_ficheiro,
            read_only=False,
            data_only=False,
            keep_vba=keep_vba,
        )
        sheets = wb.sheetnames
        if nome_sheet not in sheets:
            print(f"❌ Sheet '{nome_sheet}' não encontrada. Disponíveis: {', '.join(sheets)}")
            wb.close()
            return None

        ws = wb[nome_sheet]
        
        # Obter os cabeçalhos reais da linha 1
        headers_linha_1 = [c.value for c in ws[1]]
        headers_norm = [normalizar_coluna(h) if h is not None else "" for h in headers_linha_1]

        renames = {
            "USER": "UTILIZADOR",
            "USERNAME": "UTILIZADOR",
            "SYSTEM": "SISTEMA",
            "FUNCAO": "AGR_NAME",
            "FUNÇÃO": "AGR_NAME",
            "ROLE": "AGR_NAME",
            "NOME FUNCAO": "AGR_NAME",
            "NOME FUNÇAO": "AGR_NAME",
            "NOME FUNÇÂO": "AGR_NAME",
            "AGRNAME": "AGR_NAME",
            "TIMESTAMP": "TIMESTEMP",
        }

        resolved_headers = []
        for h in headers_norm:
            resolved_headers.append(renames.get(h, h))

        # Colunas de entrada obrigatórias
        colunas_entrada = {"ID", "UTILIZADOR", "SISTEMA", "AGR_NAME"}
        # Colunas de saída que podem ser criadas
        colunas_saida = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

        # Verificar se as colunas de entrada existem
        falta_entrada = [c for c in colunas_entrada if c not in resolved_headers]
        if falta_entrada:
            print(f"❌ Colunas obrigatórias de entrada em falta: {', '.join(falta_entrada)}")
            wb.close()
            return None

        # Se as colunas de entrada existem, podemos prosseguir e criar as colunas de saída em falta
        modificou = False
        last_col = len(headers_norm)
        for col_key, col_name in colunas_saida.items():
            if col_key not in resolved_headers:
                last_col += 1
                ws.cell(row=1, column=last_col, value=col_name)
                print(f"➕ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                modificou = True

        # Validar e remover espaços do nome da Role (AGR_NAME) diretamente no Excel
        col_idx_agr = resolved_headers.index("AGR_NAME") + 1
        modificou_roles = False
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=col_idx_agr).value
            if cell_val is not None:
                agr = str(cell_val).strip()
                if " " in agr:
                    orig = agr
                    agr = agr.replace(" ", "")
                    ws.cell(row=r, column=col_idx_agr, value=agr)
                    print(f"⚠️ Espaços detetados e removidos da Role: '{orig}' -> '{agr}' (atualizado no Excel)")
                    modificou_roles = True

        if modificou or modificou_roles:
            try:
                wb.save(caminho_ficheiro)
                print("💾 Ficheiro Excel guardado com as correções de cabeçalhos/roles.")
            except Exception as e:
                wb.close()
                print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas/roles: {e}")
                print("💡 Certifique-se de que o ficheiro Excel não está aberto no Excel e tente novamente.")
                return None

        wb.close()

        # 2. Carregar o DataFrame em pandas para uso no script
        df = pd.read_excel(caminho_ficheiro, sheet_name=nome_sheet, dtype=object)
        df.columns = [normalizar_coluna(c) for c in df.columns]

        df.rename(
            columns={
                "USER": "UTILIZADOR",
                "USERNAME": "UTILIZADOR",
                "SYSTEM": "SISTEMA",
                "FUNCAO": "AGR_NAME",
                "FUNÇÃO": "AGR_NAME",
                "ROLE": "AGR_NAME",
                "NOME FUNCAO": "AGR_NAME",
                "NOME FUNÇAO": "AGR_NAME",
                "NOME FUNÇÂO": "AGR_NAME",
                "AGRNAME": "AGR_NAME",
                "TIMESTAMP": "TIMESTEMP",
            },
            inplace=True,
        )

        obrigatorias = ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]
        falta = [c for c in obrigatorias if c not in df.columns]
        if falta:
            print(f"❌ Colunas obrigatórias em falta: {', '.join(falta)}")
            return None

        for c in ["UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]:
            df[c] = df[c].apply(texto_limpo)

        df["CHAVE_ID"] = df["ID"].apply(chave_id)

        print(f"📄 Sheet carregada: '{nome_sheet}' | Registos: {len(df)}")
        return df

    except Exception as e:
        print(f"❌ Erro ao ler a sheet: {e}")
        return None


###################################################################################
# BLOCO 6: RFC / SAP CUA
###################################################################################



def obter_credenciais_rfc_cua():
    ashost, chave_ashost = _first_env_value(
        [
            ("SAP_ASHOST_CUA", ""),
            ("SAP_ASHOST", ""),
        ]
    )
    sysnr, chave_sysnr = _first_env_value(
        [
            ("SAP_SYSNR_CUA", "00"),
            ("SAP_SYSNR", "00"),
        ]
    )
    user, chave_user = _first_env_value(
        [
            ("SAP_USER_CUA", ""),
            ("SAP_USER_SPACLNT001", ""),
            ("SAP_USER", ""),
        ]
    )
    client, chave_client = _first_env_value(
        [
            ("SAP_CLIENT_CUA", "001"),
            ("SAP_CLIENT_SPACLNT001", "001"),
            ("SAP_CLIENT", "001"),
        ]
    )
    lang, chave_lang = _first_env_value(
        [
            ("SAP_LANGUAGE_CUA", ""),
            ("SAP_CUA_LANGUAGE", ""),
            ("SAP_LANGUAGE", "PT"),
        ]
    )
    pwd_candidates = [
        ("SAP_PASSWORD_CUA", os.getenv("SAP_PASSWORD_CUA", "")),
        ("SAP_PASSWORD_SPACLNT001", os.getenv("SAP_PASSWORD_SPACLNT001", "")),
        ("SAP_PASSWD", os.getenv("SAP_PASSWD", "")),
        ("SAP_PASSWORD", os.getenv("SAP_PASSWORD", "")),
    ]
    mshost, chave_mshost = _first_env_value(
        [
            ("SAP_MSHOST_CUA", ""),
            ("SAP_MSHOST", ""),
        ]
    )
    msserv, chave_msserv = _first_env_value(
        [
            ("SAP_MSSERV_CUA", ""),
            ("SAP_MSSERV", ""),
        ]
    )
    r3name, chave_r3name = _first_env_value(
        [
            ("SAP_R3NAME_CUA", ""),
            ("SAP_R3NAME", ""),
        ]
    )
    group, chave_group = _first_env_value(
        [
            ("SAP_GROUP_CUA", ""),
            ("SAP_GROUP", ""),
        ]
    )
    saprouter, chave_saprouter = _first_env_value(
        [
            ("SAP_SAPROUTER_CUA", ""),
            ("SAP_SAPROUTER", ""),
        ]
    )

    return (
        ashost,
        sysnr,
        client,
        user,
        lang,
        pwd_candidates,
        {
            "ASHOST": chave_ashost,
            "SYSNR": chave_sysnr,
            "USER": chave_user,
            "CLIENT": chave_client,
            "LANG": chave_lang,
            "MSHOST": chave_mshost,
            "MSSERV": chave_msserv,
            "R3NAME": chave_r3name,
            "GROUP": chave_group,
            "SAPROUTER": chave_saprouter,
        },
        {
            "MSHOST": mshost,
            "MSSERV": msserv,
            "R3NAME": r3name,
            "GROUP": group,
            "SAPROUTER": saprouter,
        },
    )


def conectar_rfc_cua():
    if Connection is None:
        raise RuntimeError("pyrfc não está disponível neste ambiente.")

    ashost, sysnr, client, user, lang, pwd_candidates, origem, balanco = obter_credenciais_rfc_cua()

    print("[RFC CUA] Diagnóstico de credenciais:")
    print(f"[RFC CUA]  - ASHOST: {_flag_presente(ashost)} | origem={origem['ASHOST'] or 'n/a'}")
    print(f"[RFC CUA]  - SYSNR: {_flag_presente(sysnr)} | origem={origem['SYSNR'] or 'n/a'}")
    print(f"[RFC CUA]  - USER: {_flag_presente(user)} | origem={origem['USER'] or 'n/a'}")
    print(f"[RFC CUA]  - CLIENT: {_flag_presente(client)} | origem={origem['CLIENT'] or 'n/a'}")
    print(f"[RFC CUA]  - LANG: {_flag_presente(lang)} | origem={origem['LANG'] or 'n/a'}")
    print(f"[RFC CUA]  - MSHOST: {_flag_presente(balanco['MSHOST'])} | origem={origem['MSHOST'] or 'n/a'}")
    print(f"[RFC CUA]  - MSSERV: {_flag_presente(balanco['MSSERV'])} | origem={origem['MSSERV'] or 'n/a'}")
    print(f"[RFC CUA]  - R3NAME: {_flag_presente(balanco['R3NAME'])} | origem={origem['R3NAME'] or 'n/a'}")
    print(f"[RFC CUA]  - GROUP: {_flag_presente(balanco['GROUP'])} | origem={origem['GROUP'] or 'n/a'}")
    print(f"[RFC CUA]  - SAPROUTER: {_flag_presente(balanco['SAPROUTER'])} | origem={origem['SAPROUTER'] or 'n/a'}")
    for chave, valor in pwd_candidates:
        print(f"[RFC CUA]  - {chave}: {_resumo_secreto(valor)}")

    if not ashost or not user:
        raise RuntimeError(
            "Credenciais RFC CUA incompletas. "
            "São necessários SAP_ASHOST e SAP_USER."
        )

    passwd_candidatos = [
        (chave, texto_limpo(valor))
        for chave, valor in pwd_candidates
        if texto_limpo(valor)
    ]

    if not passwd_candidatos:
        raise RuntimeError(
            "Credenciais RFC CUA incompletas. "
            "É necessário pelo menos uma password RFC "
            "(SAP_PASSWORD_SPACLNT001, SAP_PASSWD ou SAP_PASSWORD)."
        )

    print(
        f"🛰️ A ligar via RFC ao CUA (Host: {ashost}, Client: {client}, User: {user})..."
    )
    print(
        f"[RFC CUA] Parâmetros efectivos: ASHOST={ashost} | SYSNR={sysnr} | CLIENT={client} | LANG={lang}"
    )
    ultimo_erro = None
    for chave_password, passwd in passwd_candidatos:
        print(f"[RFC CUA] Password em teste: {chave_password} (valor oculto)")
        try:
            conn_kwargs = {
                "client": client,
                "user": user,
                "passwd": passwd,
                "lang": lang,
            }

            if balanco["MSHOST"] and balanco["R3NAME"]:
                conn_kwargs.update(
                    {
                        "mshost": balanco["MSHOST"],
                        "r3name": balanco["R3NAME"],
                    }
                )
                if balanco["MSSERV"]:
                    conn_kwargs["msserv"] = balanco["MSSERV"]
                if balanco["GROUP"]:
                    conn_kwargs["group"] = balanco["GROUP"]
                if balanco["SAPROUTER"]:
                    conn_kwargs["saprouter"] = balanco["SAPROUTER"]
                print(
                    f"[RFC CUA] Modo RFC: message server | MSHOST={balanco['MSHOST']} | R3NAME={balanco['R3NAME']}"
                )
            else:
                conn_kwargs.update(
                    {
                        "ashost": ashost,
                        "sysnr": sysnr,
                    }
                )
                if balanco["SAPROUTER"]:
                    conn_kwargs["saprouter"] = balanco["SAPROUTER"]
                print(
                    f"[RFC CUA] Modo RFC: directo | ASHOST={ashost} | SYSNR={sysnr}"
                )

            return Connection(**conn_kwargs)
        except Exception as exc:
            ultimo_erro = exc
            print(f"[RFC CUA] Falha com {chave_password}: {exc}")
            continue

    raise RuntimeError(
        "Não foi possível autenticar via RFC com nenhuma das passwords configuradas."
    ) from ultimo_erro


def _bapi_has_errors(return_tab):
    for row in return_tab or []:
        tipo = normalizar_valor(row.get("TYPE", ""))
        if tipo in ("E", "A", "X"):
            return True
    return False


def _bapi_messages(return_tab):
    msgs = []
    for row in return_tab or []:
        tipo = texto_limpo(row.get("TYPE", ""))
        msg = texto_limpo(row.get("MESSAGE", "")) or texto_limpo(row.get("MESSAGE_V1", ""))
        if tipo or msg:
            msgs.append(f"{tipo}: {msg}".strip(": "))
    return " | ".join([m for m in msgs if m])


def _data_sap_yyyymmdd(valor=None):
    if not valor:
        return datetime.now().strftime("%Y%m%d")
    if isinstance(valor, datetime):
        return valor.strftime("%Y%m%d")
    if isinstance(valor, date):
        return valor.strftime("%Y%m%d")
    txt = str(valor).strip()
    for fmt in ("%Y%m%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(txt, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return datetime.now().strftime("%Y%m%d")


def ler_funcoes_existentes_rfc(conn, utilizador):
    res = conn.call(
        "BAPI_USER_LOCACTGROUPS_READ",
        USERNAME=utilizador,
        WITH_TEXT="X",
    )
    retorno = res.get("RETURN", [])
    if _bapi_has_errors(retorno):
        raise RuntimeError(_bapi_messages(retorno) or "Erro a ler funções existentes via RFC.")

    rows = res.get("ACTIVITYGROUPS", []) or []
    existentes = set()
    normalizados = []

    for row in rows:
        row_norm = {
            "SUBSYSTEM": texto_limpo(row.get("SUBSYSTEM", "")),
            "AGR_NAME": texto_limpo(row.get("AGR_NAME", "")),
            "FROM_DAT": texto_limpo(row.get("FROM_DAT", "")),
            "TO_DAT": texto_limpo(row.get("TO_DAT", "")),
            "AGR_TEXT": texto_limpo(row.get("AGR_TEXT", "")),
            "ORG_FLAG": texto_limpo(row.get("ORG_FLAG", "")),
        }
        if row_norm["SUBSYSTEM"] or row_norm["AGR_NAME"]:
            existentes.add((row_norm["SUBSYSTEM"].upper(), row_norm["AGR_NAME"].upper()))
            normalizados.append(row_norm)

    return normalizados, existentes


def atribuir_funcoes_rfc_cua(conn, utilizador, sistema, roles_list):
    sistema_norm = texto_limpo(sistema).upper()
    roles_norm = []
    for role in roles_list:
        role_norm = texto_limpo(role).upper()
        if role_norm and role_norm not in roles_norm:
            roles_norm.append(role_norm)

    existentes_rows, existentes_pairs = ler_funcoes_existentes_rfc(conn, utilizador)

    rows_envio = []
    vistos_envio = set()

    for row in existentes_rows:
        key = (texto_limpo(row["SUBSYSTEM"]).upper(), texto_limpo(row["AGR_NAME"]).upper())
        if key in vistos_envio:
            continue
        vistos_envio.add(key)
        rows_envio.append(
            {
                "SUBSYSTEM": row["SUBSYSTEM"],
                "AGR_NAME": row["AGR_NAME"],
                "FROM_DAT": _data_sap_yyyymmdd(row["FROM_DAT"]),
                "TO_DAT": _data_sap_yyyymmdd(row["TO_DAT"] or "99991231"),
                "AGR_TEXT": row["AGR_TEXT"],
                "ORG_FLAG": row["ORG_FLAG"],
            }
        )

    adicionadas = []
    ja_existentes = []

    for role_norm in roles_norm:
        key = (sistema_norm, role_norm)
        if key in existentes_pairs:
            ja_existentes.append(role_norm)
            continue

        rows_envio.append(
            {
                "SUBSYSTEM": sistema_norm,
                "AGR_NAME": role_norm,
                "FROM_DAT": datetime.now().strftime("%Y%m%d"),
                "TO_DAT": "99991231",
                "AGR_TEXT": "",
                "ORG_FLAG": "",
            }
        )
        adicionadas.append(role_norm)

    if not adicionadas and rows_envio:
        return {
            "STATUS": "CONCLUIDO",
            "MSG": "Nenhuma função nova para atribuir via RFC.",
            "ADICIONADAS": [],
            "JA_EXISTENTES": ja_existentes,
        }

    res = conn.call(
        "BAPI_USER_LOCACTGROUPS_ASSIGN",
        USERNAME=utilizador,
        NO_DB_UPDATE="",
        INCL_HR_ASSIGN="",
        DISTRIBUTE_CHANGE_ONLY="",
        ACTIVITYGROUPS=rows_envio,
    )

    retorno = res.get("RETURN", [])
    if _bapi_has_errors(retorno):
        raise RuntimeError(_bapi_messages(retorno) or "Erro a atribuir funções via RFC.")

    try:
        conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")
    except Exception:
        pass

    msg = "Atribuição via RFC concluída."
    if ja_existentes:
        msg += f" Já existentes: {', '.join(ja_existentes)}."
    if adicionadas:
        msg += f" Adicionadas: {', '.join(adicionadas)}."

    return {
        "STATUS": "CONCLUIDO",
        "MSG": msg,
        "ADICIONADAS": adicionadas,
        "JA_EXISTENTES": ja_existentes,
    }


###################################################################################
# BLOCO 7: FILTRO DE LINHAS A PROCESSAR
###################################################################################

def filtrar_pendentes(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df2 = df.copy()
    df2["STATUS_NORM"] = df2["STATUS"].apply(normalizar_valor)

    pend = df2[
        (df2["CHAVE_ID"] != "") &
        (df2["STATUS_NORM"] != "CONCLUIDO")
    ].drop(columns=["STATUS_NORM"])

    if pend.empty:
        print("\n⚠️ Nenhuma linha com STATUS ≠ 'Concluído' foi encontrada.")
    else:
        print("\n📋 Linhas a processar:")
        exibir = pend[["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME"]].copy()
        for c in exibir.columns:
            exibir[c] = exibir[c].apply(texto_limpo)
        print(exibir.to_string(index=False))
        print()

    return pend


###################################################################################
# BLOCO 8: EXECUÇÃO SAP
###################################################################################

def marcar_resultado(df_ref, idx, status, msg):
    df_ref.at[idx, "STATUS"] = texto_limpo(status)
    df_ref.at[idx, "MSG"] = texto_limpo(msg)
    df_ref.at[idx, "TIMESTEMP"] = agora_str()


def processar_atribuicoes_rfc_cua(
    df_filtrado,
    rfc_conn,
    pedir_confirmacao=True,
    modo_nao_interativo=False,
):
    """
    Fluxo RFC puro para a folha CUA_ADICIONAR.
    Nao usa SAP GUI, SE16 nem SU10.
    """
    if df_filtrado is None or df_filtrado.empty:
        return df_filtrado

    if rfc_conn is None:
        raise RuntimeError("Ligacao RFC indisponivel para o processamento CUA.")

    vistos = set()
    duplicados_marcar = []
    indices_unicos = []

    for idx_row in df_filtrado.index:
        user_val = str(df_filtrado.at[idx_row, "UTILIZADOR"]).strip()
        sys_val = str(df_filtrado.at[idx_row, "SISTEMA"]).strip()
        role_val = str(df_filtrado.at[idx_row, "AGR_NAME"]).strip()

        user_norm = user_val.upper()
        sys_norm = sys_val.upper()
        role_norm = role_val.upper()

        if not user_norm or not sys_norm or not role_norm:
            indices_unicos.append(idx_row)
            continue

        chave = (user_norm, sys_norm, role_norm)
        if chave in vistos:
            duplicados_marcar.append(idx_row)
        else:
            vistos.add(chave)
            indices_unicos.append(idx_row)

    for idx_row in duplicados_marcar:
        user = df_filtrado.at[idx_row, "UTILIZADOR"]
        sys_name = df_filtrado.at[idx_row, "SISTEMA"]
        role = df_filtrado.at[idx_row, "AGR_NAME"]
        msg_dup = (
            f"Combinação '{role}' no sistema '{sys_name}' para o utilizador "
            f"'{user}' duplicada no Excel. Tratada na primeira ocorrência."
        )
        marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg_dup)

    df_unicos = df_filtrado.loc[indices_unicos]
    grupos = df_unicos.groupby(["UTILIZADOR", "SISTEMA"], sort=False)
    total_grupos = len(grupos)
    total_linhas_pendentes = len(df_unicos)
    total_roles_distintas = df_unicos["AGR_NAME"].nunique()

    print(f"\n[INFO] Utilizadores a processar agrupados (excluindo duplicados do Excel): {total_grupos}")
    print(f"[INFO] Linhas únicas pendentes: {total_linhas_pendentes}")
    print(f"[INFO] Roles distintas: {total_roles_distintas}")

    if not modo_nao_interativo and pedir_confirmacao:
        resposta = input("Deseja lançar estas funções no SAP via RFC? [S/N]: ").strip().upper()
        if resposta != "S":
            print("[X] Lançamento cancelado pelo utilizador.")
            return df_filtrado

    print("\n[Fase 3] Processamento RFC dos Utilizadores")
    tempo_total_inicio = time.time()

    for idx_grupo, ((utilizador, sistema), df_grupo) in enumerate(grupos, 1):
        inicio = time.time()
        roles_list = list(dict.fromkeys([str(r).strip() for r in df_grupo["AGR_NAME"] if str(r).strip()]))

        print(f"\n[Utilizador {idx_grupo}/{total_grupos}] {utilizador}")

        if not utilizador or not sistema or not roles_list:
            msg = "Dados obrigatórios (UTILIZADOR/SISTEMA/ROLES) vazios."
            for idx_row in df_grupo.index:
                marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
            duracao_str = formatar_tempo(time.time() - inicio)
            print(f"🔴 ERRO: {msg} ⏱️ (Tempo: {duracao_str})")
            continue

        try:
            rfc_result = atribuir_funcoes_rfc_cua(rfc_conn, utilizador, sistema, roles_list)
            adicionadas = set(rfc_result.get("ADICIONADAS", []))
            ja_existentes = set(rfc_result.get("JA_EXISTENTES", []))
            msg_rfc = rfc_result.get("MSG", "Atribuição via RFC concluída.")

            for idx_row in df_grupo.index:
                row_role = str(df_filtrado.at[idx_row, "AGR_NAME"]).strip().upper()
                if row_role in adicionadas:
                    marcar_resultado(
                        df_filtrado,
                        idx_row,
                        "CONCLUIDO",
                        f"Função '{row_role}' atribuída via RFC ao utilizador '{utilizador}' no sistema '{sistema}'. {msg_rfc}",
                    )
                elif row_role in ja_existentes:
                    marcar_resultado(
                        df_filtrado,
                        idx_row,
                        "CONCLUIDO",
                        f"Função '{row_role}' já atribuída ao utilizador '{utilizador}' no sistema '{sistema}'. {msg_rfc}",
                    )
                else:
                    marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg_rfc)

            duracao_str = formatar_tempo(time.time() - inicio)
            print(f"✅ RFC concluído para o utilizador {utilizador}. (Tempo: {duracao_str})")
        except Exception as exc:
            msg = f"Falha ao atribuir funções via RFC para '{utilizador}': {exc}"
            for idx_row in df_grupo.index:
                marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
            duracao_str = formatar_tempo(time.time() - inicio)
            print(f"🔴 ERRO: {msg} ⏱️ (Tempo: {duracao_str})")

    tempo_total = formatar_tempo(time.time() - tempo_total_inicio)
    print(f"\n[Fase 3] Processamento RFC concluído. Tempo decorrido: {tempo_total}")
    return df_filtrado


def gravar_preservando_formatacao(caminho_ficheiro, nome_sheet, df_atualizado):
    """
    Atualiza APENAS:
    - STATUS
    - MSG
    - TIMESTEMP

    Faz match pela coluna ID.
    """
    try:
        ext = os.path.splitext(caminho_ficheiro)[1].lower()
        keep_vba = ext == ".xlsm"

        wb = load_workbook(caminho_ficheiro, keep_vba=keep_vba)
        if nome_sheet not in wb.sheetnames:
            print(f"❌ Sheet '{nome_sheet}' não existe para gravar.")
            return False

        ws = wb[nome_sheet]
        mapa_cols = mapear_cabecalhos_openpyxl(ws)

        if "TIMESTAMP" in mapa_cols and "TIMESTEMP" not in mapa_cols:
            mapa_cols["TIMESTEMP"] = mapa_cols["TIMESTAMP"]

        obrig_excel = ["ID", "STATUS", "MSG", "TIMESTEMP"]
        falta = [c for c in obrig_excel if c not in mapa_cols]
        if falta:
            print(f"❌ Cabeçalhos obrigatórios em falta na sheet para gravação: {', '.join(falta)}")
            return False

        col_id = mapa_cols["ID"]
        col_status = mapa_cols["STATUS"]
        col_msg = mapa_cols["MSG"]
        col_timestemp = mapa_cols["TIMESTEMP"]

        mapa_linhas_por_id = {}
        for r in range(2, ws.max_row + 1):
            valor_id = ws.cell(row=r, column=col_id).value
            id_chave = chave_id(valor_id)
            if id_chave:
                mapa_linhas_por_id[id_chave] = r

        atualizados = 0
        nao_encontrados = 0

        for _, row in df_atualizado.iterrows():
            id_chave = texto_limpo(row.get("CHAVE_ID", ""))
            if not id_chave:
                continue

            linha_excel = mapa_linhas_por_id.get(id_chave)
            if not linha_excel:
                nao_encontrados += 1
                print(f"⚠️ ID não encontrado na sheet para gravação: {id_chave}")
                continue

            ws.cell(row=linha_excel, column=col_status).value = texto_limpo(row.get("STATUS", ""))
            ws.cell(row=linha_excel, column=col_msg).value = texto_limpo(row.get("MSG", ""))
            ws.cell(row=linha_excel, column=col_timestemp).value = texto_limpo(row.get("TIMESTEMP", ""))

            atualizados += 1

        wb.save(caminho_ficheiro)

        print(
            f"💾 Ficheiro atualizado com formatação preservada "
            f"(sheet '{nome_sheet}') | Linhas atualizadas: {atualizados}"
        )

        if nao_encontrados:
            print(f"⚠️ IDs não encontrados na sheet: {nao_encontrados}")

        return True

    except PermissionError:
        base, ext = os.path.splitext(caminho_ficheiro)
        alternativo = f"{base}_resultado{ext}"
        try:
            wb.save(alternativo)
            print(f"⚠️ Ficheiro estava aberto. Foi criada uma cópia:\n   {alternativo}")
            return True
        except Exception as e:
            print(f"❌ Erro ao salvar cópia: {e}")
            return False

    except Exception as e:
        print(f"❌ Erro ao salvar: {e}")
        return False


###################################################################################
# BLOCO 10: API PARA O COCKPIT
###################################################################################

def executar(
    ambiente_cockpit,
    pfcg_object=None,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    **kwargs
):
    tempo_inicio_total = time.time()
    print(f"✅ Processo selecionado: {NOME_SCRIPT}")

    sheet_alvo = pfcg_object if pfcg_object else NOME_SHEET
    print(f"📄 Script atual: {NOME_SCRIPT} | Sheet alvo: '{sheet_alvo}'")

    if modo_nao_interativo:
        if not caminho_ficheiro:
            raise ValueError("Faltou o parâmetro caminho_ficheiro em modo web/não-interativo.")
    else:
        if not caminho_ficheiro:
            caminho_ficheiro = selecionar_ficheiro_excel()

    if not caminho_ficheiro:
        return False

    print("\n[Fase 1] Leitura do Excel")
    df = ler_ficheiro(caminho_ficheiro, sheet_alvo)
    if df is None:
        return False

    if str(ambiente_cockpit).upper() != "CUA":
        print("❌ Este sub-processo é RFC puro para CUA. Selecione o ambiente CUA no Cockpit.")
        return False

    try:
        rfc_conn = conectar_rfc_cua()
    except Exception as e:
        print(f"❌ RFC CUA indisponível: {e}")
        return False

    df_pend = filtrar_pendentes(df)
    if df_pend.empty:
        try:
            rfc_conn.close()
        except Exception:
            pass
        tempo_decorrido_total = time.time() - tempo_inicio_total
        print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
        print("🔁 Fim.")
        return True

    try:
        df_proc = processar_atribuicoes_rfc_cua(
            df_pend.copy(),
            rfc_conn,
            pedir_confirmacao=pedir_confirmacao,
            modo_nao_interativo=modo_nao_interativo,
        )

        print("\n[Fase 4] Gravação de Resultados")
        ok_save = gravar_preservando_formatacao(caminho_ficheiro, sheet_alvo, df_proc)
        if ok_save:
            print("💾 Resultados gravados com sucesso no Excel!")
    finally:
        if rfc_conn is not None:
            try:
                rfc_conn.close()
            except Exception:
                pass

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
    print("🔁 Fim.")
    return ok_save


###################################################################################
# BLOCO 11: EXECUÇÃO DIRETA
###################################################################################

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD", "CUA"])
    parser.add_argument("--xlsx")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    args = parser.parse_args()

    env_cli = args.ambiente or "CUA"
    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm)
    )
