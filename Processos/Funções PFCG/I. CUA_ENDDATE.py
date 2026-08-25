# -*- coding: utf-8 -*-

###################################################################################
# PROCESSO: Ajustar validade de Função via SU01 (sheet = nome do .py SEM o prefixo)
# Ex.: "G. CUA_DELROLE.py" → Sheet "CUA_DELROLE"
#
# COLUNAS ESPERADAS:
# ID | UTILIZADOR | SISTEMA | AGR_NAME | STATUS | MSG | TIMESTEMP
#
# COLUNAS OPCIONAIS:
# UPDATE_TO_DAT | VALID_TO | VALIDO_ATE | DATA_FIM
#
# REGRAS DE RETORNO:
# - STATUS    = retorno do wnd[0]/sbar no formato "TIPO - TEXTO"
# - MSG       = detalhes complementares (popup, exceção, observações)
# - TIMESTEMP = data/hora da execução
###################################################################################

###################################################################################
# BLOCO 1: IMPORTAÇÕES E CONFIGURAÇÃO
###################################################################################

import os
import time
import pandas as pd
import unicodedata
import win32com.client
import tkinter as tk

from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime, timedelta

###################################################################################
# BLOCO 2: NOME DO SCRIPT / SHEET, MAPA DE SISTEMAS
###################################################################################

try:
    NOME_SCRIPT = os.path.splitext(os.path.basename(__file__))[0]
except NameError:
    NOME_SCRIPT = "G. CUA_DELROLE"  # fallback

NOME_SHEET = NOME_SCRIPT.split(".", 1)[-1].strip() if "." in NOME_SCRIPT else NOME_SCRIPT

MAPA_SISTEMA = {
    "DEV": "S4D",
    "QAD": "S4Q",
    "PRD": "S4P",
    "CUA": "SPA",
}

# Mapeamento para o filtro SUBSYSTEM do grid de Roles na SU01
MAPA_SUBSYSTEM = {
    "DEV": "S4DCLNT100",
    "S4D": "S4DCLNT100",
    "S4DCLNT100": "S4DCLNT100",
    "DESENVOLVIMENTO": "S4DCLNT100",
    "QAD": "S4QCLNT100",
    "S4Q": "S4QCLNT100",
    "S4QCLNT100": "S4QCLNT100",
    "QUALIDADE": "S4QCLNT100",
    "PRD": "S4PCLNT100",
    "S4P": "S4PCLNT100",
    "S4PCLNT100": "S4PCLNT100",
    "PRODUCAO": "S4PCLNT100",
    "PRODUÇÃO": "S4PCLNT100",
    "CUA": "SPACLNT001",
    "SPA": "SPACLNT001",
    "SPACLNT001": "SPACLNT001",
}

###################################################################################
# BLOCO 3: UTILITÁRIOS
###################################################################################

def selecionar_ficheiro_excel():
    """Abre popup em primeiro plano usando naturalmente a última pasta utilizada."""
    try:
        root = tk.Tk()
        root.withdraw()
        root.update_idletasks()
        root.attributes("-topmost", True)

        caminho = filedialog.askopenfilename(
            title="Selecione o ficheiro Excel",
            filetypes=[("Ficheiros Excel", "*.xlsx *.xls"), ("Todos os ficheiros", "*.*")]
        )

        root.destroy()

        if not caminho:
            print("⚠️ Seleção cancelada pelo utilizador.")
            return None

        print(f"✅ Ficheiro a processar: {caminho}")
        return caminho

    except Exception as e:
        print(f"❌ Erro ao abrir o popup: {e}")
        return None


def normalizar_coluna(col):
    return unicodedata.normalize("NFKD", str(col)).encode("ASCII", "ignore").decode("utf-8").strip().upper()


def normalizar_valor(val):
    return unicodedata.normalize("NFKD", str(val)).encode("ASCII", "ignore").decode("utf-8").strip().upper()


def valor_em_branco(val):
    if val is None:
        return True
    txt = str(val).strip()
    return txt == "" or normalizar_valor(txt) in {"NAN", "NONE", "<NA>"}


def normalizar_id(val):
    if val is None:
        return ""
    try:
        if isinstance(val, float) and val.is_integer():
            return str(int(val)).strip()
    except Exception:
        pass
    return str(val).strip()


def mapear_cabecalhos_openpyxl(ws):
    mapa = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        mapa[normalizar_coluna(v)] = c
    return mapa


def obter_timempestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def obter_data_ontem_sap():
    return (datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")


def juntar_textos(*partes):
    itens = []
    vistos = set()

    for p in partes:
        if p is None:
            continue
        txt = str(p).strip()
        if not txt:
            continue
        if txt not in vistos:
            itens.append(txt)
            vistos.add(txt)

    return " | ".join(itens)


def resolver_subsystem(valor_sistema):
    """
    Aceita:
    - DEV / QAD / PRD / CUA
    - S4D / S4Q / S4P / SPA
    - S4DCLNT100 / S4QCLNT100 / S4PCLNT100 / SPACLNT001
    - Rótulos do assistente Web como 'S4P — Cliente 100 — PRODUÇÃO (S4H)'
    Retorna o código de SUBSYSTEM CUA exato do SAP (ex: S4PCLNT100).
    """
    bruto = str(valor_sistema or "").strip()
    if not bruto:
        return ""

    norm = normalizar_valor(bruto)

    if norm in MAPA_SUBSYSTEM:
        return MAPA_SUBSYSTEM[norm]

    if "S4P" in norm or "PRD" in norm or "PRODUCAO" in norm:
        return "S4PCLNT100"
    if "S4D" in norm or "DEV" in norm or "DESENVOLVIMENTO" in norm:
        return "S4DCLNT100"
    if "S4Q" in norm or "QAD" in norm or "QUALIDADE" in norm:
        return "S4QCLNT100"
    if "SPA" in norm or "CUA" in norm:
        return "SPACLNT001"

    return bruto


def resolver_data_fim(row):
    """
    Procura uma coluna opcional de data fim.
    Se não existir/estiver vazia, usa a data de ontem (hoje - 1 dia).
    """
    colunas_data = ["UPDATE_TO_DAT", "VALID_TO", "VALIDO_ATE", "DATA_FIM"]
    for c in colunas_data:
        if c in row and not valor_em_branco(row.get(c, "")):
            return str(row.get(c, "")).strip()
    return obter_data_ontem_sap()

###################################################################################
# BLOCO 4: LEITURA DO EXCEL
###################################################################################

def ler_sheet(caminho_ficheiro, nome_sheet):
    """Lê a sheet alvo, padroniza cabeçalhos e garante colunas auxiliares."""
    if not caminho_ficheiro or not os.path.exists(caminho_ficheiro):
        print("❌ Caminho inválido ou ficheiro inexistente.")
        return None

    try:
        wb = load_workbook(caminho_ficheiro, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

        if nome_sheet not in sheets:
            print(f"❌ Sheet '{nome_sheet}' não encontrada. Disponíveis: {', '.join(sheets)}")
            return None

        df = pd.read_excel(
            caminho_ficheiro,
            sheet_name=nome_sheet,
            dtype=object,
            keep_default_na=False
        )

        df.columns = [normalizar_coluna(c) for c in df.columns]

        # Harmonização tolerante
        df.rename(columns={
            "USER": "UTILIZADOR",
            "USERNAME": "UTILIZADOR",
            "SYSTEM": "SISTEMA",
            "FUNCAO": "AGR_NAME",
            "FUNÇÃO": "AGR_NAME",
            "ROLE": "AGR_NAME",
            "NOME FUNCAO": "AGR_NAME",
            "NOME FUNÇAO": "AGR_NAME",
            "NOME FUNÇÂO": "AGR_NAME",
            "TIMESTAMP": "TIMESTEMP",
            "DATA FINAL": "UPDATE_TO_DAT",
            "DATA_FIM_ROLE": "UPDATE_TO_DAT",
        }, inplace=True)

        obrigatorias = ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS"]
        em_falta = [c for c in obrigatorias if c not in df.columns]

        if em_falta:
            print(f"❌ Colunas obrigatórias em falta: {', '.join(em_falta)}")
            return None

        if "MSG" not in df.columns:
            df["MSG"] = ""
        if "TIMESTEMP" not in df.columns:
            df["TIMESTEMP"] = ""

        for c in ["STATUS", "MSG", "TIMESTEMP"]:
            df[c] = df[c].astype(str)

        print(f"📄 Sheet carregada: '{nome_sheet}' | Registos: {len(df)}")
        return df

    except Exception as e:
        print(f"❌ Erro ao ler a sheet: {e}")
        return None

###################################################################################
# BLOCO 5: CONEXÃO COM SAP GUI
###################################################################################

def conectar_sap(sistema_desejado):
    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application = sap_gui_auto.GetScriptingEngine

        for conn in application.Children:
            for sess in conn.Children:
                try:
                    if sess.Info.SystemName.upper() == sistema_desejado:
                        print(
                            f"✅ Conectado ao SAP: {sess.Info.SystemName.upper()} | "
                            f"Utilizador: {sess.Info.User} | Cliente: {sess.Info.Client}"
                        )
                        return sess
                except Exception:
                    continue

        print(f"❌ Sessão SAP não encontrada para o sistema {sistema_desejado}.")
        return None

    except Exception as e:
        print(f"❌ Erro ao conectar SAP GUI: {e}")
        return None

###################################################################################
# BLOCO 6: HELPERS SAP GUI
###################################################################################

def existe_objeto(session, obj_id):
    try:
        session.findById(obj_id)
        return True
    except Exception:
        return False


def aguardar_objeto(session, obj_id, timeout=10, intervalo=0.3):
    fim = time.time() + timeout
    while time.time() < fim:
        if existe_objeto(session, obj_id):
            return True
        time.sleep(intervalo)
    return False


def obter_status_bar(session):
    """
    Lê a status bar do SAP.
    Retorna:
    {
        "tipo": "S|W|E|A|I|''",
        "texto": "...",
        "status": "TIPO - TEXTO" ou "TEXTO" ou ""
    }
    """
    try:
        sbar = session.findById("wnd[0]/sbar")
        tipo = str(getattr(sbar, "MessageType", "") or "").strip().upper()
        texto = str(getattr(sbar, "Text", "") or "").strip()

        if tipo and texto:
            status = f"{tipo} - {texto}"
        elif texto:
            status = texto
        else:
            status = ""

        return {"tipo": tipo, "texto": texto, "status": status}

    except Exception:
        return {"tipo": "", "texto": "", "status": ""}


def _coletar_textos_recursivo(gui_obj, textos, profundidade=0, max_profundidade=5):
    if profundidade > max_profundidade or gui_obj is None:
        return

    try:
        txt = str(getattr(gui_obj, "Text", "") or "").strip()
        tipo = str(getattr(gui_obj, "Type", "") or "").strip().upper()
        if txt and "BTN" not in tipo and "BUTTON" not in tipo:
            textos.append(txt)
    except Exception:
        pass

    try:
        filhos = gui_obj.Children
        for i in range(filhos.Count):
            try:
                filho = filhos.Item(i)
                _coletar_textos_recursivo(filho, textos, profundidade + 1, max_profundidade)
            except Exception:
                continue
    except Exception:
        pass


def extrair_texto_popup(session):
    try:
        wnd1 = session.findById("wnd[1]")
    except Exception:
        return ""

    textos = []
    _coletar_textos_recursivo(wnd1, textos)

    vistos = set()
    saida = []
    for t in textos:
        if t not in vistos:
            saida.append(t)
            vistos.add(t)

    return " | ".join(saida).strip()


def tratar_popup(session, max_tentativas=5, pausa=0.4):
    """
    Fecha popup(s) de confirmação/aviso em wnd[1], capturando o texto.
    Retorna o texto consolidado dos popups encontrados.
    """
    textos_popup = []

    for _ in range(max_tentativas):
        if not existe_objeto(session, "wnd[1]"):
            break

        texto = extrair_texto_popup(session)
        if texto:
            textos_popup.append(texto)

        pressionado = False

        botoes_confirmacao = [
            "wnd[1]/tbar[0]/btn[0]",
            "wnd[1]/usr/btnSPOP-OPTION1",
            "wnd[1]/usr/btnBUTTON_1",
            "wnd[1]/usr/btnSPOP-VAROPTION1",
        ]

        for bid in botoes_confirmacao:
            try:
                session.findById(bid).press()
                pressionado = True
                break
            except Exception:
                continue

        if not pressionado:
            try:
                session.findById("wnd[1]").sendVKey(0)
                pressionado = True
            except Exception:
                pass

        if not pressionado:
            break

        time.sleep(pausa)

    vistos = set()
    unicos = []
    for t in textos_popup:
        if t not in vistos:
            unicos.append(t)
            vistos.add(t)

    return " | ".join(unicos).strip()


def reset_para_tela_inicial(session):
    """Tenta deixar a sessão limpa para o próximo item."""
    try:
        tratar_popup(session, max_tentativas=3, pausa=0.3)
    except Exception:
        pass

    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
        session.findById("wnd[0]").sendVKey(0)
        time.sleep(0.5)
    except Exception:
        pass

    try:
        tratar_popup(session, max_tentativas=2, pausa=0.3)
    except Exception:
        pass

###################################################################################
# BLOCO 7: FILTRO DE LINHAS A PROCESSAR
###################################################################################

def status_ja_processado(status):
    """
    Considera como já processado:
    - legado: CONCLUÍDO / CONCLUIDO
    - novo padrão: STATUS iniciado por S - ou W -
    """
    s = normalizar_valor(status)

    if s in {"CONCLUIDO", "CONCLUÍDO"}:
        return True

    if s.startswith("S - ") or s.startswith("W - "):
        return True

    return False


def filtrar_pendentes(df):
    """Linhas com ID preenchido e STATUS ainda não concluído/processado."""
    if df is None or df.empty:
        return pd.DataFrame()

    df2 = df.copy()

    pend = df2[
        df2["ID"].apply(lambda x: not valor_em_branco(x)) &
        ~df2["STATUS"].apply(status_ja_processado)
    ].copy()

    try:
        pend["_ID_SORT"] = pend["ID"].apply(normalizar_id)
        pend = pend.sort_values(by="_ID_SORT").drop(columns=["_ID_SORT"])
    except Exception:
        pass

    if pend.empty:
        print("\n⚠️ Nenhuma linha pendente foi encontrada.")
    else:
        cols_show = ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME"]
        for c in ["UPDATE_TO_DAT", "VALID_TO", "VALIDO_ATE", "DATA_FIM"]:
            if c in pend.columns:
                cols_show.append(c)
                break

        print("\n📋 Linhas a processar:")
        print(pend[cols_show].fillna("").to_string(index=False))
        print()

    return pend

###################################################################################
# BLOCO 8: VALIDAÇÃO DAS LINHAS
###################################################################################

def validar_linha(row):
    erros = []

    if valor_em_branco(row.get("ID", "")):
        erros.append("ID vazio")
    if valor_em_branco(row.get("UTILIZADOR", "")):
        erros.append("UTILIZADOR vazio")
    if valor_em_branco(row.get("SISTEMA", "")):
        erros.append("SISTEMA vazio")
    if valor_em_branco(row.get("AGR_NAME", "")):
        erros.append("AGR_NAME vazio")

    return "; ".join(erros)

###################################################################################
# BLOCO 9: HELPERS DO NOVO MAPEAMENTO SU01
###################################################################################

def obter_children(obj):
    try:
        children = obj.Children
        total = children.Count
        return [children.Item(i) for i in range(total)]
    except Exception:
        try:
            return list(obj.Children)
        except Exception:
            return []

def coletar_componentes_recursivo(obj, lista):
    try:
        lista.append(obj)
    except Exception:
        return
    for filho in obter_children(obj):
        coletar_componentes_recursivo(filho, lista)

def listar_campos_popup(session):
    componentes = []
    try:
        wnd1 = session.findById("wnd[1]")
    except Exception:
        return componentes
    coletar_componentes_recursivo(wnd1, componentes)
    return componentes

def descrever_componente(obj):
    try:
        obj_id = str(getattr(obj, "Id", "") or "")
    except Exception:
        obj_id = ""
    try:
        obj_type = str(getattr(obj, "Type", "") or "")
    except Exception:
        obj_type = ""
    try:
        name = str(getattr(obj, "Name", "") or "")
    except Exception:
        name = ""
    try:
        text = str(getattr(obj, "Text", "") or "")
    except Exception:
        text = ""
    try:
        changeable = getattr(obj, "Changeable")
    except Exception:
        changeable = "?"
    return {"id": obj_id, "type": obj_type, "name": name, "text": text, "changeable": changeable}

def obter_campos_low_popup(session):
    candidatos = []
    for comp in listar_campos_popup(session):
        info = descrever_componente(comp)
        obj_id = info["id"].upper()
        if "-LOW" not in obj_id:
            continue
        candidatos.append(info)

    def score(info):
        changeable = info["changeable"] is True
        obj_id = info["id"].lower()
        prioridade_tipo = 0
        if "/ctxt" in obj_id:
            prioridade_tipo = 2
        elif "/txt" in obj_id:
            prioridade_tipo = 1
        return (1 if changeable else 0, prioridade_tipo, len(obj_id))

    candidatos.sort(key=score, reverse=True)
    return candidatos

def preencher_popup_filtro(session, valor, descricao_filtro=""):
    candidatos = obter_campos_low_popup(session)
    if not candidatos:
        return {"success": False, "error": f"Nenhum campo '*-LOW' foi encontrado no wnd[1] para {descricao_filtro}."}

    for info in candidatos:
        obj_id = info["id"]
        try:
            obj = session.findById(obj_id)
            try:
                obj.setFocus()
            except Exception:
                pass
            obj.text = str(valor)
            escrito = str(getattr(obj, "Text", "") or "")
            if normalizar_valor(escrito) != normalizar_valor(valor):
                raise Exception("Valor escrito não corresponde ao esperado no campo.")

            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(0.4)
            return {"success": True, "field_id": obj_id, "current_value": escrito}
        except Exception as e:
            continue

    return {"success": False, "error": f"Não foi possível preencher o campo do filtro {descricao_filtro}."}

def aplicar_filtro_shell(session, shell, coluna, valor_filtro):
    """
    Aplica filtro no ALV/GRID dinamicamente usando preencher_popup_filtro do J. CUA_REMOVE.
    """
    shell.currentCellColumn = coluna
    shell.contextMenu()
    shell.selectContextMenuItem("&FILTER")

    res = preencher_popup_filtro(session, valor_filtro, descricao_filtro=coluna)
    if not res.get("success"):
        raise Exception(f"Falha ao aplicar filtro na coluna {coluna} com valor '{valor_filtro}': {res.get('error')}")
    time.sleep(0.3)


def abrir_su01_em_alteracao(session, utilizador):
    """
    Abre a SU01 e entra no utilizador em modo alteração.
    """
    print("\n[Etapa 2] Acesso ao SAP CUA")
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsu01"
    session.findById("wnd[0]").sendVKey(0)

    campo_utilizador = "wnd[0]/usr/ctxtSUID_ST_BNAME-BNAME"
    if aguardar_objeto(session, campo_utilizador, timeout=8, intervalo=0.3):
        session.findById(campo_utilizador).text = utilizador
        try:
            session.findById(campo_utilizador).setFocus()
            session.findById(campo_utilizador).caretPosition = len(utilizador)
        except Exception:
            pass

    # Botão "Alterar"
    session.findById("wnd[0]/tbar[1]/btn[18]").press()
    time.sleep(0.6)

    tab_roles = "wnd[0]/usr/tabsTABSTRIP1/tabpACTG"
    if not aguardar_objeto(session, tab_roles, timeout=8, intervalo=0.3):
        raise Exception("Separador 'Roles' não carregou na SU01.")

    session.findById(tab_roles).select()
    time.sleep(0.4)


def atualizar_validade_roles_lote_su01(session, utilizador, subsystem, lista_items):
    """
    Atualiza todas as roles do MESMO (utilizador, subsystem) numa única sessão da SU01:
    1. Abre SU01 para o utilizador uma só vez
    2. Entra no separador 'Roles' (ACTG)
    3. Filtra SUBSYSTEM uma só vez
    4. Para cada role: filtra AGR_NAME e altera UPDATE_TO_DAT no grid
    5. Guarda (sendVKey 11) uma só vez no final
    """
    abrir_su01_em_alteracao(session, utilizador)

    shell_id = (
        "wnd[0]/usr/tabsTABSTRIP1/tabpACTG/"
        "ssubMAINAREA:SAPLSUID_MAINTENANCE:1106/"
        "cntlG_ROLES_CONTAINER/shellcont/shell"
    )

    if not aguardar_objeto(session, shell_id, timeout=10, intervalo=0.3):
        raise Exception("Grid de roles não carregou na SU01.")

    print(f"\n[SU01 Lote] Utilizador={utilizador} | Subsystem={subsystem} | Roles a alterar={len(lista_items)}")
    shell = session.findById(shell_id)

    # Filtro 1: SUBSYSTEM (uma única vez por utilizador/subsistema)
    aplicar_filtro_shell(
        session=session,
        shell=shell,
        coluna="SUBSYSTEM",
        valor_filtro=subsystem
    )

    resultados_itens = []

    for item in lista_items:
        agr_name = item["agr_name"]
        data_fim = item["data_fim"]

        try:
            # Filtro 2: AGR_NAME
            aplicar_filtro_shell(
                session=session,
                shell=shell,
                coluna="AGR_NAME",
                valor_filtro=agr_name
            )

            # Alterar data final no grid
            shell.modifyCell(0, "UPDATE_TO_DAT", data_fim)
            shell.currentCellColumn = "UPDATE_TO_DAT"
            shell.pressEnter()
            time.sleep(0.3)

            resultados_itens.append({
                "idx": item["idx"],
                "agr_name": agr_name,
                "data_fim": data_fim,
                "sucesso_grid": True,
                "erro": None
            })
            print(f"  ✓ Grid atualizado: AGR_NAME={agr_name} | UPDATE_TO_DAT={data_fim}")

        except Exception as e_role:
            resultados_itens.append({
                "idx": item["idx"],
                "agr_name": agr_name,
                "data_fim": data_fim,
                "sucesso_grid": False,
                "erro": str(e_role)
            })
            print(f"  ❌ Falha no grid: AGR_NAME={agr_name} -> {e_role}")

    # Verificar se alguma role foi modificada com sucesso no grid
    sucesso_count = sum(1 for r in resultados_itens if r["sucesso_grid"])
    if sucesso_count == 0:
        status_sap = obter_status_bar(session)
        status_txt = status_sap["status"] or "E - Nenhuma role foi encontrada/alterada no grid"
        return [
            {
                "idx": r["idx"],
                "status": r["erro"] or status_txt,
                "msg": juntar_textos(f"SUBSYSTEM: {subsystem}", f"UPDATE_TO_DAT: {r['data_fim']}", f"ERRO: {r['erro']}"),
            }
            for r in resultados_itens
        ]

    status_antes_save = obter_status_bar(session)

    # Guardar na SU01 (sendVKey 11) UMA ÚNICA VEZ para todas as roles do utilizador!
    session.findById("wnd[0]").sendVKey(11)
    time.sleep(0.5)

    popup_txt = tratar_popup(session, max_tentativas=5, pausa=0.4)
    time.sleep(0.3)

    status_final = obter_status_bar(session)
    status_para_gravar = status_final["status"] or status_antes_save["status"] or "S - Validade atualizada na SU01"

    resultados_finais = []
    for r in resultados_itens:
        if r["sucesso_grid"]:
            msg = juntar_textos(
                f"SUBSYSTEM: {subsystem}",
                f"UPDATE_TO_DAT: {r['data_fim']}",
                f"ANTES_SAVE: {status_antes_save['status']}" if status_antes_save["status"] else "",
                f"POPUP: {popup_txt}" if popup_txt else ""
            )
            resultados_finais.append({
                "idx": r["idx"],
                "status": status_para_gravar,
                "msg": msg
            })
        else:
            msg = juntar_textos(
                f"SUBSYSTEM: {subsystem}",
                f"UPDATE_TO_DAT: {r['data_fim']}",
                f"ERRO_GRID: {r['erro']}"
            )
            resultados_finais.append({
                "idx": r["idx"],
                "status": f"E - {r['erro']}",
                "msg": msg
            })

    return resultados_finais

###################################################################################
# BLOCO 10: EXECUÇÃO (AJUSTAR UPDATE_TO_DAT NO SU01 EM LOTE)
###################################################################################

def remover_funcao_usuario(df_filtrado, session, sistema_desejado):
    """
    Executa a atualização de validade das funções agrupando por (UTILIZADOR, SUBSYSTEM)
    para abrir a SU01 apenas UMA vez por utilizador/subsistema e guardar todas as roles juntas.
    """
    df_proc = df_filtrado.copy()

    for col in ["STATUS", "MSG", "TIMESTEMP"]:
        if col in df_proc.columns:
            df_proc[col] = df_proc[col].astype(str)
        else:
            df_proc[col] = ""

    total = len(df_proc)
    tempo_total_inicio = time.time()

    # Pre-validar linhas e agrupar pendentes por (UTILIZADOR, SUBSYSTEM)
    grupos = {}
    for idx, row in df_proc.iterrows():
        erro_validacao = validar_linha(row)
        if erro_validacao:
            status_val = f"E - {erro_validacao}"
            df_proc.at[idx, "STATUS"] = status_val
            df_proc.at[idx, "MSG"] = "Validação da linha falhou antes da SU01."
            df_proc.at[idx, "TIMESTEMP"] = obter_timempestamp()
            print(f"❌ ID={row.get('ID', '')}: {status_val}")
            continue

        utilizador = str(row.get("UTILIZADOR", "")).strip()
        sistema_excel = str(row.get("SISTEMA", "")).strip()
        subsystem = resolver_subsystem(sistema_excel)
        agr_name = str(row.get("AGR_NAME", "")).strip()
        data_fim = resolver_data_fim(row)

        chave = (utilizador, subsystem)
        if chave not in grupos:
            grupos[chave] = []

        grupos[chave].append({
            "idx": idx,
            "id": normalizar_id(row.get("ID", "")),
            "utilizador": utilizador,
            "sistema_excel": sistema_excel,
            "subsystem": subsystem,
            "agr_name": agr_name,
            "data_fim": data_fim
        })

    sistema_conectado = str(session.Info.SystemName).strip().upper()
    if sistema_conectado != sistema_desejado:
        print(f"❌ Sistema SAP incorreto: esperado {sistema_desejado}, conectado a {sistema_conectado}")
        return df_proc

    for (utilizador, subsystem), items in grupos.items():
        inicio_grupo = time.time()
        print(f"\n🔧 Processando lote: Utilizador={utilizador} | SUBSYSTEM={subsystem} ({len(items)} role(s))")

        try:
            reset_para_tela_inicial(session)

            resultados = atualizar_validade_roles_lote_su01(
                session=session,
                utilizador=utilizador,
                subsystem=subsystem,
                lista_items=items
            )

            for res in resultados:
                idx = res["idx"]
                df_proc.at[idx, "STATUS"] = res["status"]
                df_proc.at[idx, "MSG"] = res["msg"]
                df_proc.at[idx, "TIMESTEMP"] = obter_timempestamp()

            duracao = time.time() - inicio_grupo
            print(f"✅ Lote de {len(items)} role(s) concluído em {duracao:.1f}s")

            try:
                session.findById("wnd[0]/tbar[0]/btn[3]").press()
                time.sleep(0.3)
            except Exception:
                pass

            reset_para_tela_inicial(session)

        except Exception as e:
            popup_txt = ""
            try:
                popup_txt = tratar_popup(session, max_tentativas=3, pausa=0.3)
            except Exception:
                pass

            status_sap = obter_status_bar(session)
            status_para_gravar = status_sap["status"] or f"E - {str(e)}"
            
            for item in items:
                idx = item["idx"]
                msg_para_gravar = juntar_textos(
                    f"SUBSYSTEM: {subsystem}",
                    f"UPDATE_TO_DAT: {item['data_fim']}",
                    f"POPUP: {popup_txt}" if popup_txt else "",
                    f"EXCECAO: {str(e)}"
                )
                df_proc.at[idx, "STATUS"] = status_para_gravar
                df_proc.at[idx, "MSG"] = msg_para_gravar
                df_proc.at[idx, "TIMESTEMP"] = obter_timempestamp()

            print(f"❌ STATUS LOTE={status_para_gravar} | Erro: {e}")
            reset_para_tela_inicial(session)

    tempo_total = time.time() - tempo_total_inicio
    print(f"\n⏱️ Tempo total: {tempo_total:.1f}s")

    total_sucesso = df_proc["STATUS"].apply(lambda x: normalizar_valor(x).startswith("S - "))
    total_aviso = df_proc["STATUS"].apply(lambda x: normalizar_valor(x).startswith("W - "))
    total_erro = ~(total_sucesso | total_aviso)

    print(
        f"📊 S={int(total_sucesso.sum())} | "
        f"W={int(total_aviso.sum())} | "
        f"Outros/Erro={int(total_erro.sum())}"
    )

    return df_proc

###################################################################################
# BLOCO 11: GRAVAR SOMENTE STATUS / MSG / TIMESTEMP
###################################################################################

def gravar_retorno_preservando_formatacao(caminho_ficheiro, nome_sheet, df_atualizado):
    """
    Atualiza apenas as colunas STATUS / MSG / TIMESTEMP por ID,
    sem limpar a sheet inteira e preservando a formatação.
    """
    wb = None

    try:
        wb = load_workbook(caminho_ficheiro)

        if nome_sheet not in wb.sheetnames:
            print(f"❌ Sheet '{nome_sheet}' não existe para gravar.")
            return

        ws = wb[nome_sheet]
        mapa_cols = mapear_cabecalhos_openpyxl(ws)

        for cab in ["STATUS", "MSG", "TIMESTEMP"]:
            if cab not in mapa_cols:
                nova_col = ws.max_column + 1
                ws.cell(row=1, column=nova_col).value = cab
                mapa_cols[cab] = nova_col
                print(f"ℹ️ Cabeçalho '{cab}' criado na coluna {nova_col}.")

        if "ID" not in mapa_cols:
            print("❌ Cabeçalho 'ID' não encontrado na sheet para atualizar os resultados.")
            return

        col_id = mapa_cols["ID"]
        linhas_por_id = {}

        for r in range(2, ws.max_row + 1):
            valor_id = ws.cell(row=r, column=col_id).value
            chave = normalizar_id(valor_id)
            if chave:
                linhas_por_id[chave] = r

        qtd_atualizada = 0

        for _, row in df_atualizado.iterrows():
            chave_id = normalizar_id(row.get("ID", ""))
            if not chave_id:
                continue

            linha_excel = linhas_por_id.get(chave_id)
            if not linha_excel:
                continue

            ws.cell(row=linha_excel, column=mapa_cols["STATUS"]).value = str(row.get("STATUS", "") or "")
            ws.cell(row=linha_excel, column=mapa_cols["MSG"]).value = str(row.get("MSG", "") or "")
            ws.cell(row=linha_excel, column=mapa_cols["TIMESTEMP"]).value = str(row.get("TIMESTEMP", "") or "")
            qtd_atualizada += 1

        wb.save(caminho_ficheiro)
        print(f"💾 Ficheiro atualizado com {qtd_atualizada} linha(s) na sheet '{nome_sheet}'.")

    except PermissionError:
        base, ext = os.path.splitext(caminho_ficheiro)
        alternativo = f"{base}_resultado{ext}"
        try:
            if wb is None:
                wb = load_workbook(caminho_ficheiro)
            wb.save(alternativo)
            print(f"⚠️ Ficheiro estava aberto. Foi criada uma cópia:\n   {alternativo}")
        except Exception as e:
            print(f"❌ Erro ao salvar cópia: {e}")

    except Exception as e:
        print(f"❌ Erro ao salvar: {e}")

###################################################################################
# BLOCO 12: API PARA O COCKPIT (executar)
###################################################################################

def executar(ambiente, caminho_ficheiro=None):
    print(f"✅ Processo selecionado: {NOME_SCRIPT}")
    print(f"📄 Script atual: {NOME_SCRIPT} | Sheet alvo: '{NOME_SHEET}'")
    print("▶️ Ação: Atualizar UPDATE_TO_DAT da função via SU01")
    print("ℹ️ STATUS será preenchido com o retorno do wnd[0]/sbar.")

    if not caminho_ficheiro:
        caminho_ficheiro = selecionar_ficheiro_excel()
    if not caminho_ficheiro:
        return False

    print("\n[Etapa 1] Leitura do Excel")
    df = ler_sheet(caminho_ficheiro, NOME_SHEET)
    if df is None:
        return False

    sistema_desejado = MAPA_SISTEMA.get(ambiente)
    if not sistema_desejado:
        print(f"❌ Ambiente inválido: {ambiente}. Use: {', '.join(MAPA_SISTEMA.keys())}")
        return False

    session = conectar_sap(sistema_desejado)
    if not session:
        return False

    df_pend = filtrar_pendentes(df)
    if df_pend.empty:
        return True

    df_proc = remover_funcao_usuario(df_pend, session, sistema_desejado)

    print("\n[Etapa 4] Gravação de Resultados")
    gravar_retorno_preservando_formatacao(caminho_ficheiro, NOME_SHEET, df_proc)

    return True

###################################################################################
# BLOCO 13: EXECUÇÃO DIRETA (opcional)
###################################################################################

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD", "CUA"])
    parser.add_argument("--xlsx")
    args = parser.parse_args()

    env_cli = args.ambiente or "CUA"
    executar(env_cli, caminho_ficheiro=args.xlsx)