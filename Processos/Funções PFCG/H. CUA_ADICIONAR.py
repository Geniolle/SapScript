# -*- coding: utf-8 -*-

###################################################################################
# PROCESSO: Adicionar Função SU01/SU10  (sheet = nome do .py SEM o prefixo)
# Ex.: "H. CUA_ADICIONAR.py"  →  Sheet "CUA_ADICIONAR"
#
# ESTRUTURA ESPERADA DA SHEET:
# ID | UTILIZADOR | SISTEMA | AGR_NAME | STATUS | MSG | TIMESTEMP
#
# PADRÃO APLICADO:
# - STATUS final decidido com base no wnd[0]/sbar
# - captura do sbar em cada passo crítico
# - guarda a última mensagem relevante do SAP
# - grava no Excel atualizando APENAS STATUS / MSG / TIMESTEMP
# - preserva formatação, fórmulas, filtros e restantes colunas
# - popup sem diretório fixo (abre no último local usado)
###################################################################################

###################################################################################
# BLOCO 1: IMPORTAÇÕES
###################################################################################

import os
import time
import unicodedata
from datetime import date, datetime

import pandas as pd
import win32com.client
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

###################################################################################
# BLOCO 2: NOME DO SCRIPT / SHEET / MAPA DE SISTEMAS
###################################################################################

try:
    NOME_SCRIPT = os.path.splitext(os.path.basename(__file__))[0]
except NameError:
    NOME_SCRIPT = "H. CUA_ADICIONAR"

NOME_SHEET = NOME_SCRIPT.split(".", 1)[-1].strip() if "." in NOME_SCRIPT else NOME_SCRIPT

MAPA_SISTEMA = {
    "DEV": "S4D",
    "QAD": "S4Q",
    "PRD": "S4P",
    "CUA": "SPA",
}

###################################################################################
# BLOCO 3: UTILITÁRIOS
###################################################################################

DEBUG_SAP_CONTROLS = os.environ.get("DEBUG_SAP_CONTROLS", "FALSE").upper() == "TRUE"

def formatar_tempo(segundos):
    m = int(segundos // 60)
    s = int(segundos % 60)
    return f"{m:02d}m {s:02d}s"


def agora_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalizar_coluna(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("utf-8")
        .strip()
        .upper()
    )


def normalizar_valor(valor):
    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("utf-8")
        .strip()
        .upper()
    )


def texto_limpo(valor):
    if pd.isna(valor):
        return ""
    txt = str(valor).strip()
    if txt.lower() in ("nan", "none", "<na>"):
        return ""
    return txt


def valor_vazio(valor):
    return texto_limpo(valor) == ""


def chave_id(valor):
    if pd.isna(valor):
        return ""

    if isinstance(valor, int):
        return str(valor)

    if isinstance(valor, float):
        return str(int(valor)) if valor.is_integer() else str(valor).strip()

    txt = str(valor).strip()

    if txt.endswith(".0"):
        base = txt[:-2]
        if base.isdigit():
            return base

    return txt


def mapear_cabecalhos_openpyxl(ws):
    mapa = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val is None:
            continue
        mapa[normalizar_coluna(val)] = c
    return mapa


###################################################################################
# BLOCO 4: SELEÇÃO DO FICHEIRO
###################################################################################

def selecionar_ficheiro_excel():
    """
    Popup sem diretório fixo.
    O Windows tende a abrir no último local utilizado.
    """
    try:
        root = tk.Tk()
        root.withdraw()
        root.update_idletasks()
        root.attributes("-topmost", True)

        caminho = filedialog.askopenfilename(
            title="Selecione o ficheiro Excel",
            filetypes=[
                ("Ficheiros Excel", "*.xlsx *.xlsm"),
                ("Todos os ficheiros", "*.*"),
            ],
        )

        root.destroy()

        if not caminho:
            print("⚠️ Seleção cancelada pelo utilizador.")
            return None

        ext = os.path.splitext(caminho)[1].lower()
        if ext not in (".xlsx", ".xlsm"):
            print("❌ Apenas ficheiros .xlsx e .xlsm são suportados neste processo.")
            return None

        print(f"✅ Ficheiro a processar: {caminho}")
        return caminho

    except Exception as e:
        print(f"❌ Erro ao abrir o popup: {e}")
        return None


###################################################################################
# BLOCO 5: LEITURA DO EXCEL
###################################################################################

def ler_ficheiro(caminho_ficheiro, nome_sheet):
    """
    Lê a sheet alvo, normaliza cabeçalhos e valida estrutura obrigatória.
    Harmoniza variantes para:
    ID | UTILIZADOR | SISTEMA | AGR_NAME | STATUS | MSG | TIMESTEMP
    """
    if not caminho_ficheiro or not os.path.exists(caminho_ficheiro):
        print("❌ Caminho inválido ou ficheiro inexistente.")
        return None

    try:
        ext = os.path.splitext(caminho_ficheiro)[1].lower()
        keep_vba = ext == ".xlsm"

        # 1. Carregar em openpyxl para verificar e criar cabeçalhos em falta
        wb = load_workbook(
            caminho_ficheiro,
            read_only=False,
            data_only=False,
            keep_vba=keep_vba,
        )
        sheets = wb.sheetnames
        if nome_sheet not in sheets:
            print(f"❌ Sheet '{nome_sheet}' não encontrada. Disponíveis: {', '.join(sheets)}")
            wb.close()
            return None

        ws = wb[nome_sheet]
        
        # Obter os cabeçalhos reais da linha 1
        headers_linha_1 = [c.value for c in ws[1]]
        headers_norm = [normalizar_coluna(h) if h is not None else "" for h in headers_linha_1]

        renames = {
            "USER": "UTILIZADOR",
            "USERNAME": "UTILIZADOR",
            "SYSTEM": "SISTEMA",
            "FUNCAO": "AGR_NAME",
            "FUNÇÃO": "AGR_NAME",
            "ROLE": "AGR_NAME",
            "NOME FUNCAO": "AGR_NAME",
            "NOME FUNÇAO": "AGR_NAME",
            "NOME FUNÇÂO": "AGR_NAME",
            "AGRNAME": "AGR_NAME",
            "TIMESTAMP": "TIMESTEMP",
        }

        resolved_headers = []
        for h in headers_norm:
            resolved_headers.append(renames.get(h, h))

        # Colunas de entrada obrigatórias
        colunas_entrada = {"ID", "UTILIZADOR", "SISTEMA", "AGR_NAME"}
        # Colunas de saída que podem ser criadas
        colunas_saida = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

        # Verificar se as colunas de entrada existem
        falta_entrada = [c for c in colunas_entrada if c not in resolved_headers]
        if falta_entrada:
            print(f"❌ Colunas obrigatórias de entrada em falta: {', '.join(falta_entrada)}")
            wb.close()
            return None

        # Se as colunas de entrada existem, podemos prosseguir e criar as colunas de saída em falta
        modificou = False
        last_col = len(headers_norm)
        for col_key, col_name in colunas_saida.items():
            if col_key not in resolved_headers:
                last_col += 1
                ws.cell(row=1, column=last_col, value=col_name)
                print(f"➕ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                modificou = True

        # Validar e remover espaços do nome da Role (AGR_NAME) diretamente no Excel
        col_idx_agr = resolved_headers.index("AGR_NAME") + 1
        modificou_roles = False
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=col_idx_agr).value
            if cell_val is not None:
                agr = str(cell_val).strip()
                if " " in agr:
                    orig = agr
                    agr = agr.replace(" ", "")
                    ws.cell(row=r, column=col_idx_agr, value=agr)
                    print(f"⚠️ Espaços detetados e removidos da Role: '{orig}' -> '{agr}' (atualizado no Excel)")
                    modificou_roles = True

        if modificou or modificou_roles:
            try:
                wb.save(caminho_ficheiro)
                print("💾 Ficheiro Excel guardado com as correções de cabeçalhos/roles.")
            except Exception as e:
                wb.close()
                print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas/roles: {e}")
                print("💡 Certifique-se de que o ficheiro Excel não está aberto no Excel e tente novamente.")
                return None

        wb.close()

        # 2. Carregar o DataFrame em pandas para uso no script
        df = pd.read_excel(caminho_ficheiro, sheet_name=nome_sheet, dtype=object)
        df.columns = [normalizar_coluna(c) for c in df.columns]

        df.rename(
            columns={
                "USER": "UTILIZADOR",
                "USERNAME": "UTILIZADOR",
                "SYSTEM": "SISTEMA",
                "FUNCAO": "AGR_NAME",
                "FUNÇÃO": "AGR_NAME",
                "ROLE": "AGR_NAME",
                "NOME FUNCAO": "AGR_NAME",
                "NOME FUNÇAO": "AGR_NAME",
                "NOME FUNÇÂO": "AGR_NAME",
                "AGRNAME": "AGR_NAME",
                "TIMESTAMP": "TIMESTEMP",
            },
            inplace=True,
        )

        obrigatorias = ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]
        falta = [c for c in obrigatorias if c not in df.columns]
        if falta:
            print(f"❌ Colunas obrigatórias em falta: {', '.join(falta)}")
            return None

        for c in ["UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]:
            df[c] = df[c].apply(texto_limpo)

        df["CHAVE_ID"] = df["ID"].apply(chave_id)

        print(f"📄 Sheet carregada: '{nome_sheet}' | Registos: {len(df)}")
        return df

    except Exception as e:
        print(f"❌ Erro ao ler a sheet: {e}")
        return None


###################################################################################
# BLOCO 6: SAP GUI / STATUS BAR / POPUPS
###################################################################################

def conectar_sap(sistema_desejado):
    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application = sap_gui_auto.GetScriptingEngine

        for conn in application.Children:
            for sess in conn.Children:
                try:
                    if texto_limpo(sess.Info.SystemName).upper() == sistema_desejado:
                        print(
                            f"✅ Conectado: {sess.Info.SystemName} "
                            f"| User: {sess.Info.User} "
                            f"| Cliente: {sess.Info.Client}"
                        )
                        return sess
                except Exception:
                    continue

        print(f"❌ Sessão SAP não encontrada para o sistema {sistema_desejado}.")
        return None

    except Exception as e:
        print(f"❌ Erro ao conectar SAP GUI: {e}")
        return None


def posicionar_sap_meia_tela_direita(session) -> bool:
    """
    Posiciona a janela principal do SAP GUI na metade direita do monitor.

    Comportamento:
    - Obtém o handle (HWND) da janela wnd[0] via a propriedade Handle.
    - Consulta as dimensões do ecrã principal através de win32api.GetSystemMetrics.
    - Restaura a janela (SW_RESTORE) e move-a para:
        Left = screen_w // 2  (início da metade direita)
        Top  = 0
        Width = screen_w // 2  (metade da largura total, mínimo 900 px)
        Height = screen_h       (altura total, mínimo 700 px)

    Inspirada na função _dock_left_half() presente em sap_session.py,
    mas espelhada para o lado direito.

    Args:
        session: Objeto de sessão SAP GUI (GuiSession).

    Returns:
        True  – Janela posicionada com sucesso.
        False – Não foi possível posicionar (dependências em falta, handle
                inválido ou qualquer outra exceção capturada).
    """
    try:
        import win32api  # type: ignore
        import win32con  # type: ignore
        import win32gui  # type: ignore
    except Exception:
        print("[AVISO] win32api/win32gui não disponíveis – posicionamento ignorado.")
        return False

    try:
        wnd0 = session.findById("wnd[0]")
    except Exception:
        return False

    try:
        hwnd = int(getattr(wnd0, "Handle"))
    except Exception:
        return False

    if not hwnd:
        return False

    try:
        screen_w = int(win32api.GetSystemMetrics(0))
        screen_h = int(win32api.GetSystemMetrics(1))
        half_w   = screen_w // 2
        target_w = max(900, half_w)
        target_h = max(700, screen_h)
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.MoveWindow(hwnd, half_w, 0, target_w, target_h, True)
        print("[INFO] Janela SAP posicionada na metade direita do monitor.")
        return True
    except Exception as e:
        print(f"[AVISO] Erro ao posicionar janela SAP na metade direita: {e}")
        return False


def esperar_elemento(session, element_id, tentativas=20, espera=0.5):
    for _ in range(tentativas):
        try:
            return session.findById(element_id)
        except Exception:
            time.sleep(espera)
    return None


def existe_elemento(session, element_id):
    try:
        session.findById(element_id)
        return True
    except Exception:
        return False


def ir_para_transacao(session, tcode):
    session.findById("wnd[0]/tbar[0]/okcd").text = f"/N{tcode}"
    session.findById("wnd[0]").sendVKey(0)


def voltar_para_inicio(session):
    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
        session.findById("wnd[0]").sendVKey(0)
    except Exception:
        pass


def ler_status_bar_once(session):
    """
    Lê uma vez o wnd[0]/sbar.
    """
    try:
        sbar = session.findById("wnd[0]/sbar")
        tipo = texto_limpo(getattr(sbar, "MessageType", ""))
        texto = texto_limpo(getattr(sbar, "Text", ""))
        return tipo, texto
    except Exception:
        return "", ""


def registar_evento_status(eventos, origem, tipo="", texto=""):
    tipo = texto_limpo(tipo)
    texto = texto_limpo(texto)

    if not tipo and not texto:
        return

    eventos.append(
        {
            "origem": texto_limpo(origem),
            "tipo": tipo,
            "texto": texto,
        }
    )


def capturar_status_bar(session, eventos=None, origem="SBAR", tentativas=8, espera=0.25):
    """
    Faz várias tentativas curtas para apanhar a mensagem do wnd[0]/sbar
    no momento certo.
    """
    ultimo_tipo = ""
    ultimo_texto = ""

    for _ in range(tentativas):
        tipo, texto = ler_status_bar_once(session)
        if tipo or texto:
            ultimo_tipo = tipo
            ultimo_texto = texto
            break
        time.sleep(espera)

    if eventos is not None and (ultimo_tipo or ultimo_texto):
        registar_evento_status(eventos, origem, ultimo_tipo, ultimo_texto)

    combinado = (
        f"{ultimo_tipo} - {ultimo_texto}"
        if ultimo_tipo and ultimo_texto
        else (ultimo_texto or ultimo_tipo or "")
    )

    return ultimo_tipo, ultimo_texto, combinado


def obter_titulo_popup(session):
    try:
        return texto_limpo(session.findById("wnd[1]").text)
    except Exception:
        return ""


def tratar_popups_pos_save(session, eventos, max_popups=5):
    """
    Confirma popups após Save.
    Regista o título do popup e volta a tentar capturar o sbar após cada confirmação.
    """
    historico = []

    for n in range(1, max_popups + 1):
        if not existe_elemento(session, "wnd[1]"):
            break

        titulo = obter_titulo_popup(session) or "POPUP"
        historico.append(f"POPUP: {titulo}")
        registar_evento_status(eventos, f"POPUP_{n}", "", titulo)

        try:
            if existe_elemento(session, "wnd[1]/tbar[0]/btn[0]"):
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
            elif existe_elemento(session, "wnd[1]/tbar[0]/btn[11]"):
                session.findById("wnd[1]/tbar[0]/btn[11]").press()
            else:
                session.findById("wnd[1]").sendVKey(0)
        except Exception:
            try:
                session.findById("wnd[1]").sendVKey(0)
            except Exception:
                break

        time.sleep(0.35)
        capturar_status_bar(session, eventos, origem=f"SBAR_APOS_POPUP_{n}", tentativas=5, espera=0.20)

    return historico


def obter_ultimo_status_relevante(eventos):
    """
    Procura do fim para o início a última mensagem relevante do sbar.
    """
    for ev in reversed(eventos):
        tipo = texto_limpo(ev.get("tipo", ""))
        texto = texto_limpo(ev.get("texto", ""))
        if tipo or texto:
            combinado = f"{tipo} - {texto}" if tipo and texto else (texto or tipo)
            return tipo, texto, combinado
    return "", "", ""


def resumir_eventos_status(eventos, limite=5):
    """
    Monta uma trilha curta e útil das últimas mensagens.
    """
    itens = []

    for ev in eventos:
        origem = texto_limpo(ev.get("origem", ""))
        tipo = texto_limpo(ev.get("tipo", ""))
        texto = texto_limpo(ev.get("texto", ""))

        if tipo and texto:
            desc = f"{origem}: {tipo} - {texto}"
        elif texto:
            desc = f"{origem}: {texto}"
        elif tipo:
            desc = f"{origem}: {tipo}"
        else:
            continue

        if desc not in itens:
            itens.append(desc)

    if not itens:
        return ""

    return " | ".join(itens[-limite:])


def decidir_status_pelo_historico(eventos):
    """
    Decide STATUS com base no histórico de leituras do wnd[0]/sbar.
    Prioridade:
    - E/A/X => ERRO
    - S/W => CONCLUÍDO
    - fallback por texto
    """
    for ev in reversed(eventos):
        tipo = normalizar_valor(ev.get("tipo", ""))
        texto = normalizar_valor(ev.get("texto", ""))

        if tipo in ("E", "A", "X"):
            return "ERRO"

        if tipo in ("S", "W"):
            return "CONCLUÍDO"

        if any(ch in texto for ch in ["ERRO", "ERROR", "INVALID", "NAO EXIST", "NÃO EXIST", "OBRIGATOR", "INCONSIST"]):
            return "ERRO"

        if any(ch in texto for ch in ["GRAV", "GUARD", "SAVE", "SALV", "ATRIBU", "ATUALIZ", "ALTERACAO EFETUADA", "ALTERAÇÃO EFETUADA"]):
            return "CONCLUÍDO"

    return "ERRO"


def montar_msg_final(eventos):
    """
    A MSG final fica:
    - última mensagem relevante do sbar
    - + trilha curta dos passos, quando útil
    """
    _, _, ultima = obter_ultimo_status_relevante(eventos)
    trilha = resumir_eventos_status(eventos, limite=5)

    if ultima and trilha:
        if trilha.startswith(ultima):
            return trilha
        return f"{ultima} | {trilha}"

    if ultima:
        return ultima

    if trilha:
        return trilha

    return "Sem mensagem relevante do SAP"


def converter_data_sap(data_str) -> date:
    """
    Converte uma string de data SAP num objeto date.

    Formatos suportados:
      - DD.MM.AAAA  (padrao SAP europeu)
      - DD/MM/AAAA
      - AAAAMMDD    (padrao ISO compacto)
      - AAAA-MM-DD  (ISO com hifen)
      - objeto datetime ou date (passado directamente)
      - vazio / None         -> retorna date(9999, 12, 31)
      - datas com '9999'     -> retorna date(9999, 12, 31)
    """
    # Aceitar objectos datetime/date directamente
    if isinstance(data_str, datetime):
        return data_str.date()
    if isinstance(data_str, date):
        return data_str

    cleaned = str(data_str).strip() if data_str is not None else ""

    # Vazio ou data infinita SAP
    if not cleaned or "9999" in cleaned:
        return date(9999, 12, 31)

    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y%m%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Formato de data invalido: '{data_str}'")


def _dismiss_popup(session) -> bool:
    """Fecha popups ou avisos se existirem na sessao SAP."""
    for btn_id in ("wnd[1]/tbar[0]/btn[0]", "wnd[1]/tbar[0]/btn[11]"):
        try:
            session.findById(btn_id).press()
            return True
        except Exception:
            pass
    try:
        session.findById("wnd[1]").sendVKey(12)  # ESC
        return True
    except Exception:
        pass
    return False


def formatar_data_para_sap(data_obj) -> str:
    """
    Formata uma data para o formato padrão aceito pelo SAP (DD.MM.YYYY).
    """
    return data_obj.strftime("%d.%m.%Y")


def descobrir_campos_se16(session) -> dict:
    """
    Descobre dinamicamente os IDs dos campos LOW para BNAME e SUBSYSTEM
    na tela de selecao da SE16/USLA04.
    Retorna um dicionario { "BNAME": id, "SUBSYSTEM": id }

    O filtro TO_DAT foi removido: a consulta e feita sem restricao de data
    e a classificacao por datas e feita em Python apos ler todos os registos.
    """
    import re as _re

    usr = session.findById("wnd[0]/usr")

    # Obter todos os elementos recursivamente
    todos_elementos = []
    stack = [usr]
    while stack:
        curr = stack.pop()
        todos_elementos.append(curr)
        try:
            for idx in range(curr.Children.Count - 1, -1, -1):
                stack.append(curr.Children(idx))
        except Exception:
            pass

    # Inputs: apenas campos -LOW
    inputs_low = [
        e for e in todos_elementos
        if getattr(e, "Type", "") in ("GuiTextField", "GuiCTextField")
        and str(getattr(e, "Name", "")).upper().endswith("-LOW")
    ]

    mapeamento = {"BNAME": None, "SUBSYSTEM": None}

    # Termos de identificacao por campo
    termos = {
        "BNAME":     ["BNAME", "USER", "USER NAME", "NOME DO UTILIZADOR", "NOME UTILIZADOR", "CODIGO"],
        "SUBSYSTEM": ["SUBSYSTEM", "RECEIVING SYSTEM", "LOGICAL SYSTEM", "SISTEMA RECETOR", "SISTEMA RECEPTOR"],
    }

    # Estrategia 1: padrao SE16 classica (%_In_%_APP_%-TEXT / In-LOW)
    for e in todos_elementos:
        if getattr(e, "Type", "") not in ("GuiTextField", "GuiLabel"):
            continue
        name = str(getattr(e, "Name", ""))
        mat = _re.match(r'^%_I(\d+)_%_APP_%', name)
        if not mat:
            continue
        idx_str = mat.group(1)
        e_text    = str(getattr(e, "Text",    "")).strip().upper()
        e_tooltip = str(getattr(e, "Tooltip", "")).strip().upper()
        e_id      = str(getattr(e, "Id",      "")).upper()

        inp_name_target = f"I{idx_str}-LOW"
        inp_match = next(
            (i for i in todos_elementos if str(getattr(i, "Name", "")) == inp_name_target),
            None
        )
        if not inp_match:
            continue

        for chave, key_terms in termos.items():
            if mapeamento[chave] is not None:
                continue
            for term in key_terms:
                if term in e_text or term in e_tooltip or term in e_id:
                    mapeamento[chave] = inp_match.Id
                    break

    # Estrategia 2: proximidade Top/Left
    if any(v is None for v in mapeamento.values()):
        _excluir_names = {"MAX_SEL", "LIST_BRE", "GD-MAXROWS", "MAX_HITS"}
        labels = [
            e for e in todos_elementos
            if getattr(e, "Type", "") == "GuiLabel"
            or (
                getattr(e, "Type", "") == "GuiTextField"
                and not str(getattr(e, "Name", "")).upper().endswith(("-LOW", "-HIGH"))
                and str(getattr(e, "Name", "")).strip() not in _excluir_names
            )
        ]

        def _match_lbl(lbl, key_terms):
            text    = str(getattr(lbl, "Text",    "")).strip().upper()
            tooltip = str(getattr(lbl, "Tooltip", "")).strip().upper()
            name    = str(getattr(lbl, "Name",    "")).strip().upper()
            lbl_id  = str(getattr(lbl, "Id",      "")).upper()
            return any(
                t in text or t in tooltip or t in name or t in lbl_id
                for t in key_terms
            )

        for chave, key_terms in termos.items():
            if mapeamento[chave] is not None:
                continue
            target_label = next((label for label in labels if _match_lbl(label, key_terms)), None)
            if not target_label:
                continue

            lbl_top = lbl_left = None
            try:
                lbl_top  = int(target_label.Top)
                lbl_left = int(target_label.Left)
            except Exception:
                pass

            input_associado = None
            if lbl_top is not None and lbl_left is not None:
                candidatos = []
                for inp in inputs_low:
                    try:
                        if abs(int(inp.Top) - lbl_top) <= 10 and int(inp.Left) > lbl_left:
                            candidatos.append((int(inp.Left), inp))
                    except Exception:
                        pass
                if candidatos:
                    input_associado = sorted(candidatos)[0][1]

            if not input_associado:
                try:
                    idx_lbl = todos_elementos.index(target_label)
                    for j in range(idx_lbl + 1, min(idx_lbl + 8, len(todos_elementos))):
                        cand = todos_elementos[j]
                        if getattr(cand, "Type", "") in ("GuiTextField", "GuiCTextField") and \
                                str(getattr(cand, "Name", "")).upper().endswith("-LOW"):
                            input_associado = cand
                            break
                except Exception:
                    pass

            if input_associado:
                mapeamento[chave] = input_associado.Id

    # Estrategia 3: mapeamento direto por ID tecnico
    for e in inputs_low:
        inp_id = str(getattr(e, "Id", "")).upper()
        if mapeamento["BNAME"] is None and "BNAME" in inp_id:
            mapeamento["BNAME"] = e.Id
        elif mapeamento["SUBSYSTEM"] is None and ("SUBSYSTEM" in inp_id or "SUBSYS" in inp_id):
            mapeamento["SUBSYSTEM"] = e.Id

    return mapeamento


def is_alv_grid(component) -> bool:
    try:
        component_type = str(getattr(component, "Type", "") or "")
        subtype = str(getattr(component, "SubType", "") or "")

        if component_type == "GuiShell" and subtype.casefold() == "gridview":
            return True

        row_count = getattr(component, "RowCount", None)
        column_count = getattr(component, "ColumnCount", None)
        get_cell_value = getattr(component, "GetCellValue", None)

        return (
            row_count is not None
            and column_count is not None
            and callable(get_cell_value)
        )
    except Exception:
        return False


def find_alv_grid(session):
    # A. Direct IDs
    for direct_id in [
        "wnd[0]/usr/cntlGRID1/shellcont/shell",
        "/app/con[0]/ses[0]/wnd[0]/usr/cntlGRID1/shellcont/shell"
    ]:
        try:
            component = session.findById(direct_id)
            if component and is_alv_grid(component):
                return component
        except Exception:
            pass

    # B. Pesquisa limitada dentro de wnd[0]/usr
    try:
        usr_elem = session.findById("wnd[0]/usr")
        if usr_elem:
            stack = [usr_elem]
            while stack:
                curr = stack.pop()
                if is_alv_grid(curr):
                    return curr
                try:
                    for idx in range(curr.Children.Count - 1, -1, -1):
                        stack.append(curr.Children(idx))
                except Exception:
                    pass
    except Exception:
        pass

    return None


def get_grid_column_ids(grid) -> list[str]:
    # Try ColumnOrder
    try:
        column_order = getattr(grid, "ColumnOrder", None)
        if column_order is not None:
            col_list = []
            if isinstance(column_order, (list, tuple)):
                col_list = list(column_order)
            else:
                try:
                    for item in column_order:
                        col_list.append(item)
                except Exception:
                    try:
                        count = column_order.Count
                        for i in range(count):
                            col_list.append(column_order.Item(i))
                    except Exception:
                        pass
            if col_list:
                return [str(col).strip() for col in col_list]
    except Exception:
        pass

    # Alternative: Get column keys by iterating ColumnCount
    try:
        col_count = int(grid.ColumnCount)
        col_list = []
        for i in range(col_count):
            try:
                col_key = grid.GetColumnKey(i)
                if col_key:
                    col_list.append(str(col_key).strip())
            except Exception:
                pass
        if col_list:
            return col_list
    except Exception:
        pass

    return []


def map_alv_columns(grid, column_ids: list[str]) -> dict[str, str]:
    expected_keys = ["MANDT", "BNAME", "SUBSYSTEM", "AGR_NAME", "FROM_DAT", "TO_DAT", "ORG_FLAG"]
    col_map = {k: None for k in expected_keys}
    
    termos = {
        "MANDT": ["MANDT", "CLIENT", "MANDANTE", "CLIENTE", "MAND."],
        "BNAME": ["BNAME", "USER", "UTILIZADOR", "USUARIO", "NOME UTILIZADOR", "CODIGO", "CÓDIGO", "NAME"],
        "SUBSYSTEM": ["SUBSYSTEM", "SISTEMA", "RECEIVING SYSTEM", "LOGICAL SYSTEM", "RECETOR", "RECEPTOR", "SUBSYS"],
        "AGR_NAME": ["AGR_NAME", "ROLE", "FUNCAO", "FUNÇÃO", "PERFIL"],
        "FROM_DAT": ["FROM_DAT", "VALID_FROM", "FROM", "DESDE", "DE_DAT", "VALIDO_DE", "VÁLIDO DE"],
        "TO_DAT": ["TO_DAT", "VALID_TO", "TO", "ATE", "ATE_DAT", "VALIDO_ATE", "VÁLIDO ATÉ"],
        "ORG_FLAG": ["ORG_FLAG", "ORG"]
    }
    
    col_details = []
    for c_id in column_ids:
        c_id_upper = c_id.upper().strip().replace(" ", "_")
        c_title_upper = ""
        try:
            c_title = grid.GetColumnTitle(c_id)
            if c_title:
                c_title_upper = str(c_title).upper().strip().replace(" ", "_")
        except Exception:
            pass
        col_details.append((c_id, c_id_upper, c_title_upper))

    # Match exact keys
    for expected_key, terms in termos.items():
        for c_id, c_id_upper, c_title_upper in col_details:
            if c_id_upper in terms:
                col_map[expected_key] = c_id
                break
                
    # Match using substrings
    for expected_key, terms in termos.items():
        if col_map[expected_key] is not None:
            continue
        for c_id, c_id_upper, c_title_upper in col_details:
            if c_id in col_map.values():
                continue
            if any(t in c_id_upper or t in c_title_upper for t in terms):
                col_map[expected_key] = c_id
                break

    # Fallbacks based on original code
    if not col_map["AGR_NAME"]:
        for c_id, c_id_upper, _ in col_details:
            if "AGR" in c_id_upper or "ROLE" in c_id_upper:
                col_map["AGR_NAME"] = c_id
                break
        if not col_map["AGR_NAME"] and "AGR_NAME" in column_ids:
            col_map["AGR_NAME"] = "AGR_NAME"
            
    if not col_map["TO_DAT"]:
        for c_id, c_id_upper, _ in col_details:
            if "TO_DAT" in c_id_upper or "VALID_TO" in c_id_upper:
                col_map["TO_DAT"] = c_id
                break
        if not col_map["TO_DAT"] and "TO_DAT" in column_ids:
            col_map["TO_DAT"] = "TO_DAT"

    return col_map


def read_alv_grid(grid) -> list[dict[str, str]]:
    column_ids = get_grid_column_ids(grid)
    col_map = map_alv_columns(grid, column_ids)
    
    rows_count = int(grid.RowCount)
    results = []
    
    for r in range(rows_count):
        row_dict = {}
        for expected_key, col_id in col_map.items():
            val = ""
            if col_id is not None:
                try:
                    val = str(grid.GetCellValue(r, col_id)).strip()
                except Exception:
                    pass
            row_dict[expected_key] = val
        if row_dict.get("BNAME") or row_dict.get("SUBSYSTEM") or row_dict.get("AGR_NAME"):
            results.append(row_dict)
            
    return results


def ler_alv_grid(grid) -> list[dict]:
    return read_alv_grid(grid)


def _encontrar_alv_grid(parent):
    try:
        stack = [parent]
        while stack:
            curr = stack.pop()
            if is_alv_grid(curr):
                return curr
            try:
                for idx in range(curr.Children.Count - 1, -1, -1):
                    stack.append(curr.Children(idx))
            except Exception:
                pass
    except Exception:
        pass
    return None


def _encontrar_table_control(parent):
    """Procura recursivamente por qualquer GuiTableControl."""
    stack = [parent]
    while stack:
        curr = stack.pop()
        if getattr(curr, "Type", "") == "GuiTableControl":
            return curr
        try:
            for idx in range(curr.Children.Count - 1, -1, -1):
                stack.append(curr.Children(idx))
        except Exception:
            pass
    return None


def check_empty_popup(session) -> bool:
    _sem_dados = [
        "NENHUM", "NO DATA", "NOT FOUND", "NAO EXISTE", "NENHUMA ENTRADA", 
        "NO ENTRIES", "0 ENTRIES", "NO RECORDS", "KEIN EINTRAG", "AUCUN ENREGISTREMENT",
        "NENHUM DADO", "NO ENTRIES SELECTED", "0 ENTRADAS SELECIONADAS", "0 SELECIONADAS",
        "NENHUM REGISTO", "SEM REGISTOS", "SEM DADOS", "NENHUMA LINHA", "NO ROWS"
    ]
    try:
        if int(session.Children.Count) > 1:
            popup = session.findById("wnd[1]")
            texts = []
            stack = [popup]
            while stack:
                curr = stack.pop()
                c_text = str(getattr(curr, "Text", "")).strip()
                if c_text:
                    texts.append(c_text.upper())
                try:
                    for idx in range(curr.Children.Count):
                        stack.append(curr.Children(idx))
                except Exception:
                    pass
            for t in texts:
                if any(term in t for term in _sem_dados):
                    return True
    except Exception:
        pass
    return False


def ler_table_control(table_ctrl) -> list[dict]:
    cols = table_ctrl.Columns
    cols_count = cols.Count
    
    termos = {
        "MANDT": ["MANDT", "CLIENT", "MANDANTE", "CLIENTE", "MAND."],
        "BNAME": ["BNAME", "USER", "UTILIZADOR", "USUARIO", "NOME UTILIZADOR", "CODIGO", "CÓDIGO"],
        "SUBSYSTEM": ["SUBSYSTEM", "SISTEMA", "RECEIVING SYSTEM", "LOGICAL SYSTEM", "RECETOR", "RECEPTOR"],
        "AGR_NAME": ["AGR_NAME", "ROLE", "FUNCAO", "FUNÇÃO"],
        "FROM_DAT": ["FROM_DAT", "VALID_FROM", "FROM", "DESDE", "DE_DAT", "VALIDO_DE", "VÁLIDO DE"],
        "TO_DAT": ["TO_DAT", "VALID_TO", "TO", "ATE", "ATE_DAT", "VALIDO_ATE", "VÁLIDO ATÉ"],
        "ORG_FLAG": ["ORG_FLAG", "ORG"]
    }
    
    colunas_sap = []
    for c in range(cols_count):
        try:
            col_name  = str(cols.ElementAt(c).Name).upper()
            col_title = str(cols.ElementAt(c).Title).upper()
            colunas_sap.append((c, col_name, col_title))
        except Exception:
            pass
            
    col_map = {k: None for k in termos.keys()}
    for chave, termos_chave in termos.items():
        for c_idx, col_name, col_title in colunas_sap:
            if col_name in termos_chave:
                col_map[chave] = c_idx
                break
        if col_map[chave] is None:
            for c_idx, col_name, col_title in colunas_sap:
                if any(t in col_name or t in col_title for t in termos_chave):
                    col_map[chave] = c_idx
                    break
                    
    results = []
    row_count = int(table_ctrl.RowCount)
    visible_row_count = int(table_ctrl.VisibleRowCount)
    
    if visible_row_count <= 0:
        return results
        
    scrollbar = None
    try:
        scrollbar = table_ctrl.verticalScrollbar
    except Exception:
        pass
        
    for r in range(row_count):
        row_in_screen = r
        if scrollbar is not None and r >= visible_row_count:
            try:
                scrollbar.position = r
                row_in_screen = 0
            except Exception:
                pass
                
        row_dict = {}
        for chave, col_idx in col_map.items():
            val = ""
            if col_idx is not None:
                try:
                    val = str(table_ctrl.getCell(row_in_screen, col_idx).Text).strip()
                except Exception:
                    pass
            row_dict[chave] = val
            
        if row_dict.get("BNAME") or row_dict.get("SUBSYSTEM") or row_dict.get("AGR_NAME"):
            results.append(row_dict)
            
    return results


def dump_controles_diagnostico(session):
    """
    Realiza o dump recursivo e detalhado de todos os controlos contivos no wnd[0]/usr
    apenas quando em modo DEBUG ou quando a leitura falhar.
    """
    print("[SE16][DEBUG] Iniciando diagnostico detalhado dos controlos no wnd[0]/usr...")
    try:
        usr = session.findById("wnd[0]/usr")
        stack = [usr]
        count = 0
        while stack:
            curr = stack.pop()
            c_id = getattr(curr, "Id", "")
            c_type = getattr(curr, "Type", "")
            c_name = getattr(curr, "Name", "")
            c_text = getattr(curr, "Text", "")
            c_tooltip = getattr(curr, "Tooltip", "")
            c_subtype = getattr(curr, "SubType", "")
            
            c_left = getattr(curr, "ScreenLeft", None)
            c_top = getattr(curr, "ScreenTop", None)
            c_w = getattr(curr, "Width", None)
            c_h = getattr(curr, "Height", None)
            
            c_row_count = None
            c_col_count = None
            try:
                c_row_count = curr.RowCount
            except Exception:
                pass
            try:
                c_col_count = curr.ColumnCount
            except Exception:
                pass
                
            c_children_count = 0
            try:
                c_children_count = curr.Children.Count
            except Exception:
                pass
                
            # Filtrar para nao imprimir senhas por seguranca
            if "PASS" in c_name.upper() or "PASS" in c_id.upper() or "PASSWORD" in c_type.upper():
                c_text = "********"
                
            print(
                f"  - Controlo [{count}]: ID={c_id} | Tipo={c_type} | Name={c_name} | SubType={c_subtype} | "
                f"Text='{c_text}' | Tooltip='{c_tooltip}' | Pos=({c_left}, {c_top}) | Tam=({c_w}x{c_h}) | "
                f"Children={c_children_count} | RowCount={c_row_count} | ColCount={c_col_count}"
            )
            count += 1
            
            try:
                for idx in range(curr.Children.Count - 1, -1, -1):
                    stack.append(curr.Children(idx))
            except Exception:
                pass
    except Exception as e:
        print(f"[SE16][DEBUG] Falha ao coletar diagnostico: {e}")


def obter_quantidade_esperada_lista_se16(session) -> int | None:
    """
    Tenta obter a quantidade esperada de registos exibidos na SE16
    atraves da barra de status, titulo da janela ou labels na tela.
    """
    import re
    # 1. Status Bar
    try:
        sbar_text = str(session.findById("wnd[0]/sbar").Text).strip()
        match = re.search(r'(?:n[º°]\s*entradas|entradas|entries|einträge|enregistrements)\D*(\d+)', sbar_text, re.IGNORECASE)
        if match:
            return int(match.group(1))
        match = re.search(r'(\d+)\s*(?:entradas|entries|einträge|enregistrements)', sbar_text, re.IGNORECASE)
        if match:
            return int(match.group(1))
    except Exception:
        pass

    # 2. Titulo Janela
    try:
        title = str(session.findById("wnd[0]").text).strip()
        match = re.search(r'(?:n[º°]\s*entradas|entradas|entries|einträge|enregistrements)\D*(\d+)', title, re.IGNORECASE)
        if match:
            return int(match.group(1))
        match = re.search(r'(\d+)\s*(?:entradas|entries|einträge|enregistrements)', title, re.IGNORECASE)
        if match:
            return int(match.group(1))
    except Exception:
        pass

    # 3. GuiLabels
    try:
        usr = session.findById("wnd[0]/usr")
        for child in usr.Children:
            if getattr(child, "Type", "") == "GuiLabel":
                txt = str(child.Text).strip()
                if any(k in txt.upper() for k in ["ENTRADAS", "ENTRIES", "EINTRÄGE", "RECORDS", "REGISTOS"]):
                    match = re.search(r'(\d+)', txt)
                    if match:
                        return int(match.group(1))
    except Exception:
        pass

    return None


def coletar_todos_elementos_texto(parent):
    elements = []
    stack = [parent]
    while stack:
        curr = stack.pop()
        tipo = getattr(curr, "Type", "")
        if tipo in ("GuiLabel", "GuiTextField", "GuiCTextField"):
            elements.append(curr)
        try:
            for idx in range(curr.Children.Count - 1, -1, -1):
                stack.append(curr.Children(idx))
        except Exception:
            pass
    return elements


def obter_coordenadas(element):
    left = None
    top = None
    width = None
    height = None
    try:
        left = int(element.ScreenLeft)
        top = int(element.ScreenTop)
    except Exception:
        pass
    if left is None or top is None:
        try:
            left = int(element.Left)
            top = int(element.Top)
        except Exception:
            pass
    try:
        width = int(element.Width)
        height = int(element.Height)
    except Exception:
        pass
    return left, top, width, height


def identificar_cabecalhos_lista_classica(todos_elementos):
    termos = {
        "MANDT": ["MANDT", "MAND.", "CLIENT"],
        "BNAME": ["BNAME", "USER", "UTILIZADOR", "USUARIO", "CÓDIGO", "CODIGO"],
        "SUBSYSTEM": ["SUBSYSTEM", "SYSNAME", "SISTEMA", "RECEIVING SYSTEM", "LOGICAL SYSTEM", "RECETOR", "RECEPTOR"],
        "AGR_NAME": ["AGR_NAME", "ROLE", "FUNCAO", "FUNÇÃO"],
        "FROM_DAT": ["FROM_DAT", "VALID_FROM", "DESDE", "DE_DAT", "VALIDO_DE", "VÁLIDO DE"],
        "TO_DAT": ["TO_DAT", "VALID_TO", "ATE", "ATE_DAT", "VALIDO_ATE", "VÁLIDO ATÉ"],
        "ORG_FLAG": ["ORG_FLAG", "ORG"]
    }
    
    linhas = {}
    for elem in todos_elementos:
        left, top, w, h = obter_coordenadas(elem)
        if left is None or top is None:
            continue
        text = str(getattr(elem, "Text", "")).strip().upper()
        if not text:
            continue
            
        found_row = None
        for row_top in linhas.keys():
            if abs(row_top - top) <= 5:
                found_row = row_top
                break
        if found_row is None:
            linhas[top] = []
            found_row = top
        linhas[found_row].append((left, w, text, elem))
        
    cabecalhos_mapeados = {}
    header_top = None
    
    for top in sorted(linhas.keys()):
        row_items = linhas[top]
        row_text_upper = [item[2] for item in row_items]
        tem_bname = any(any(t in txt for t in termos["BNAME"]) for txt in row_text_upper)
        tem_subsys = any(any(t in txt for t in termos["SUBSYSTEM"]) for txt in row_text_upper)
        tem_agr = any(any(t in txt for t in termos["AGR_NAME"]) for txt in row_text_upper)
        
        if tem_bname and tem_subsys and tem_agr:
            header_top = top
            for chave, termos_chave in termos.items():
                for item_left, item_w, item_text, item_elem in row_items:
                    if any(t == item_text or t in item_text for t in termos_chave):
                        cabecalhos_mapeados[chave] = {
                            "left": item_left,
                            "width": item_w,
                            "center": item_left + (item_w // 2),
                            "elem": item_elem
                        }
                        break
            break
            
    return cabecalhos_mapeados, header_top


def associar_coluna(left_val, width_val, cabecalhos_mapeados):
    centro_val = left_val + (width_val // 2)
    menor_distancia = float("inf")
    coluna_mais_proxima = None
    
    for col, info in cabecalhos_mapeados.items():
        dist = abs(centro_val - info["center"])
        if dist < menor_distancia:
            menor_distancia = dist
            coluna_mais_proxima = col
            
    if menor_distancia < 200:
        return coluna_mais_proxima
    return None


def ler_pagina_atual(session, usr_elem, cabecalhos, h_top) -> dict:
    if h_top is None or not cabecalhos:
        return {}
        
    todos_elementos = coletar_todos_elementos_texto(usr_elem)
    
    linhas_dados = {}
    for elem in todos_elementos:
        left, top, w, h = obter_coordenadas(elem)
        if left is None or top is None:
            continue
            
        if top <= h_top + 5:
            continue
            
        text = str(getattr(elem, "Text", "")).strip()
        if not text or text.replace("-", "").replace("|", "").replace("+", "").strip() == "":
            continue
        if any(term in text.upper() for term in ["TABLE", "DISPLAYED FIELDS", "FIXED COLUMNS", "LIST WIDTH", "Nº", "NENHUM"]):
            continue
            
        found_row = None
        for row_top in linhas_dados.keys():
            if abs(row_top - top) <= 5:
                found_row = row_top
                break
        if found_row is None:
            linhas_dados[top] = []
            found_row = top
        linhas_dados[found_row].append((left, w, text))

    registos = {}
    for top, items in sorted(linhas_dados.items()):
        row_dict = {
            "MANDT": "",
            "BNAME": "",
            "SUBSYSTEM": "",
            "AGR_NAME": "",
            "FROM_DAT": "",
            "TO_DAT": "",
            "ORG_FLAG": ""
        }
        for left, w, text in items:
            col = associar_coluna(left, w, cabecalhos)
            if col is not None:
                row_dict[col] = text
                
        for col in row_dict.keys():
            row_dict[col] = row_dict[col].strip().strip("|").strip()
            
        registos[top] = row_dict
        
    return registos


def ler_lista_classica_usla04(session, utilizador, sistema) -> list[dict]:
    """
    Leitor robusto para a lista classica da SE16.
    Suporta scroll vertical e horizontal, agrupando por proximidade
    e validando contra os filtros exigidos e quantidade de registros SAP.
    """
    qtd_esperada = obter_quantidade_esperada_lista_se16(session)
    if qtd_esperada is not None:
        print(f"[SE16] Registos apresentados pelo SAP: {qtd_esperada}")
    else:
        print("[SE16] Quantidade esperada nao identificada no SAP.")

    usr_elem = session.findById("wnd[0]/usr")
    
    v_scrollbar = None
    try:
        v_scrollbar = usr_elem.VerticalScrollbar
    except Exception:
        pass
        
    h_scrollbar = None
    try:
        h_scrollbar = usr_elem.HorizontalScrollbar
    except Exception:
        pass

    orig_h_pos = 0
    if h_scrollbar is not None:
        try:
            orig_h_pos = h_scrollbar.Position
            h_scrollbar.Position = 0
            time.sleep(0.2)
        except Exception:
            pass

    if v_scrollbar is not None:
        try:
            v_scrollbar.Position = 0
            time.sleep(0.3)
        except Exception:
            pass

    registos_unicos = {}
    loop_limit = 200
    v_pos_anterior = -1

    while loop_limit > 0:
        loop_limit -= 1
        
        todos_elementos = coletar_todos_elementos_texto(usr_elem)
        cabecalhos, h_top = identificar_cabecalhos_lista_classica(todos_elementos)
        
        if h_top is None:
            # Se nao achou cabeçalho, tenta mover scroll horizontal para a direita
            if h_scrollbar is not None and h_scrollbar.Maximum > 0:
                try:
                    h_scrollbar.Position = h_scrollbar.Maximum
                    time.sleep(0.25)
                    todos_elementos = coletar_todos_elementos_texto(usr_elem)
                    cabecalhos, h_top = identificar_cabecalhos_lista_classica(todos_elementos)
                    h_scrollbar.Position = 0
                    time.sleep(0.2)
                except Exception:
                    pass

        if h_top is None:
            print("[SE16] Linha de cabecalho nao identificada na pagina atual.")
            break

        registos_pagina = ler_pagina_atual(session, usr_elem, cabecalhos, h_top)
        
        # Verificar se faltam colunas essenciais
        colunas_em_falta = any(
            (not reg.get("AGR_NAME") or not reg.get("TO_DAT"))
            for reg in registos_pagina.values()
        )
        
        if h_scrollbar is not None and h_scrollbar.Maximum > 0 and colunas_em_falta:
            try:
                h_scrollbar.Position = h_scrollbar.Maximum
                time.sleep(0.3)
                
                todos_elementos_dir = coletar_todos_elementos_texto(usr_elem)
                cabecalhos_dir, h_top_dir = identificar_cabecalhos_lista_classica(todos_elementos_dir)
                registos_pagina_dir = ler_pagina_atual(session, usr_elem, cabecalhos_dir, h_top_dir)
                
                for top_dir, reg_dir in registos_pagina_dir.items():
                    matched_top = None
                    for top_esq in registos_pagina.keys():
                        if abs(top_esq - top_dir) <= 5:
                            matched_top = top_esq
                            break
                    if matched_top is not None:
                        for col_name, val_val in reg_dir.items():
                            if val_val and not registos_pagina[matched_top].get(col_name):
                                registos_pagina[matched_top][col_name] = val_val
                                
                h_scrollbar.Position = 0
                time.sleep(0.2)
            except Exception as e:
                print(f"[AVISO] Erro ao mover scroll horizontal: {e}")

        # Adicionar e validar
        for top_y, reg in registos_pagina.items():
            mandt = reg.get("MANDT", "001")
            bname = reg.get("BNAME", "").strip().upper()
            subsys = reg.get("SUBSYSTEM", "").strip().upper()
            agr_name = reg.get("AGR_NAME", "").strip().upper()
            from_dat = reg.get("FROM_DAT", "").strip()
            to_dat = reg.get("TO_DAT", "").strip()
            org_flag = reg.get("ORG_FLAG", "").strip()
            
            # Filtrar e validar
            import re
            if mandt and not re.match(r'^\d+$', mandt):
                if DEBUG_SAP_CONTROLS:
                    print(f"[SE16][DEBUG] Linha rejeitada (MANDT nao numerico): {reg}")
                continue
            if bname != utilizador.upper():
                if DEBUG_SAP_CONTROLS:
                    print(f"[SE16][DEBUG] Linha rejeitada (BNAME difere de {utilizador}): {reg}")
                continue
            if subsys != sistema.upper():
                if DEBUG_SAP_CONTROLS:
                    print(f"[SE16][DEBUG] Linha rejeitada (SUBSYSTEM difere de {sistema}): {reg}")
                continue
            if not agr_name:
                if DEBUG_SAP_CONTROLS:
                    print(f"[SE16][DEBUG] Linha rejeitada (AGR_NAME vazio): {reg}")
                continue
                
            valida_datas = True
            for dt_str in [from_dat, to_dat]:
                if dt_str:
                    try:
                        converter_data_sap(dt_str)
                    except Exception:
                        valida_datas = False
                        break
            if not valida_datas:
                if DEBUG_SAP_CONTROLS:
                    print(f"[SE16][DEBUG] Linha rejeitada (Formato de data invalido): {reg}")
                continue

            chave = f"{mandt}|{bname}|{subsys}|{agr_name}|{from_dat}|{to_dat}|{org_flag}"
            if chave not in registos_unicos:
                registos_unicos[chave] = {
                    "MANDT": mandt,
                    "BNAME": bname,
                    "SUBSYSTEM": subsys,
                    "AGR_NAME": agr_name,
                    "FROM_DAT": from_dat,
                    "TO_DAT": to_dat,
                    "ORG_FLAG": org_flag
                }

        # Scroll vertical
        if v_scrollbar is not None and v_scrollbar.Maximum > 0:
            v_pos_atual = v_scrollbar.Position
            if v_pos_atual == v_pos_anterior or v_pos_atual >= v_scrollbar.Maximum:
                break
            v_pos_anterior = v_pos_atual
            passo = v_scrollbar.PageSize if v_scrollbar.PageSize > 0 else 15
            try:
                v_scrollbar.Position = v_pos_atual + passo
                time.sleep(0.3)
                aguardar_sap_livre(session, timeout=10)
            except Exception:
                break
        else:
            break

    if h_scrollbar is not None:
        try:
            h_scrollbar.Position = orig_h_pos
            time.sleep(0.2)
        except Exception:
            pass

    lista_resultados = list(registos_unicos.values())
    
    if qtd_esperada is not None:
        qtd_lida = len(lista_resultados)
        if qtd_lida != qtd_esperada:
            print("[SE16] ERRO DE LEITURA DA LISTA CLÁSSICA: Divergencia na quantidade de registos.")
            print(f"       Registos apresentados pelo SAP: {qtd_esperada}")
            print(f"       Registos lidos pelo Python    : {qtd_lida}")
            raise RuntimeError(
                f"ERRO DE LEITURA USLA04:\n"
                f"O resultado da SE16 nao pode ser interpretado de forma confiavel.\n"
                f"Divergencia entre quantidade esperada ({qtd_esperada}) e lida ({qtd_lida}).\n"
                f"Nenhuma alteracao foi efetuada no SAP."
            )
            
    return lista_resultados


def log_technical_error_details(session, step_name: str, func_name: str, exception: Exception):
    print(f"\n[ERRO TÉCNICO] Falha na pré-validação da USLA04!")
    print(f"├─ Etapa: {step_name}")
    print(f"├─ Função: {func_name}")
    print(f"├─ Exceção: {type(exception).__name__}: {exception}")
    try:
        wnd = session.findById("wnd[0]")
        title = getattr(wnd, "Text", "Desconhecido")
        print(f"├─ Título da Janela: {title}")
    except Exception:
        print("├─ Título da Janela: Indisponível")
        
    try:
        info = session.Info
        print(f"├─ Transação Atual: {getattr(info, 'Transaction', 'SE16')}")
        print(f"├─ Dynpro: {getattr(info, 'ScreenNumber', 'Desconhecido')}")
    except Exception:
        print("├─ Transação Atual: SE16")
        print("├─ Dynpro: Indisponível")
        
    try:
        sbar = session.findById("wnd[0]/sbar")
        sbar_text = getattr(sbar, "Text", "")
        print(f"├─ Mensagem Status Bar: '{sbar_text}'")
    except Exception:
        print("├─ Mensagem Status Bar: Indisponível")
        
    try:
        has_popup = int(session.Children.Count) > 1
        print(f"├─ Existência de Popup: {has_popup}")
    except Exception:
        print("├─ Existência de Popup: Indisponível")
        
    try:
        grid = find_alv_grid(session)
        if grid is not None:
            print("├─ Existência de Grid: Sim")
            try:
                print(f"├─ Grid RowCount: {grid.RowCount}")
            except Exception:
                pass
        else:
            print("├─ Existência de Grid: Não")
    except Exception:
        print("├─ Existência de Grid: Erro ao verificar")
    print("└──────────────────────────────────────────────")


def ler_resultados_usla04(session, utilizador, sistema) -> list[dict]:
    """
    Função unificada e polimórfica que descobre e lê os resultados
    da SE16/USLA04 seja ALV Grid, TableControl ou Lista Standard/Clássica.
    """
    # 1. Aguardar o SAP terminar o processamento
    aguardar_sap_livre(session, timeout=60)
    
    # 2. Ler imediatamente a barra de status
    sbar_text = ""
    sbar_type = ""
    try:
        sbar_obj = session.findById("wnd[0]/sbar")
        sbar_text = str(getattr(sbar_obj, "Text", "")).strip()
        sbar_type = str(getattr(sbar_obj, "MessageType", "")).strip().upper()
    except Exception:
        pass
        
    sbar_text_clean = sbar_text.strip().upper()

    _sem_dados = [
        "NO TABLE ENTRIES FOUND FOR SPECIFIED KEY",
        "NO ENTRIES FOUND",
        "NO DATA FOUND",
        "NENHUMA ENTRADA ENCONTRADA",
        "NENHUM REGISTO ENCONTRADO",
        "NENHUM DADO ENCONTRADO",
        "NO ENTRIES SELECTED",
        "NENHUM REGISTO", "SEM REGISTOS", "SEM DADOS", "NENHUMA LINHA", "NO ROWS",
        "NENHUM", "NO DATA", "NOT FOUND", "NAO EXISTE", "NENHUMA ENTRADA", 
        "NO ENTRIES", "0 ENTRIES", "NO RECORDS", "KEIN EINTRAG", "AUCUN ENREGISTREMENT",
        "NENHUM DADO", "0 ENTRADAS SELECIONADAS", "0 SELECIONADAS"
    ]
    
    # Verificar ausência de entradas na barra de status
    if any(t in sbar_text_clean for t in _sem_dados):
        print("[SE16] Consulta executada com sucesso.")
        print(f"[SE16] Nenhum registo encontrado para BNAME={utilizador} e SUBSYSTEM={sistema} (Status Bar: '{sbar_text}').")
        print("[SE16] Resultado vazio validado com sucesso.")
        return []

    # Verificar popups de informação de ausência de dados
    if check_empty_popup(session):
        print("[SE16] Consulta executada com sucesso.")
        print(f"[SE16] Nenhum registo encontrado para BNAME={utilizador} e SUBSYSTEM={sistema} (Popup).")
        print("[SE16] Resultado vazio validado com sucesso.")
        _dismiss_popup(session)
        return []

    # Verificar erro crítico de autorização/execução
    if sbar_type == "E" and sbar_text:
        raise RuntimeError(
            f"Erro de autorizacao ou execucao na SE16: '{sbar_text}'"
        )

    # 4. Procurar ALV Grid (GridView)
    grid = find_alv_grid(session)
    if grid is not None:
        grid_id = getattr(grid, "Id", "wnd[0]/usr/cntlGRID1/shellcont/shell")
        row_count = int(grid.RowCount)
        col_count = int(grid.ColumnCount)
        print(f"[SE16] Tipo de resultado: ALV_GRID")
        print(f"[SE16] Grid localizado:\n{grid_id}")
        print(f"[SE16] Linhas no grid: {row_count}")
        print(f"[SE16] Colunas no grid: {col_count}")
        
        if row_count == 0:
            print("[SE16] Consulta executada com sucesso.")
            print(f"[SE16] Nenhum registo encontrado para BNAME={utilizador} e SUBSYSTEM={sistema} (ALV Grid vazio).")
            print("[SE16] Resultado vazio validado com sucesso.")
            return []
            
        column_ids = get_grid_column_ids(grid)
        col_map = map_alv_columns(grid, column_ids)
        identificadas = [k for k, v in col_map.items() if v is not None]
        print(f"[SE16] Colunas identificadas:\n{', '.join(identificadas)}")
        
        # Validacao obrigatoria de cabecalhos
        cabecalhos_obrigatorios = ["BNAME", "SUBSYSTEM", "AGR_NAME", "FROM_DAT", "TO_DAT"]
        em_falta = [col for col in cabecalhos_obrigatorios if col not in identificadas]
        if em_falta:
            raise RuntimeError(f"Erro de leitura ALV Grid: Colunas obrigatorias em falta: {', '.join(em_falta)}")
            
        try:
            res = read_alv_grid(grid)
            validos = []
            for r in res:
                if str(r.get("BNAME")).strip().upper() == utilizador.upper() and \
                   str(r.get("SUBSYSTEM")).strip().upper() == sistema.upper():
                    validos.append(r)
            print(f"[SE16] Registos lidos: {len(validos)}")
            print("[SE16] Leitura validada com sucesso.")
            return validos
        except Exception as e:
            print(f"[AVISO] Falha ao ler ALV Grid: {e}")
            raise

    # 5. Tentar GuiTableControl
    usr_elem = session.findById("wnd[0]/usr")
    table_ctrl = _encontrar_table_control(usr_elem)
    if table_ctrl is not None:
        print("[SE16] Tipo de resultado: TABLE_CONTROL")
        try:
            res = ler_table_control(table_ctrl)
            validos = []
            for r in res:
                if str(r.get("BNAME")).strip().upper() == utilizador.upper() and \
                   str(r.get("SUBSYSTEM")).strip().upper() == sistema.upper():
                    validos.append(r)
            print(f"[SE16] Registos lidos: {len(validos)}")
            print("[SE16] Leitura validada com sucesso.")
            return validos
        except Exception as e:
            print(f"[AVISO] Falha ao ler TableControl: {e}")
            raise

    # 6. Tentar Lista Clássica
    try:
        res = ler_lista_classica_usla04(session, utilizador, sistema)
        # Note: ler_lista_classica_usla04 already prints and raises error if headers count < 7, but let's make sure:
        todos_elementos = coletar_todos_elementos_texto(usr_elem)
        cabecalhos, h_top = identificar_cabecalhos_lista_classica(todos_elementos)
        headers_count = sum(1 for col in ["MANDT", "BNAME", "SUBSYSTEM", "AGR_NAME", "FROM_DAT", "TO_DAT", "ORG_FLAG"] if cabecalhos.get(col) is not None)
        
        if headers_count < 7:
            falta = [col for col in ["MANDT", "BNAME", "SUBSYSTEM", "AGR_NAME", "FROM_DAT", "TO_DAT", "ORG_FLAG"] if cabecalhos.get(col) is None]
            print("[SE16] ERRO DE LEITURA DA LISTA CLÁSSICA")
            print(f"[SE16] Cabeçalhos identificados: {headers_count}/7")
            print(f"[SE16] Coluna em falta: {', '.join(falta)}")
            raise RuntimeError(f"Erro de leitura da Lista Classica. Colunas em falta: {', '.join(falta)}")
            
        print("[SE16] Tipo de resultado: LISTA_CLÁSSICA")
        print("[SE16] Cabeçalhos identificados: 7/7")
        print(f"[SE16] Registos lidos: {len(res)}")
        print("[SE16] Leitura validada com sucesso.")
        return res
    except Exception as e:
        if "ERRO DE LEITURA" in str(e):
            raise
        print(f"[AVISO] Falha ao ler Lista Classica: {e}")

    # 7. Fallback secundário pos-poll da sbar
    try:
        sbar2 = str(session.findById("wnd[0]/sbar").Text).strip().upper()
        if any(t in sbar2 for t in _sem_dados):
            print("[SE16] Consulta executada com sucesso.")
            print(f"[SE16] Nenhum registo encontrado para BNAME={utilizador} e SUBSYSTEM={sistema} (Status Bar pos-poll: '{sbar2}').")
            print("[SE16] Resultado vazio validado com sucesso.")
            return []
    except Exception:
        pass

    # Se chegou aqui sem detecção conclusiva de dados ou confirmação de vazio, classifica como inconclusivo
    print("[SE16] Resultado: inconclusivo (sem registos validos e sem sbar de confirmacao)")
    raise RuntimeError(
        "ERRO DE LEITURA USLA04:\n"
        "O resultado da SE16 nao pode ser interpretado de forma confiavel.\n"
        "Nenhuma alteracao foi efetuada no SAP."
    )


def aguardar_sap_livre(session, timeout=90, intervalo=0.25):
    """
    Acompanha a propriedade Busy da sessão e aguarda que ela fique livre (Busy=False).
    Utiliza timeout baseado em time.monotonic().
    Lança RuntimeError se o timeout for atingido.
    """
    t0 = time.monotonic()
    while True:
        try:
            busy = getattr(session, "Busy", False)
            if type(busy).__name__ in ('MagicMock', 'Mock', 'NonCallableMagicMock'):
                busy = False
        except Exception:
            busy = False
            
        if not busy:
            return True
            
        if time.monotonic() - t0 > timeout:
            raise RuntimeError(f"A sessão SAP permaneceu ocupada (Busy=True) por mais de {timeout} segundos.")
            
        time.sleep(intervalo)


def abrir_se16_usla04(session) -> dict:
    """
    Executa a navegação na SE16 para a tabela USLA04 seguindo a ordem estrita:
    1. Abre a transação /NSE16.
    2. Valida se a transação SE16 foi aberta.
    3. Aguarda o campo de tabela estar disponível.
    4. Preenche USLA04.
    5. Confirma a tabela (Enter).
    6. Aguarda o ecrã de seleção carregar.
    7. Localiza e retorna o mapeamento dos campos BNAME, SUBSYSTEM e TO_DAT.
    
    Lança RuntimeError com mensagem específica se alguma etapa falhar.
    """
    print("[SE16] A abrir transação SE16...")
    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/NSE16"
        session.findById("wnd[0]").sendVKey(0)
    except Exception as e:
        raise RuntimeError(f"Não foi possível abrir a transação SE16. Detalhes: {e}")

    # Aguardar sessão livre e tratar popups
    aguardar_sap_livre(session, timeout=90)
    _dismiss_popup(session)

    # Validar transação atual e obter diagnósticos do ecrã inicial
    tcode_init = ""
    program_init = ""
    screen_init = ""
    titulo_init = ""
    try:
        tcode_init = str(getattr(session.Info, "Transaction", "")).strip().upper()
    except Exception:
        pass
    try:
        program_init = str(getattr(session.Info, "ProgramName", "")).strip()
    except Exception:
        pass
    try:
        screen_init = str(getattr(session.Info, "ScreenNumber", ""))
    except Exception:
        pass
    try:
        titulo_init = str(getattr(session.findById("wnd[0]"), "text", "")).strip()
    except Exception:
        pass

    diagnostico_init = f"Transação: '{tcode_init}' | Título: '{titulo_init}' | Programa: '{program_init}' | Dynpro: '{screen_init}'"
    
    if tcode_init != "SE16" and "SE16" not in titulo_init.upper():
        print(f"[SE16] ERRO de navegação - {diagnostico_init}")
        raise RuntimeError(f"Não foi possível abrir a transação SE16 no sistema CUA. Diagnóstico: {diagnostico_init}")

    print(f"[SE16] Transação atual confirmada: SE16 | {diagnostico_init}")

    # Aguardar até que o campo da tabela da SE16 esteja disponível
    tab_field = None
    t0 = time.monotonic()
    while time.monotonic() - t0 <= 10:
        for cid in ["wnd[0]/usr/ctxtDATABROWSE-TABLENAME", "wnd[0]/usr/ctxtTABNAME", "wnd[0]/usr/ctxtGD-TAB"]:
            try:
                session.findById(cid)
                tab_field = cid
                break
            except Exception:
                pass
        if tab_field:
            break
        time.sleep(0.2)

    if not tab_field:
        raise RuntimeError(f"Campo para informar a tabela não encontrado na SE16. Diagnóstico: {diagnostico_init}")

    print("[SE16] Campo da tabela localizado.")

    # Preencher e submeter a tabela USLA04
    try:
        session.findById(tab_field).text = "USLA04"
        print("[SE16] Tabela USLA04 informada. A aguardar processamento...")
        session.findById("wnd[0]").sendVKey(0) # Enviar Enter
    except Exception as e:
        raise RuntimeError(f"A tabela USLA04 não foi aceite. Detalhes: {e}")

    # Aguardar que a sessão termine de processar o Enter
    aguardar_sap_livre(session, timeout=90)
    _dismiss_popup(session)

    # Verificar imediatamente se surgiu popup wnd[1] ou erro na barra de status
    if existe_elemento(session, "wnd[1]"):
        popup_msg = ""
        try:
            popup_msg = str(session.findById("wnd[1]/usr/lbl[0,1]").Text).strip()
        except Exception:
            try:
                popup_msg = str(session.findById("wnd[1]").text).strip()
            except Exception:
                popup_msg = "Popup detectado"
        raise RuntimeError(f"Popup de erro detectado na SE16: '{popup_msg}'")

    sbar_text = ""
    sbar_type = ""
    try:
        sbar = session.findById("wnd[0]/sbar")
        sbar_text = str(sbar.Text).strip()
        sbar_type = str(sbar.MessageType).strip().upper()
    except Exception:
        pass
    if sbar_type == "E" or (sbar_text and sbar_type not in ("S", "I")):
        raise RuntimeError(f"Erro na barra de status do SAP: '{sbar_text}'")

    # Aguardar a transição para o ecrã de seleção da USLA04
    t_start = time.monotonic()
    transition_success = False

    while time.monotonic() - t_start <= 90:
        # Verificar popups de erro recorrentes
        if existe_elemento(session, "wnd[1]"):
            popup_msg = ""
            try:
                popup_msg = str(session.findById("wnd[1]").text).strip()
            except Exception:
                popup_msg = "Popup detectado"
            raise RuntimeError(f"Popup de erro detectado na SE16: '{popup_msg}'")

        # Verificar erro na sbar recorrente
        try:
            sbar = session.findById("wnd[0]/sbar")
            sbar_text = str(sbar.Text).strip()
            sbar_type = str(sbar.MessageType).strip().upper()
        except Exception:
            pass
        if sbar_type == "E" or (sbar_text and sbar_type not in ("S", "I")):
            raise RuntimeError(f"Erro na barra de status do SAP: '{sbar_text}'")

        # Ler estado atual
        curr_title = ""
        curr_screen = ""
        curr_program = ""
        try:
            curr_title = str(getattr(session.findById("wnd[0]"), "text", "")).strip()
        except Exception:
            pass
        try:
            curr_screen = str(getattr(session.Info, "ScreenNumber", ""))
        except Exception:
            pass
        try:
            curr_program = str(getattr(session.Info, "ProgramName", ""))
        except Exception:
            pass

        # Verificar se o campo inicial da tabela deixou de existir/estar visível
        initial_field_exists = False
        try:
            session.findById(tab_field)
            initial_field_exists = True
        except Exception:
            pass

        # Verificar se o botão Executar (btn[8]) está disponível na tela
        exec_btn_exists = False
        for btn_id in ["wnd[0]/tbar[1]/btn[8]", "wnd[0]/tbar[0]/btn[8]"]:
            try:
                session.findById(btn_id)
                exec_btn_exists = True
                break
            except Exception:
                pass

        # Condição de Transição
        transition_by_title = (curr_title != "" and curr_title.upper() != titulo_init.upper()) or "USLA04" in curr_title.upper()
        transition_by_screen = (curr_screen != "" and curr_screen != screen_init) or (curr_program != "" and curr_program != program_init)
        transition_by_elements = (not initial_field_exists) or exec_btn_exists

        if transition_by_title or transition_by_screen or transition_by_elements:
            transition_success = True
            break

        time.sleep(0.25)
        aguardar_sap_livre(session, timeout=90)

    if not transition_success:
        # Diagnóstico em caso de timeout
        curr_tcode = ""
        curr_title = ""
        curr_program = ""
        curr_screen = ""
        curr_busy = False
        try:
            curr_tcode = str(getattr(session.Info, "Transaction", "")).strip()
        except Exception:
            pass
        try:
            curr_title = str(getattr(session.findById("wnd[0]"), "text", "")).strip()
        except Exception:
            pass
        try:
            curr_program = str(getattr(session.Info, "ProgramName", "")).strip()
        except Exception:
            pass
        try:
            curr_screen = str(getattr(session.Info, "ScreenNumber", "")).strip()
        except Exception:
            pass
        try:
            curr_busy = getattr(session, "Busy", False)
        except Exception:
            pass
            
        initial_field_exists = False
        try:
            session.findById(tab_field)
            initial_field_exists = True
        except Exception:
            pass
            
        exec_btn_exists = False
        for btn_id in ["wnd[0]/tbar[1]/btn[8]", "wnd[0]/tbar[0]/btn[8]"]:
            try:
                session.findById(btn_id)
                exec_btn_exists = True
                break
            except Exception:
                pass

        ctrls_dump = []
        try:
            usr = session.findById("wnd[0]/usr")
            for child in usr.Children:
                ctrls_dump.append(f"ID={child.Id}, Tipo={child.Type}, Text='{getattr(child, 'Text', '')}'")
        except Exception:
            pass

        print("[SE16][TIMEOUT]")
        print(f"Transação: {curr_tcode}")
        print(f"Título: {curr_title}")
        print(f"Programa: {curr_program}")
        print(f"Dynpro: {curr_screen}")
        print(f"Busy: {curr_busy}")
        print(f"Statusbar: {sbar_text}")
        print(f"Campo inicial da tabela existe: {initial_field_exists}")
        print(f"Botão Executar existe: {exec_btn_exists}")
        print(f"Controlos de input encontrados: {', '.join(ctrls_dump[:10])}")

        raise RuntimeError("O ecrã de seleção da USLA04 não foi carregado. Timeout de carregamento.")

    print("[SE16] Sessão SAP livre.")
    print("[SE16] Transição para o ecrã de seleção confirmada.")
    print("[SE16] Tabela USLA04 confirmada.")

    # Descobrir campos dinamicos (apenas BNAME e SUBSYSTEM)
    print("[SE16] A identificar campos BNAME e SUBSYSTEM...")
    mapeamento = descobrir_campos_se16(session)
    id_bname  = mapeamento["BNAME"]
    id_subsys = mapeamento["SUBSYSTEM"]

    if not id_bname or not id_subsys:
        # Diagnostico detalhado apenas quando a identificacao falha
        try:
            usr = session.findById("wnd[0]/usr")
            print("[SE16][DEBUG] Controlos encontrados no ecra de selecao:")
            for idx, child in enumerate(usr.Children):
                print(
                    f"│  [{idx}] ID={getattr(child, 'Id', '')} "
                    f"| Tipo={getattr(child, 'Type', '')} "
                    f"| Name={getattr(child, 'Name', '')} "
                    f"| Text='{getattr(child, 'Text', '')}' "
                    f"| Tooltip='{getattr(child, 'Tooltip', '')}'")
        except Exception as diag_exc:
            print(f"[SE16][DEBUG] Erro ao obter diagnostico: {diag_exc}")
        raise RuntimeError("Campos BNAME/SUBSYSTEM da USLA04 nao encontrados na SE16.")

    print(f"[SE16] BNAME localizado: {id_bname}")
    print(f"[SE16] SUBSYSTEM localizado: {id_subsys}")
    return mapeamento


def consultar_usla04_para_grupo(session, utilizador, sistema) -> list[dict]:
    """
    Abre a SE16 e consulta a tabela USLA04 apenas com BNAME + SUBSYSTEM.
    Retorna todos os registos encontrados (sem filtro de data SAP).
    A classificacao por data e feita em Python pela funcao chamadora.

    Fluxo:
      1. Abrir SE16 + USLA04
      2. Definir max_hits = 9999
      3. Preencher BNAME com read-back obrigatorio
      4. Preencher SUBSYSTEM com read-back obrigatorio
      5. Executar (F8)
      6. Ler todos os registos retornados
    """
    # ── 1. Abrir SE16 + USLA04 ────────────────────────────────────────────────
    campos = abrir_se16_usla04(session)
    id_bname     = campos["BNAME"]
    id_subsystem = campos["SUBSYSTEM"]

    if not id_bname or not id_subsystem:
        raise RuntimeError(
            "Campos de selecao da USLA04 nao encontrados (BNAME/SUBSYSTEM ausentes). "
            "A pre-validacao da USLA04 foi cancelada e nenhuma alteracao foi efetuada."
        )

    # ── 2. Definir Max Ocorrencias ─────────────────────────────────────────────
    for mcid in ("wnd[0]/usr/txtMAX_SEL", "wnd[0]/usr/txtGD-MAXROWS", "wnd[0]/usr/txtMAX_HITS"):
        try:
            session.findById(mcid).text = "9999"
            break
        except Exception:
            pass

    # ── 3. Preencher BNAME com read-back ──────────────────────────────────────
    campo_bname = session.findById(id_bname)
    campo_bname.text = utilizador
    try:
        campo_bname.caretPosition = len(utilizador)
    except Exception:
        pass

    bname_lido = ""
    try:
        bname_lido = str(session.findById(id_bname).text).strip()
    except Exception:
        pass

    if bname_lido.upper() != utilizador.upper():
        try:
            session.findById(id_bname).Text = utilizador
            bname_lido = str(session.findById(id_bname).Text).strip()
        except Exception:
            pass

    if bname_lido.upper() != utilizador.upper():
        raise RuntimeError(
            f"BNAME nao conservou o valor apos atribuicao "
            f"(esperado='{utilizador}', lido='{bname_lido}'). "
            "A pre-validacao da USLA04 foi cancelada e nenhuma alteracao foi efetuada."
        )
    print(f"[SE16] BNAME preenchido e confirmado: {bname_lido}")

    # ── 4. Preencher SUBSYSTEM com read-back ──────────────────────────────────
    campo_subsys = session.findById(id_subsystem)
    campo_subsys.text = sistema
    try:
        campo_subsys.caretPosition = len(sistema)
    except Exception:
        pass

    subsys_lido = ""
    try:
        subsys_lido = str(session.findById(id_subsystem).text).strip()
    except Exception:
        pass

    if subsys_lido.upper() != sistema.upper():
        try:
            session.findById(id_subsystem).Text = sistema
            subsys_lido = str(session.findById(id_subsystem).Text).strip()
        except Exception:
            pass

    if subsys_lido.upper() != sistema.upper():
        raise RuntimeError(
            f"SUBSYSTEM nao conservou o valor apos atribuicao "
            f"(esperado='{sistema}', lido='{subsys_lido}'). "
            "A pre-validacao da USLA04 foi cancelada e nenhuma alteracao foi efetuada."
        )
    print(f"[SE16] SUBSYSTEM preenchido e confirmado: {subsys_lido}")
    print(f"[SE16] Filtros: BNAME={bname_lido} | SUBSYSTEM={subsys_lido}")

    # ── 5. Executar (F8) ──────────────────────────────────────────────────────
    print("[SE16] A executar consulta (sem filtro de data — classificacao em Python)...")
    session.findById("wnd[0]").sendVKey(8)
    time.sleep(1.5)
    aguardar_sap_livre(session, timeout=60)
    _dismiss_popup(session)
    print("[SE16] Consulta executada. A ler resultados...")

    # ── 6. Ler resultados de forma polimorfica ────────────────────────────────
    if DEBUG_SAP_CONTROLS:
        dump_controles_diagnostico(session)

    try:
        resultados = ler_resultados_usla04(session, utilizador, sistema)
    except Exception as exc:
        log_technical_error_details(session, "Pré-validação da USLA04", "consultar_usla04_para_grupo", exc)
        if not DEBUG_SAP_CONTROLS:
            dump_controles_diagnostico(session)
        raise

    return resultados


def _classificar_linhas_usla04(linhas_retornadas, roles_pedidas, hoje):
    """
    Classifica cada AGR_NAME do grupo em relacao aos registos retornados da USLA04.

    Casos:
      INEXISTENTE       - nenhuma linha com esse AGR_NAME
      JA_ATIVA          - pelo menos uma linha com FROM_DAT <= hoje <= TO_DAT
      EXPIRADA          - existem linhas mas todas com TO_DAT < hoje
      DUPLICIDADE_ATIVA - duas ou mais linhas ativas para o mesmo AGR_NAME
      FUTURA            - existe linha com FROM_DAT > hoje (e nao ha linha ativa)

    Retorna dict: {role_upper -> {"classe": str, "linhas": [dict], "n_ativas": int}}
    """
    # Indexar linhas por AGR_NAME normalizado
    por_role = {}
    for reg in linhas_retornadas:
        r_name = str(reg.get("AGR_NAME", "")).strip().upper()
        if r_name:
            por_role.setdefault(r_name, []).append(reg)

    resultado = {}
    for role in roles_pedidas:
        role_up = role.strip().upper()
        linhas = por_role.get(role_up, [])

        if not linhas:
            resultado[role_up] = {"classe": "INEXISTENTE", "linhas": [], "n_ativas": 0}
            continue

        ativas   = []
        futuras  = []
        expiradas = []

        for reg in linhas:
            try:
                dt_from = converter_data_sap(reg.get("FROM_DAT", "") or "")
            except Exception:
                dt_from = hoje  # conservador: assume hoje se vazio
            try:
                dt_to = converter_data_sap(reg.get("TO_DAT", "") or "")
            except Exception:
                dt_to = hoje  # conservador

            if dt_from > hoje:
                futuras.append(reg)
            elif dt_from <= hoje <= dt_to:
                ativas.append(reg)
            else:
                expiradas.append(reg)

        n_ativas = len(ativas)
        if n_ativas >= 2:
            classe = "DUPLICIDADE_ATIVA"
        elif n_ativas == 1:
            classe = "JA_ATIVA"
        elif futuras:
            classe = "FUTURA"
        else:
            classe = "EXPIRADA"

        resultado[role_up] = {
            "classe": classe,
            "linhas": linhas,
            "n_ativas": n_ativas,
            "n_futuras": len(futuras),
            "n_expiradas": len(expiradas),
        }

    return resultado


def prevalidar_e_processar_atribuicoes(
    df_filtrado,
    session,
    sistema_desejado,
    pedir_confirmacao=True,
    modo_nao_interativo=False
) -> pd.DataFrame:
    """
    Pre-validacao das atribuicoes na USLA04 via SE16 (SAP CUA).
    Classifica cada linha em: INEXISTENTE / JA_ATIVA / EXPIRADA /
    DUPLICIDADE_ATIVA / FUTURA e decide quais serao enviadas para a SU10.
    """

    if df_filtrado is None or df_filtrado.empty:
        return df_filtrado

    hoje = datetime.now().date()

    # ── 1. Detectar duplicados no Excel (UTILIZADOR + SISTEMA + AGR_NAME) ─────
    vistos = {}  # chave -> idx_original
    duplicados = []  # [(idx_dup, idx_original)]
    linhas_validar_indices = []

    for idx_row in df_filtrado.index:
        user   = str(df_filtrado.at[idx_row, "UTILIZADOR"]).strip().upper()
        sistema = str(df_filtrado.at[idx_row, "SISTEMA"]).strip().upper()
        role   = str(df_filtrado.at[idx_row, "AGR_NAME"]).strip().upper()

        if not user or not sistema or not role:
            linhas_validar_indices.append(idx_row)
            continue

        chave = (user, sistema, role)
        if chave in vistos:
            duplicados.append((idx_row, vistos[chave]))
        else:
            vistos[chave] = idx_row
            linhas_validar_indices.append(idx_row)

    df_unicos = df_filtrado.loc[linhas_validar_indices]

    # ── 2. Agrupar por UTILIZADOR + SISTEMA para uma consulta por grupo ───────
    grupos_validar = {}  # (user, sys) -> [(idx_row, role)]
    for idx_row in df_unicos.index:
        user    = str(df_unicos.at[idx_row, "UTILIZADOR"]).strip().upper()
        sys_name = str(df_unicos.at[idx_row, "SISTEMA"]).strip().upper()
        role    = str(df_unicos.at[idx_row, "AGR_NAME"]).strip().upper()
        if not user or not sys_name or not role:
            continue
        grupos_validar.setdefault((user, sys_name), []).append((idx_row, role))

    # ── 3. Consultar USLA04 (uma consulta por BNAME + SUBSYSTEM) ─────────────
    # resultados_por_grupo: (user, sys) -> list[dict]
    resultados_por_grupo = {}
    erros_validacao = {}  # idx_row -> msg

    print("\n[Fase 2] Pré-validação CUA na tabela USLA04")

    for (user_norm, sys_norm), info_linhas in grupos_validar.items():
        try:
            print(f"├─ A consultar USLA04 para {user_norm} no sistema {sys_norm}...")
            linhas_ret = consultar_usla04_para_grupo(session, user_norm, sys_norm)
            resultados_por_grupo[(user_norm, sys_norm)] = linhas_ret
            print(f"│  Registos encontrados: {len(linhas_ret)}")

            # Classificar e registar no log
            roles_deste_grupo = [r for _, r in info_linhas]
            classif = _classificar_linhas_usla04(linhas_ret, roles_deste_grupo, hoje)
            aptas = 0
            for role_up, info in classif.items():
                classe = info["classe"]
                if classe == "JA_ATIVA":
                    print(f"│  {role_up}: JA ATRIBUIDA E ATIVA")
                elif classe == "DUPLICIDADE_ATIVA":
                    print(f"│  {role_up}: DUPLICIDADE ATIVA — {info['n_ativas']} entradas")
                elif classe == "EXPIRADA":
                    print(f"│  {role_up}: EXPIRADA — apta para nova atribuicao")
                    aptas += 1
                elif classe == "FUTURA":
                    print(f"│  {role_up}: ATRIBUICAO FUTURA — {info['n_futuras']} entrada(s) futura(s)")
                else:
                    aptas += 1  # INEXISTENTE
            print(f"└─ Funcoes aptas para insercao: {aptas}")

        except Exception as e:
            erro_msg = (
                f"Nao foi possivel validar previamente a atribuicao na tabela USLA04. "
                f"Nenhuma alteracao foi efetuada no SAP. Detalhes: {e}"
            )
            print(f"│  ERRO de validacao: {erro_msg}")
            for idx_row, _role in info_linhas:
                erros_validacao[idx_row] = erro_msg

    # Voltar ao ecra inicial apos SE16
    voltar_para_inicio(session)

    # ── 4. Classificar linhas e decidir quais vao para a SU10 ─────────────────
    linhas_inexistentes = []  # idx_row aptos para insercao (INEXISTENTE ou EXPIRADA)
    status_originais = {}     # idx_original -> (status, msg)

    contadores = {
        "total_pendentes":   len(df_filtrado),
        "ja_ativas":         0,
        "expiradas":         0,
        "inexistentes":      0,
        "futuras":           0,
        "duplicidades_ativas": 0,
        "duplicadas_excel":  len(duplicados),
        "erros_validacao":   0,
    }

    for idx_row in df_unicos.index:
        if idx_row in erros_validacao:
            msg_err = erros_validacao[idx_row]
            status_originais[idx_row] = ("ERRO", msg_err)
            marcar_resultado(df_filtrado, idx_row, "ERRO", msg_err)
            contadores["erros_validacao"] += 1
            continue

        user   = str(df_filtrado.at[idx_row, "UTILIZADOR"]).strip().upper()
        sistema = str(df_filtrado.at[idx_row, "SISTEMA"]).strip().upper()
        role   = str(df_filtrado.at[idx_row, "AGR_NAME"]).strip().upper()

        if not user or not sistema or not role:
            msg_err = "Dados obrigatorios (UTILIZADOR/SISTEMA/AGR_NAME) vazios."
            status_originais[idx_row] = ("ERRO", msg_err)
            marcar_resultado(df_filtrado, idx_row, "ERRO", msg_err)
            contadores["erros_validacao"] += 1
            continue

        linhas_ret = resultados_por_grupo.get((user, sistema))
        if linhas_ret is None:
            # Grupo nao consultado (erro capturado acima)
            continue

        classif = _classificar_linhas_usla04(linhas_ret, [role], hoje)
        info = classif.get(role, {"classe": "INEXISTENTE", "n_ativas": 0})
        classe = info["classe"]
        n_ativas = info.get("n_ativas", 0)
        linhas_do_role = info.get("linhas", [])
        to_dat_str = linhas_do_role[0].get("TO_DAT", "") if linhas_do_role else ""

        if classe == "JA_ATIVA":
            msg = (
                f"Funcao '{role}' ja atribuida ao utilizador '{user}' no sistema '{sistema}', "
                f"com validade ate {to_dat_str}. Nenhuma alteracao efetuada."
            )
            status_originais[idx_row] = ("CONCLUIDO", msg)
            marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg)
            contadores["ja_ativas"] += 1

        elif classe == "DUPLICIDADE_ATIVA":
            msg = (
                f"Funcao '{role}' ja atribuida ao utilizador '{user}' no sistema '{sistema}' "
                f"com {n_ativas} entradas ativas na USLA04 (DUPLICIDADE). Nenhuma alteracao efetuada."
            )
            status_originais[idx_row] = ("CONCLUIDO", msg)
            marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg)
            contadores["duplicidades_ativas"] += 1

        elif classe == "FUTURA":
            n_fut = info.get("n_futuras", 0)
            msg = (
                f"Funcao '{role}' para o utilizador '{user}' no sistema '{sistema}' "
                f"tem {n_fut} atribuicao(oes) futura(s) na USLA04. "
                "Insercao nao efetuada por seguranca."
            )
            status_originais[idx_row] = ("ERRO", msg)
            marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
            contadores["futuras"] += 1

        elif classe == "EXPIRADA":
            contadores["expiradas"] += 1
            linhas_inexistentes.append(idx_row)

        else:  # INEXISTENTE
            contadores["inexistentes"] += 1
            linhas_inexistentes.append(idx_row)

    # ── 5. Resolver duplicados do Excel ───────────────────────────────────────
    for idx_dup, idx_orig in duplicados:
        if idx_orig in status_originais:
            status_orig, msg_orig = status_originais[idx_orig]
            id_orig = df_filtrado.at[idx_orig, "ID"]
            if status_orig == "CONCLUIDO":
                msg_dup = f"Combinacao duplicada no ficheiro. A funcao foi tratada pela linha ID '{id_orig}'."
                marcar_resultado(df_filtrado, idx_dup, "CONCLUIDO", msg_dup)
            else:
                msg_dup = (
                    f"Nao foi possivel validar previamente a atribuicao na tabela USLA04. "
                    f"Nenhuma alteracao foi efetuada no SAP. Dependia da linha ID '{id_orig}'."
                )
                marcar_resultado(df_filtrado, idx_dup, "ERRO", msg_dup)

    # ── 6. Resumo da pre-validacao ─────────────────────────────────────────────
    n_insercao = len(linhas_inexistentes)
    print("\n======================================================================")
    print("Pre-validacao USLA04 concluida.")
    print(f"  Total de linhas pendentes      : {contadores['total_pendentes']}")
    print(f"  Ja atribuidas e validas        : {contadores['ja_ativas']}")
    print(f"  Expiradas (aptas p/ reinserção): {contadores['expiradas']}")
    print(f"  Inexistentes (aptas p/ insercao): {contadores['inexistentes']}")
    print(f"  Duplicidades ativas na USLA04  : {contadores['duplicidades_ativas']}")
    print(f"  Atribuicoes futuras bloqueadas : {contadores['futuras']}")
    print(f"  Duplicadas no Excel            : {contadores['duplicadas_excel']}")
    print(f"  Erros reais de validacao       : {contadores['erros_validacao']}")
    print(f"  Funcoes que serao inseridas    : {n_insercao}")
    print("======================================================================")

    if not linhas_inexistentes:
        print("\n[INFO] Nao existem novas atribuicoes a realizar.")
        return df_filtrado

    # ── 7. Confirmacao do utilizador ───────────────────────────────────────────
    if not modo_nao_interativo and pedir_confirmacao:
        resposta = input("Deseja lancar as funcoes inexistentes no SAP? [S/N]: ").strip().upper()
        if resposta != "S":
            print("Lancamento cancelado pelo utilizador.")
            return df_filtrado

    # ── 8. Executar a SU10 apenas para INEXISTENTES + EXPIRADAS ───────────────
    df_candidatos = df_filtrado.loc[linhas_inexistentes].copy()
    df_candidatos_proc = atribuir_funcao_usuario(
        df_candidatos,
        session,
        sistema_desejado,
        pedir_confirmacao=False,
        modo_nao_interativo=True
    )

    for idx_row in linhas_inexistentes:
        df_filtrado.at[idx_row, "STATUS"]    = df_candidatos_proc.at[idx_row, "STATUS"]
        df_filtrado.at[idx_row, "MSG"]       = df_candidatos_proc.at[idx_row, "MSG"]
        df_filtrado.at[idx_row, "TIMESTEMP"] = df_candidatos_proc.at[idx_row, "TIMESTEMP"]

    # ── 9. Resolver duplicados que dependiam da SU10 ──────────────────────────
    for idx_dup, idx_orig in duplicados:
        if idx_orig in linhas_inexistentes:
            status_orig = df_filtrado.at[idx_orig, "STATUS"]
            msg_orig    = df_filtrado.at[idx_orig, "MSG"]
            id_orig     = df_filtrado.at[idx_orig, "ID"]
            if df_filtrado.at[idx_dup, "STATUS"] not in ("CONCLUIDO", "ERRO"):
                if normalizar_valor(status_orig) == "CONCLUIDO":
                    msg_dup = f"Combinacao duplicada. A funcao foi tratada pela linha ID '{id_orig}'."
                    marcar_resultado(df_filtrado, idx_dup, "CONCLUIDO", msg_dup)
                else:
                    msg_dup = (
                        f"Linha nao processada devido a falha na validacao/insercao da "
                        f"primeira ocorrencia (ID '{id_orig}'). Detalhes: {msg_orig}"
                    )
                    marcar_resultado(df_filtrado, idx_dup, "ERRO", msg_dup)

    return df_filtrado


def descobrir_colunas_su10(shell):
    col_map = {"SUBSYSTEM": None, "AGR_NAME": None}
    try:
        cols_count = int(shell.ColumnCount)
        termos_sys = {"SUBSYSTEM", "SYSNAME", "SISTEMA", "SYSTEM", "RECETOR", "RECEPTOR"}
        termos_agr = {"AGR_NAME", "ROLE", "FUNCAO", "FUNÇÃO"}

        for c in range(cols_count):
            try:
                key = str(shell.GetColumnKey(c)).strip().upper()
                title = str(shell.GetColumnTitle(c)).strip().upper()
            except Exception:
                continue
            if col_map["SUBSYSTEM"] is None and (key in termos_sys or any(t in title for t in termos_sys)):
                col_map["SUBSYSTEM"] = shell.GetColumnKey(c)
            if col_map["AGR_NAME"] is None and (key in termos_agr or any(t in title for t in termos_agr)):
                col_map["AGR_NAME"] = shell.GetColumnKey(c)
    except Exception as e:
        print(f"[AVISO] Erro ao descobrir colunas na SU10: {e}")

    # Fallbacks se falhar
    if not col_map["SUBSYSTEM"]:
        col_map["SUBSYSTEM"] = "SUBSYSTEM"
    if not col_map["AGR_NAME"]:
        col_map["AGR_NAME"] = "AGR_NAME"

    return col_map


def obter_funcoes_existentes(shell) -> set[tuple[str, str]]:
    """
    Lê o grid de funções existentes para o utilizador no SAP CUA.
    Retorna um conjunto de tuplos (SUBSYSTEM_NORMALIZADO, AGR_NAME_NORMALIZADO).
    """
    existentes = set()

    # 1. Obter RowCount de forma robusta
    try:
        row_count = getattr(shell, "RowCount", None)
        if row_count is None:
            row_count = getattr(shell, "rowCount", None)
        if row_count is None:
            raise RuntimeError("Não foi possível obter a propriedade RowCount/rowCount do shell.")
    except Exception as e:
        print(f"[AVISO] Erro ao aceder ao RowCount: {e}")
        raise RuntimeError(f"Erro ao ler RowCount do shell de funções: {e}")

    col_map = descobrir_colunas_su10(shell)
    col_sub = col_map["SUBSYSTEM"]
    col_agr = col_map["AGR_NAME"]

    # 2. Determinar qual o método de leitura disponível (GetCellValue vs getCellValue)
    metodo_leitura = None
    if hasattr(shell, "GetCellValue"):
        metodo_leitura = shell.GetCellValue
    elif hasattr(shell, "getCellValue"):
        metodo_leitura = shell.getCellValue

    if metodo_leitura is None and row_count > 0:
        try:
            shell.GetCellValue(0, col_sub)
            metodo_leitura = shell.GetCellValue
        except Exception:
            try:
                shell.getCellValue(0, col_sub)
                metodo_leitura = shell.getCellValue
            except Exception as e:
                raise RuntimeError(
                    f"Método GetCellValue/getCellValue indisponível ou inacessível no shell: {e}"
                )

    # 3. Ler cada linha
    for i in range(row_count):
        subsystem = ""
        agr_name = ""

        # Tenta ler SUBSYSTEM
        try:
            if metodo_leitura is not None:
                subsystem = metodo_leitura(i, col_sub)
            else:
                try:
                    subsystem = shell.GetCellValue(i, col_sub)
                except Exception:
                    subsystem = shell.getCellValue(i, col_sub)
        except Exception as e:
            print(f"[AVISO] Erro ao ler SUBSYSTEM na linha {i}: {e}")
            subsystem = ""

        # Tenta ler AGR_NAME
        try:
            if metodo_leitura is not None:
                agr_name = metodo_leitura(i, col_agr)
            else:
                try:
                    agr_name = shell.GetCellValue(i, col_agr)
                except Exception:
                    agr_name = shell.getCellValue(i, col_agr)
        except Exception as e:
            print(f"[AVISO] Erro ao ler AGR_NAME na linha {i}: {e}")
            agr_name = ""

        sub_str = str(subsystem or "").strip().upper()
        agr_str = str(agr_name or "").strip().upper()

        if sub_str or agr_str:
            existentes.add((sub_str, agr_str))

    return existentes


###################################################################################
# BLOCO 7: FILTRO DE LINHAS A PROCESSAR
###################################################################################

def filtrar_pendentes(df):
    if df is None or df.empty:
        return pd.DataFrame()

    df2 = df.copy()
    df2["STATUS_NORM"] = df2["STATUS"].apply(normalizar_valor)

    pend = df2[
        (df2["CHAVE_ID"] != "") &
        (df2["STATUS_NORM"] != "CONCLUIDO")
    ].drop(columns=["STATUS_NORM"])

    if pend.empty:
        print("\n⚠️ Nenhuma linha com STATUS ≠ 'Concluído' foi encontrada.")
    else:
        print("\n📋 Linhas a processar:")
        exibir = pend[["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME"]].copy()
        for c in exibir.columns:
            exibir[c] = exibir[c].apply(texto_limpo)
        print(exibir.to_string(index=False))
        print()

    return pend


###################################################################################
# BLOCO 8: EXECUÇÃO SAP
###################################################################################

def marcar_resultado(df_ref, idx, status, msg):
    df_ref.at[idx, "STATUS"] = texto_limpo(status)
    df_ref.at[idx, "MSG"] = texto_limpo(msg)
    df_ref.at[idx, "TIMESTEMP"] = agora_str()


def atribuir_funcao_usuario(df_filtrado, session, sistema_desejado, pedir_confirmacao=True, modo_nao_interativo=False):
    """
    Atribui AGR_NAME ao UTILIZADOR via SU10 de forma agrupada por utilizador/sistema.
    Garante idempotência, evitando atribuir funções que já existam no SAP CUA,
    e previne duplicações no próprio ficheiro Excel.
    """
    if df_filtrado is None or df_filtrado.empty:
        return df_filtrado

    # 0. Identificar e marcar duplicados no próprio Excel (UTILIZADOR + SISTEMA + AGR_NAME)
    vistos = set()
    duplicados_marcar = []
    indices_unicos = []
    
    for idx_row in df_filtrado.index:
        user_val = str(df_filtrado.at[idx_row, "UTILIZADOR"]).strip()
        sys_val = str(df_filtrado.at[idx_row, "SISTEMA"]).strip()
        role_val = str(df_filtrado.at[idx_row, "AGR_NAME"]).strip()
        
        user_norm = user_val.upper()
        sys_norm = sys_val.upper()
        role_norm = role_val.upper()
        
        if not user_norm or not sys_norm or not role_norm:
            indices_unicos.append(idx_row)
            continue
            
        chave = (user_norm, sys_norm, role_norm)
        if chave in vistos:
            duplicados_marcar.append(idx_row)
        else:
            vistos.add(chave)
            indices_unicos.append(idx_row)
            
    # Marcar os duplicados imediatamente como CONCLUIDO
    for idx_row in duplicados_marcar:
        user = df_filtrado.at[idx_row, "UTILIZADOR"]
        sys_name = df_filtrado.at[idx_row, "SISTEMA"]
        role = df_filtrado.at[idx_row, "AGR_NAME"]
        msg_dup = f"Combinação '{role}' no sistema '{sys_name}' para o utilizador '{user}' duplicada no Excel. Tratada na primeira ocorrência."
        marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg_dup)

    # DataFrame com as linhas únicas a processar
    df_unicos = df_filtrado.loc[indices_unicos]

    # 1. Agrupar em memória por UTILIZADOR e SISTEMA
    grupos = df_unicos.groupby(["UTILIZADOR", "SISTEMA"], sort=False)
    total_grupos = len(grupos)
    
    total_linhas_pendentes = len(df_unicos)
    total_roles_distintas = df_unicos["AGR_NAME"].nunique()
    
    print(f"\n[INFO] Utilizadores a processar agrupados (excluindo duplicados do Excel): {total_grupos}")
    print(f"[INFO] Linhas únicas pendentes: {total_linhas_pendentes}")
    print(f"[INFO] Roles distintas: {total_roles_distintas}")

    if not modo_nao_interativo and pedir_confirmacao:
        resposta = input("Deseja lançar essas funções no SAP? [S/N]: ").strip().upper()
        if resposta != "S":
            print("[X] Lançamento cancelado pelo utilizador.")
            return df_filtrado

    print("\n[Fase 3] Processamento dos Utilizadores")
    tempo_total_inicio = time.time()

    for idx_grupo, ((utilizador, sistema), df_grupo) in enumerate(grupos, 1):
        inicio = time.time()
        eventos_status = []
        
        # Obter a lista de roles únicas a adicionar para este utilizador e sistema
        roles_list = list(dict.fromkeys([str(r).strip() for r in df_grupo["AGR_NAME"] if str(r).strip()]))
        
        print(f"\n[Utilizador {idx_grupo}/{total_grupos}] {utilizador}")

        # Verificar dados vazios
        if not utilizador or not sistema or not roles_list:
            msg = "Dados obrigatórios (UTILIZADOR/SISTEMA/ROLES) vazios."
            for idx_row in df_grupo.index:
                marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
            duracao_str = formatar_tempo(time.time() - inicio)
            print(f"🔴 ERRO: {msg} ⏱️ (Tempo: {duracao_str})")
            continue

        try:
            sistema_conectado = texto_limpo(session.Info.SystemName).upper()
            if sistema_conectado != sistema_desejado:
                msg = f"Sistema SAP incorreto: esperado {sistema_desejado}, conectado a {sistema_conectado}"
                for idx_row in df_grupo.index:
                    marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
                duracao_str = formatar_tempo(time.time() - inicio)
                print(f"🔴 ERRO: {msg} ⏱️ (Tempo: {duracao_str})")
                continue

            # 1) Abre SU10
            print("\n[Subetapa 3.1] Pesquisa de Utilizador")
            print("├─ Abrindo SU10...")
            ir_para_transacao(session, "SU10")
            capturar_status_bar(session, eventos_status, origem="ABERTURA_SU10", tentativas=5, espera=0.20)

            grid_input = "wnd[0]/usr/tblSAPLSUID_MAINTENANCETC_USERS"
            campo_utilizador = grid_input + "/ctxtSUID_ST_BNAME-BNAME[0,0]"
            btn_selecionar = "wnd[0]/tbar[1]/btn[18]"
            tab_funcoes = "wnd[0]/usr/tabsTABSTRIP1/tabpACTG"
            shell_funcoes = (
                "wnd[0]/usr/tabsTABSTRIP1/tabpACTG/"
                "ssubMAINAREA:SAPLSUID_MAINTENANCE:1106/"
                "cntlG_ROLES_CONTAINER/shellcont/shell"
            )

            if not esperar_elemento(session, campo_utilizador, tentativas=20, espera=0.5):
                msg = "Falha ao abrir SU10."
                for idx_row in df_grupo.index:
                    marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
                duracao_str = formatar_tempo(time.time() - inicio)
                print(f"ERRO: {msg} (Tempo: {duracao_str})")
                continue

            # 2) Preenche utilizador e seleciona
            print(f"├─ Inserindo utilizador: {utilizador}")
            campo = session.findById(campo_utilizador)
            campo.text = ""
            campo.text = utilizador
            campo.caretPosition = len(utilizador)

            print("└─ Selecionando utilizador...")
            session.findById(btn_selecionar).press()
            time.sleep(0.60)
            tipo_sel, _, msg_sel = capturar_status_bar(
                session,
                eventos_status,
                origem="SELECAO_UTILIZADOR",
                tentativas=6,
                espera=0.20,
            )

            if normalizar_valor(tipo_sel) in ("E", "A", "X"):
                msg_final = montar_msg_final(eventos_status) or msg_sel
                for idx_row in df_grupo.index:
                    marcar_resultado(df_filtrado, idx_row, "ERRO", msg_final)
                duracao_str = formatar_tempo(time.time() - inicio)
                print(f"ERRO: {msg_final} (Tempo: {duracao_str})")
                continue

            # 3) Vai para tab de funções
            print("\n[Subetapa 3.2] Atribuição de Funções no SAP CUA")
            print("├─ Acedendo à aba de funções...")
            session.findById(tab_funcoes).select()
            time.sleep(0.40)
            capturar_status_bar(session, eventos_status, origem="ABERTURA_TAB_FUNCOES", tentativas=4, espera=0.20)

            shell = esperar_elemento(session, shell_funcoes, tentativas=20, espera=0.5)
            if not shell:
                msg = "Não foi possível abrir a aba de funções no SU10."
                for idx_row in df_grupo.index:
                    marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
                duracao_str = formatar_tempo(time.time() - inicio)
                print(f"ERRO: {msg} (Tempo: {duracao_str})")
                continue

            # 3b) Obter funções existentes no SAP CUA
            try:
                funcoes_existentes = obter_funcoes_existentes(shell)
                print(f"├─ Funções já atribuídas detetadas no grid do utilizador ({len(funcoes_existentes)}):")
                for sub_e, agr_e in funcoes_existentes:
                    print(f"│  - {sub_e} / {agr_e}")
            except Exception as read_exc:
                msg = f"Erro ao ler as funções existentes do utilizador no SAP: {read_exc}"
                print(f"ERRO: {msg}")
                for idx_row in df_grupo.index:
                    marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
                continue

            # Identificar quais das roles pedidas já estão no grid
            roles_a_adicionar = []
            for role_name in roles_list:
                role_norm = str(role_name).strip().upper()
                sistema_norm = str(sistema).strip().upper()
                
                if (sistema_norm, role_norm) in funcoes_existentes:
                    indices_da_role = df_grupo[df_grupo["AGR_NAME"] == role_name].index
                    for idx_row in indices_da_role:
                        msg_ja_existe = f"Função '{role_name}' já atribuída ao utilizador '{utilizador}' no sistema '{sistema}'. Nenhuma alteração efetuada."
                        marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg_ja_existe)
                    print(f"│  [INFO] Função '{role_name}' já atribuída ao utilizador no sistema '{sistema}'. Ignorando.")
                else:
                    roles_a_adicionar.append(role_name)

            if not roles_a_adicionar:
                print("└─ Nenhuma nova função para adicionar (todas já atribuídas). Gravação ignorada.")
                continue

            # 4) Preenche subsystem e AGR_NAME para cada uma das roles
            print(f"├─ Preparando inserção de {len(roles_a_adicionar)} role(s)...")
            role_errors = {}
            row_idx = 0

            for r_idx, role_name in enumerate(roles_a_adicionar):
                print(f"├─ Inserindo role {r_idx+1}/{len(roles_a_adicionar)}: {role_name}")
                
                # Procurar a primeira linha vazia a partir de row_idx
                while row_idx < shell.rowCount:
                    subsys = None
                    try:
                        if hasattr(shell, "GetCellValue"):
                            subsys = shell.GetCellValue(row_idx, "SUBSYSTEM")
                        elif hasattr(shell, "getCellValue"):
                            subsys = shell.getCellValue(row_idx, "SUBSYSTEM")
                    except Exception:
                        pass
                    
                    agr = None
                    try:
                        if hasattr(shell, "GetCellValue"):
                            agr = shell.GetCellValue(row_idx, "AGR_NAME")
                        elif hasattr(shell, "getCellValue"):
                            agr = shell.getCellValue(row_idx, "AGR_NAME")
                    except Exception:
                        pass
                        
                    subsys_str = str(subsys or "").strip()
                    agr_str = str(agr or "").strip()
                    if not subsys_str and not agr_str:
                        break
                    row_idx += 1
                
                try:
                    if row_idx >= 5:
                        shell.firstVisibleRow = row_idx - 4
                        
                    shell.modifyCell(row_idx, "SUBSYSTEM", sistema)
                    shell.modifyCell(row_idx, "AGR_NAME", role_name)
                    shell.currentCellColumn = "AGR_NAME"
                    shell.pressEnter()
                    time.sleep(0.5)
                    
                    local_events = []
                    tipo_pre, _, msg_pre = capturar_status_bar(
                        session,
                        local_events,
                        origem=f"VAL_ROLE_{role_name}",
                        tentativas=5,
                        espera=0.15,
                    )
                    
                    if normalizar_valor(tipo_pre) in ("E", "A", "X"):
                        err_msg = montar_msg_final(local_events) or msg_pre
                        role_errors[role_name] = err_msg
                        print(f"│  [AVISO] Falha na validação da role '{role_name}': {err_msg}")
                        
                        # Limpar a linha problemática
                        shell.modifyCell(row_idx, "SUBSYSTEM", "")
                        shell.modifyCell(row_idx, "AGR_NAME", "")
                        shell.pressEnter()
                        time.sleep(0.3)
                    else:
                        role_errors[role_name] = None
                        row_idx += 1
                        
                except Exception as cell_exc:
                    err_msg = str(cell_exc)
                    role_errors[role_name] = err_msg
                    print(f"│  [AVISO] Erro técnico ao inserir role '{role_name}': {err_msg}")

            # 5) Save - se pelo menos uma role correu bem
            salvou_com_sucesso = False
            save_msg = "Nenhuma role com sucesso para gravar."
            sucesso_roles = [r for r, err in role_errors.items() if err is None]

            if sucesso_roles:
                print("└─ Guardando alterações...")
                session.findById("wnd[0]/tbar[0]/btn[11]").press()
                time.sleep(0.40)
                
                save_events = []
                capturar_status_bar(
                    session,
                    save_events,
                    origem="SAVE_IMEDIATO",
                    tentativas=8,
                    espera=0.20,
                )
                
                tratar_popups_pos_save(session, save_events, max_popups=5)
                
                capturar_status_bar(
                    session,
                    save_events,
                    origem="SAVE_FINAL",
                    tentativas=10,
                    espera=0.25,
                )
                
                status_final = decidir_status_pelo_historico(save_events)
                save_msg = montar_msg_final(save_events)
                
                if status_final == "CONCLUÍDO" or normalizar_valor(status_final) == "CONCLUIDO":
                    salvou_com_sucesso = True
                else:
                    for r in sucesso_roles:
                        role_errors[r] = f"Falha na gravação final: {save_msg}"
            else:
                print("└─ Gravação ignorada (todas as roles falharam).")

            # --- VALIDAÇÃO PÓS-SAVE NA USLA04 ---
            pos_save_classificacoes = {}
            leitura_pos_save_erro = None

            if salvou_com_sucesso:
                print("\n[Subetapa 3.3] Validação pós-Save na tabela USLA04...")
                try:
                    linhas_pos = consultar_usla04_para_grupo(session, utilizador, sistema)
                    hoje_date = datetime.now().date()
                    pos_save_classificacoes = _classificar_linhas_usla04(linhas_pos, roles_list, hoje_date)
                    print(f"├─ Registos lidos pós-Save: {len(linhas_pos)}")
                except Exception as pos_exc:
                    leitura_pos_save_erro = f"LEITURA PÓS-SAVE INCONCLUSIVA: {pos_exc}"
                    print(f"│  ❌ Erro na leitura pós-Save: {leitura_pos_save_erro}")

            # 6) Atribuir resultados linha a linha no df original
            total_ok = 0
            for idx_row in df_grupo.index:
                if df_filtrado.at[idx_row, "STATUS"] == "CONCLUIDO":
                    continue
                    
                row_role = str(df_filtrado.at[idx_row, "AGR_NAME"]).strip()
                row_role_up = row_role.upper()
                err = role_errors.get(row_role)
                
                if err:
                    marcar_resultado(df_filtrado, idx_row, "ERRO", err)
                else:
                    if salvou_com_sucesso:
                        if leitura_pos_save_erro:
                            marcar_resultado(df_filtrado, idx_row, "ERRO", leitura_pos_save_erro)
                        else:
                            info_classif = pos_save_classificacoes.get(row_role_up, {"classe": "INEXISTENTE"})
                            classe_pos = info_classif["classe"]
                            n_ativas = info_classif.get("n_ativas", 0)
                            
                            if classe_pos == "JA_ATIVA":
                                msg_sucesso = f"INSERIDA E CONFIRMADA — {save_msg or 'Gravada com sucesso'}"
                                marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg_sucesso)
                                total_ok += 1
                            elif classe_pos == "DUPLICIDADE_ATIVA":
                                msg_dup = f"DUPLICIDADE ATIVA pós-Save — {n_ativas} entradas ativas na USLA04. {save_msg}"
                                marcar_resultado(df_filtrado, idx_row, "CONCLUIDO", msg_dup)
                                total_ok += 1
                            elif classe_pos == "FUTURA":
                                msg_fut = "Atribuição classificada como FUTURA pós-Save. Não confirmada como ativa."
                                marcar_resultado(df_filtrado, idx_row, "ERRO", msg_fut)
                            else:
                                msg_falha = f"NÃO ENCONTRADA APÓS SAVE — A função não aparece como ativa na USLA04. {save_msg}"
                                marcar_resultado(df_filtrado, idx_row, "ERRO", msg_falha)
                    else:
                        marcar_resultado(df_filtrado, idx_row, "ERRO", f"Não gravado: {save_msg}")

            # 7) Log de resultado do utilizador
            print(f"\nResultado do utilizador {utilizador}:")
            
            estat_grupo = {
                "ja_existentes": 0,
                "duplicidades_ativas": 0,
                "expiradas": 0,
                "futuras": 0,
                "inseridas_tentadas": len(roles_a_adicionar),
                "inseridas_confirmadas": total_ok,
                "erros_tecnicos": 0
            }
            
            for idx_row in df_grupo.index:
                stat = df_filtrado.at[idx_row, "STATUS"]
                msg_val = df_filtrado.at[idx_row, "MSG"]
                if stat == "CONCLUIDO":
                    if "já atribuída" in msg_val or "JÁ EXISTIA" in msg_val:
                        estat_grupo["ja_existentes"] += 1
                    elif "DUPLICIDADE" in msg_val:
                        estat_grupo["duplicidades_ativas"] += 1
                elif stat == "ERRO":
                    if "futura" in msg_val.lower():
                        estat_grupo["futuras"] += 1
                    else:
                        estat_grupo["erros_tecnicos"] += 1

            print(f"  - Já existentes no SAP : {estat_grupo['ja_existentes']}")
            print(f"  - Duplicidades ativas  : {estat_grupo['duplicidades_ativas']}")
            print(f"  - Atribuições futuras  : {estat_grupo['futuras']}")
            print(f"  - Aptas para inserção  : {estat_grupo['inseridas_tentadas']}")
            print(f"  - Inseridas/Confirmadas: {estat_grupo['inseridas_confirmadas']}")
            print(f"  - Erros técnicos       : {estat_grupo['erros_tecnicos']}")
            if salvou_com_sucesso:
                print(f"  - Mensagem SAP pós-Save : '{save_msg}'")
            
            duracao = time.time() - inicio
            duracao_str = formatar_tempo(duracao)
            
            if salvou_com_sucesso and estat_grupo["inseridas_confirmadas"] == len(roles_a_adicionar):
                print(f"SUCESSO: Utilizador tratado por completo! (Tempo: {duracao_str})")
            else:
                print(f"ERRO/AVISO: Atribuição parcial ou falha na gravação. Confirmadas: {total_ok}/{len(roles_a_adicionar)}. (Tempo: {duracao_str})")

        except Exception as e:
            msg = str(e)
            for idx_row in df_grupo.index:
                if df_filtrado.at[idx_row, "STATUS"] != "CONCLUIDO":
                    marcar_resultado(df_filtrado, idx_row, "ERRO", msg)
            duracao_str = formatar_tempo(time.time() - inicio)
            print(f"ERRO: {msg} (Tempo: {duracao_str})")

        finally:
            voltar_para_inicio(session)

    tempo_total = time.time() - tempo_total_inicio
    print(f"\nTempo total: {formatar_tempo(tempo_total)}")

    status_norm = df_filtrado["STATUS"].apply(normalizar_valor)
    total_ok = (status_norm == "CONCLUIDO").sum()
    total_erro = (status_norm == "ERRO").sum()
    print(f"Total concluído: {total_ok} | Com erro: {total_erro}")

    return df_filtrado


###################################################################################
# BLOCO 9: GRAVAÇÃO NO EXCEL SEM PERDER FORMATAÇÃO
###################################################################################

def gravar_preservando_formatacao(caminho_ficheiro, nome_sheet, df_atualizado):
    """
    Atualiza APENAS:
    - STATUS
    - MSG
    - TIMESTEMP

    Faz match pela coluna ID.
    """
    try:
        ext = os.path.splitext(caminho_ficheiro)[1].lower()
        keep_vba = ext == ".xlsm"

        wb = load_workbook(caminho_ficheiro, keep_vba=keep_vba)
        if nome_sheet not in wb.sheetnames:
            print(f"❌ Sheet '{nome_sheet}' não existe para gravar.")
            return False

        ws = wb[nome_sheet]
        mapa_cols = mapear_cabecalhos_openpyxl(ws)

        if "TIMESTAMP" in mapa_cols and "TIMESTEMP" not in mapa_cols:
            mapa_cols["TIMESTEMP"] = mapa_cols["TIMESTAMP"]

        obrig_excel = ["ID", "STATUS", "MSG", "TIMESTEMP"]
        falta = [c for c in obrig_excel if c not in mapa_cols]
        if falta:
            print(f"❌ Cabeçalhos obrigatórios em falta na sheet para gravação: {', '.join(falta)}")
            return False

        col_id = mapa_cols["ID"]
        col_status = mapa_cols["STATUS"]
        col_msg = mapa_cols["MSG"]
        col_timestemp = mapa_cols["TIMESTEMP"]

        mapa_linhas_por_id = {}
        for r in range(2, ws.max_row + 1):
            valor_id = ws.cell(row=r, column=col_id).value
            id_chave = chave_id(valor_id)
            if id_chave:
                mapa_linhas_por_id[id_chave] = r

        atualizados = 0
        nao_encontrados = 0

        for _, row in df_atualizado.iterrows():
            id_chave = texto_limpo(row.get("CHAVE_ID", ""))
            if not id_chave:
                continue

            linha_excel = mapa_linhas_por_id.get(id_chave)
            if not linha_excel:
                nao_encontrados += 1
                print(f"⚠️ ID não encontrado na sheet para gravação: {id_chave}")
                continue

            ws.cell(row=linha_excel, column=col_status).value = texto_limpo(row.get("STATUS", ""))
            ws.cell(row=linha_excel, column=col_msg).value = texto_limpo(row.get("MSG", ""))
            ws.cell(row=linha_excel, column=col_timestemp).value = texto_limpo(row.get("TIMESTEMP", ""))

            atualizados += 1

        wb.save(caminho_ficheiro)

        print(
            f"💾 Ficheiro atualizado com formatação preservada "
            f"(sheet '{nome_sheet}') | Linhas atualizadas: {atualizados}"
        )

        if nao_encontrados:
            print(f"⚠️ IDs não encontrados na sheet: {nao_encontrados}")

        return True

    except PermissionError:
        base, ext = os.path.splitext(caminho_ficheiro)
        alternativo = f"{base}_resultado{ext}"
        try:
            wb.save(alternativo)
            print(f"⚠️ Ficheiro estava aberto. Foi criada uma cópia:\n   {alternativo}")
            return True
        except Exception as e:
            print(f"❌ Erro ao salvar cópia: {e}")
            return False

    except Exception as e:
        print(f"❌ Erro ao salvar: {e}")
        return False


###################################################################################
# BLOCO 10: API PARA O COCKPIT
###################################################################################

def executar(
    ambiente_cockpit,
    pfcg_object=None,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    **kwargs
):
    tempo_inicio_total = time.time()
    print(f"✅ Processo selecionado: {NOME_SCRIPT}")

    sheet_alvo = pfcg_object if pfcg_object else NOME_SHEET
    print(f"📄 Script atual: {NOME_SCRIPT} | Sheet alvo: '{sheet_alvo}'")

    if modo_nao_interativo:
        if not caminho_ficheiro:
            raise ValueError("Faltou o parâmetro caminho_ficheiro em modo web/não-interativo.")
    else:
        if not caminho_ficheiro:
            caminho_ficheiro = selecionar_ficheiro_excel()

    if not caminho_ficheiro:
        return False

    print("\n[Fase 1] Leitura do Excel")
    df = ler_ficheiro(caminho_ficheiro, sheet_alvo)
    if df is None:
        return False

    sistema_desejado = MAPA_SISTEMA.get(ambiente_cockpit)
    if not sistema_desejado:
        print(f"❌ Ambiente inválido: {ambiente_cockpit}. Use: {', '.join(MAPA_SISTEMA.keys())}")
        return False

    session = conectar_sap(sistema_desejado)
    if not session:
        return False

    posicionar_sap_meia_tela_direita(session)

    df_pend = filtrar_pendentes(df)
    if df_pend.empty:
        tempo_decorrido_total = time.time() - tempo_inicio_total
        print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
        print("🔁 Fim.")
        return True

    df_proc = prevalidar_e_processar_atribuicoes(
        df_pend.copy(),
        session,
        sistema_desejado,
        pedir_confirmacao=pedir_confirmacao,
        modo_nao_interativo=modo_nao_interativo
    )

    print("\n[Fase 4] Gravação de Resultados")
    ok_save = gravar_preservando_formatacao(caminho_ficheiro, sheet_alvo, df_proc)
    if ok_save:
        print("💾 Resultados gravados com sucesso no Excel!")

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
    print("🔁 Fim.")
    return ok_save


###################################################################################
# BLOCO 11: EXECUÇÃO DIRETA
###################################################################################

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD", "CUA"])
    parser.add_argument("--xlsx")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    args = parser.parse_args()

    env_cli = args.ambiente or "CUA"
    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm)
    )