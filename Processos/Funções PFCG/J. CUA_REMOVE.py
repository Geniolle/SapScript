# -*- coding: utf-8 -*-

###################################################################################
# PROCESSO: Remover FunÃ§Ãµes SU01 (sheet = nome do ficheiro .py sem prefixo)
###################################################################################
# Ex.: "I. CUA_REMOVE.py"  â†’  Sheet "CUA_REMOVE"
#
# COLUNAS ESPERADAS NA SHEET:
# ID | UTILIZADOR | SISTEMA | AGR_NAME | STATUS | MSG | TIMESTEMP
#
# MODO TEMPORÃRIO DE VALIDAÃ‡ÃƒO:
# - SÃ³ avanÃ§a quando o utilizador carregar Enter
# - Mostra no terminal cada passo antes de executar
# - Faz varredura real dos campos do popup de filtro
###################################################################################

###################################################################################
# BLOCO 1: IMPORTAÃ‡Ã•ES E CONFIGURAÃ‡ÃƒO
###################################################################################

import os
import time
import pandas as pd
import unicodedata
import win32com.client
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime

###################################################################################
# BLOCO 2: DETETAR NOME DO SCRIPT E MAPA DE SISTEMAS
###################################################################################

try:
    NOME_SCRIPT = os.path.splitext(os.path.basename(__file__))[0]
except NameError:
    NOME_SCRIPT = "I. CUA_REMOVE"  # fallback

NOME_SHEET = NOME_SCRIPT.split(".", 1)[-1].strip() if "." in NOME_SCRIPT else NOME_SCRIPT

MAPA_SISTEMA = {
    "DEV": "S4D",
    "QAD": "S4Q",
    "PRD": "S4P",
    "CUA": "SPA"
}

# TEMPORÃRIO: modo passo a passo
MODO_DEBUG_PASSO_A_PASSO = True

###################################################################################
# BLOCO 3: UTILITÃRIOS
###################################################################################

def normalizar_coluna(valor):
    return unicodedata.normalize("NFKD", str(valor)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

def normalizar_valor(valor):
    texto = texto_limpo(valor)
    if not texto:
        return ""
    return texto.strip().upper()

def linha_corresponde_ao_processamento(
    agr_name_grid,
    subsystem_grid,
    agr_name_excel,
    sistema_excel,
    opcao_processamento,
):
    agr_grid = normalizar_valor(agr_name_grid)
    subsystem = normalizar_valor(subsystem_grid)
    agr_excel = normalizar_valor(agr_name_excel)
    sistema = normalizar_valor(sistema_excel)

    if not subsystem or not sistema:
        return False

    if opcao_processamento == "sistema":
        return subsystem == sistema

    if opcao_processamento == "sistema_user":
        if not agr_grid or not agr_excel:
            return False
        return subsystem == sistema and agr_grid == agr_excel

    return False

def agora_timestamp():
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")

def texto_limpo(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()

def pausar(msg):
    if MODO_DEBUG_PASSO_A_PASSO:
        input(f"\nâ¸ï¸ {msg}\nPrima Enter para continuar...")

def selecionar_ficheiro_excel():
    """
    Abre o popup SEM diretoria predefinida.
    O Windows abrirÃ¡ naturalmente na Ãºltima pasta utilizada.
    """
    try:
        root = tk.Tk()
        root.withdraw()
        root.update_idletasks()
        root.attributes("-topmost", True)

        caminho = filedialog.askopenfilename(
            title="Selecione o ficheiro Excel",
            filetypes=[("Ficheiros Excel", "*.xlsx *.xls"), ("Todos os ficheiros", "*.*")]
        )

        root.destroy()

        if not caminho:
            print("âš ï¸ SeleÃ§Ã£o de ficheiro cancelada pelo utilizador.")
            return None

        print(f"âœ… Ficheiro selecionado: {caminho}")
        return caminho

    except Exception as e:
        print(f"âŒ Erro ao abrir a janela de seleÃ§Ã£o: {e}")
        return None

def obter_status_bar(session):
    try:
        sbar = session.findById("wnd[0]/sbar")
        tipo = str(getattr(sbar, "MessageType", "") or "").strip().upper()
        texto = str(getattr(sbar, "Text", "") or "").strip()
        return tipo, texto
    except Exception:
        return "", ""

def tipo_sbar_para_status(tipo_sbar):
    tipo = (tipo_sbar or "").strip().upper()

    if tipo == "S":
        return "CONCLUÃDO"
    if tipo == "W":
        return "AVISO"
    if tipo in ("E", "A"):
        return "ERRO"
    if tipo == "I":
        return "INFO"
    return "ERRO"

def registar_resultado(df, idx, status, msg):
    df.loc[idx, "STATUS"] = status
    df.loc[idx, "MSG"] = msg
    df.loc[idx, "TIMESTEMP"] = agora_timestamp()

def existe_objeto(session, obj_id):
    try:
        session.findById(obj_id)
        return True
    except Exception:
        return False

def obter_objeto(session, obj_id):
    try:
        return session.findById(obj_id)
    except Exception as e:
        raise Exception(f"Objeto SAP nÃ£o encontrado: {obj_id} | {e}")

def setar_texto_debug(session, obj_id, valor, descricao="campo"):
    print(f"ðŸ“ A preencher {descricao}: {obj_id} = '{valor}'")
    pausar(f"Validar antes de preencher {descricao}")

    try:
        obj = obter_objeto(session, obj_id)
        try:
            obj.setFocus()
        except Exception:
            pass

        try:
            obj.text = ""
        except Exception:
            pass

        obj.text = str(valor)

        try:
            obj.caretPosition = len(str(valor))
        except Exception:
            pass

    except Exception as e:
        raise Exception(f"Falha ao preencher {descricao} ({obj_id}) com valor '{valor}': {e}")

def pressionar_botao_debug(session, obj_id, descricao="botÃ£o"):
    print(f"ðŸ–±ï¸ A pressionar {descricao}: {obj_id}")
    pausar(f"Validar antes de pressionar {descricao}")

    try:
        obj = obter_objeto(session, obj_id)
        obj.press()
    except Exception as e:
        raise Exception(f"Falha ao pressionar {descricao} ({obj_id}): {e}")

def enviar_vkey_debug(session, wnd_id, vkey, descricao="aÃ§Ã£o"):
    print(f"âŒ¨ï¸ A executar {descricao}: {wnd_id}.sendVKey({vkey})")
    pausar(f"Validar antes de executar {descricao}")

    try:
        wnd = obter_objeto(session, wnd_id)
        wnd.sendVKey(vkey)
    except Exception as e:
        raise Exception(f"Falha ao executar {descricao} em {wnd_id} com VKey={vkey}: {e}")

def selecionar_tab_debug(session, obj_id, descricao="tab"):
    print(f"ðŸ“‘ A selecionar {descricao}: {obj_id}")
    pausar(f"Validar antes de selecionar {descricao}")

    try:
        obj = obter_objeto(session, obj_id)
        obj.select()
    except Exception as e:
        raise Exception(f"Falha ao selecionar {descricao} ({obj_id}): {e}")

def obter_children(obj):
    try:
        children = obj.Children
        total = children.Count
        return [children.Item(i) for i in range(total)]
    except Exception:
        try:
            return list(obj.Children)
        except Exception:
            return []

def coletar_componentes_recursivo(obj, lista):
    """
    Percorre recursivamente os componentes SAP a partir de um nÃ³.
    """
    try:
        lista.append(obj)
    except Exception:
        return

    for filho in obter_children(obj):
        coletar_componentes_recursivo(filho, lista)

def listar_campos_popup(session):
    """
    Lista todos os componentes do wnd[1], para debug.
    """
    componentes = []
    try:
        wnd1 = session.findById("wnd[1]")
    except Exception:
        return componentes

    coletar_componentes_recursivo(wnd1, componentes)
    return componentes

def descrever_componente(obj):
    try:
        obj_id = str(getattr(obj, "Id", "") or "")
    except Exception:
        obj_id = ""

    try:
        obj_type = str(getattr(obj, "Type", "") or "")
    except Exception:
        obj_type = ""

    try:
        name = str(getattr(obj, "Name", "") or "")
    except Exception:
        name = ""

    try:
        text = str(getattr(obj, "Text", "") or "")
    except Exception:
        text = ""

    try:
        changeable = getattr(obj, "Changeable")
    except Exception:
        changeable = "?"

    return {
        "id": obj_id,
        "type": obj_type,
        "name": name,
        "text": text,
        "changeable": changeable
    }

def obter_campos_low_popup(session):
    """
    Procura no popup todos os campos candidatos que terminem com -LOW.
    Ordena por prioridade:
      1) changeable=True
      2) ids contendo 'ctxt'
      3) ids contendo 'txt'
    """
    candidatos = []

    for comp in listar_campos_popup(session):
        info = descrever_componente(comp)
        obj_id = info["id"].upper()

        if "-LOW" not in obj_id:
            continue

        candidatos.append(info)

    def score(info):
        changeable = info["changeable"] is True
        obj_id = info["id"].lower()
        prioridade_tipo = 0
        if "/ctxt" in obj_id:
            prioridade_tipo = 2
        elif "/txt" in obj_id:
            prioridade_tipo = 1
        return (1 if changeable else 0, prioridade_tipo, len(obj_id))

    candidatos.sort(key=score, reverse=True)
    return candidatos

def preencher_popup_filtro(
    session,
    valor,
    descricao_filtro="",
    campos_protegidos=None,
    valores_protegidos=None,
):
    """
    Procura todos os campos LOW do popup e tenta escrever no primeiro que
    realmente aceite .text. Protege campos e valores anteriores.
    """
    pausar(f"Validar popup de filtro {descricao_filtro} antes de procurar os campos reais")

    campos_protegidos = campos_protegidos or {}
    valores_protegidos = valores_protegidos or set()

    candidatos = obter_campos_low_popup(session)

    if not candidatos:
        return {
            "success": False,
            "field_id": None,
            "candidates_tested": 0,
            "rejected_candidates": 0,
            "error": "Nenhum campo '*-LOW' foi encontrado no wnd[1]."
        }

    erros = []
    candidatos_rejeitados = 0

    for info in candidatos:
        obj_id = info["id"]

        # Ignorar e preservar campos protegidos por ID ou por valor
        if obj_id in campos_protegidos:
            valor_esperado = campos_protegidos[obj_id]
            try:
                campo_obj = session.findById(obj_id)
                valor_atual = normalizar_valor(campo_obj.Text)
            except Exception as e:
                valor_atual = ""

            print(f"ðŸ”’ Campo preservado por ID: {obj_id} | valor atual='{valor_atual}' | valor esperado='{valor_esperado}'")
            if valor_atual != normalizar_valor(valor_esperado):
                return {
                    "success": False,
                    "field_id": None,
                    "candidates_tested": candidatos_rejeitados,
                    "rejected_candidates": candidatos_rejeitados,
                    "error": f"O filtro protegido foi alterado. Campo='{obj_id}' | esperado='{valor_esperado}' | atual='{valor_atual}'"
                }
            continue

        try:
            obj = session.findById(obj_id)
            val_atual_bruto = str(getattr(obj, "Text", "") or "")
            val_atual_norm = normalizar_valor(val_atual_bruto)
        except Exception:
            val_atual_norm = ""
            val_atual_bruto = ""

        # ProteÃ§Ã£o por valor
        if val_atual_norm in valores_protegidos:
            print(f"ðŸ”’ Campo preservado por valor: {obj_id} | contem valor protegido='{val_atual_bruto}'")
            continue

        try:
            try:
                obj.setFocus()
            except Exception:
                pass

            try:
                obj.text = ""
            except Exception:
                pass

            obj.text = str(valor)

            try:
                escrito = str(getattr(obj, "Text", "") or "")
            except Exception:
                escrito = str(valor)

            # Validar que a escrita funcionou no campo candidato
            if normalizar_valor(escrito) != normalizar_valor(valor):
                raise Exception("Valor escrito nÃ£o corresponde ao esperado no campo.")

            # Validar todos os campos protegidos e o prÃ³prio valor escrito antes de clicar em Confirmar
            for cp_id, cp_val in campos_protegidos.items():
                try:
                    cp_obj = session.findById(cp_id)
                    cp_atual = normalizar_valor(cp_obj.Text)
                except Exception:
                    cp_atual = ""
                if cp_atual != normalizar_valor(cp_val):
                    raise Exception(f"ValidaÃ§Ã£o prÃ©-confirmaÃ§Ã£o falhou: Campo protegido '{cp_id}' foi corrompido.")

            print(f"âœ… Campo aceite: {obj_id} | antes='{val_atual_bruto}' | depois='{escrito}'")
            print(f"âœ… ValidaÃ§Ã£o antes de confirmar: {descricao_filtro}='{valor}'")
            pressionar_botao_debug(session, "wnd[1]/tbar[0]/btn[0]", "confirmar filtro")
            
            return {
                "success": True,
                "field_id": obj_id,
                "previous_value": val_atual_bruto,
                "current_value": escrito,
                "candidates_tested": candidatos_rejeitados + 1,
                "rejected_candidates": candidatos_rejeitados,
            }

        except Exception as e:
            erros.append(f"{obj_id} => {e}")
            candidatos_rejeitados += 1
            print(f"ðŸ”Ž Candidato nÃ£o corresponde ao filtro solicitado. A testar prÃ³ximo campo.")
            print(f"[DEBUG] Campo ignorado: id='{obj_id}' | erro='{e}'")

    return {
        "success": False,
        "field_id": None,
        "candidates_tested": candidatos_rejeitados,
        "rejected_candidates": candidatos_rejeitados,
        "error": f"Nenhum dos campos LOW do popup aceitou escrita. Erros: {' | '.join(erros)}"
    }

def obter_grid_roles(session):
    return session.findById(
        "wnd[0]/usr/tabsTABSTRIP1/tabpACTG/"
        "ssubMAINAREA:SAPLSUID_MAINTENANCE:1106/"
        "cntlG_ROLES_CONTAINER/shellcont/shell"
    )

def obter_row_count_grid(shell):
    for attr in ("RowCount", "rowCount"):
        try:
            return int(getattr(shell, attr))
        except Exception:
            pass
    return 0

def obter_valor_celula_grid(shell, row, coluna):
    for metodo in ("GetCellValue", "getCellValue"):
        try:
            fn = getattr(shell, metodo)
            return str(fn(row, coluna)).strip()
        except Exception:
            pass
    return ""

###################################################################################
# BLOCO 4: LEITURA DO EXCEL
###################################################################################

def ler_ficheiro_excel(caminho_ficheiro, nome_sheet):
    if not caminho_ficheiro or not os.path.exists(caminho_ficheiro):
        print("âŒ Ficheiro nÃ£o encontrado ou caminho invÃ¡lido.")
        return None

    try:
        wb = load_workbook(caminho_ficheiro, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

        if nome_sheet not in sheets:
            print(f"âŒ Sheet '{nome_sheet}' nÃ£o encontrada. DisponÃ­veis: {', '.join(sheets)}")
            return None

        df = pd.read_excel(caminho_ficheiro, sheet_name=nome_sheet, dtype=object)
        df.columns = [normalizar_coluna(c) for c in df.columns]

        df.rename(columns={
            "USER": "UTILIZADOR",
            "USERNAME": "UTILIZADOR",
            "SYSTEM": "SISTEMA",
            "NOME FUNCAO": "AGR_NAME",
            "FUNCAO": "AGR_NAME",
            "ROLE": "AGR_NAME",
            "TIMESTAMP": "TIMESTEMP"
        }, inplace=True)

        obrigatorias = ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME"]
        faltantes = [c for c in obrigatorias if c not in df.columns]
        if faltantes:
            print(f"âŒ Colunas obrigatÃ³rias em falta na sheet '{nome_sheet}': {', '.join(faltantes)}")
            return None

        for col in ["STATUS", "MSG", "TIMESTEMP"]:
            if col not in df.columns:
                df[col] = ""

        df["_LINHA_EXCEL"] = df.index + 2

        for col in ["ID", "UTILIZADOR", "SISTEMA", "AGR_NAME", "STATUS", "MSG", "TIMESTEMP"]:
            df[col] = df[col].fillna("").astype(str)

        df["STATUS"] = df["STATUS"].str.strip().str.upper()
        df["MSG"] = df["MSG"].str.strip()
        df["TIMESTEMP"] = df["TIMESTEMP"].str.strip()

        print(f"ðŸ“„ Sheet carregada: '{nome_sheet}' | Registos: {len(df)}")
        return df

    except Exception as e:
        print(f"âŒ Erro ao ler o ficheiro/sheet: {e}")
        return None

###################################################################################
# BLOCO 5: CONEXÃƒO COM SAP GUI
###################################################################################

def conectar_sap(sistema_desejado):
    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application = sap_gui_auto.GetScriptingEngine

        for conn in application.Children:
            for sess in conn.Children:
                try:
                    if sess.Info.SystemName.upper() == sistema_desejado.upper():
                        print(
                            f"âœ… Conectado ao SAP: {sess.Info.SystemName.upper()} | "
                            f"Utilizador: {sess.Info.User} | Cliente: {sess.Info.Client}"
                        )
                        return sess
                except Exception:
                    continue

        print(f"âŒ SessÃ£o SAP nÃ£o encontrada para o sistema '{sistema_desejado}'.")
        return None

    except Exception as e:
        print(f"âŒ Erro na conexÃ£o SAP GUI: {e}")
        return None

###################################################################################
# BLOCO 6: EXECUÃ‡ÃƒO (REMOVER FUNÃ‡ÃƒO NO CUA)
###################################################################################

def remover_funcao_usuario(df, session, opcao_processamento="sistema_user"):
    opcao_processamento = texto_limpo(opcao_processamento or "sistema_user").lower()
    if opcao_processamento not in {"sistema_user", "sistema"}:
        print(f"âŒ OpÃ§Ã£o de processamento invÃ¡lida: {opcao_processamento}")
        return df

    # Normalizar strings para agrupamento seguro
    df["_UTILIZADOR_NORM"] = df["UTILIZADOR"].apply(lambda x: normalizar_coluna(texto_limpo(x)))
    df["_SISTEMA_NORM"] = df["SISTEMA"].apply(lambda x: normalizar_coluna(texto_limpo(x)))

    if opcao_processamento == "sistema":
        # Agrupar por utilizador e sistema normalizados
        grupos = [g for _, g in df.groupby(["_UTILIZADOR_NORM", "_SISTEMA_NORM"], dropna=False)]
        total_grupos = len(grupos)
        print(f"Linhas Excel recebidas: {len(df)}")
        print(f"CombinaÃ§Ãµes Ãºnicas UTILIZADOR + SISTEMA: {total_grupos}")

        for i, grp in enumerate(grupos, 1):
            row_exemplo = grp.iloc[0]
            utilizador = texto_limpo(row_exemplo.get("UTILIZADOR", ""))
            sistema = texto_limpo(row_exemplo.get("SISTEMA", ""))
            
            print(
                f"\nðŸ”§ Grupo {i}/{total_grupos} | "
                f"UTILIZADOR='{utilizador}' | SISTEMA='{sistema}' | Linhas Excel relacionadas: {len(grp)}"
            )

            inicio = time.time()

            if not utilizador:
                msg = "UTILIZADOR vazio no grupo."
                print(f"âŒ {msg}")
                for idx in grp.index:
                    registar_resultado(df, idx, "ERRO", msg)
                continue

            if not sistema:
                msg = "SISTEMA vazio no grupo."
                print(f"âŒ {msg}")
                for idx in grp.index:
                    registar_resultado(df, idx, "ERRO", msg)
                continue

            try:
                print("âž¡ï¸ Passo 1: Entrar na SU01")
                print("\n[Etapa 2] Acesso ao SAP CUA")
                setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/nSU01", "campo de comando")
                enviar_vkey_debug(session, "wnd[0]", 0, "confirmar entrada na SU01")
                time.sleep(0.3)

                print("âž¡ï¸ Passo 2: Informar utilizador")
                setar_texto_debug(session, "wnd[0]/usr/ctxtSUID_ST_BNAME-BNAME", utilizador, "campo utilizador")
                enviar_vkey_debug(session, "wnd[0]", 0, "confirmar utilizador")
                time.sleep(0.4)

                tipo_sbar, texto_sbar = obter_status_bar(session)
                print(f"ðŸ“£ STATUS BAR apÃ³s abrir utilizador: tipo='{tipo_sbar}' | texto='{texto_sbar}'")

                if tipo_sbar in ("E", "A"):
                    msg = texto_sbar or f"Erro ao abrir o utilizador '{utilizador}'."
                    print(f"âŒ {msg}")
                    for idx in grp.index:
                        registar_resultado(df, idx, tipo_sbar_para_status(tipo_sbar), msg)
                    continue

                print("âž¡ï¸ Passo 3: Entrar em modo alteraÃ§Ã£o")
                pressionar_botao_debug(session, "wnd[0]/tbar[1]/btn[18]", "botÃ£o alterar")
                time.sleep(0.3)

                print("âž¡ï¸ Passo 4: Selecionar tab de funÃ§Ãµes")
                selecionar_tab_debug(session, "wnd[0]/usr/tabsTABSTRIP1/tabpACTG", "tab ACTG")
                time.sleep(0.3)

                print("âž¡ï¸ Passo 5: Obter grid de funÃ§Ãµes")
                pausar("Validar grid de funÃ§Ãµes antes de obter o objeto")
                shell = obter_grid_roles(session)

                print("âž¡ï¸ Passo 6: Filtrar SUBSYSTEM")
                pausar("Validar antes de abrir filtro SUBSYSTEM")
                shell.currentCellColumn = "SUBSYSTEM"
                shell.contextMenu()
                shell.selectContextMenuItem("&FILTER")
                res_filtro = preencher_popup_filtro(session, sistema, descricao_filtro="SUBSYSTEM")
                time.sleep(0.3)

                if not res_filtro.get("success"):
                    msg = f"NÃ£o foi possÃ­vel preencher o filtro SUBSYSTEM='{sistema}'. Erro: {res_filtro.get('error')}"
                    print(f"âŒ {msg}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "ERRO", msg)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o apÃ³s falha no filtro")
                    continue

                row_count = obter_row_count_grid(shell)
                print(f"ðŸ“Š RowCount do grid apÃ³s filtros: {row_count}")

                if row_count <= 0:
                    msg = f"Nenhuma funÃ§Ã£o encontrada para o utilizador {utilizador} no sistema {sistema}."
                    print(f"âš ï¸ {msg}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "AVISO", msg)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o")
                    continue

                # Validar cada linha individualmente
                print("\n[Etapa 3] ValidaÃ§Ã£o e RemoÃ§Ã£o de Perfis")
                indices_para_remover = []
                ignoradas = 0

                # Validar colunas obrigatÃ³rias
                subsystem_legivel = True
                try:
                    _ = obter_valor_celula_grid(shell, 0, "SUBSYSTEM")
                except Exception:
                    subsystem_legivel = False

                if not subsystem_legivel:
                    msg = "NÃ£o foi possÃ­vel validar as colunas necessÃ¡rias do grid. Nenhuma linha foi eliminada."
                    print(f"âŒ {msg}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "ERRO", msg)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o")
                    continue

                for row_index in range(row_count):
                    agr_grid = texto_limpo(obter_valor_celula_grid(shell, row_index, "AGR_NAME"))
                    sub_grid = texto_limpo(obter_valor_celula_grid(shell, row_index, "SUBSYSTEM"))

                    corresponde = linha_corresponde_ao_processamento(
                        agr_name_grid=agr_grid,
                        subsystem_grid=sub_grid,
                        agr_name_excel="",
                        sistema_excel=sistema,
                        opcao_processamento="sistema"
                    )

                    if corresponde:
                        print(
                            f"   Linha {row_index} validada:\n"
                            f"      AGR_NAME='{agr_grid}'\n"
                            f"      SUBSYSTEM='{sub_grid}'\n"
                            f"      CORRESPONDÃŠNCIA=SIM\n"
                            f"      AÃ‡ÃƒO=ELIMINAR"
                        )
                        indices_para_remover.append(row_index)
                    else:
                        motivo = "SUBSYSTEM diferente do sistema indicado no Excel" if normalizar_valor(sub_grid) != normalizar_valor(sistema) else "NÃ£o foi possÃ­vel validar os dados da linha"
                        print(
                            f"   Linha {row_index} preservada:\n"
                            f"      AGR_NAME='{agr_grid}'\n"
                            f"      SUBSYSTEM='{sub_grid}'\n"
                            f"      CORRESPONDÃŠNCIA=NÃƒO\n"
                            f"      MOTIVO={motivo}"
                        )
                        ignoradas += 1

                total_correspondencias = len(indices_para_remover)
                print(
                    f"\n   Entradas no grid: {row_count}"
                    f" | CorrespondÃªncias do sistema: {total_correspondencias}"
                    f" | Ignoradas: {ignoradas}"
                )

                if total_correspondencias == 0:
                    msg = f"Nenhuma correspondÃªncia encontrada | Sistema: {sistema} | Entradas eliminadas: 0"
                    print(f"âš ï¸ {msg}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "AVISO", msg)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o")
                    continue

                # Eliminar em ordem decrescente
                entradas_eliminadas = 0
                falhas_remocao = 0
                erros_remocao = []
                grid_inconsistente = False
                estado_grid_inseguro = False

                for row_index in sorted(indices_para_remover, reverse=True):
                    # RevalidaÃ§Ã£o imediatamente antes de eliminar
                    try:
                        agr_atual = texto_limpo(obter_valor_celula_grid(shell, row_index, "AGR_NAME"))
                        sub_atual = texto_limpo(obter_valor_celula_grid(shell, row_index, "SUBSYSTEM"))
                    except Exception:
                        print(f"âŒ Grid deixou de responder ao tentar revalidar a linha {row_index} antes da remoÃ§Ã£o.")
                        grid_inconsistente = True
                        estado_grid_inseguro = True
                        break

                    if not linha_corresponde_ao_processamento(
                        agr_name_grid=agr_atual,
                        subsystem_grid=sub_atual,
                        agr_name_excel="",
                        sistema_excel=sistema,
                        opcao_processamento="sistema"
                    ):
                        print(
                            f"   Linha {row_index} deixou de corresponder apÃ³s atualizaÃ§Ã£o do grid. "
                            f"RemoÃ§Ã£o cancelada para esta linha."
                        )
                        continue

                    try:
                        shell.setCurrentCell(row_index, "AGR_NAME")
                        shell.selectedRows = str(row_index)
                        shell.pressToolbarButton("DEL_LINE")
                        entradas_eliminadas += 1
                        time.sleep(0.2)

                        # RevalidaÃ§Ã£o da consistÃªncia apÃ³s DEL_LINE
                        try:
                            _ = obter_row_count_grid(shell)
                        except Exception as grid_err:
                            print(f"âŒ Grid deixou de responder apÃ³s DEL_LINE na linha {row_index}: {grid_err}")
                            grid_inconsistente = True
                            estado_grid_inseguro = True
                            break
                    except Exception as del_err:
                        tipo_sbar_del, texto_sbar_del = obter_status_bar(session)
                        detalhe_del = f" | SAP: {texto_sbar_del}" if texto_sbar_del else ""
                        err_msg = f"Linha {row_index}: {del_err}{detalhe_del}"
                        print(f"âŒ Falha ao eliminar {err_msg}")
                        erros_remocao.append(err_msg)
                        falhas_remocao += 1

                        try:
                            _ = obter_row_count_grid(shell)
                        except Exception:
                            print("âŒ Grid inacessÃ­vel apÃ³s falha. Interrompendo remoÃ§Ãµes.")
                            grid_inconsistente = True
                            estado_grid_inseguro = True
                            break

                duracao = time.time() - inicio

                def _resumo_contadores():
                    return (
                        f"Entradas no grid: {row_count}"
                        f" | CorrespondÃªncias do sistema: {total_correspondencias}"
                        f" | Entradas eliminadas: {entradas_eliminadas}"
                        f" | Falhas: {falhas_remocao}"
                        f" | Entradas ignoradas: {ignoradas}"
                        f" | Tempo: {duracao:.1f}s"
                    )

                if estado_grid_inseguro:
                    msg_final = f"Processamento interrompido por seguranÃ§a | O estado do grid deixou de ser confiÃ¡vel | Entradas eliminadas antes da interrupÃ§Ã£o: {entradas_eliminadas} | Nenhuma eliminaÃ§Ã£o adicional foi executada"
                    print(f"âŒ {msg_final}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "ERRO", msg_final)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o apÃ³s estado inseguro")
                    continue

                if entradas_eliminadas == 0:
                    msg_final = f"Falha na remoÃ§Ã£o | {_resumo_contadores()}"
                    print(f"âŒ {msg_final}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "ERRO", msg_final)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o")
                    continue

                print("âž¡ï¸ Passo 9: Gravar")
                enviar_vkey_debug(session, "wnd[0]", 11, "gravar alteraÃ§Ã£o")
                time.sleep(0.5)

                tipo_sbar, texto_sbar = obter_status_bar(session)
                print(f"ðŸ“£ STATUS BAR final: tipo='{tipo_sbar}' | texto='{texto_sbar}'")

                if tipo_sbar in ("E", "A"):
                    msg_sap = texto_sbar or "Falha ao gravar alteraÃ§Ãµes"
                    msg_final = f"{msg_sap} | Estado SAP nÃ£o confirmado | {_resumo_contadores()}"
                    print(f"âŒ {msg_final}")
                    for idx in grp.index:
                        registar_resultado(df, idx, "ERRO", msg_final)
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "sair da transaÃ§Ã£o apÃ³s falha de gravaÃ§Ã£o")
                    continue

                msg_sap = texto_sbar or f"User {utilizador} has changed"

                if entradas_eliminadas == total_correspondencias and falhas_remocao == 0:
                    status_final = "CONCLUÃDO"
                elif entradas_eliminadas > 0 and falhas_remocao > 0:
                    status_final = "AVISO"
                else:
                    status_final = tipo_sbar_para_status(tipo_sbar if tipo_sbar else "S")

                msg_final = (
                    f"{msg_sap} | Modo: Sistema | "
                    f"Sistema: {sistema} | "
                    f"Entradas no grid: {row_count} | "
                    f"CorrespondÃªncias do sistema: {total_correspondencias} | "
                    f"Entradas eliminadas: {entradas_eliminadas} | "
                    f"Entradas preservadas: {ignoradas} | "
                    f"Falhas: {falhas_remocao}"
                )
                icone = "âœ…" if status_final == "CONCLUÃDO" else "âš ï¸" if status_final == "AVISO" else "âŒ"
                print(f"{icone} {msg_final}")
                
                for idx in grp.index:
                    registar_resultado(df, idx, status_final, msg_final)

                print("âž¡ï¸ Passo 10: Sair da transaÃ§Ã£o")
                setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                enviar_vkey_debug(session, "wnd[0]", 0, "confirmar saÃ­da da transaÃ§Ã£o")

            except Exception as e:
                tipo_sbar, texto_sbar = obter_status_bar(session)
                detalhe_sap = f" | SAP: {texto_sbar}" if texto_sbar else ""
                err_txt = f"{str(e).strip()}{detalhe_sap}"
                print(f"âŒ Erro ao remover funÃ§Ãµes do sistema '{sistema}' para o utilizador '{utilizador}': {err_txt}")
                for idx in grp.index:
                    registar_resultado(df, idx, "ERRO", err_txt)

                try:
                    print("â†©ï¸ Tentativa de sair da transaÃ§Ã£o apÃ³s erro")
                    setar_texto_debug(session, "wnd[0]/tbar[0]/okcd", "/N", "campo de comando")
                    enviar_vkey_debug(session, "wnd[0]", 0, "confirmar saÃ­da apÃ³s erro")
                except Exception:
                    pass

    else:
        print("Modo de processamento: Sistema / User")
        return processar_grupos_sistema_user(df, session)
###################################################################################
# BLOCO 7: GUARDAR RESULTADOS PRESERVANDO FORMATAÃ‡ÃƒO
###################################################################################

def mapear_cabecalhos(ws):
    mapa = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor is None:
            continue
        mapa[normalizar_coluna(valor)] = col
    return mapa

def garantir_coluna_sheet(ws, mapa_cols, nome_coluna):
    chave = normalizar_coluna(nome_coluna)
    if chave in mapa_cols:
        return mapa_cols[chave]

    nova_col = ws.max_column + 1
    ws.cell(row=1, column=nova_col).value = nome_coluna
    mapa_cols[chave] = nova_col
    return nova_col

def salvar_resultado(df, caminho_ficheiro, nome_sheet):
    try:
        wb = load_workbook(caminho_ficheiro)

        if nome_sheet not in wb.sheetnames:
            print(f"âŒ Sheet '{nome_sheet}' nÃ£o existe para gravar.")
            return

        ws = wb[nome_sheet]
        mapa_cols = mapear_cabecalhos(ws)

        col_status = garantir_coluna_sheet(ws, mapa_cols, "STATUS")
        col_msg = garantir_coluna_sheet(ws, mapa_cols, "MSG")
        col_timestamp = garantir_coluna_sheet(ws, mapa_cols, "TIMESTEMP")

        total_atualizadas = 0

        for _, row in df.iterrows():
            linha_excel = row.get("_LINHA_EXCEL")
            if pd.isna(linha_excel):
                continue

            linha_excel = int(linha_excel)
            ws.cell(row=linha_excel, column=col_status).value = texto_limpo(row.get("STATUS", ""))
            ws.cell(row=linha_excel, column=col_msg).value = texto_limpo(row.get("MSG", ""))
            ws.cell(row=linha_excel, column=col_timestamp).value = texto_limpo(row.get("TIMESTEMP", ""))
            total_atualizadas += 1

        wb.save(caminho_ficheiro)
        print(
            f"ðŸ’¾ Resultados atualizados na sheet '{nome_sheet}' "
            f"(apenas STATUS / MSG / TIMESTEMP). Linhas atualizadas: {total_atualizadas}"
        )

    except PermissionError:
        base, ext = os.path.splitext(caminho_ficheiro)
        alternativo = f"{base}_resultado{ext}"
        wb.save(alternativo)
        print(f"âš ï¸ Ficheiro estava aberto. Foi criada uma cÃ³pia:\n   {alternativo}")

    except Exception as e:
        print(f"âŒ Erro ao salvar preservando formataÃ§Ã£o: {e}")

###################################################################################
# BLOCO 8: EXECUTAR PROCESSO
###################################################################################

def executar(
    ambiente,
    caminho_ficheiro=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    opcao_processamento="sistema_user",
):
    global MODO_DEBUG_PASSO_A_PASSO
    MODO_DEBUG_PASSO_A_PASSO = bool(pedir_confirmacao and not modo_nao_interativo)

    print(f"ðŸ“„ Script atual: {NOME_SCRIPT} | Sheet alvo: '{NOME_SHEET}'")

    # Resolver o caminho do ficheiro
    caminho = texto_limpo(caminho_ficheiro)

    if caminho:
        if not os.path.exists(caminho):
            print(f"âŒ Ficheiro recebido nÃ£o encontrado no disco: {caminho}")
            return False
        print(f"âœ… Ficheiro recebido para processamento: {caminho}")
    else:
        if modo_nao_interativo:
            print("âŒ Caminho do ficheiro nÃ£o recebido pela pÃ¡gina web.")
            return False
        caminho = selecionar_ficheiro_excel()
        if not caminho:
            return False

    print("\n[Etapa 1] Leitura do Excel")
    df = ler_ficheiro_excel(caminho, NOME_SHEET)
    if df is None:
        return False

    sistema_desejado = MAPA_SISTEMA.get(str(ambiente).strip().upper())
    if not sistema_desejado:
        print(f"âŒ Ambiente invÃ¡lido: {ambiente}. Use: {', '.join(MAPA_SISTEMA.keys())}")
        return False

    session = conectar_sap(sistema_desejado)
    if not session:
        return False

    opcao_proc_normalizado = texto_limpo(opcao_processamento or "sistema_user").lower()
    print(f"Modo de processamento: {'Sistema / User' if opcao_proc_normalizado == 'sistema_user' else 'Sistema'}")

    df_final = remover_funcao_usuario(df, session, opcao_processamento=opcao_proc_normalizado)
    
    # Gerar resumo final do job
    print("\n==================================================")
    print("Resumo CUA_REMOVE")
    if opcao_proc_normalizado == "sistema":
        # Agrupamento simulado para contagem
        df_temp = df_final.copy()
        df_temp["_U"] = df_temp["UTILIZADOR"].apply(lambda x: normalizar_coluna(texto_limpo(x)))
        df_temp["_S"] = df_temp["SISTEMA"].apply(lambda x: normalizar_coluna(texto_limpo(x)))
        total_grupos = len(df_temp.groupby(["_U", "_S"]))
        
        concluidos = 0
        avisos = 0
        erros = 0
        tot_eliminadas = 0

        # Para cada grupo, ler o status e somar eliminadas
        for _, grp in df_temp.groupby(["_U", "_S"]):
            st = grp.iloc[0].get("STATUS", "")
            if st == "CONCLUÃDO":
                concluidos += 1
            elif st == "AVISO":
                avisos += 1
            else:
                erros += 1
            
            # Tentar extrair do MSG "Entradas eliminadas: N"
            msg = grp.iloc[0].get("MSG", "")
            import re
            m = re.search(r"Entradas eliminadas:\s*(\d+)", msg)
            if m:
                tot_eliminadas += int(m.group(1))

        print("Modo: Sistema")
        print(f"Linhas do Excel: {len(df_final)}")
        print(f"Grupos UTILIZADOR + SISTEMA: {total_grupos}")
        print(f"Grupos concluÃ­dos: {concluidos}")
        print(f"Grupos com aviso: {avisos}")
        print(f"Grupos com erro: {erros}")
        print(f"FunÃ§Ãµes eliminadas: {tot_eliminadas}")
    else:
        concluidas = len(df_final[df_final["STATUS"] == "CONCLUÃDO"])
        avisos = len(df_final[df_final["STATUS"] == "AVISO"])
        erros = len(df_final[df_final["STATUS"] == "ERRO"])
        
        tot_eliminadas = 0
        for _, row in df_final.iterrows():
            import re
            m = re.search(r"Entradas eliminadas:\s*(\d+)", row.get("MSG", ""))
            if m:
                tot_eliminadas += int(m.group(1))

        print("Modo: Sistema / User")
        print(f"Linhas processadas: {len(df_final)}")
        print(f"ConcluÃ­das: {concluidas}")
        print(f"Avisos: {avisos}")
        print(f"Erros: {erros}")
        print(f"FunÃ§Ãµes eliminadas: {tot_eliminadas}")
    print("==================================================")

    print("\n[Etapa 4] GravaÃ§Ã£o de Resultados")
    salvar_resultado(df_final, caminho, NOME_SHEET)

    # Determinar status final de forma estruturada a partir dos contadores
    if opcao_proc_normalizado == "sistema":
        if erros > 0:
            status_final = "ERRO"
            sucesso_final = False
        elif avisos > 0:
            status_final = "AVISO"
            sucesso_final = True
        else:
            status_final = "CONCLUIDO"
            sucesso_final = True
    else:
        if erros > 0:
            status_final = "ERRO"
            sucesso_final = False
        elif avisos > 0:
            status_final = "AVISO"
            sucesso_final = True
        else:
            status_final = "CONCLUIDO"
            sucesso_final = True

    return {
        "success": sucesso_final,
        "status": status_final,
        "message": f"RemoÃ§Ã£o de funÃ§Ãµes concluÃ­da com estado: {status_final}",
        "summary": {
            "grupos_concluidos": concluidos if opcao_proc_normalizado == "sistema" else concluidas,
            "grupos_aviso": avisos,
            "grupos_erro": erros,
            "funcoes_eliminadas": tot_eliminadas,
        }
    }

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD", "CUA"])
    parser.add_argument("--xlsx")
    args = parser.parse_args()

    env_cli = args.ambiente or "CUA"
    executar(env_cli, caminho_ficheiro=args.xlsx)
