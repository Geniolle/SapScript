# -*- coding: utf-8 -*-

###################################################################################
# A. PFCG_CREATE_RFC.py
# PFCG - Criar/Modificar Roles + Atribuir TCODEs + Perfil + Transporte via RFC
#
# Regras:
#  - Logger visual estruturado por Etapas
#  - Criação rápida 100% via RFC
#  - Gravação segura no Excel sem corrupção
#  - Suporte a Transport Request
####################################################################################
import sys
if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

import functools
print = functools.partial(print, flush=True)


def executar(
    ambiente_cockpit,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    nome_pasta=None
):
    import os
    import time
    import re
    import unicodedata
    import tkinter as tk
    from tkinter import filedialog
    from datetime import datetime
    from openpyxl import load_workbook
    from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn
    from pyrfc import Connection
    from dotenv import load_dotenv

    tempo_inicio_total = time.time()

    # --- CORREÇÃO DA ESTRUTURA DE PASTAS ---
    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual)
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)
    # ---------------------------------------

    # Load .env
    env_path = os.path.join(os.getcwd(), ".env")
    load_dotenv(env_path)

    # Sheet base do processo; o sufixo _RFC do ficheiro não altera o nome da aba.
    NOME_SHEET = "PFCG_CREATE"
    SEARCH_HEADER_IN_FIRST_ROWS = 20

    COLUNAS_OBRIGATORIAS = {"AGR_NAME", "TEXT", "TCODE", "STATUS", "MSG", "TIMESTEMP"}

    MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    CLIENTES_POR_AMBIENTE = {"DEV": "100", "QAD": "100", "PRD": "100"}
    cliente_esperado = CLIENTES_POR_AMBIENTE.get(ambiente_cockpit, "100")

    def now_ts():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def norm_col(s):
        if s is None:
            s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def norm_txt(s):
        if s is None:
            s = ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def formatar_tempo(segundos):
        h, resto = divmod(segundos, 3600)
        m, s = divmod(resto, 60)
        if h > 0:
            return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
        return f"{int(m):02d}m {int(s):02d}s"

    def selecionar_ficheiro():
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
            filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*"))
        )
        root.destroy()
        return path

    def split_tcodes(raw):
        if not raw:
            return []
        s = str(raw).replace("\r", "\n").replace("\t", " ").strip().upper()
        parts = re.split(r"[;, \n]+", s)
        out = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if p.startswith("/N") or p.startswith("/O"):
                p = p[2:].strip()
            if p:
                out.append(p)
        return list(dict.fromkeys(out))

    def gravar_resultados_excel(caminho_ficheiro, sheet_name, header_map, records, resultados):
        col_st, col_ms, col_tm = header_map.get("STATUS"), header_map.get("MSG"), header_map.get("TIMESTEMP")
        # 1. Tentar via Excel COM primeiro (mais seguro no Windows para não corromper abas complexas)
        pythoncom = None
        excel_app = None
        wb_excel = None
        try:
            import win32com.client
            import pythoncom as _pythoncom

            pythoncom = _pythoncom
            pythoncom.CoInitialize()
            
            try:
                excel_app = win32com.client.GetActiveObject("Excel.Application")
            except Exception:
                excel_app = win32com.client.Dispatch("Excel.Application")
            
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            abs_path = os.path.abspath(caminho_ficheiro)
            wb_excel = excel_app.Workbooks.Open(abs_path)
            ws_excel = wb_excel.Worksheets(sheet_name)
            
            for rec in records:
                chave_busca = str(rec["AGR_NAME"]).strip()
                res = resultados.get(chave_busca)
                if res:
                    if col_st:
                        ws_excel.Cells(rec["_row"], col_st).Value = res["STATUS"]
                    if col_ms:
                        ws_excel.Cells(rec["_row"], col_ms).Value = res["MSG"]
                    if col_tm:
                        ws_excel.Cells(rec["_row"], col_tm).Value = res["TIMESTEMP"]
            
            wb_excel.Save()
            return True
        except Exception as e_com:
            print(f"  [DEBUG] Falha ao gravar via Excel COM ({e_com}). Usando openpyxl como fallback...")
        finally:
            try:
                if wb_excel is not None:
                    wb_excel.Close(SaveChanges=True)
            except:
                pass
            try:
                if excel_app is not None:
                    excel_app.Quit()
            except:
                pass
            try:
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
            except:
                pass
            
        # 2. Fallback usando openpyxl
        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(caminho_ficheiro)
            ws = wb[sheet_name]
            for rec in records:
                chave_busca = str(rec["AGR_NAME"]).strip()
                res = resultados.get(chave_busca)
                if res:
                    if col_st:
                        ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
                    if col_ms:
                        ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
                    if col_tm:
                        ws.cell(row=rec["_row"], column=col_tm).value = res["TIMESTEMP"]
            wb.save(caminho_ficheiro)
            return True
        except Exception as e_openpyxl:
            print(f"  ❌ Falha crítica ao gravar Excel com openpyxl: {e_openpyxl}")
            return False
        finally:
            try:
                if wb is not None:
                    wb.close()
            except:
                pass

    ###################################################################################
    # BLOCO 1: LER EXCEL
    ###################################################################################
    if not caminho_ficheiro:
        if modo_nao_interativo:
            raise ValueError("Faltou o parâmetro --xlsx em modo não-interativo.")
        print("📂 Selecione o ficheiro Excel…")
        caminho_ficheiro = selecionar_ficheiro()
        if not caminho_ficheiro:
            print("❌ Operação cancelada.")
            return

    if not os.path.exists(caminho_ficheiro):
        print(f"❌ Ficheiro não encontrado: {caminho_ficheiro}")
        return

    try:
        wb = load_workbook(caminho_ficheiro)
    except Exception as e:
        print(f"❌ Não consegui abrir o Excel: {e}")
        return

    if NOME_SHEET in wb.sheetnames:
        ws = wb[NOME_SHEET]
    else:
        if len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
        else:
            print(f"❌ Sheet '{NOME_SHEET}' não encontrada.")
            wb.close()
            return

    header_row = None
    header_map = {}
    
    COLUNAS_ENTRADA = {"AGR_NAME", "TEXT", "TCODE"}
    COLUNAS_SAIDA = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [norm_col(c.value) for c in ws[r]]
        colunas_entrada_encontradas = set(row_vals).intersection(COLUNAS_ENTRADA)

        if len(colunas_entrada_encontradas) == len(COLUNAS_ENTRADA):
            header_row = r
            for idx, name in enumerate(row_vals, start=1):
                if name:
                    header_map[name] = idx
            
            modificou = False
            last_col = 0
            for idx, val in enumerate(row_vals, start=1):
                if val:
                    last_col = idx

            for col_key, col_name in COLUNAS_SAIDA.items():
                if col_key not in header_map:
                    last_col += 1
                    ws.cell(row=r, column=last_col, value=col_name)
                    header_map[col_key] = last_col
                    print(f"➕ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                    modificou = True
            
            if modificou:
                try:
                    wb.save(caminho_ficheiro)
                    print("💾 Ficheiro Excel guardado com as novas colunas.")
                except Exception as e:
                    wb.close()
                    print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas: {e}")
                    return
            break

    if not header_row:
        wb.close()
        print("\n❌ Não encontrei a linha de cabeçalho completa.")
        return

    col_agr = header_map.get("AGR_NAME")
    col_text = header_map.get("TEXT")
    col_tcode = header_map.get("TCODE")
    col_status = header_map.get("STATUS")
    col_msg = header_map.get("MSG")
    col_ts = header_map.get("TIMESTEMP")

    records = []

    for r in range(header_row + 1, ws.max_row + 1):
        agr_val = ws.cell(row=r, column=col_agr).value if col_agr else None
        agr = "" if agr_val is None else str(agr_val).strip()
        if not agr:
            continue

        if " " in agr:
            orig = agr
            agr = agr.replace(" ", "")
            ws.cell(row=r, column=col_agr, value=agr)
            print(f"⚠️ Espaços detetados e removidos da Role: '{orig}' -> '{agr}' (atualizado no Excel)")
            try:
                wb.save(caminho_ficheiro)
            except Exception:
                pass

        text_val = ws.cell(row=r, column=col_text).value if col_text else None
        tcode_val = ws.cell(row=r, column=col_tcode).value if col_tcode else None
        status_val = ws.cell(row=r, column=col_status).value if col_status else None
        msg_val = ws.cell(row=r, column=col_msg).value if col_msg else None
        ts_val = ws.cell(row=r, column=col_ts).value if col_ts else None

        records.append({
            "_row": r,
            "AGR_NAME": agr,
            "TEXT": "" if text_val is None else str(text_val).strip(),
            "TCODE": "" if tcode_val is None else str(tcode_val).strip(),
            "STATUS": "" if status_val is None else str(status_val).strip(),
            "MSG": "" if msg_val is None else str(msg_val).strip(),
            "TIMESTEMP": "" if ts_val is None else str(ts_val).strip(),
        })

    if not records:
        wb.close()
        print("⚠️ Não encontrei linhas para processar.")
        return

    roles_map = {}

    for rec in records:
        status_norm = norm_txt(rec["STATUS"])
        if status_norm == "CONCLUIDO":
            continue

        agr = rec["AGR_NAME"].strip()
        if not agr:
            continue

        if agr not in roles_map:
            roles_map[agr] = {
                "AGR_NAME": agr,
                "TEXT": rec["TEXT"].strip(),
                "TCODE_LIST": []
            }

        if not roles_map[agr]["TEXT"] and rec["TEXT"].strip():
            roles_map[agr]["TEXT"] = rec["TEXT"].strip()

        roles_map[agr]["TCODE_LIST"].extend(split_tcodes(rec["TCODE"]))

    if not roles_map:
        wb.close()
        print("⚠️ Nada para processar (tudo CONCLUIDO).")
        return

    roles_agrupadas = []
    for item in roles_map.values():
        item["TCODE_LIST"] = list(dict.fromkeys(item["TCODE_LIST"]))
        roles_agrupadas.append(item)

    roles_agrupadas.sort(key=lambda x: x["AGR_NAME"])

    if not request_transporte:
        env_req = os.getenv("SAP_REQUEST_NUMBER", "").strip().upper()
        if env_req:
            request_transporte = env_req
            print(f"📦 Request de transporte carregada do ambiente: {request_transporte}")

    if not request_transporte and not modo_nao_interativo:
        print("\n============================================================")
        print("🚚 Opções de configuração de Transporte.\n")
        print("   1 - Escreva o número da Request")
        print("   2 - Criar nova ordem de transporte via RFC")
        print("   3 - Pesquisar suas requests criadas via RFC")
        print("   4 - Prima [Enter] vazio para NÃO transportar")
        print("============================================================")

        while True:
            req_input = input("\n👉 Opção: ").strip()
            if req_input in ("1", "2", "3", "4", ""):
                if req_input == "":
                    req_input = "4"
                break
            print("❌ Opção inválida. Use 1, 2, 3, 4 ou apenas pressione Enter.")

        if req_input == "1":
            request_transporte = input("🔢 Numero da Request (ex: S4QK900396): ").strip().upper()

        elif req_input == "2":
            try:
                from criar_request_rfc import criar_nova_request_rfc
                req_rfc, task_rfc = criar_nova_request_rfc(
                    ambiente=ambiente_cockpit or "DEV",
                    tipo="customizing",
                    descricao="ROLES SIMPLES PFCG RFC"
                )
                request_transporte = req_rfc
            except Exception as e_req:
                print(f"❌ Erro ao criar request via RFC: {e_req}")

        elif req_input == "3":
            try:
                import pesquisar_request_rfc
                print("\n🔍 A pesquisar requests abertas via RFC...")
                resultados_pesquisa = pesquisar_request_rfc.listar_requests(
                    system_name=SISTEMA_ESPERADO,
                    include_requests=True
                )
                if resultados_pesquisa:
                    escolha = input("\n👉 Digite o número (N) da Request que deseja utilizar (ou Enter para cancelar): ").strip()
                    if escolha.isdigit() and 1 <= int(escolha) <= len(resultados_pesquisa):
                        request_transporte = resultados_pesquisa[int(escolha) - 1][0]
                        print(f"✔️ Selecionou a Request: {request_transporte}")
                    else:
                        print("❌ Seleção cancelada. Não haverá transporte.")
                else:
                    print("⚠️ Não foram encontradas Requests abertas para o seu utilizador.")
            except Exception as e:
                print(f"❌ Erro ao pesquisar request: {e}")

        elif req_input == "4":
            print("⏭️  Nenhuma request selecionada (Transporte ignorado).")
            request_transporte = None
        print("============================================================")

    if request_transporte:
        print(f"📦 Request de Transporte ativa para gravação: {request_transporte}")
    else:
        print("ℹ️ Nenhuma Request de Transporte selecionada (alterações gravadas apenas localmente no SAP).")

    ambiente_up = str(ambiente_cockpit).upper().strip()
    ashost = (
        os.getenv(f"SAP_ASHOST_{ambiente_up}")
        or os.getenv("SAP_ASHOST")
        or ""
    ).strip()
    sysnr = (
        os.getenv(f"SAP_SYSNR_{ambiente_up}")
        or os.getenv("SAP_SYSNR")
        or "00"
    ).strip() or "00"
    user = (
        os.getenv(f"SAP_USER_{ambiente_up}")
        or os.getenv("SAP_USER")
        or ""
    ).strip()
    client = (
        os.getenv(f"SAP_CLIENT_{ambiente_up}")
        or os.getenv("SAP_CLIENT")
        or cliente_esperado
    ).strip() or cliente_esperado
    lang = (
        os.getenv(f"SAP_LANGUAGE_{ambiente_up}")
        or os.getenv("SAP_CUA_LANGUAGE" if ambiente_up == "CUA" else "SAP_LANGUAGE")
        or "PT"
    ).strip() or "PT"

    env_pw_key = f"SAP_PASSWORD_{ambiente_up}"
    env_pw_key_legacy = f"SAP_PASSWORD_{SISTEMA_ESPERADO}CLNT{cliente_esperado}"
    passwd = (
        os.getenv(env_pw_key)
        or os.getenv(env_pw_key_legacy)
        or os.getenv("SAP_PASSWD")
        or os.getenv("SAP_PASSWORD")
        or ""
    ).strip()

    if not ashost or not user or not passwd:
        wb.close()
        print(
            "❌ Faltam credenciais SAP no ficheiro .env "
            f"(esperado: SAP_ASHOST_{ambiente_up} / SAP_USER_{ambiente_up} / SAP_PASSWORD_{ambiente_up} "
            "ou fallback antigo)"
        )
        return

    print(f"\n📋 Roles a processar (agrupadas): {len(roles_agrupadas)}")
    for r in roles_agrupadas:
        print(f"  - {r['AGR_NAME']}: {r['TEXT']} (TCODEs: {len(r['TCODE_LIST'])})")

    if pedir_confirmacao and not modo_nao_interativo:
        conf = input("\nDeseja lançar esses dados no SAP? [S/N]: ").strip().upper()
        if conf != "S":
            wb.close()
            print("❌ Execução cancelada pelo utilizador.")
            return

    resultados = {}
    role_metrics = {}

    with Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        TextColumn("({task.completed}/{task.total})"),
        TimeElapsedColumn(),
    ) as progress:
        task_roles = progress.add_task("Processando roles via RFC...", total=len(roles_agrupadas))

        for idx_role, r_item in enumerate(roles_agrupadas, start=1):
            nome = r_item["AGR_NAME"]
            desc = r_item["TEXT"]
            tcodes = r_item["TCODE_LIST"]
            tempo_inicio_role = time.time()

            print("======================================================================")
            print(f"▶ [{idx_role}/{len(roles_agrupadas)}] INICIANDO ROLE: {nome} | TCODEs: {len(tcodes)}")
            print("======================================================================")

            try:
                # 1. Ligação RFC
                print(f"  ├─ Ligando via RFC ao sistema {SISTEMA_ESPERADO} (Host: {ashost}, Client: {client}, User: {user})...")
                conn = Connection(ashost=ashost, sysnr=sysnr, client=client, user=user, passwd=passwd, lang=lang)

                tcodes_param = [{"TCODE": tc} for tc in tcodes]
                modo_operacao = None

                max_retries = 3
                for tentativa in range(1, max_retries + 1):
                    try:
                        # 2a. Tentar CRIAR a Role (PRGN_RFC_CREATE_ACTIVITY_GROUP)
                        print("  ├─ Chamando RFC para criar a role e associar TCODEs...")
                        try:
                            res_create = conn.call(
                                "PRGN_RFC_CREATE_ACTIVITY_GROUP",
                                ACTIVITY_GROUP=nome,
                                ACTIVITY_GROUP_TEXT=desc,
                                TCODES=tcodes_param,
                                REQUEST=request_transporte or ""
                            )
                            ret_create = res_create.get("RETURN", [])
                            err_msgs = [row.get("MESSAGE", "") for row in ret_create if row.get("TYPE") == "E"]
                            if err_msgs:
                                raise RuntimeError("Erro ao criar role: " + " | ".join(err_msgs))
                            modo_operacao = "CRIADA"

                        except Exception as e_create:
                            err_str = str(e_create)

                            # Tratar TCODEs inválidas na criação
                            if "ILLEGAL_TCODES" in err_str and len(tcodes_param) > 0:
                                print(f"  ⚠️ TCODEs inválidas detetadas na role {nome}. A filtrar TCODEs inexistentes via TSTC...")
                                validadas, invalidas = [], []
                                for tc in tcodes:
                                    try:
                                        r_t = conn.call("RFC_READ_TABLE", QUERY_TABLE="TSTC", OPTIONS=[{"TEXT": f"TCODE = '{tc}'"}], FIELDS=[{"FIELDNAME": "TCODE"}], ROWCOUNT=1)
                                        if r_t.get("DATA"):
                                            validadas.append(tc)
                                        else:
                                            invalidas.append(tc)
                                    except Exception:
                                        validadas.append(tc)
                                if invalidas:
                                    print(f"  ⚠️ TCODEs inexistentes no SAP ignoradas: {invalidas}")
                                    tcodes_param = [{"TCODE": tc} for tc in validadas]
                                    res_create = conn.call(
                                        "PRGN_RFC_CREATE_ACTIVITY_GROUP",
                                        ACTIVITY_GROUP=nome,
                                        ACTIVITY_GROUP_TEXT=desc,
                                        TCODES=tcodes_param,
                                        REQUEST=request_transporte or ""
                                    )
                                    modo_operacao = "CRIADA"
                                else:
                                    raise

                            # 2b. Role já existe → verificar se as TCODEs mudaram antes de atualizar e transportar
                            elif "ACTIVITY_GROUP_ALREADY_EXISTS" in err_str:
                                tcodes_sap = []
                                try:
                                    res_read = conn.call("PRGN_RFC_READ_TRANSACTIONS", ACTIVITY_GROUP=nome)
                                    tcodes_sap = [r.get("TCODE", "").strip().upper() for r in res_read.get("TCODES", []) if r.get("TCODE")]
                                except Exception:
                                    pass

                                set_excel = set(t.upper() for t in tcodes)
                                set_sap = set(tcodes_sap)

                                if set_excel == set_sap and tcodes_sap:
                                    print("  ├─ Role já contém as mesmas TCODEs no SAP — Sem alterações necessárias.")
                                    modo_operacao = "SEM ALTERACOES"
                                else:
                                    print("  ├─ Alteração detetada nas TCODEs — a atualizar via RFC (PRGN_RFC_CHANGE_TRANSACTIONS)...")
                                    try:
                                        res_upd = conn.call(
                                            "PRGN_RFC_CHANGE_TRANSACTIONS",
                                            ACTIVITY_GROUP=nome,
                                            TCODES=tcodes_param,
                                            NO_DIALOG="X",
                                            REQUEST=request_transporte or ""
                                        )
                                        ret_upd = res_upd.get("RETURN", [])
                                        err_upd_msgs = [row.get("MESSAGE", "") for row in ret_upd if row.get("TYPE") == "E"]
                                        if err_upd_msgs:
                                            raise RuntimeError("Erro ao actualizar TCODEs: " + " | ".join(err_upd_msgs))
                                        modo_operacao = "ATUALIZADA"
                                    except Exception as e_upd:
                                        if "ILLEGAL_TCODES" in str(e_upd) and len(tcodes_param) > 0:
                                            print(f"  ⚠️ TCODEs inválidas detetadas na atualização da role {nome}. A filtrar TCODEs via TSTC...")
                                            validadas, invalidas = [], []
                                            for tc in tcodes:
                                                try:
                                                    r_t = conn.call("RFC_READ_TABLE", QUERY_TABLE="TSTC", OPTIONS=[{"TEXT": f"TCODE = '{tc}'"}], FIELDS=[{"FIELDNAME": "TCODE"}], ROWCOUNT=1)
                                                    if r_t.get("DATA"):
                                                        validadas.append(tc)
                                                    else:
                                                        invalidas.append(tc)
                                                except Exception:
                                                    validadas.append(tc)
                                            if invalidas:
                                                print(f"  ⚠️ TCODEs inexistentes no SAP ignoradas: {invalidas}")
                                                tcodes_param = [{"TCODE": tc} for tc in validadas]
                                                res_upd = conn.call(
                                                    "PRGN_RFC_CHANGE_TRANSACTIONS",
                                                    ACTIVITY_GROUP=nome,
                                                    TCODES=tcodes_param,
                                                    NO_DIALOG="X",
                                                    REQUEST=request_transporte or ""
                                                )
                                                modo_operacao = "ATUALIZADA"
                                            else:
                                                raise
                                        else:
                                            raise
                            else:
                                raise

                        if modo_operacao == "SEM ALTERACOES":
                            print("  ├─ Perfil e Request ignorados (role não foi modificada).")
                        else:
                            # 3. Geração / Regeneração de Perfil de Autorizações
                            print(f"  ├─ Chamando RFC para gerar/regenerar o perfil ({modo_operacao})...")
                            res_gen = conn.call(
                                "PRGN_AUTO_GENERATE_PROFILE_NEW",
                                ACTIVITY_GROUP=nome,
                                GENERATE_PROFILE="X",
                                FILL_EMPTY_FIELDS_WITH_STAR="X",
                                ORG_LEVELS_WITH_STAR="X",
                                REQUEST=request_transporte or ""
                            )

                            ret_gen = res_gen.get("RETURN", [])
                            err_gen_msgs = [row.get("MESSAGE", "") for row in ret_gen if row.get("TYPE") == "E"]
                            if err_gen_msgs:
                                raise RuntimeError("Erro ao gerar perfil: " + " | ".join(err_gen_msgs))

                            # 4. Gravar objeto da role na Transport Request (E071) apenas se houve alteração/criação
                            if request_transporte:
                                try:
                                    conn.call(
                                        "TR_EXT_INSERT_IN_REQUEST",
                                        IV_REQ_ID=request_transporte,
                                        IT_OBJECTS=[{"PGMID": "R3TR", "OBJECT": "ACGR", "OBJ_NAME": nome}]
                                    )
                                    print(f"  ├─ Role {nome} gravada e associada com sucesso na Request {request_transporte} (E071)!")
                                except Exception as e_tr:
                                    print(f"  ⚠️ Aviso ao gravar objeto na Request {request_transporte}: {e_tr}")

                        break  # Sucesso! Sair do loop de tentativas

                    except Exception as e_attempt:
                        err_str_attempt = str(e_attempt)
                        if "ACTIVITY_GROUP_ENQUEUED" in err_str_attempt and tentativa < max_retries:
                            print(f"  ⚠️ Role {nome} está bloqueada no SAP (ACTIVITY_GROUP_ENQUEUED). Aguardando 2s (tentativa {tentativa}/{max_retries})...")
                            time.sleep(2)
                            continue
                        raise

                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)
                msg_transporte = f" | Add Req {request_transporte}" if request_transporte else ""

                resultados[nome] = {
                    "STATUS": "CONCLUIDO",
                    "MSG": f"{modo_operacao} (RFC) | {len(tcodes)} TCODEs | Perfil Gerado{msg_transporte}.",
                    "TIMESTEMP": now_ts()
                }
                icone = "🟢" if modo_operacao == "CRIADA" else "🔵"
                print(f"\n{icone} SUCESSO [{modo_operacao}]: Role processada via RFC! ⏱️ (Tempo: {str_tempo})")
                print("----------------------------------------------------------------------")

                role_metrics[nome] = {
                    "total": int(tempo_decorrido_role)
                }

                # Fecha ligação RFC
                conn.close()

            except Exception as e:
                tempo_decorrido_role = time.time() - tempo_inicio_role
                str_tempo = formatar_tempo(tempo_decorrido_role)
                err = str(e)
                resultados[nome] = {
                    "STATUS": "ERRO",
                    "MSG": err,
                    "TIMESTEMP": now_ts()
                }
                print(f"\n🔴 ERRO: {err} ⏱️ (Tempo: {str_tempo})")
                print("----------------------------------------------------------------------")
                role_metrics[nome] = {
                    "total": int(tempo_decorrido_role)
                }

            finally:
                # Grava checkpoint no Excel de forma segura
                try:
                    gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados)
                except Exception as cp_exc:
                    print(f"  ⚠️ Erro ao salvar checkpoint do Excel: {cp_exc}")

                progress.advance(task_roles)

    ###################################################################################
    # BLOCO 5: GRAVAR EXCEL E TEMPO TOTAL
    ###################################################################################
    if gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, records, resultados):
        print("\n💾 Resultados gravados com sucesso no Excel!")
    else:
        print("\n❌ Erro ao gravar resultados finais no Excel.")

    # Imprimir Resumo Comparativo das Roles
    if role_metrics:
        try:
            fastest_role = min(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            slowest_role = max(role_metrics.keys(), key=lambda k: role_metrics[k]["total"])
            
            print("\n=======================================================")
            print("📊 RESUMO COMPARATIVO DE PERFORMANCE DAS ROLES (RFC)")
            print("=======================================================")
            print(f" - Role mais rápida: {fastest_role} ({role_metrics[fastest_role]['total']}s)")
            print(f" - Role mais lenta: {slowest_role} ({role_metrics[slowest_role]['total']}s)")
            print("=======================================================")
        except Exception:
            pass

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
    print("🔁 Fim.")
    return True


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD"])
    parser.add_argument("--xlsx")
    parser.add_argument("--request", help="Número da Request de Transporte (Opcional)")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    args = parser.parse_args()

    env_cli = args.ambiente or (input("Ambiente (DEV/QAD/PRD): ").strip().upper() or "DEV")

    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        request_transporte=args.request,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm)
    )
