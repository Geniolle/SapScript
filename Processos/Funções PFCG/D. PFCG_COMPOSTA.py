# -*- coding: utf-8 -*-

###################################################################################
# D. PFCG_COMPOSTA.py
# PFCG - Criar/Modificar Roles Compostas + Atribuir Roles Componentes + Transporte
#
# Regras:
#  - Logger visual estruturado por Etapas
#  - Integração com 'pesquisar_request.py'
#  - Inserção direta e rápida de Roles componentes (Aba Funções) e de Ordem de Transporte
#  - Menu de Request Unificado
#  - Barra de progresso por Role
#  - Etapa 1 de performance: esperas inteligentes
#  - Etapa 2 de performance: sem pandas
#  - Etapa 3 de performance: cache de IDs SAP
####################################################################################

import sys
import os
import time
import re
import unicodedata
import tkinter as tk
import win32com.client
from tkinter import filedialog
from datetime import datetime
from math import ceil
from openpyxl import load_workbook
from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn

# --- Variáveis a nível de Módulo ---
session = None
sap_id_cache = {}
roles_existentes_cache = {}

# --- Constantes do Script ---
NOME_SHEET = "PFCG_COMPOSTA"
SEARCH_HEADER_IN_FIRST_ROWS = 20
COLUNAS_OBRIGATORIAS_MINIMAS = {"TEXT", "STATUS", "MSG", "TIMESTEMP"}
OPCOES_ROLES = {"ROLES", "TCODE", "ROLES_FILHAS", "AGR_NAME_FILHA"}
MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
SLEEP_UI = 0.08
SLEEP_ACTION = 0.15
ROLES_BLOCK_SIZE = 10  # Standard table control size

# --- Helpers de Utilidades ---
def now_ts():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _sleep(t=SLEEP_UI):
    time.sleep(t)

def norm_col(s):
    if s is None:
        s = ""
    return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

def norm_txt(s):
    if s is None:
        s = ""
    return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

def formatar_tempo(segundos):
    h, resto = divmod(segundos, 3600)
    m, s = divmod(resto, 60)
    if h > 0:
        return f"{int(h):02d}h {int(m):02d}m {int(s):02d}s"
    return f"{int(m):02d}m {int(s):02d}s"

def split_roles(raw):
    if not raw:
        return []
    s = str(raw).replace("\r", "\n").replace("\t", " ").strip().upper()
    parts = re.split(r"[;, \n]+", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if p.startswith("/N") or p.startswith("/O"):
            p = p[2:].strip()
        if p:
            out.append(p)
    return list(dict.fromkeys(out))

# --- Helpers do SAP GUI ---
def _safe_find(sap_id):
    try:
        return session.findById(sap_id)
    except:
        return None

def _sap_busy():
    try:
        return bool(getattr(session, "Busy", False))
    except:
        return False

def _esperar_sap_livre(timeout=8.0, pausa=0.05):
    limite = time.time() + timeout
    while time.time() < limite:
        if not _sap_busy():
            return True
        time.sleep(pausa)
    return False

def _esperar_objeto(sap_id, timeout=5.0, pausa=0.05):
    limite = time.time() + timeout
    while time.time() < limite:
        obj = _safe_find(sap_id)
        if obj:
            return obj
        time.sleep(pausa)
    return None

def _esperar_sumir(sap_id, timeout=5.0, pausa=0.05):
    limite = time.time() + timeout
    while time.time() < limite:
        if not _safe_find(sap_id):
            return True
        time.sleep(pausa)
    return False

def _send_vkey(vkey, wait_after=True):
    session.findById("wnd[0]").sendVKey(vkey)
    if wait_after:
        _esperar_sap_livre()

def _press_if_exists(sap_id, timeout=2.0):
    obj = _esperar_objeto(sap_id, timeout=timeout)
    if not obj:
        return False
    try:
        obj.press()
        _esperar_sap_livre()
        return True
    except:
        return False

def _resolver_id(cache_key, candidatos):
    sap_id = sap_id_cache.get(cache_key)
    if sap_id:
        obj = _safe_find(sap_id)
        if obj:
            return sap_id, obj
        sap_id_cache.pop(cache_key, None)

    for sap_id in candidatos:
        obj = _safe_find(sap_id)
        if obj:
            sap_id_cache[cache_key] = sap_id
            return sap_id, obj
    return None, None

def _resolver_id_esperando(cache_key, candidatos, timeout=3.0, pausa=0.05):
    limite = time.time() + timeout
    while time.time() < limite:
        sap_id, obj = _resolver_id(cache_key, candidatos)
        if obj:
            return sap_id, obj
        time.sleep(pausa)
    return None, None

def is_fatal_error(msg_type, msg_text):
    if not msg_text:
        return False
    msg_text_norm = norm_txt(msg_text)
    non_fatal_patterns = [
        "JA ESTA SELECIONADA",
        "JA SELECIONADA",
        "NAO FORAM MODIFICADOS DADOS",
        "SEM ALTERACOES",
        "JA EXISTE"
    ]
    for pat in non_fatal_patterns:
        if pat in msg_text_norm:
            return False
    fatal_patterns = [
        "NAO EXISTE",
        "NAO FOI POSSIVEL",
        "NAO AUTORIZADO",
        "ERRO"
    ]
    for pat in fatal_patterns:
        if pat in msg_text_norm:
            return True
    if msg_type in ("E", "A"):
        return True
    return False

def get_statusbar():
    try:
        sbar = session.findById("wnd[0]/sbar")
        tipo = getattr(sbar, "MessageType", "").strip().upper()
        texto = (sbar.Text or "").strip()
        if texto:
            print(f"[SAP_SBAR] {texto}")
        return (tipo, texto)
    except:
        return ("", "")

def try_actions(actions):
    for a in actions:
        try:
            ctrl = session.findById(a["path"])
            if a["op"] == "text":
                ctrl.setFocus()
                ctrl.text = a["val"]
                _esperar_sap_livre()
                return True
            elif a["op"] == "press":
                if hasattr(ctrl, "Enabled") and not ctrl.Enabled:
                    continue
                ctrl.press()
                _esperar_sap_livre()
                return True
            elif a["op"] == "select":
                ctrl.select()
                _esperar_sap_livre()
                return True
        except:
            continue
    return False

def tratar_popup_modal(max_loops=6):
    for _ in range(max_loops):
        try:
            _esperar_sap_livre(timeout=2.0, pausa=0.05)
            if session.ActiveWindow.Type != "GuiModalWindow":
                return False
            candidatos = [
                "wnd[1]/usr/btnBUTTON_1",
                "wnd[1]/usr/btnSPOP-OPTION1",
                "wnd[1]/tbar[0]/btn[0]",
                "wnd[1]/tbar[0]/btn[19]",
                "wnd[1]/tbar[0]/btn[11]"
            ]
            for p in candidatos:
                if try_actions([{"path": p, "op": "press"}]):
                    return True
            return True
        except:
            return False
    return True

def wait_wnd1_close(timeout=3.0):
    return _esperar_sumir("wnd[1]", timeout=timeout, pausa=0.05)

def encontrar_grids_alv(root):
    grids = []
    try:
        stack = [root]
        while stack:
            obj = stack.pop()
            try:
                obj_id = obj.Id
                obj_type = obj.Type
                obj_subtype = getattr(obj, "SubType", "")
                obj_name = obj.Name
            except:
                continue
                
            is_toolbar = "toolbar" in str(obj_type).lower() or "toolbar" in str(obj_subtype).lower() or "toolbar" in str(obj_name).lower()
            
            has_row_count = False
            try:
                rc = int(obj.RowCount)
                if rc >= 0:
                    has_row_count = True
            except:
                pass
                
            if has_row_count and not is_toolbar:
                print(f"├─ Candidato ALV/Grid encontrado: Id={obj_id} | Type={obj_type} | SubType={obj_subtype} | Name={obj_name}")
                grids.append(obj)
                
            try:
                for idx in range(obj.Children.Count):
                    stack.append(obj.Children(idx))
            except:
                pass
    except Exception as e_find:
        print(f"  ⚠️ Erro ao procurar ALV: {e_find}")
    return grids

def consultar_componentes_agr_agrs(sess_principal, agr_name):
    res = {
        "ok": False,
        "fonte": "AGR_AGRS",
        "componentes": set(),
        "qtd": 0,
        "mensagem": "Inicializado",
        "erro_tecnico": False,
        "debug": {}
    }
    new_session = None
    try:
        print(f"├─ Consultando AGR_AGRS para a role composta {agr_name}...")
        connection = sess_principal.Parent
        before_ids = set()
        for i in range(connection.Children.Count):
            before_ids.add(connection.Children(i).Id)
            
        sess_principal.findById("wnd[0]/tbar[0]/okcd").text = "/ose16h"
        sess_principal.findById("wnd[0]").sendVKey(0)
        
        t0 = time.time()
        while time.time() - t0 <= 10:
            for i in range(connection.Children.Count):
                c = connection.Children(i)
                if c.Id not in before_ids:
                    new_session = c
                    break
            if new_session:
                break
            time.sleep(0.2)
            
        if not new_session:
            res["mensagem"] = "Não foi possível abrir uma nova sessão SAP (SE16H)."
            res["erro_tecnico"] = True
            print("  ⚠️ Não foi possível abrir uma nova sessão SAP para consultar a AGR_AGRS.")
            return res

        print("├─ SE16H aberta em novo modo.")
        res["debug"]["session_id"] = new_session.Id

        t_wait = time.time()
        while time.time() - t_wait <= 8:
            if not getattr(new_session, "Busy", False):
                break
            time.sleep(0.1)
            
        tab_field_candidates = [
            "wnd[0]/usr/ctxtGD-TAB",
            "wnd[0]/usr/ctxtDATABROWSE-TABLENAME",
            "wnd[0]/usr/ctxtTABNAME",
        ]
        tab_field_id = None
        for cid in tab_field_candidates:
            try:
                if new_session.findById(cid):
                    tab_field_id = cid
                    break
            except:
                pass
                
        if not tab_field_id:
            res["mensagem"] = "Não encontrou o campo de nome da tabela na SE16H."
            res["erro_tecnico"] = True
            print("  ⚠️ Não encontrou o campo de nome da tabela na SE16H.")
            return res
            
        new_session.findById(tab_field_id).text = "AGR_AGRS"
        new_session.findById("wnd[0]").sendVKey(0)
        print("├─ Tabela informada: AGR_AGRS")
        
        t_wait = time.time()
        tbl_control = None
        while time.time() - t_wait <= 5:
            try:
                root = new_session.findById("wnd[0]/usr")
                stack = [root]
                while stack:
                    obj = stack.pop()
                    if obj.Name == "SAPLSE16NSELFIELDS_TC" or obj.Id.endswith("tblSAPLSE16NSELFIELDS_TC"):
                        tbl_control = obj
                        break
                    for idx in range(obj.Children.Count):
                        stack.append(obj.Children(idx))
                if tbl_control:
                    break
            except:
                pass
            time.sleep(0.1)
            
        if not tbl_control:
            res["mensagem"] = "Tabela de critérios da SE16H não carregou."
            res["erro_tecnico"] = True
            print("  ⚠️ Tabela de critérios da SE16H não carregou.")
            return res
            
        max_cids = ["wnd[0]/usr/txtMAX_SEL", "wnd[0]/usr/txtGD-MAXROWS", "wnd[0]/usr/txtMAX_HITS"]
        for mcid in max_cids:
            try:
                new_session.findById(mcid).text = "9999"
                break
            except:
                pass
                
        row_count = int(tbl_control.RowCount)
        visible_rows = int(tbl_control.VisibleRowCount)
        
        col_fieldname_orig = 13
        col_fieldname_prefix = "txtGS_SELFIELDS-FIELDNAME"
        col_low = 2
        col_low_prefix = "ctxtGS_SELFIELDS-LOW"
        
        try:
            for idx in range(tbl_control.Children.Count):
                child = tbl_control.Children(idx)
                id_str = child.Id
                if "[" in id_str and "]" in id_str:
                    bracket_part = id_str.rsplit("[", 1)[-1].split("]")[0]
                    parts = bracket_part.split(",")
                    if len(parts) == 2 and int(parts[1]) == 0:
                        col_idx = int(parts[0])
                        prefix_path = id_str.rsplit("[", 1)[0]
                        prefix = prefix_path.split("/")[-1]
                        name = child.Name.upper()
                        if name.endswith("-LOW") or name == "GS_SELFIELDS-LOW":
                            col_low = col_idx
                            col_low_prefix = prefix
                        elif "FIELDNAME" in name:
                            if col_idx > 10:
                                col_fieldname_orig = col_idx
                                col_fieldname_prefix = prefix
        except:
            pass

        row_agr_name = None
        for r in range(min(row_count, visible_rows)):
            try:
                fname_id = f"{tbl_control.Id}/{col_fieldname_prefix}[{col_fieldname_orig},{r}]"
                fieldname = new_session.findById(fname_id).text.strip().upper()
                if fieldname == "AGR_NAME":
                    row_agr_name = r
                    break
            except:
                continue
                
        if row_agr_name is None:
            res["mensagem"] = "Não encontrou o critério 'AGR_NAME' na tabela SE16H."
            res["erro_tecnico"] = True
            print("  ⚠️ Não encontrou o critério 'AGR_NAME' na tabela SE16H.")
            return res
            
        low_id = f"{tbl_control.Id}/{col_low_prefix}[{col_low},{row_agr_name}]"
        new_session.findById(low_id).text = agr_name
        print(f"├─ Filtro aplicado: AGR_NAME = {agr_name}")
        
        new_session.findById("wnd[0]/tbar[1]/btn[8]").press()
        
        t_wait = time.time()
        while time.time() - t_wait <= 10:
            if not getattr(new_session, "Busy", False):
                break
            time.sleep(0.15)
            
        try:
            sbar = new_session.findById("wnd[0]/sbar")
            sbar_text = str(sbar.Text).strip()
            sbar_type = str(sbar.MessageType).strip().upper()
            if sbar_text:
                print(f"├─ [SE16H_SBAR] Tipo: {sbar_type} | Texto: {sbar_text}")
                res["debug"]["sbar"] = {"type": sbar_type, "text": sbar_text}
                
                sbar_text_upper = norm_txt(sbar_text)
                if sbar_type in ("E", "A"):
                    # Regra especial: se a composta existe mas retornar "Nenhum valor encontrado", tratamos como ok com 0 componentes
                    if "NENHUM" in sbar_text_upper or "NO DATA" in sbar_text_upper or "ZERO" in sbar_text_upper or "NOT FOUND" in sbar_text_upper or "FOUND" in sbar_text_upper:
                        res["ok"] = True
                        res["qtd"] = 0
                        res["mensagem"] = f"Consulta com sucesso: {sbar_text}"
                        print("├─ AGR_AGRS sem registos para role composta existente.")
                        return res
                    else:
                        res["mensagem"] = f"Erro no statusbar da SE16H: {sbar_text}"
                        return res
        except:
            pass

        print("├─ Consulta executada.")

        grid = None
        try:
            root_wnd = new_session.findById("wnd[0]")
            grids_found = encontrar_grids_alv(root_wnd)
            if grids_found:
                grid = grids_found[0]
        except Exception as e_grid:
            print(f"  ⚠️ Exceção ao varrer tela por ALV/Grid: {e_grid}")
            
        if not grid:
            try:
                sbar = new_session.findById("wnd[0]/sbar")
                sbar_text = norm_txt(sbar.Text)
                if "NENHUM" in sbar_text or "NO DATA" in sbar_text or "ZERO" in sbar_text or "NOT FOUND" in sbar_text or "FOUND" in sbar_text:
                    res["ok"] = True
                    res["qtd"] = 0
                    res["mensagem"] = "Consulta com sucesso, 0 registos encontrados."
                    print("├─ Linhas retornadas: 0 (Nenhum dado selecionado)")
                    return res
            except:
                pass
            
            res["mensagem"] = "Não consegui localizar o ALV/Grid de resultados da SE16H."
            res["erro_tecnico"] = True
            print("  ⚠️ Não consegui localizar o ALV/Grid de resultados da SE16H.")
            return res

        print(f"├─ ALV/Grid encontrado: {grid.Id}")
        res["debug"]["grid_id"] = grid.Id

        r_count = int(grid.RowCount)
        print(f"├─ Linhas retornadas: {r_count}")
        res["debug"]["row_count"] = r_count

        comp_col = None
        colunas_candidatas = ["CHILD_AGR", "AGR_NAME_CHILD", "CHILD_AGR_NAME", "AGR_NAME_COMP", "ROLE", "AGR_NAME"]
        
        alv_cols = []
        try:
            for col in grid.ColumnOrder:
                alv_cols.append(str(col))
        except:
            pass
            
        print(f"├─ Colunas encontradas: {', '.join(alv_cols) if alv_cols else 'não listadas'}")
        res["debug"]["colunas"] = alv_cols
        
        for c_cand in colunas_candidatas:
            if c_cand in alv_cols or c_cand.upper() in [c.upper() for c in alv_cols]:
                if c_cand.upper() == "AGR_NAME":
                    tem_outra = any(x in [c.upper() for c in alv_cols] for x in ["CHILD_AGR", "AGR_NAME_CHILD", "CHILD_AGR_NAME", "AGR_NAME_COMP", "ROLE"])
                    if tem_outra:
                        continue
                comp_col = c_cand
                break
                
        if not comp_col and r_count > 0:
            for c_cand in colunas_candidatas:
                try:
                    val = grid.GetCellValue(0, c_cand)
                    if val is not None and str(val).strip().upper() != agr_name.upper():
                        comp_col = c_cand
                        break
                except:
                    pass
                    
        if not comp_col and r_count > 0 and alv_cols:
            for col in alv_cols:
                try:
                    val = str(grid.GetCellValue(0, col)).strip()
                    if val and val.upper() != agr_name.upper() and len(val) <= 30 and re.match(r"^[A-Z0-9_]+$", val, re.IGNORECASE):
                        comp_col = col
                        break
                except:
                    pass

        if r_count > 0 and not comp_col:
            res["mensagem"] = "Não foi possível identificar a coluna de função componente no ALV/Grid."
            res["erro_tecnico"] = True
            print("  ⚠️ Não foi possível identificar a coluna de função componente no ALV/Grid.")
            return res

        if comp_col:
            print(f"├─ Coluna componente identificada: {comp_col}")

        componentes_set = set()
        for r in range(r_count):
            try:
                val = str(grid.GetCellValue(r, comp_col)).strip().upper()
                if val and val != agr_name.upper():
                    componentes_set.add(val)
            except:
                pass

        res["componentes"] = componentes_set
        res["qtd"] = len(componentes_set)
        res["ok"] = True
        res["mensagem"] = f"Consulta concluída com sucesso. {len(componentes_set)} componentes lidos."

        comp_log = sorted(list(componentes_set))
        print("├─ Componentes encontrados na AGR_AGRS:")
        if comp_log:
            for comp_item in comp_log[:10]:
                print(f"│  └─ {comp_item}")
            if len(comp_log) > 10:
                print(f"│  └─ ... (+ {len(comp_log) - 10} adicionais)")
        else:
            print("│  └─ nenhuma")

    except Exception as e_agr:
        res["mensagem"] = f"Exceção técnica: {e_agr}"
        res["erro_tecnico"] = True
        print(f"  ⚠️ Erro técnico ao consultar a tabela AGR_AGRS: {e_agr}")
        
    finally:
        if new_session:
            try:
                new_session.findById("wnd[0]").close()
                time.sleep(0.3)
                try:
                    if new_session.ActiveWindow.Type == "GuiModalWindow":
                        candidatos_btn = [
                            "wnd[1]/usr/btnSPOP-OPTION1",
                            "wnd[1]/usr/btnBUTTON_1",
                            "wnd[1]/tbar[0]/btn[0]",
                        ]
                        for c_btn in candidatos_btn:
                            try:
                                new_session.findById(c_btn).press()
                                break
                            except:
                                pass
                except:
                    pass
            except:
                pass
                
    return res

def verificar_role_existe_agr_define(sess_principal, role_name):
    new_session = None
    try:
        print(f"├─ Verificando existência de {role_name} na AGR_DEFINE...")
        connection = sess_principal.Parent
        before_ids = set()
        for i in range(connection.Children.Count):
            before_ids.add(connection.Children(i).Id)
            
        sess_principal.findById("wnd[0]/tbar[0]/okcd").text = "/ose16h"
        sess_principal.findById("wnd[0]").sendVKey(0)
        
        t0 = time.time()
        while time.time() - t0 <= 10:
            for i in range(connection.Children.Count):
                c = connection.Children(i)
                if c.Id not in before_ids:
                    new_session = c
                    break
            if new_session:
                break
            time.sleep(0.2)
            
        if not new_session:
            print("  ⚠️ Não foi possível abrir uma nova sessão SAP para consultar a AGR_DEFINE.")
            return False

        t_wait = time.time()
        while time.time() - t_wait <= 8:
            if not getattr(new_session, "Busy", False):
                break
            time.sleep(0.1)
            
        tab_field_candidates = [
            "wnd[0]/usr/ctxtGD-TAB",
            "wnd[0]/usr/ctxtDATABROWSE-TABLENAME",
            "wnd[0]/usr/ctxtTABNAME",
        ]
        tab_field_id = None
        for cid in tab_field_candidates:
            try:
                if new_session.findById(cid):
                    tab_field_id = cid
                    break
            except:
                pass
                
        if not tab_field_id:
            return False
            
        new_session.findById(tab_field_id).text = "AGR_DEFINE"
        new_session.findById("wnd[0]").sendVKey(0)
        
        t_wait = time.time()
        tbl_control = None
        while time.time() - t_wait <= 5:
            try:
                root = new_session.findById("wnd[0]/usr")
                stack = [root]
                while stack:
                    obj = stack.pop()
                    if obj.Name == "SAPLSE16NSELFIELDS_TC" or obj.Id.endswith("tblSAPLSE16NSELFIELDS_TC"):
                        tbl_control = obj
                        break
                    for idx in range(obj.Children.Count):
                        stack.append(obj.Children(idx))
                if tbl_control:
                    break
            except:
                pass
            time.sleep(0.1)
            
        if not tbl_control:
            return False
            
        max_cids = ["wnd[0]/usr/txtMAX_SEL", "wnd[0]/usr/txtGD-MAXROWS", "wnd[0]/usr/txtMAX_HITS"]
        for mcid in max_cids:
            try:
                new_session.findById(mcid).text = "1"
                break
            except:
                pass
                
        row_count = int(tbl_control.RowCount)
        visible_rows = int(tbl_control.VisibleRowCount)
        
        col_fieldname_orig = 13
        col_fieldname_prefix = "txtGS_SELFIELDS-FIELDNAME"
        col_low = 2
        col_low_prefix = "ctxtGS_SELFIELDS-LOW"
        
        try:
            for idx in range(tbl_control.Children.Count):
                child = tbl_control.Children(idx)
                id_str = child.Id
                if "[" in id_str and "]" in id_str:
                    bracket_part = id_str.rsplit("[", 1)[-1].split("]")[0]
                    parts = bracket_part.split(",")
                    if len(parts) == 2 and int(parts[1]) == 0:
                        col_idx = int(parts[0])
                        prefix_path = id_str.rsplit("[", 1)[0]
                        prefix = prefix_path.split("/")[-1]
                        name = child.Name.upper()
                        if name.endswith("-LOW") or name == "GS_SELFIELDS-LOW":
                            col_low = col_idx
                            col_low_prefix = prefix
                        elif "FIELDNAME" in name:
                            if col_idx > 10:
                                col_fieldname_orig = col_idx
                                col_fieldname_prefix = prefix
        except:
            pass

        row_agr_name = None
        for r in range(min(row_count, visible_rows)):
            try:
                fname_id = f"{tbl_control.Id}/{col_fieldname_prefix}[{col_fieldname_orig},{r}]"
                fieldname = new_session.findById(fname_id).text.strip().upper()
                if fieldname == "AGR_NAME":
                    row_agr_name = r
                    break
            except:
                continue
                
        if row_agr_name is None:
            return False
            
        low_id = f"{tbl_control.Id}/{col_low_prefix}[{col_low},{row_agr_name}]"
        new_session.findById(low_id).text = role_name
        
        new_session.findById("wnd[0]/tbar[1]/btn[8]").press()
        
        t_wait = time.time()
        while time.time() - t_wait <= 10:
            if not getattr(new_session, "Busy", False):
                break
            time.sleep(0.15)
            
        try:
            sbar = new_session.findById("wnd[0]/sbar")
            sbar_text = str(sbar.Text).strip()
            if sbar_text:
                sbar_text_norm = norm_txt(sbar_text)
                if "NENHUM" in sbar_text_norm or "NO DATA" in sbar_text_norm or "ZERO" in sbar_text_norm or "NOT FOUND" in sbar_text_norm:
                    return False
        except:
            pass
            
        grid = None
        try:
            root_wnd = new_session.findById("wnd[0]")
            grids_found = encontrar_grids_alv(root_wnd)
            if grids_found:
                grid = grids_found[0]
        except:
            pass
            
        if grid and int(grid.RowCount) > 0:
            return True
        return False
    except Exception as e:
        print(f"  ⚠️ Erro ao verificar existência de {role_name}: {e}")
        return False
    finally:
        if new_session:
            try:
                new_session.findById("wnd[0]").close()
                time.sleep(0.3)
                try:
                    if new_session.ActiveWindow.Type == "GuiModalWindow":
                        new_session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
                except:
                    pass
            except:
                pass

# --- Classe Page Object ---
class PFCGPage:
    def __init__(self, sess):
        self.sess = sess

    def open(self):
        print("  ├─ Abrindo a transação /NPFCG...")
        self.sess.findById("wnd[0]/tbar[0]/okcd").text = "/NPFCG"
        _send_vkey(0)
        tratar_popup_modal()

    def set_role_name(self, nome):
        print(f"  ├─ Inserindo o nome da Role: {nome}")
        sap_id, obj = _resolver_id(
            "role_name_field",
            ["wnd[0]/usr/ctxtAGR_NAME_NEU", "wnd[0]/usr/ctxtAGR_NAME"]
        )
        if not obj:
            return False
        try:
            obj.setFocus()
            obj.text = nome
            _esperar_sap_livre()
            return True
        except:
            return False

    def open_for_edit(self):
        print("  ├─ Tentando abrir em modo de 'Criação Composta'...")
        if not try_actions([
            {"path": "wnd[0]/usr/btn%#AUTOTEXT004", "op": "press"},
            {"path": "wnd[0]/tbar[1]/btn[5]", "op": "press"}
        ]):
            raise Exception("Não consegui clicar em Criar Função Composta.")

        tratar_popup_modal()

        mt, sb = get_statusbar()
        if "EXISTE" in norm_txt(sb) or "EXISTS" in norm_txt(sb):
            print("  ├─ A Role já existe. Alterando para modo de 'Alteração'...")
            if not try_actions([
                {"path": "wnd[0]/usr/btn%#AUTOTEXT001", "op": "press"},
                {"path": "wnd[0]/tbar[1]/btn[2]", "op": "press"}
            ]):
                raise Exception("Role já existe, mas não consegui abrir Alterar.")
            tratar_popup_modal()
            return "CHANGE"
        return "CREATE"

    def set_description(self, desc):
        print("  ├─ Preenchendo a descrição da Role...")
        sap_id, obj = _resolver_id(
            "role_desc_field",
            ["wnd[0]/usr/txtS_AGR_TEXTS-TEXT", "wnd[0]/usr/txtS_AGR_TEXTS-TEXT1", "wnd[0]/usr/txtAGR_TEXTS-TEXT"]
        )
        if not obj:
            return False
        try:
            obj.text = desc
            _send_vkey(0)
            tratar_popup_modal()
            return True
        except:
            return False

    def save(self, log_msg="  └─ Guardando alterações..."):
        print(log_msg)
        try:
            self.sess.findById("wnd[0]").sendVKey(11)
        except:
            try_actions([{"path": "wnd[0]/tbar[0]/btn[11]", "op": "press"}])

        _esperar_sap_livre()
        tratar_popup_modal()

        mt, sb = get_statusbar()
        if sb:
            if mt in ("E", "A"):
                print(f"     ❌ SAP Erro: {sb}")
                raise Exception(f"Falha ao guardar: {sb}")
            else:
                print(f"     ✔️ SAP: {sb}")
        else:
            print("     ✔️ SAP: Operação concluída sem mensagem do sistema.")

    def goto_roles_tab(self):
        print("  ├─ Acedendo à aba 'Funções' (TAB8)...")
        sap_id, obj = _resolver_id("roles_tab", ["wnd[0]/usr/tabsTABSTRIP1/tabpTAB8"])
        if not obj:
            raise Exception("Não consegui abrir a aba Funções (TAB8).")
        try:
            obj.select()
            _esperar_sap_livre()
        except:
            raise Exception("Não consegui abrir a aba Funções (TAB8).")
        tratar_popup_modal()

    def read_existing_roles(self):
        table_id, table_obj = _resolver_id(
            "roles_table",
            [
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0600/tblSAPLPRGN_TREECTRL_AGRLIST2",
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0610/tblSAPLPRGN_TREECTRL_AGRLIST2",
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0620/tblSAPLPRGN_TREECTRL_AGRLIST2",
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0330/tblSAPLPRGN_TREECTRL_AGRLIST2",
            ]
        )
        if not table_obj:
            raise Exception("Não encontrei a tabela de funções componentes.")

        existing_roles = set()
        visible_rows = 10
        try:
            visible_rows = int(table_obj.VisibleRowCount)
        except:
            pass
            
        total_rows = int(table_obj.RowCount)
        
        # 1. Ler todas as rows existentes fazendo scroll
        try:
            scroll_pos = 0
            while scroll_pos < total_rows:
                table_obj.VerticalScrollbar.Position = scroll_pos
                _esperar_sap_livre()
                
                for row_in_page in range(visible_rows):
                    r_idx = scroll_pos + row_in_page
                    if r_idx >= total_rows:
                        break
                    
                    cell_id = f"{table_id}/ctxtI_ACTGROUPS-AGR_NAME[0,{row_in_page}]"
                    cell_obj = _safe_find(cell_id)
                    if cell_obj:
                        val = str(cell_obj.Text).strip().upper()
                        if val:
                            existing_roles.add(val)
                scroll_pos += visible_rows
                if visible_rows <= 0:
                    break
        except Exception as read_exc:
            print(f"  ⚠️ Erro ao ler componentes existentes: {read_exc}")

        try:
            table_obj.VerticalScrollbar.Position = 0
            _esperar_sap_livre()
        except:
            pass

        # 2. Encontrar o índice da primeira linha vazia no table control
        first_empty_row = 0
        try:
            scroll_pos = 0
            found = False
            while scroll_pos < total_rows and not found:
                table_obj.VerticalScrollbar.Position = scroll_pos
                _esperar_sap_livre()
                for row_in_page in range(visible_rows):
                    r_idx = scroll_pos + row_in_page
                    if r_idx >= total_rows:
                        break
                    cell_id = f"{table_id}/ctxtI_ACTGROUPS-AGR_NAME[0,{row_in_page}]"
                    cell_obj = _safe_find(cell_id)
                    if cell_obj and not str(cell_obj.Text).strip():
                        first_empty_row = r_idx
                        found = True
                        break
                if found:
                    break
                scroll_pos += visible_rows
            if not found:
                first_empty_row = total_rows
        except:
            first_empty_row = len(existing_roles)

        try:
            table_obj.VerticalScrollbar.Position = 0
            _esperar_sap_livre()
        except:
            pass

        return existing_roles, first_empty_row

    def add_only_missing_roles(self, roles_to_insert, first_empty_row):
        table_id, table_obj = _resolver_id(
            "roles_table",
            [
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0600/tblSAPLPRGN_TREECTRL_AGRLIST2",
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0610/tblSAPLPRGN_TREECTRL_AGRLIST2",
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0620/tblSAPLPRGN_TREECTRL_AGRLIST2",
                "wnd[0]/usr/tabsTABSTRIP1/tabpTAB8/ssubSUB1:SAPLPRGN_TREE:0330/tblSAPLPRGN_TREECTRL_AGRLIST2",
            ]
        )
        if not table_obj:
            raise Exception("Não encontrei a tabela de funções componentes.")

        visible_rows = 10
        try:
            visible_rows = int(table_obj.VisibleRowCount)
        except:
            pass

        for idx, role in enumerate(roles_to_insert):
            write_row = first_empty_row + idx
            row_in_page = write_row % visible_rows
            
            try:
                table_obj.VerticalScrollbar.Position = (write_row // visible_rows) * visible_rows
                _esperar_sap_livre()
            except:
                pass
            
            cell_id = f"{table_id}/ctxtI_ACTGROUPS-AGR_NAME[0,{row_in_page}]"
            cell = _esperar_objeto(cell_id, timeout=2.0)
            if not cell:
                cell = _safe_find(cell_id)
            
            if cell:
                cell.text = role
                print(f"     ├─ Inserindo {role} na linha {write_row}")
            else:
                print(f"     ⚠️ Não consegui encontrar o campo para a linha {write_row} (ID: {cell_id})")

        _esperar_sap_livre()
        # Press enter to validate
        try:
            self.sess.findById("wnd[0]").sendVKey(0)
        except:
            pass
        _esperar_sap_livre()
        tratar_popup_modal()

        mt, sb = get_statusbar()
        if is_fatal_error(mt, sb):
            raise Exception(f"Erro na validação de componentes: {sb}")

    def execute_transport_and_exit(self, req_num):
        if req_num:
            print("  ├─ Recuando para a base da PFCG para pedir Transporte (F3)...")
            try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
            tratar_popup_modal()

            print("  ├─ Acedendo ao Menu Função -> Transporte...")
            try_actions([{"path": "wnd[0]/mbar/menu[0]/menu[9]", "op": "select"}])
            tratar_popup_modal()

            print("  ├─ Clicando em Executar transporte...")
            try_actions([{"path": "wnd[0]/tbar[1]/btn[8]", "op": "press"}])

            field_id, req_field = _resolver_id_esperando(
                "transport_req_field",
                ["wnd[1]/usr/ctxtKO008-TRKORR"],
                timeout=3.0
            )

            print(f"  ├─ Injetando a Request ({req_num}) diretamente no popup...")
            if req_field:
                req_field.text = str(req_num)

            try_actions([{"path": "wnd[1]/tbar[0]/btn[0]", "op": "press"}])
            tratar_popup_modal()

            mt, sb = get_statusbar()
            if sb and mt not in ("E", "A"):
                print(f"     ✔️ SAP: {sb}")
            else:
                print("     ✔️ SAP: Transporte associado com sucesso!")

        print("  └─ Regressando em segurança ao ecrã principal SAP Easy Access (F3)...")
        for _ in range(3):
            try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
            tratar_popup_modal()

# --- Função Modular de Chamada Externa ---
def executar_grupos_compostos(
    ambiente_cockpit,
    grupos_compostos,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=True,
    pedir_confirmacao=False,
    origem="PFCG_AUTHORITY",
    sess_externa=None
):
    global session
    if sess_externa:
        session = sess_externa
    else:
        MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
        SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
        try:
            SapGuiAuto = win32com.client.GetObject("SAPGUI")
            application = SapGuiAuto.GetScriptingEngine
            session = next((sess for conn in application.Children for sess in conn.Children if sess.Info.SystemName.upper() == SISTEMA_ESPERADO), None)
        except Exception:
            session = None
            
    if not session:
        print(f"❌ Não encontrei sessão do ambiente '{ambiente_cockpit}'.")
        return {}
        
    pfcg = PFCGPage(session)
    resultados = {}
    
    for gc in grupos_compostos:
        nome = gc["agr_name_composta"]
        desc = gc["text_composta"]
        componentes = gc["componentes"]
        
        print("\n======================================================================")
        print(f"▶ [Fase Composta] Iniciar role composta: {nome} | Componentes: {len(componentes)}")
        print("======================================================================")
        
        try:
            pfcg.open()
            if not pfcg.set_role_name(nome):
                raise Exception("Falha ao escrever nome da Função Composta.")
            modo = pfcg.open_for_edit()
            if modo == "CHANGE":
                print("├─ Role composta já existe; abrindo em alteração...")
            else:
                print("├─ Role composta não existe; abrindo em criação...")
            pfcg.set_description(desc)
            pfcg.save("  └─ Guardando alterações iniciais da Composta...")
            
            print("\n[Etapa 2] Validação de Funções Componentes")
            
            existing_roles = set()
            fonte_final = "AGR_AGRS"
            first_empty_row = 0
            
            if modo == "CHANGE":
                res_agr = consultar_componentes_agr_agrs(session, nome)
                if res_agr["ok"]:
                    existing_roles = res_agr["componentes"]
                    fonte_final = "AGR_AGRS"
                else:
                    print(f"├─ ERRO técnico na consulta AGR_AGRS: {res_agr['mensagem']}")
                    raise Exception("Validação inconclusiva: não foi possível ler AGR_AGRS. Inserção bloqueada para evitar duplicação.")
            
            componentes_excel = [r.strip().upper() for r in componentes if r and str(r).strip()]
            componentes_existentes = [r for r in componentes_excel if r in existing_roles]
            componentes_para_inserir = [r for r in componentes_excel if r not in existing_roles]
            
            # Validar que todos os componentes_para_inserir existem no SAP
            roles_inexistentes = []
            for comp in componentes_para_inserir:
                if comp not in roles_existentes_cache:
                    exists = verificar_role_existe_agr_define(session, comp)
                    roles_existentes_cache[comp] = exists
                
                if not roles_existentes_cache[comp]:
                    roles_inexistentes.append(comp)
            
            if roles_inexistentes:
                raise Exception(f"Componente inexistente: {roles_inexistentes[0]}. Composta não atualizada.")

            print(f"├─ Componentes encontrados na {fonte_final}:")
            if existing_roles:
                for comp_item in sorted(list(existing_roles))[:10]:
                    print(f"│  └─ {comp_item}")
                if len(existing_roles) > 10:
                    print(f"│  └─ ... (+ {len(existing_roles)-10} adicionais)")
            else:
                print("│  └─ nenhuma")
                
            print("├─ Componentes no ficheiro:")
            for comp_item in sorted(list(componentes_excel))[:10]:
                print(f"│  └─ {comp_item}")
            if len(componentes_excel) > 10:
                print(f"│  └─ ... (+ {len(componentes_excel)-10} adicionais)")
                
            print("├─ Já existentes:")
            if componentes_existentes:
                for comp_item in sorted(list(componentes_existentes))[:10]:
                    print(f"│  └─ {comp_item}")
                if len(componentes_existentes) > 10:
                    print(f"│  └─ ... (+ {len(componentes_existentes)-10} adicionais)")
            else:
                print("│  └─ nenhuma")
                
            print("├─ Novos a inserir:")
            if componentes_para_inserir:
                for comp_item in sorted(list(componentes_para_inserir))[:10]:
                    print(f"│  └─ {comp_item}")
                if len(componentes_para_inserir) > 10:
                    print(f"│  └─ ... (+ {len(componentes_para_inserir)-10} adicionais)")
            else:
                print("│  └─ nenhuma")
                
            print(f"└─ Fonte final da validação: {fonte_final}")
            
            skip_rest = False
            if modo == "CHANGE" and not componentes_para_inserir:
                print("\n🟢 SUCESSO: Sem alterações necessárias. Todas as funções componentes já estavam atribuídas na AGR_AGRS.")
                resultados[nome] = {
                    "STATUS": "CONCLUIDO",
                    "MSG": "Sem alterações: todas as funções componentes já estavam atribuídas na AGR_AGRS.",
                    "modo": "CHANGE",
                    "inseridos": 0,
                    "existentes": len(componentes_existentes)
                }
                skip_rest = True
                
                print("  └─ Regressando ao ecrã principal SAP Easy Access (F3)...")
                for _ in range(2):
                    try_actions([{"path": "wnd[0]/tbar[0]/btn[3]", "op": "press"}])
                    tratar_popup_modal()
            else:
                print("\n[Etapa 3] Atribuir Componentes")
                pfcg.goto_roles_tab()
                _, first_empty_row = pfcg.read_existing_roles()
                
                if componentes_para_inserir:
                    pfcg.add_only_missing_roles(componentes_para_inserir, first_empty_row)
                    pfcg.save("  └─ Guardando Funções inseridas...")
                    
            if not skip_rest:
                print("\n[Etapa 4] Ordem de Transporte")
                pfcg.execute_transport_and_exit(request_transporte)
                
                msg_final = f"Sucesso ({modo}) | Inseridos {len(componentes_para_inserir)}/{len(componentes_excel)} Componentes | Já existentes {len(componentes_existentes)}/{len(componentes_excel)} | Fonte validação: AGR_AGRS."
                
                resultados[nome] = {
                    "STATUS": "CONCLUIDO",
                    "MSG": msg_final,
                    "modo": modo,
                    "inseridos": len(componentes_para_inserir),
                    "existentes": len(componentes_existentes)
                }
                print(f"🟢 SUCESSO: Role composta {nome} tratada por completo!")
                
        except Exception as e:
            err = str(e)
            if "Validação inconclusiva" in err:
                err = "Validação inconclusiva: não foi possível ler AGR_AGRS. Inserção bloqueada para evitar duplicação."
            else:
                mt, sb = get_statusbar()
                if mt in ("E", "A"):
                    err = sb
            resultados[nome] = {"STATUS": "ERRO", "MSG": err, "modo": "CHANGE", "inseridos": 0, "existentes": 0}
            print(f"🔴 ERRO na Composta {nome}: {err}")
            
            try:
                for _ in range(3):
                    if session.ActiveWindow.Type == "GuiModalWindow":
                        session.ActiveWindow.close()
                        tratar_popup_modal()
                session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
                _send_vkey(0)
            except:
                pass
                
    return resultados

# --- Executar Original do Script (Via Ficheiro) ---
def executar(
    ambiente_cockpit,
    caminho_ficheiro=None,
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True
):
    global session
    tempo_inicio_total = time.time()

    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual)
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)

    if not caminho_ficheiro:
        if modo_nao_interativo:
            raise ValueError("Faltou o parâmetro --xlsx em modo não-interativo.")
        print("📂 Selecione o ficheiro Excel…")
        caminho_ficheiro = selecionar_ficheiro()
        if not caminho_ficheiro:
            print("❌ Operação cancelada.")
            return

    if not os.path.exists(caminho_ficheiro):
        print(f"❌ Ficheiro não encontrado: {caminho_ficheiro}")
        return

    try:
        wb = load_workbook(caminho_ficheiro)
    except Exception as e:
        print(f"❌ Não consegui abrir o Excel: {e}")
        return

    if NOME_SHEET in wb.sheetnames:
        ws = wb[NOME_SHEET]
    else:
        if len(wb.sheetnames) == 1:
            ws = wb[wb.sheetnames[0]]
        else:
            print(f"❌ Sheet '{NOME_SHEET}' não encontrada.")
            wb.close()
            return

    header_row = None
    header_map = {}
    COLUNAS_SAIDA = {"STATUS": "STATUS", "MSG": "MSG", "TIMESTEMP": "TIMESTEMP"}

    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [norm_col(c.value) for c in ws[r]]
        row_set = set(row_vals)
        
        if "TEXT" not in row_set or "AGR_NAME" not in row_set:
            continue
            
        is_format_1 = "AGR_NAME_COMPOSTA" in row_set
        is_format_2 = len(OPCOES_ROLES.intersection(row_set)) >= 1
        
        if is_format_1 or is_format_2:
            header_row = r
            for idx, name in enumerate(row_vals, start=1):
                if name:
                    header_map[name] = idx
            
            modificou = False
            last_col = 0
            for idx, val in enumerate(row_vals, start=1):
                if val:
                    last_col = idx

            for col_key, col_name in COLUNAS_SAIDA.items():
                if col_key not in header_map:
                    last_col += 1
                    ws.cell(row=r, column=last_col, value=col_name)
                    header_map[col_key] = last_col
                    print(f"➕ Coluna '{col_name}' criada automaticamente na coluna {last_col}.")
                    modificou = True

            if modificou:
                try:
                    wb.save(caminho_ficheiro)
                    print("💾 Ficheiro Excel guardado com as novas colunas.")
                except Exception as e:
                    wb.close()
                    print(f"❌ Erro ao guardar o ficheiro Excel com as novas colunas: {e}")
                    return
            break

    if not header_row:
        wb.close()
        print("\n❌ Não encontrei a linha de cabeçalho completa com as colunas obrigatórias.")
        return

    col_text = header_map.get("TEXT")
    col_status = header_map.get("STATUS")
    col_msg = header_map.get("MSG")
    col_ts = header_map.get("TIMESTEMP")

    if "AGR_NAME_COMPOSTA" in header_map:
        col_agr = header_map["AGR_NAME_COMPOSTA"]
        col_roles = header_map["AGR_NAME"]
    else:
        col_agr = header_map.get("AGR_NAME")
        col_roles = None
        for opt in ["ROLES", "TCODE", "ROLES_FILHAS", "AGR_NAME_FILHA"]:
            if opt in header_map:
                col_roles = header_map[opt]
                break

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        agr_val = ws.cell(row=r, column=col_agr).value if col_agr else None
        agr = "" if agr_val is None else str(agr_val).strip()
        if not agr:
            continue

        if " " in agr:
            orig = agr
            agr = agr.replace(" ", "")
            ws.cell(row=r, column=col_agr, value=agr)
            print(f"⚠️ Espaços detetados e removidos da Role: '{orig}' -> '{agr}' (atualizado no Excel)")
            try:
                wb.save(caminho_ficheiro)
            except:
                pass

        text_val = ws.cell(row=r, column=col_text).value if col_text else None
        roles_val = ws.cell(row=r, column=col_roles).value if col_roles else None
        status_val = ws.cell(row=r, column=col_status).value if col_status else None
        msg_val = ws.cell(row=r, column=col_msg).value if col_msg else None
        ts_val = ws.cell(row=r, column=col_ts).value if col_ts else None

        records.append({
            "_row": r,
            "AGR_NAME": agr,
            "TEXT": "" if text_val is None else str(text_val).strip(),
            "ROLES": "" if roles_val is None else str(roles_val).strip(),
            "STATUS": "" if status_val is None else str(status_val).strip(),
            "MSG": "" if msg_val is None else str(msg_val).strip(),
            "TIMESTEMP": "" if ts_val is None else str(ts_val).strip(),
        })

    if not records:
        wb.close()
        print("⚠️ Não encontrei linhas para processar.")
        return

    roles_map = {}
    for rec in records:
        status_norm = norm_txt(rec["STATUS"])
        if status_norm == "CONCLUIDO":
            continue

        agr = rec["AGR_NAME"].strip()
        if not agr:
            continue

        if agr not in roles_map:
            roles_map[agr] = {
                "AGR_NAME": agr,
                "TEXT": rec["TEXT"].strip(),
                "ROLES_LIST": []
            }

        if not roles_map[agr]["TEXT"] and rec["TEXT"].strip():
            roles_map[agr]["TEXT"] = rec["TEXT"].strip()

        roles_map[agr]["ROLES_LIST"].extend(split_roles(rec["ROLES"]))

    if not roles_map:
        wb.close()
        print("⚠️ Nada para processar (tudo CONCLUIDO).")
        return

    roles_agrupadas = []
    for item in roles_map.values():
        item["ROLES_LIST"] = list(dict.fromkeys(item["ROLES_LIST"]))
        roles_agrupadas.append(item)

    roles_agrupadas.sort(key=lambda x: x["AGR_NAME"])

    SISTEMA_ESPERADO = MAPA_SISTEMA.get(str(ambiente_cockpit).upper().strip() or "", None)
    if not SISTEMA_ESPERADO:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    try:
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = SapGuiAuto.GetScriptingEngine
        session = next((sess for conn in application.Children for sess in conn.Children if sess.Info.SystemName.upper() == SISTEMA_ESPERADO), None)
    except:
        session = None

    if not session:
        wb.close()
        print(f"❌ Não encontrei sessão do ambiente '{ambiente_cockpit}'.")
        return

    if not request_transporte and not modo_nao_interativo:
        print("\n============================================================")
        print("🚚 Opções de configuração de Transporte.\n")
        print("   1 - Escreva o número da Request")
        print("   2 - Criar nova ordem de transporte")
        print("   3 - Pesquisar suas request criadas.")
        print("   4 - Prima [Enter] vazio para NÃO transportar")
        print("============================================================")

        while True:
            req_input = input("\n👉 Opção: ").strip()
            if req_input in ("1", "2", "3", "4", ""):
                if req_input == "":
                    req_input = "4"
                break
            print("❌ Opção inválida. Use 1, 2, 3, 4 ou apenas pressione Enter.")

        if req_input == "1":
            request_transporte = input("🔢 Numero da Request (ex: S4QK900396): ").strip().upper()
        elif req_input == "2":
            request_transporte = _criar_nova_request_no_sap_local(session)
        elif req_input == "3":
            try:
                import pesquisar_request
                print("\n🔍 A abrir nova sessão em segundo plano para pesquisar (SE16H)...")
                resultados_pesquisa = pesquisar_request.listar_requests(
                    system_name=SISTEMA_ESPERADO,
                    include_requests=True,
                    use_new_mode=True,
                    minimize=False,
                    close_after=True,
                    session=session
                )
                if resultados_pesquisa:
                    escolha = input("\n👉 Digite o número (N) da Request que deseja utilizar (ou Enter para cancelar): ").strip()
                    if escolha.isdigit() and 1 <= int(escolha) <= len(resultados_pesquisa):
                        request_transporte = resultados_pesquisa[int(escolha) - 1][0]
                        print(f"✔️ Selecionou a Request: {request_transporte}")
                    else:
                        print("❌ Seleção cancelada. Não haverá transporte.")
                else:
                    print("⚠️ Não foram encontradas Requests abertas para o seu utilizador.")
            except ImportError as e:
                print(f"❌ Erro de Importação: {e}")

        elif req_input == "4":
            print("⏭️  Nenhuma request selecionada (Transporte ignorado).")
            request_transporte = None

    if pedir_confirmacao and not modo_nao_interativo:
        if input("\nDeseja lançar esses dados no SAP? [S/N]: ").strip().upper() != "S":
            wb.close()
            return

    grupos_para_exec = []
    for rr in roles_agrupadas:
        grupos_para_exec.append({
            "agr_name_composta": rr["AGR_NAME"],
            "text_composta": rr["TEXT"],
            "componentes": rr["ROLES_LIST"]
        })

    resultados = executar_grupos_compostos(
        ambiente_cockpit=ambiente_cockpit,
        grupos_compostos=grupos_para_exec,
        caminho_ficheiro=caminho_ficheiro,
        request_transporte=request_transporte,
        modo_nao_interativo=modo_nao_interativo,
        pedir_confirmacao=False,
        origem="PFCG_COMPOSTA",
        sess_externa=session
    )

    try:
        col_st, col_ms, col_tm = header_map.get("STATUS"), header_map.get("MSG"), header_map.get("TIMESTEMP")
        for rec in records:
            res = resultados.get(str(rec["AGR_NAME"]).strip())
            if res:
                if col_st:
                    ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
                if col_ms:
                    ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
                if col_tm:
                    ws.cell(row=rec["_row"], column=col_tm).value = now_ts()
        wb.save(caminho_ficheiro)
        wb.close()
        print("\n💾 Resultados gravados com sucesso no Excel!")
    except Exception as e:
        print(f"\n❌ Erro a gravar Excel final: {e}")
        try:
            wb.close()
        except:
            pass

    tempo_decorrido_total = time.time() - tempo_inicio_total
    print(f"\n⏱️ Tempo total da operação: {formatar_tempo(tempo_decorrido_total)}")
    print("🔁 Fim.")
    return True

def selecionar_ficheiro():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = filedialog.askopenfilename(
        title=f"Selecione o ficheiro Excel (sheet '{NOME_SHEET}')",
        filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*"))
    )
    root.destroy()
    return path

def _criar_nova_request_no_sap_local(sess):
    okcd = _safe_find("wnd[0]/tbar[0]/okcd")
    if okcd:
        okcd.text = "/nSE10"
        sess.findById("wnd[0]").sendVKey(0)
        time.sleep(0.8)

    print("\nTipo da ordem:")
    print("1 - Ordem customizing")
    print("2 - Ordem workbench")

    while True:
        tipo = input("Digite a opção (1/2): ").strip()
        if tipo in ("1", "2"):
            break
        print("❌ Opção inválida. Use apenas 1 ou 2.")

    desc = input("Descrição da request (máx 60): ").strip()
    desc = desc[:60] if desc else "REQUEST CRIADA VIA SCRIPT"

    sess.findById("wnd[0]/tbar[1]/btn[6]").press()
    time.sleep(0.4)

    if tipo == "2":
        try:
            sess.findById("wnd[1]/usr/radKO042-REQ_CONS_K").select()
        except:
            pass

    sess.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(0.4)

    try:
        sess.findById("wnd[1]/usr/txtKO013-AS4TEXT").text = desc
    except:
        pass
    sess.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(0.6)

    trkorr = None
    for sap_id in ["wnd[0]/usr/lbl[20,9]", "wnd[0]/usr/lbl[1,1]"]:
        try:
            txt = sess.findById(sap_id).Text
            match = re.search(r"\b[A-Z0-9]{3,4}K\d{6,}\b", txt)
            if match:
                trkorr = match.group(0)
        except:
            pass
        if trkorr:
            break

    if okcd:
        okcd.text = "/n"
        sess.findById("wnd[0]").sendVKey(0)

    tipo_txt = "Customizing" if tipo == "1" else "Workbench"
    print("\n✔️ Request criada.")
    print(f"Tipo: {tipo_txt} | Descrição: {desc}")

    if not trkorr:
        trkorr = input("Não consegui extrair a request automaticamente. Cole aqui: ").strip().upper()

    print(f"Request: {trkorr}")
    return trkorr

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--ambiente", choices=["DEV", "QAD", "PRD"])
    parser.add_argument("--xlsx")
    parser.add_argument("--request", help="Número da Request de Transporte (Opcional)")
    parser.add_argument("--auto", action="store_true")
    parser.add_argument("--no-confirm", action="store_true")
    args = parser.parse_args()

    env_cli = args.ambiente or (input("Ambiente (DEV/QAD/PRD): ").strip().upper() or "DEV")

    executar(
        ambiente_cockpit=env_cli,
        caminho_ficheiro=args.xlsx,
        request_transporte=args.request,
        modo_nao_interativo=bool(args.auto),
        pedir_confirmacao=(not args.no_confirm)
    )
