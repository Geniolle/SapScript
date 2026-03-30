# -*- coding: utf-8 -*-

###################################################################################
# BLOCO: ELIMINAÇÃO EM MASSA DE FUNÇÕES VIA /NPFCGMASSDELETE
# Mantém a formatação do Excel intacta escrevendo célula a célula via openpyxl.
# Com barra de progresso Rich.
###################################################################################
def executar(
    ambiente_cockpit,
    pfcg_object,       # <-- Obrigatório: Cockpit deteta a aba pelo nome do script
    caminho_ficheiro,  # <-- Obrigatório: Cockpit abre a janela do Windows para o Excel
    request_transporte=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    **kwargs
):
    import os, time, sys, re
    import unicodedata
    from datetime import datetime
    import win32com.client
    import pyperclip
    from openpyxl import load_workbook
    from rich.progress import Progress, BarColumn, TextColumn, TimeElapsedColumn

    # --- CORREÇÃO DA ESTRUTURA DE PASTAS PARA IMPORTAR O PESQUISAR_REQUEST ---
    dir_atual = os.path.dirname(os.path.abspath(__file__))
    dir_processos = os.path.dirname(dir_atual)
    if dir_processos not in sys.path:
        sys.path.insert(0, dir_processos)

    ###################################################################################
    # CONFIG INICIAL E LOGGING
    ###################################################################################
    tempo_inicio = time.time()
    mapa_sistema = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P", "CUA": "SPA"}
    sistema_desejado = mapa_sistema.get(ambiente_cockpit)

    NOME_SHEET = pfcg_object if pfcg_object else "PFCG_DELETE"
    SEARCH_HEADER_IN_FIRST_ROWS = 20

    TIMEOUT_SAP_BUSY = 180.0

    COL_ID        = "ID"
    COL_AGR_NAME  = "AGR_NAME"
    COL_TEXT      = "TEXT"
    COL_STATUS    = "STATUS"
    COL_MSG       = "MSG"
    COL_TIMESTAMP = "TIMESTEMP"

    COLUNAS_OBRIGATORIAS = {COL_ID, COL_AGR_NAME, COL_STATUS, COL_MSG, COL_TIMESTAMP}

    def agora_ts():
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def log(msg):
        print(f"{agora_ts()} | {msg}", flush=True)

    ###################################################################################
    # HELPERS: NORMALIZAÇÃO
    ###################################################################################
    def norm_col(s):
        if s is None:
            return ""
        return unicodedata.normalize("NFKD", str(s)).encode("ASCII", "ignore").decode("utf-8").strip().upper()

    def traduzir_nome_coluna(s):
        s = norm_col(s)
        if s in ["NOME FUNCAO", "NOME FUNÇÂO", "NOME FUNÇAO"]:
            return COL_AGR_NAME
        if s in ["DESCRITIVO", "DESCRITIVO FUNCAO", "DESCRICAO", "DESCRIÇÃO"]:
            return COL_TEXT
        if s == "TIMESTAMP":
            return COL_TIMESTAMP
        return s

    ###################################################################################
    # HELPERS: SAP
    ###################################################################################
    def existe(session, obj_id):
        try:
            session.findById(obj_id)
            return True
        except:
            return False

    def esperar_sap_livre(session, timeout=120.0, pausa=0.2):
        limite = time.time() + timeout
        while time.time() < limite:
            try:
                busy = bool(getattr(session, "Busy", False))
            except:
                busy = False
            if not busy:
                return True
            time.sleep(pausa)
        return False

    def esperar_janela(session, wnd_idx, timeout=10.0, pausa=0.2):
        limite = time.time() + timeout
        while time.time() < limite:
            if existe(session, f"wnd[{wnd_idx}]"):
                return True
            time.sleep(pausa)
        return False

    def fechar_popups(session, timeout=10.0, pausa=0.2, prefer_yes=True):
        limite = time.time() + timeout
        while time.time() < limite:
            if any(existe(session, f"wnd[{i}]") for i in (1, 2, 3)):
                break
            time.sleep(pausa)

        fechou = False
        for _ in range(40):
            algum = False
            for i in (3, 2, 1):
                if not existe(session, f"wnd[{i}]"):
                    continue
                algum = True
                try:
                    if existe(session, f"wnd[{i}]/tbar[0]/btn[0]"):
                        session.findById(f"wnd[{i}]/tbar[0]/btn[0]").press()
                    elif prefer_yes and existe(session, f"wnd[{i}]/usr/btnSPOP-OPTION1"):
                        session.findById(f"wnd[{i}]/usr/btnSPOP-OPTION1").press()
                    elif existe(session, f"wnd[{i}]/usr/btnSPOP-OPTION2"):
                        session.findById(f"wnd[{i}]/usr/btnSPOP-OPTION2").press()
                    else:
                        session.findById(f"wnd[{i}]").sendVKey(0)
                    fechou = True
                except:
                    pass
                time.sleep(pausa)
            if not algum:
                break
        return fechou

    def mensagem_sem_resultado(msg):
        m = (msg or "").lower()
        return (
            ("nenhum" in m or "nenhuma" in m or "nenhumas" in m)
            and ("funç" in m or "regist" in m or "obj") and ("encontrad" in m)
        )

    ###################################################################################
    # LEITURA DO EXCEL VIA OPENPYXL (PRESERVA A FORMATAÇÃO)
    ###################################################################################
    if not caminho_ficheiro or not os.path.exists(caminho_ficheiro):
        log(f"❌ Ficheiro Excel não encontrado: '{caminho_ficheiro}'.")
        return "voltar"

    try:
        wb = load_workbook(caminho_ficheiro, data_only=False)
    except Exception as e:
        log(f"❌ Erro ao abrir o ficheiro Excel: {e}")
        return "voltar"

    if NOME_SHEET in wb.sheetnames:
        ws = wb[NOME_SHEET]
    else:
        log(f"❌ Aba '{NOME_SHEET}' não encontrada no Excel.")
        log(f"💡 Abas disponíveis: {', '.join(wb.sheetnames)}")
        wb.close()
        return "voltar"

    log(f"📑 Aba (Sheet) lida com sucesso: '{NOME_SHEET}'")

    # Localizar o cabeçalho
    header_row = None
    header_map = {}
    for r in range(1, SEARCH_HEADER_IN_FIRST_ROWS + 1):
        row_vals = [traduzir_nome_coluna(c.value) for c in ws[r]]
        colunas_encontradas = set(row_vals).intersection(COLUNAS_OBRIGATORIAS)

        if len(colunas_encontradas) >= len(COLUNAS_OBRIGATORIAS):
            header_row = r
            for idx, name in enumerate(row_vals, start=1):
                if name:
                    header_map[name] = idx
            break

    if not header_row:
        wb.close()
        print(f"\n❌ Não encontrei as colunas obrigatórias nas primeiras {SEARCH_HEADER_IN_FIRST_ROWS} linhas.")
        print(f"   Esperado: {', '.join(COLUNAS_OBRIGATORIAS)}")
        return "voltar"

    # Extrair os dados para processamento
    records = []

    def get_cell(row_idx, col_name):
        if col_name not in header_map:
            return ""
        v = ws.cell(row=row_idx, column=header_map[col_name]).value
        return "" if v is None else str(v).strip()

    for r in range(header_row + 1, ws.max_row + 1):
        agr = get_cell(r, COL_AGR_NAME)
        status = get_cell(r, COL_STATUS).upper()

        # Filtra logo linhas vazias ou já concluídas
        if not agr or status == "CONCLUÍDO" or status == "CONCLUIDO":
            continue

        records.append({
            "_row": r,
            COL_AGR_NAME: agr
        })

    if not records:
        wb.close()
        log("⚠️ Nenhuma linha válida pendente encontrada na aba.")
        return "voltar"

    # Extrair lista de roles e copiar para o Clipboard
    funcoes = [rec[COL_AGR_NAME] for rec in records]
    pyperclip.copy("\r\n".join(funcoes))

    ###################################################################################
    # CAPTURA SESSÃO SAP
    ###################################################################################
    try:
        log("🔌 A localizar sessão SAP...")
        SapGuiAuto = win32com.client.GetObject("SAPGUI")
        app = SapGuiAuto.GetScriptingEngine
        session = next(
            (sess for conn in app.Children for sess in conn.Children if sess.Info.SystemName.upper() == (sistema_desejado or "")),
            None
        )
    except Exception:
        session = None

    if not session:
        log(f"❌ Sessão SAP não encontrada para '{ambiente_cockpit}' (esperado: {sistema_desejado}).")
        wb.close()
        return "voltar"

    ###################################################################################
    # LÓGICA DA REQUEST DE TRANSPORTE
    ###################################################################################
    def _criar_nova_request_no_sap_local():
        try:
            session.findById("wnd[0]/tbar[0]/okcd").text = "/nSE10"
            session.findById("wnd[0]").sendVKey(0)
            time.sleep(0.8)

            print("\nTipo da ordem:")
            print("1 - Ordem customizing")
            print("2 - Ordem workbench")
            while True:
                tipo = input("Digite a opção (1/2): ").strip()
                if tipo in ("1", "2"):
                    break
                print("❌ Opção inválida. Use apenas 1 ou 2.")

            desc = input("Descrição da request (máx 60): ").strip()
            desc = desc[:60] if desc else "REQUEST CRIADA VIA SCRIPT"

            session.findById("wnd[0]/tbar[1]/btn[6]").press()
            time.sleep(0.4)

            if tipo == "2":
                try:
                    session.findById("wnd[1]/usr/radKO042-REQ_CONS_K").select()
                except:
                    pass

            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(0.4)

            try:
                session.findById("wnd[1]/usr/txtKO013-AS4TEXT").text = desc
            except:
                pass

            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(0.6)

            trkorr = None
            for sap_id in ["wnd[0]/usr/lbl[20,9]", "wnd[0]/usr/lbl[1,1]"]:
                try:
                    txt = session.findById(sap_id).Text
                    match = re.search(r"\b[A-Z0-9]{3,4}K\d{6,}\b", txt)
                    if match:
                        trkorr = match.group(0)
                except:
                    pass
                if trkorr:
                    break

            session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
            session.findById("wnd[0]").sendVKey(0)

            print(f"\n✅ Request criada: {trkorr}")
            return trkorr
        except Exception as e:
            print(f"❌ Falha ao criar request: {e}")
            return None

    if not request_transporte and not modo_nao_interativo:
        print("\n============================================================")
        print("🚚 Opções de configuração de Transporte.\n")
        print("   1 - Escreva o número da Request")
        print("   2 - Criar nova ordem de transporte")
        print("   3 - Pesquisar suas request criadas.")
        print("   4 - Prima [Enter] vazio para NÃO transportar")
        print("============================================================")

        while True:
            req_input = input("\n👉 Opção: ").strip()
            if req_input in ("1", "2", "3", "4", ""):
                if req_input == "":
                    req_input = "4"
                break
            print("❌ Opção inválida. Use 1, 2, 3, 4 ou apenas pressione Enter.")

        if req_input == "1":
            request_transporte = input("🔢 Numero da Request (ex: S4QK900396): ").strip().upper()

        elif req_input == "2":
            request_transporte = _criar_nova_request_no_sap_local()

        elif req_input == "3":
            try:
                import pesquisar_request
                print("\n🔍 A abrir nova sessão em segundo plano para pesquisar (SE16H)...")
                resultados_pesquisa = pesquisar_request.listar_requests(
                    system_name=sistema_desejado,
                    include_requests=True,
                    use_new_mode=True,
                    minimize=True,
                    close_after=True
                )
                if resultados_pesquisa:
                    escolha = input("\n👉 Digite o número (N) da Request que deseja utilizar (ou Enter para cancelar): ").strip()
                    if escolha.isdigit() and 1 <= int(escolha) <= len(resultados_pesquisa):
                        request_transporte = resultados_pesquisa[int(escolha) - 1][0]
                        print(f"✅ Selecionou a Request: {request_transporte}")
                    else:
                        print("❌ Seleção cancelada. Não haverá transporte.")
                else:
                    print("⚠️ Não foram encontradas Requests abertas.")
            except Exception as e:
                print(f"❌ Erro na pesquisa: {e}")

        elif req_input == "4":
            print("⏭️  Nenhuma request selecionada (Transporte ignorado).")
            request_transporte = None

        print("============================================================")

    log(f"📂 Ficheiro: {caminho_ficheiro}")
    log(f"📋 AGR_NAME a eliminar ({len(funcoes)}):")
    for i, n in enumerate(funcoes, 1):
        print(f" {i:02d}. {n}", flush=True)

    if not modo_nao_interativo and pedir_confirmacao:
        if input("\nDeseja eliminar essas funções no SAP? [S/N]: ").strip().upper() != "S":
            log("❌ Processo cancelado.")
            wb.close()
            return "voltar"

    ###################################################################################
    # EXECUÇÃO NO SAP (Eliminação em Massa via PFCGMASSDELETE)
    ###################################################################################
    status_geral = "ERRO"
    msg_final = "Erro desconhecido."

    try:
        session.findById("wnd[0]").maximize()
        if not esperar_sap_livre(session, timeout=TIMEOUT_SAP_BUSY):
            raise RuntimeError("SAP bloqueado antes de iniciar.")

        log("➡️ A abrir /NPFCGMASSDELETE ...")
        session.findById("wnd[0]/tbar[0]/okcd").text = "/NPFCGMASSDELETE"
        session.findById("wnd[0]").sendVKey(0)

        if existe(session, "wnd[0]/usr/radMOD_EXE"):
            session.findById("wnd[0]/usr/radMOD_EXE").select()

        # Abrir lista múltipla e colar as funções
        log("🧾 A carregar lista de funções...")
        session.findById("wnd[0]/usr/btn%_ROLE_%_APP_%-VALU_PUSH").press()
        time.sleep(0.5)
        session.findById("wnd[1]").sendVKey(24)  # Shift+F12 (Colar)
        time.sleep(0.3)
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        time.sleep(0.3)

        log("▶️ Executar eliminação...")
        session.findById("wnd[0]/tbar[1]/btn[8]").press()

        # Gestão de Popups
        timeout = time.time() + 15.0
        while time.time() < timeout:
            time.sleep(0.5)
            if existe(session, "wnd[1]/usr/ctxtKO008-TRKORR"):
                if request_transporte:
                    log(f"🚚 A injetar Request de Transporte: {request_transporte}")
                    session.findById("wnd[1]/usr/ctxtKO008-TRKORR").text = request_transporte
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                continue
            if existe(session, "wnd[1]/usr/btnSPOP-OPTION1"):
                session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
                continue
            if existe(session, "wnd[1]/tbar[0]/btn[0]"):
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                continue
            if existe(session, "wnd[0]/usr/cntlGRID1/shellcont/shell"):
                break

        # Validar sucesso no ALV ou na Barra de Status
        msg_alv = ""
        try:
            grid = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell")
            if grid.RowCount > 0:
                for col in ["MESSAGE", "TEXT", "MSG"]:
                    try:
                        v = str(grid.GetCellValue(0, col)).strip()
                        if v:
                            msg_alv = v
                            break
                    except:
                        pass
        except:
            pass

        msg_barra = ""
        try:
            msg_barra = session.findById("wnd[0]/sbar").Text.strip()
        except:
            pass

        msg_final = msg_alv or msg_barra or "Execução concluída (Verificada no Log ALV)"
        msg_transporte = f" [Req: {request_transporte}]" if request_transporte else ""

        # Voltar ao ecrã base
        try:
            if existe(session, "wnd[0]/tbar[0]/btn[3]"):
                session.findById("wnd[0]/tbar[0]/btn[3]").press()
            session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
            session.findById("wnd[0]").sendVKey(0)
        except:
            pass

        if mensagem_sem_resultado(msg_final):
            status_geral = "ERRO"
            msg_final = f"{msg_final} - SAP não encontrou as roles informadas."
            log(f"❌ SAP: {msg_final}")
        else:
            status_geral = "CONCLUÍDO"
            msg_final = f"{msg_final}{msg_transporte}"
            log(f"✅ SAP status final: {msg_final}")

    except Exception as e:
        status_geral = "ERRO"
        msg_final = f"Erro no processo SAP: {e}"
        log(f"❌ Erro crítico no SAP: {e}")
        try:
            session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
            session.findById("wnd[0]").sendVKey(0)
        except:
            pass

    ###################################################################################
    # GRAVAÇÃO CÉLULA A CÉLULA VIA OPENPYXL (Com barra de progresso)
    ###################################################################################
    log("💾 A gravar resultados no Excel preservando formatações...")
    try:
        col_st = header_map.get(COL_STATUS)
        col_ms = header_map.get(COL_MSG)
        col_tm = header_map.get(COL_TIMESTAMP)

        ts_final = agora_ts()

        with Progress(
            TextColumn("[bold cyan]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TextColumn("({task.completed}/{task.total})"),
            TimeElapsedColumn(),
            transient=False,
        ) as progress:
            task_excel = progress.add_task("A atualizar Excel...", total=len(records))

            for rec in records:
                linha_excel = rec["_row"]
                nome_role = rec[COL_AGR_NAME]

                progress.update(task_excel, description=f"A atualizar Excel: {nome_role}")

                if col_st:
                    ws.cell(row=linha_excel, column=col_st).value = status_geral
                if col_ms:
                    ws.cell(row=linha_excel, column=col_ms).value = msg_final
                if col_tm:
                    ws.cell(row=linha_excel, column=col_tm).value = ts_final

                progress.advance(task_excel)

        wb.save(caminho_ficheiro)
        print(f"✅ Ficheiro atualizado com os resultados na aba '{NOME_SHEET}'.")
    except Exception as e:
        print(f"❌ Erro ao salvar o ficheiro: {e}")
        print("⚠️ Verifica se o ficheiro está aberto e bloqueado no Excel.")
    finally:
        wb.close()

    ###################################################################################
    # TEMPO TOTAL
    ###################################################################################
    mm, ss = divmod(int(time.time() - tempo_inicio), 60)
    log(f"⏱️ Tempo total: {mm:02d}:{ss:02d}")
    return "voltar"