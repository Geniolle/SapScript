# -*- coding: utf-8 -*-
from __future__ import annotations

import functools
import os
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

print = functools.partial(print, flush=True)

MAPA_SISTEMA = {"DEV": "S4D", "QAD": "S4Q", "PRD": "S4P"}
NOME_SHEET = "CRIAR_CODIGO_IVA"
HEADER_SEARCH_LIMIT = 30
SAP_GUI_PENDING_MESSAGE = (
    "Integração SAP GUI FTXP/OB40 pendente: gravações SAP GUI Scripting não disponíveis."
)

COLUNAS_OBRIGATORIAS = {
    "ACTION",
    "COUNTRY",
    "VAT_CODE",
    "DESCRIPTION_PT",
    "RATE",
    "VAT_TYPE",
    "TAX_TYPE",
    "STATUS",
    "MSG",
    "TIMESTEMP",
}

COLUNAS_OPCIONAIS = {
    "DESCRIPTION_EN",
    "DESCRIPTION_ES",
    "OPERATION_DEBIT",
    "GL_ACCOUNT_DEBIT",
    "OPERATION_CREDIT",
    "GL_ACCOUNT_CREDIT",
    "REPORTING_COUNTRY",
}

COLUNAS_RESULTADO = {"STATUS", "MSG", "TIMESTEMP"}
VAT_OPERACOES_CONHECIDAS = ("VST", "MWS", "ESE", "ESA", "NAV", "NVV")


class ValidacaoGrupo:
    def __init__(self, ok: bool, status: str, msg: str, dados: dict[str, Any]):
        self.ok = ok
        self.status = status
        self.msg = msg
        self.dados = dados


class CodigoIvaSapGuiUnavailable(RuntimeError):
    pass


class CodigoIvaSapGuiClient:
    def __init__(self, session, request_ctx: dict[str, str] | None = None, *, recordings_available: bool = False):
        self.session = session
        self.request_ctx = request_ctx or {}
        self.recordings_available = recordings_available

    def _garantir_disponibilidade(self) -> None:
        if not self.recordings_available:
            raise CodigoIvaSapGuiUnavailable(SAP_GUI_PENDING_MESSAGE)

    def ping(self) -> bool:
        self._garantir_disponibilidade()
        return True

    def consultar_codigo(self, payload: dict[str, Any]) -> dict[str, Any]:
        self._garantir_disponibilidade()
        return {}

    def criar_codigo(self, payload: dict[str, Any]) -> dict[str, Any]:
        self._garantir_disponibilidade()
        return {}

    def consultar_conta(self, payload: dict[str, Any]) -> dict[str, Any]:
        self._garantir_disponibilidade()
        return {}

    def configurar_conta(self, payload: dict[str, Any]) -> dict[str, Any]:
        self._garantir_disponibilidade()
        return {}


def obter_sessao_sap(ambiente_cockpit: str):
    ambiente_up = str(ambiente_cockpit or "").strip().upper()
    sistema_esperado = MAPA_SISTEMA.get(ambiente_up)
    if not sistema_esperado:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    try:
        import win32com.client  # type: ignore
    except Exception as exc:  # pragma: no cover - dependência Windows
        raise RuntimeError("SAP GUI Scripting não está disponível neste ambiente.") from exc

    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application = sap_gui_auto.GetScriptingEngine
    except Exception as exc:
        raise RuntimeError("Não foi possível obter o motor de scripting do SAP GUI.") from exc

    for conn in getattr(application, "Children", []):
        for sess in getattr(conn, "Children", []):
            try:
                if str(sess.Info.SystemName).strip().upper() == sistema_esperado:
                    return sess
            except Exception:
                continue

    raise RuntimeError(f"Não foi encontrada sessão SAP aberta para o ambiente '{ambiente_up}'.")


def obter_mensagem_sbar(session) -> dict[str, str]:
    try:
        sbar = session.findById("wnd[0]/sbar")
        return {
            "tipo": str(getattr(sbar, "MessageType", "") or "").strip().upper(),
            "texto": str(getattr(sbar, "Text", "") or "").strip(),
            "codigo": str(getattr(sbar, "MessageId", "") or "").strip(),
            "numero": str(getattr(sbar, "MessageNumber", "") or "").strip(),
        }
    except Exception:
        return {"tipo": "", "texto": "", "codigo": "", "numero": ""}


def classificar_mensagem_sap(tipo: str, texto: str) -> str:
    tipo_norm = str(tipo or "").strip().upper()
    texto_norm = str(texto or "").strip()
    if tipo_norm in {"E", "A", "X"}:
        return f"{tipo_norm} - {texto_norm}" if texto_norm else tipo_norm
    if tipo_norm == "W":
        return f"Aviso SAP: {texto_norm}" if texto_norm else "Aviso SAP"
    return texto_norm


def existe_objeto(session, obj_id: str) -> bool:
    try:
        session.findById(obj_id)
        return True
    except Exception:
        return False


def aguardar_sap_livre(session, timeout: float = 30.0, pausa: float = 0.2) -> bool:
    limite = time.time() + timeout
    while time.time() < limite:
        try:
            if not bool(getattr(session, "Busy", False)):
                return True
        except Exception:
            return True
        time.sleep(pausa)
    return False


def aguardar_elemento(session, obj_id: str, timeout: float = 10.0, pausa: float = 0.2) -> bool:
    limite = time.time() + timeout
    while time.time() < limite:
        if existe_objeto(session, obj_id):
            return True
        time.sleep(pausa)
    return False


def aguardar_janela(session, wnd_idx: int, timeout: float = 10.0, pausa: float = 0.2) -> bool:
    return aguardar_elemento(session, f"wnd[{wnd_idx}]", timeout=timeout, pausa=pausa)


def aguardar_desaparecer_janela(session, wnd_idx: int, timeout: float = 10.0, pausa: float = 0.2) -> bool:
    limite = time.time() + timeout
    while time.time() < limite:
        if not existe_objeto(session, f"wnd[{wnd_idx}]"):
            return True
        time.sleep(pausa)
    return False


def obter_objeto(session, obj_id: str):
    try:
        return session.findById(obj_id)
    except Exception:
        return None


def preencher_texto(session, obj_id: str, valor: Any, caret_pos: int | None = None) -> bool:
    obj = obter_objeto(session, obj_id)
    if obj is None:
        return False
    obj.text = "" if valor is None else str(valor)
    if caret_pos is not None:
        try:
            obj.caretPosition = caret_pos
        except Exception:
            pass
    return True


def pressionar_botao(session, obj_id: str) -> bool:
    obj = obter_objeto(session, obj_id)
    if obj is None:
        return False
    obj.press()
    return True


def enviar_vkey(session, wnd_id: str, vkey: int) -> bool:
    wnd = obter_objeto(session, wnd_id)
    if wnd is None:
        return False
    wnd.sendVKey(vkey)
    return True


def selecionar_checkbox_ou_radio(session, obj_id: str) -> bool:
    obj = obter_objeto(session, obj_id)
    if obj is None:
        return False
    try:
        obj.select()
        return True
    except Exception:
        try:
            obj.Selected = True
            return True
        except Exception:
            return False


def ler_celula_grelha(grelha, linha: int, coluna: str) -> str:
    for metodo in ("GetCellValue", "getCellValue"):
        if hasattr(grelha, metodo):
            try:
                return str(getattr(grelha, metodo)(linha, coluna) or "").strip()
            except Exception:
                pass
    return ""


def localizar_linha_por_chave(grelha, coluna_chave: str, chave: str, max_linhas: int | None = None) -> int | None:
    try:
        row_count = int(getattr(grelha, "RowCount", 0) or 0)
    except Exception:
        row_count = 0
    limite = row_count if max_linhas is None or max_linhas <= 0 else min(max_linhas, row_count)
    chave_norm = str(chave or "").strip().upper()
    for idx in range(limite):
        valor = ler_celula_grelha(grelha, idx, coluna_chave)
        if valor and valor.strip().upper() == chave_norm:
            return idx
    return None


def fechar_popups_inesperados(session, timeout: float = 5.0, pausa: float = 0.2) -> bool:
    fechou = False
    limite = time.time() + timeout
    while time.time() < limite:
        encontrou = False
        for idx in (3, 2, 1):
            wnd_id = f"wnd[{idx}]"
            if not existe_objeto(session, wnd_id):
                continue
            encontrou = True
            if existe_objeto(session, f"{wnd_id}/tbar[0]/btn[0]"):
                try:
                    pressionar_botao(session, f"{wnd_id}/tbar[0]/btn[0]")
                    fechou = True
                    continue
                except Exception:
                    pass
            if existe_objeto(session, f"{wnd_id}/usr/btnSPOP-OPTION1"):
                try:
                    pressionar_botao(session, f"{wnd_id}/usr/btnSPOP-OPTION1")
                    fechou = True
                    continue
                except Exception:
                    pass
            if existe_objeto(session, f"{wnd_id}/usr/btnSPOP-OPTION2"):
                try:
                    pressionar_botao(session, f"{wnd_id}/usr/btnSPOP-OPTION2")
                    fechou = True
                    continue
                except Exception:
                    pass
            try:
                enviar_vkey(session, wnd_id, 0)
                fechou = True
            except Exception:
                pass
        if not encontrou:
            break
        time.sleep(pausa)
    return fechou


def regressar_com_seguranca_a_inicio(session) -> None:
    try:
        fechar_popups_inesperados(session)
    except Exception:
        pass
    try:
        preencher_texto(session, "wnd[0]/tbar[0]/okcd", "/n")
        enviar_vkey(session, "wnd[0]", 0)
    except Exception:
        pass


def formatar_tempo(segundos: float) -> str:
    segundos = max(0, int(round(segundos)))
    minutos, resto = divmod(segundos, 60)
    horas, minutos = divmod(minutos, 60)
    if horas:
        return f"{horas:02d}h {minutos:02d}m {resto:02d}s"
    return f"{minutos:02d}m {resto:02d}s"


def limpar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, str):
        texto = valor
    elif isinstance(valor, float) and valor.is_integer():
        texto = str(int(valor))
    else:
        texto = str(valor)
    texto = texto.replace("\xa0", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def valor_em_texto_preservado(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, str):
        return valor.strip()
    if isinstance(valor, bool):
        return "X" if valor else ""
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        texto = format(valor, "f").rstrip("0").rstrip(".")
        return texto
    return str(valor).strip()


def parse_rate(valor: Any) -> float:
    if valor is None or valor == "":
        raise ValueError("RATE deve estar preenchido.")
    if isinstance(valor, bool):
        raise ValueError("RATE inválido.")
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(" ", "")
    if not texto:
        raise ValueError("RATE deve estar preenchido.")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except Exception as exc:
        raise ValueError(f"RATE inválido: '{valor}'.") from exc


def agora_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalizar_cabecalho(valor: Any) -> str:
    texto = limpar_texto(valor)
    texto = texto.replace("\r", " ").replace("\n", " ")
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto).strip().upper()
    return texto


def _sheet_names(workbook) -> list[str]:
    return list(getattr(workbook, "sheetnames", []) or [])


def _normalizar_nome_folha(valor: Any) -> str:
    return normalizar_cabecalho(valor).replace(" ", "")


def _resolver_nome_folha(workbook, sheet_name: str) -> str:
    sheet_names = _sheet_names(workbook)
    if sheet_name in sheet_names:
        return sheet_name
    alvo = _normalizar_nome_folha(sheet_name)
    for nome in sheet_names:
        if _normalizar_nome_folha(nome) == alvo:
            return nome
    raise ValueError(
        f"Sheet '{sheet_name}' não encontrada. Sheets disponíveis: {', '.join(sheet_names) if sheet_names else 'nenhuma'}."
    )


def localizar_folha_cabecalho(workbook, sheet_name: str = NOME_SHEET):
    nome_folha = _resolver_nome_folha(workbook, sheet_name)
    ws = workbook[nome_folha]
    melhor_row = None
    melhor_map: dict[str, int] = {}
    melhor_match_count = -1
    for row_idx in range(1, min(getattr(ws, "max_row", 0), HEADER_SEARCH_LIMIT) + 1):
        header_map: dict[str, int] = {}
        for cell in ws[row_idx]:
            valor = normalizar_cabecalho(cell.value)
            if valor:
                header_map[valor] = cell.column
        match_count = len(COLUNAS_OBRIGATORIAS.intersection(header_map))
        if match_count > melhor_match_count:
            melhor_row = row_idx
            melhor_map = header_map
            melhor_match_count = match_count
        if COLUNAS_OBRIGATORIAS.issubset(header_map):
            return ws, row_idx, header_map
    missing = sorted(COLUNAS_OBRIGATORIAS.difference(melhor_map))
    if missing:
        raise ValueError(
            f"Cabeçalho não encontrado nas primeiras {HEADER_SEARCH_LIMIT} linhas da sheet '{nome_folha}'. "
            f"Colunas em falta: {', '.join(missing)}."
        )
    raise ValueError(
        f"Cabeçalho não encontrado nas primeiras {HEADER_SEARCH_LIMIT} linhas da sheet '{nome_folha}'."
    )


def ler_registros(ws, header_row: int, header_map: dict[str, int]) -> list[dict[str, Any]]:
    registros: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = ws[row_idx]
        reg: dict[str, Any] = {"_row": row_idx}
        vazio_total = True
        for header, col_idx in header_map.items():
            valor = ws.cell(row=row_idx, column=col_idx).value
            reg[header] = valor
            if limpar_texto(valor) != "":
                vazio_total = False
        if vazio_total:
            continue
        reg["RATE_RAW"] = reg.get("RATE")
        action = limpar_texto(reg.get("ACTION")).upper()
        status = limpar_texto(reg.get("STATUS")).upper()
        vat_code = limpar_texto(reg.get("VAT_CODE"))
        if action not in {"CRIAR", "VALIDAR"} or status == "CONCLUIDO" or not vat_code:
            continue
        registros.append(reg)
    return registros


def agrupar_registros(registros: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    grupos: dict[tuple[str, str], dict[str, Any]] = {}
    for reg in registros:
        action = limpar_texto(reg.get("ACTION")).upper()
        status = limpar_texto(reg.get("STATUS")).upper()
        country = limpar_texto(reg.get("COUNTRY"))
        vat_code = limpar_texto(reg.get("VAT_CODE"))
        if action not in {"CRIAR", "VALIDAR"} or status == "CONCLUIDO" or not vat_code:
            continue
        key = (country.upper(), vat_code.upper())
        grupo = grupos.setdefault(
            key,
            {
                "COUNTRY": country,
                "VAT_CODE": vat_code,
                "rows": [],
            },
        )
        grupo["rows"].append(reg)
    return [grupos[key] for key in sorted(grupos.keys())]


def _coalesce_unico(rows: list[dict[str, Any]], campo: str) -> str:
    valores: list[str] = []
    for reg in rows:
        texto = limpar_texto(reg.get(campo))
        if texto and texto not in valores:
            valores.append(texto)
    return valores[0] if valores else ""


def _valor_rate_registo(reg: dict[str, Any]) -> Any:
    if reg.get("RATE_RAW") not in (None, ""):
        return reg.get("RATE_RAW")
    return reg.get("RATE")


def _validar_pares_operacao_conta(rows: list[dict[str, Any]]) -> tuple[list[dict[str, str]], str]:
    pares: list[dict[str, str]] = []
    for operacao_chave, conta_chave, direction in (
        ("OPERATION_DEBIT", "GL_ACCOUNT_DEBIT", "DEBIT"),
        ("OPERATION_CREDIT", "GL_ACCOUNT_CREDIT", "CREDIT"),
    ):
        operacao = limpar_texto(_coalesce_unico(rows, operacao_chave))
        conta = valor_em_texto_preservado(_coalesce_unico(rows, conta_chave))
        if operacao and not conta:
            return [], f"Uma operação informada deve possuir a respetiva conta quando a determinação de conta for necessária ({operacao_chave})."
        if conta and not operacao:
            return [], f"Uma operação informada deve possuir a respetiva conta quando a determinação de conta for necessária ({conta_chave})."
        if operacao and conta:
            pares.append({"OPERATION": operacao.upper(), "GL_ACCOUNT": conta, "DIRECTION": direction})
    return pares, ""


def validar_grupo(grupo: dict[str, Any]) -> ValidacaoGrupo:
    rows = grupo["rows"]
    country = limpar_texto(grupo.get("COUNTRY"))
    vat_code = limpar_texto(grupo.get("VAT_CODE"))

    def fail(msg: str, status: str = "ERRO") -> ValidacaoGrupo:
        return ValidacaoGrupo(False, status, msg, {})

    if len(country) != 2:
        return fail("Código país inválido: deve possuir exatamente dois caracteres.")
    if len(vat_code) != 2:
        return fail("Código IVA inválido: deve possuir exatamente dois caracteres.")
    if "?" in vat_code:
        return fail("Código IVA inválido: '?' não é permitido.")

    actions = {limpar_texto(r.get("ACTION")).upper() for r in rows if limpar_texto(r.get("ACTION"))}
    if len(actions) != 1:
        return fail("Linhas do mesmo país/código têm ACTION diferente.")
    action = next(iter(actions))
    if action not in {"CRIAR", "VALIDAR"}:
        return fail("ACTION inválida: use CRIAR ou VALIDAR.")

    campos_texto = [
        "DESCRIPTION_PT",
        "DESCRIPTION_EN",
        "DESCRIPTION_ES",
        "VAT_TYPE",
        "TAX_TYPE",
        "REPORTING_COUNTRY",
        "OPERATION_DEBIT",
        "GL_ACCOUNT_DEBIT",
        "OPERATION_CREDIT",
        "GL_ACCOUNT_CREDIT",
    ]
    for campo in campos_texto:
        distintos = []
        for reg in rows:
            valor = limpar_texto(reg.get(campo))
            if valor and valor not in distintos:
                distintos.append(valor)
        if len(distintos) > 1:
            return fail(f"Campo '{campo}' divergente dentro do mesmo país/código.")

    desc_pt = _coalesce_unico(rows, "DESCRIPTION_PT")
    desc_en = _coalesce_unico(rows, "DESCRIPTION_EN")
    desc_es = _coalesce_unico(rows, "DESCRIPTION_ES")
    vat_type = _coalesce_unico(rows, "VAT_TYPE")
    tax_type = _coalesce_unico(rows, "TAX_TYPE")
    reporting_country = _coalesce_unico(rows, "REPORTING_COUNTRY")

    if not desc_pt:
        return fail("DESCRIPTION_PT é obrigatória.")
    if len(desc_pt) > 50:
        return fail("DESCRIPTION_PT não pode ultrapassar 50 caracteres.")
    if desc_en and len(desc_en) > 50:
        return fail("DESCRIPTION_EN não pode ultrapassar 50 caracteres.")
    if desc_es and len(desc_es) > 50:
        return fail("DESCRIPTION_ES não pode ultrapassar 50 caracteres.")
    if not tax_type:
        return fail("TAX_TYPE deve estar preenchido.")
    if reporting_country and len(reporting_country) != 2:
        return fail("REPORTING_COUNTRY deve possuir exatamente dois caracteres.")

    rate_raw = None
    for reg in rows:
        if _valor_rate_registo(reg) not in (None, ""):
            rate_raw = _valor_rate_registo(reg)
            break
    try:
        rate = parse_rate(rate_raw)
    except Exception as exc:
        return fail(str(exc))
    if rate < 0 or rate > 100:
        return fail("RATE deve estar entre 0 e 100.")

    pares_ob40, msg_ob40 = _validar_pares_operacao_conta(rows)
    if msg_ob40:
        return fail(msg_ob40)

    dados = {
        "action": action,
        "country": country,
        "vat_code": vat_code,
        "description_pt": desc_pt,
        "description_en": desc_en,
        "description_es": desc_es,
        "rate": rate,
        "vat_type": vat_type,
        "tax_type": tax_type,
        "reporting_country": reporting_country,
        "ob40_pairs": pares_ob40,
        "rows": rows,
    }
    return ValidacaoGrupo(True, "OK", "Validação concluída.", dados)


def _resultado_para_linhas(grupo: dict[str, Any], status: str, msg: str) -> dict[str, dict[str, str]]:
    resultado: dict[str, dict[str, str]] = {}
    ts = agora_ts()
    for reg in grupo["rows"]:
        resultado[str(reg["_row"])] = {"STATUS": status, "MSG": msg, "TIMESTEMP": ts}
    return resultado


def atualizar_resultados_em_memoria(grupos: list[dict[str, Any]], resultados: dict[str, dict[str, str]]) -> None:
    for grupo in grupos:
        for reg in grupo["rows"]:
            row_key = str(reg["_row"])
            if row_key not in resultados:
                resultados[row_key] = {"STATUS": "", "MSG": "", "TIMESTEMP": ""}


def gravar_resultados_excel(caminho_ficheiro, sheet_name, header_map, records, resultados):
    col_st = header_map.get("STATUS")
    col_ms = header_map.get("MSG")
    col_tm = header_map.get("TIMESTEMP")

    pythoncom = None
    excel_app = None
    wb_excel = None
    try:
        import pythoncom as _pythoncom  # type: ignore
        import win32com.client  # type: ignore

        pythoncom = _pythoncom
        pythoncom.CoInitialize()
        try:
            excel_app = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        wb_excel = excel_app.Workbooks.Open(os.path.abspath(caminho_ficheiro))
        ws_excel = wb_excel.Worksheets(sheet_name)
        for rec in records:
            res = resultados.get(str(rec["_row"]))
            if not res:
                continue
            if col_st:
                ws_excel.Cells(rec["_row"], col_st).Value = res["STATUS"]
            if col_ms:
                ws_excel.Cells(rec["_row"], col_ms).Value = res["MSG"]
            if col_tm:
                ws_excel.Cells(rec["_row"], col_tm).Value = res["TIMESTEMP"]
        wb_excel.Save()
        return True
    except Exception:
        pass
    finally:
        try:
            if wb_excel is not None:
                wb_excel.Close(SaveChanges=True)
        except Exception:
            pass
        try:
            if excel_app is not None:
                excel_app.Quit()
        except Exception:
            pass
        try:
            if pythoncom is not None:
                pythoncom.CoUninitialize()
        except Exception:
            pass

    try:
        from openpyxl import load_workbook

        wb = load_workbook(caminho_ficheiro)
        ws = wb[sheet_name]
        for rec in records:
            res = resultados.get(str(rec["_row"]))
            if not res:
                continue
            if col_st:
                ws.cell(row=rec["_row"], column=col_st).value = res["STATUS"]
            if col_ms:
                ws.cell(row=rec["_row"], column=col_ms).value = res["MSG"]
            if col_tm:
                ws.cell(row=rec["_row"], column=col_tm).value = res["TIMESTEMP"]
        wb.save(caminho_ficheiro)
        wb.close()
        return True
    except Exception:
        return False


def _normalizar_request_ctx(request_ctx: dict[str, Any] | None) -> dict[str, str]:
    if not isinstance(request_ctx, dict):
        return {"request_number": "", "request_option": "", "request_desc": ""}
    return {
        "request_number": limpar_texto(request_ctx.get("request_number")).upper(),
        "request_option": limpar_texto(request_ctx.get("request_option")),
        "request_desc": limpar_texto(request_ctx.get("request_desc")),
    }


def _exigir_request_para_criar(request_ctx: dict[str, str], ambiente_cockpit: str) -> dict[str, str]:
    numero = limpar_texto(request_ctx.get("request_number")).upper()
    opcao = limpar_texto(request_ctx.get("request_option"))
    ambiente_up = str(ambiente_cockpit or "").strip().upper()
    if ambiente_up != "DEV":
        raise ValueError("A criação de códigos IVA está limitada ao ambiente DEV neste processo.")
    if opcao == "4":
        raise ValueError("Execução sem transporte não é permitida para ACTION = CRIAR.")
    if not numero:
        raise ValueError("É obrigatória uma request de Customizing para criar o código IVA.")
    if opcao and opcao not in {"1", "2"}:
        raise ValueError("Apenas request de Customizing é aceite para este processo.")
    return {
        "request_number": numero,
        "request_option": opcao or "1",
        "request_desc": limpar_texto(request_ctx.get("request_desc")),
    }


def _confirmar_execucao(
    ambiente_cockpit: str,
    request_ctx: dict[str, str],
    grupos: list[dict[str, Any]],
    total_ob40: int,
) -> bool:
    paises = sorted({g["COUNTRY"] for g in grupos})
    print("\nResumo antes da execução")
    print(f"Ambiente: {ambiente_cockpit}")
    print(f"Request: {request_ctx['request_number'] or 'N/A'}")
    print(f"Quantidade de códigos: {len(grupos)}")
    print(f"Países encontrados: {', '.join(paises) if paises else 'N/A'}")
    print(f"Quantidade de operações OB40: {total_ob40}")
    resposta = input("\nConfirmar execução global? [S/N]: ").strip().upper()
    return resposta == "S"


def _payload_grupo(grupo: dict[str, Any]) -> dict[str, Any]:
    rows = grupo["rows"]
    return {
        "action": limpar_texto(rows[0].get("ACTION")).upper(),
        "country": grupo["COUNTRY"],
        "vat_code": grupo["VAT_CODE"],
        "description_pt": _coalesce_unico(rows, "DESCRIPTION_PT"),
        "description_en": _coalesce_unico(rows, "DESCRIPTION_EN"),
        "description_es": _coalesce_unico(rows, "DESCRIPTION_ES"),
        "rate": next((r for r in (_valor_rate_registo(reg) for reg in rows) if r not in (None, "")), ""),
        "vat_type": _coalesce_unico(rows, "VAT_TYPE"),
        "tax_type": _coalesce_unico(rows, "TAX_TYPE"),
        "reporting_country": _coalesce_unico(rows, "REPORTING_COUNTRY"),
        "ob40_pairs": grupo.get("ob40_pairs", []),
    }


def _normalizar_texto_rate(valor: Any) -> float:
    return parse_rate(valor)


def _rate_para_msg(valor: Any) -> str:
    try:
        rate = _normalizar_texto_rate(valor)
        if float(rate).is_integer():
            return str(int(rate))
        return str(rate).rstrip("0").rstrip(".")
    except Exception:
        return valor_em_texto_preservado(valor)


def _comparar_listas_dicionarios(excel_itens: list[dict[str, str]], sap_itens: list[dict[str, str]], chaves: tuple[str, ...]) -> list[str]:
    def _norm(itens: list[dict[str, str]]) -> list[tuple[str, ...]]:
        normalizados = []
        for item in itens:
            normalizados.append(tuple(limpar_texto(item.get(chave)).upper() for chave in chaves))
        return sorted(normalizados)

    a = _norm(excel_itens)
    b = _norm(sap_itens)
    if a == b:
        return []
    return [f"Lista diferente para {', '.join(chaves)}."]


def _comparar_configuracao_com_excel(grupo: dict[str, Any], consulta: dict[str, Any]) -> list[str]:
    rows = grupo["rows"]
    difs: list[str] = []

    def _campo_excel(nome: str) -> str:
        return _coalesce_unico(rows, nome)

    def _campo_sap(nome: str) -> str:
        return limpar_texto(consulta.get(nome))

    if limpar_texto(consulta.get("country")).upper() != grupo["COUNTRY"].upper():
        difs.append(f"País diferente: SAP={consulta.get('country')}; Excel={grupo['COUNTRY']}.")
    if limpar_texto(consulta.get("vat_code")).upper() != grupo["VAT_CODE"].upper():
        difs.append(f"Código diferente: SAP={consulta.get('vat_code')}; Excel={grupo['VAT_CODE']}.")
    if _campo_sap("description_pt") != _campo_excel("DESCRIPTION_PT"):
        difs.append(f"Descrição diferente: SAP={consulta.get('description_pt')}; Excel={_campo_excel('DESCRIPTION_PT')}.")
    if _campo_sap("vat_type") != _campo_excel("VAT_TYPE"):
        difs.append(f"Tipo de imposto diferente: SAP={consulta.get('vat_type')}; Excel={_campo_excel('VAT_TYPE')}.")
    if _campo_sap("tax_type") != _campo_excel("TAX_TYPE"):
        difs.append(f"Tax type diferente: SAP={consulta.get('tax_type')}; Excel={_campo_excel('TAX_TYPE')}.")

    try:
        rate_sap = float(consulta.get("rate")) if consulta.get("rate") not in (None, "") else None
    except Exception:
        rate_sap = None
    try:
        rate_excel = parse_rate(_valor_rate_registo(rows[0]))
    except Exception:
        rate_excel = None
    if rate_sap is not None and rate_excel is not None and abs(rate_sap - rate_excel) > 1e-9:
        difs.append(f"Código existente com taxa diferente: SAP={_rate_para_msg(rate_sap)}; Excel={_rate_para_msg(rate_excel)}.")

    cond_sap = consulta.get("conditions") or []
    cond_excel = _coalesce_unico(rows, "VAT_TYPE")
    if cond_sap:
        keys_sap = sorted(limpar_texto(item.get("condition_key")).upper() for item in cond_sap if limpar_texto(item.get("condition_key")))
        if cond_excel and cond_excel.upper() not in keys_sap and len(keys_sap) > 0:
            difs.append(
                f"Condição fiscal {cond_excel} não encontrada no procedimento do país {grupo['COUNTRY']}."
            )

    ob40_excel = grupo.get("ob40_pairs") or []
    contas_sap = consulta.get("accounts") or []
    if ob40_excel and contas_sap:
        contas_sap_map = {
            (limpar_texto(item.get("operation")).upper(), limpar_texto(item.get("direction")).upper()): valor_em_texto_preservado(item.get("gl_account"))
            for item in contas_sap
        }
        for pair in ob40_excel:
            chave = (limpar_texto(pair.get("OPERATION")).upper(), limpar_texto(pair.get("DIRECTION")).upper())
            conta_sap = contas_sap_map.get(chave)
            if conta_sap is not None and conta_sap != valor_em_texto_preservado(pair.get("GL_ACCOUNT")):
                difs.append(
                    f"Determinação de conta diferente para {pair.get('OPERATION')}: SAP={conta_sap}; Excel={pair.get('GL_ACCOUNT')}."
                )
    return difs


def _resultado_sap_para_linhas(
    grupo: dict[str, Any],
    status: str,
    msg: str,
) -> dict[str, dict[str, str]]:
    return _resultado_para_linhas(grupo, status, msg)


def _processar_consulta(
    cliente: CodigoIvaSapGuiClient,
    grupo: dict[str, Any],
    dados: dict[str, Any],
) -> tuple[str, str, dict[str, Any]]:
    consulta = cliente.consultar_codigo(dados)
    if not isinstance(consulta, dict):
        consulta = {}
    if not consulta.get("exists"):
        return "IGNORADO", "Código IVA não encontrado.", consulta
    difs = _comparar_configuracao_com_excel(grupo, consulta)
    if difs:
        return "ERRO", " ".join(difs), consulta
    return "CONCLUIDO", "Código IVA já existia com configuração equivalente. Nenhuma alteração efetuada.", consulta


def _processar_ob40(
    cliente: CodigoIvaSapGuiClient,
    grupo: dict[str, Any],
    dados: dict[str, Any],
) -> tuple[bool, str]:
    pares = dados.get("ob40_pairs") or grupo.get("ob40_pairs") or []
    if not pares:
        return True, ""
    for pair in pares:
        consulta = cliente.consultar_conta(pair)
        if not isinstance(consulta, dict):
            consulta = {}
        if consulta.get("exists") and consulta.get("equivalent") is True:
            continue
        if consulta.get("exists") and consulta.get("equivalent") is False:
            conta_sap = valor_em_texto_preservado(consulta.get("gl_account") or consulta.get("current_account"))
            return False, (
                f"Determinação de conta diferente para {pair.get('OPERATION')}: "
                f"SAP={conta_sap}; Excel={pair.get('GL_ACCOUNT')}."
            )
        resposta = cliente.configurar_conta(pair)
        if not isinstance(resposta, dict):
            resposta = {}
        if resposta.get("success") is False:
            mensagem = limpar_texto(resposta.get("message")) or "Falha ao configurar conta OB40."
            return False, mensagem
    return True, ""


def _processar_criacao(
    cliente: CodigoIvaSapGuiClient,
    grupo: dict[str, Any],
    dados: dict[str, Any],
    request_ctx: dict[str, str],
) -> tuple[str, str]:
    grupo_ctx = {**grupo, "ob40_pairs": dados.get("ob40_pairs", [])}
    consulta = cliente.consultar_codigo(dados)
    if not isinstance(consulta, dict):
        consulta = {}
    if consulta.get("exists"):
        difs = _comparar_configuracao_com_excel(grupo, consulta)
        if difs:
            return "ERRO", " ".join(difs)
        return "CONCLUIDO", "Código IVA já existia com configuração equivalente. Nenhuma alteração efetuada."

    resposta_criacao = cliente.criar_codigo({**dados, "request": request_ctx.get("request_number", "")})
    if not isinstance(resposta_criacao, dict):
        resposta_criacao = {}
    if resposta_criacao.get("success") is False:
        mensagem = limpar_texto(resposta_criacao.get("message")) or "Erro ao criar código IVA."
        return "ERRO", mensagem

    consulta_final = cliente.consultar_codigo(dados)
    if not isinstance(consulta_final, dict) or not consulta_final.get("exists"):
        return "ERRO", "Código IVA criado, mas a releitura final não confirmou a gravação."
    difs = _comparar_configuracao_com_excel(grupo, consulta_final)
    if difs:
        return "ERRO", " ".join(difs)

    ok_ob40, msg_ob40 = _processar_ob40(cliente, grupo_ctx, dados)
    if not ok_ob40:
        return "ERRO", f"Código IVA criado/configurado na FTXP; determinação de contas pendente: {msg_ob40}"

    if grupo_ctx.get("ob40_pairs"):
        return "CONCLUIDO", "Código IVA criado com sucesso na FTXP e contas configuradas na OB40."
    return "CONCLUIDO", "Código IVA criado com sucesso na FTXP."


def _aplicar_resultado_grupo(
    grupo: dict[str, Any],
    resultados: dict[str, dict[str, str]],
    status: str,
    msg: str,
) -> None:
    resultados.update(_resultado_sap_para_linhas(grupo, status, msg))


def _resumo_totais(resultados: dict[str, dict[str, str]]) -> dict[str, int]:
    total_concluido = sum(1 for r in resultados.values() if r["STATUS"] == "CONCLUIDO")
    total_erro = sum(1 for r in resultados.values() if r["STATUS"] == "ERRO")
    total_ignorado = sum(1 for r in resultados.values() if r["STATUS"] == "IGNORADO")
    return {
        "CONCLUIDO": total_concluido,
        "ERRO": total_erro,
        "IGNORADO": total_ignorado,
    }


def executar(
    ambiente_cockpit,
    request_ctx,
    caminho_ficheiro=None,
    modo_nao_interativo=False,
    pedir_confirmacao=True,
    nome_pasta=None,
):
    from openpyxl import load_workbook

    tempo_inicio_total = time.time()
    ambiente = str(ambiente_cockpit or "").strip().upper()
    sistema_esperado = MAPA_SISTEMA.get(ambiente)
    if not sistema_esperado:
        raise ValueError(f"Ambiente inválido: '{ambiente_cockpit}'. Use DEV, QAD ou PRD.")

    request_ctx_normalizado = _normalizar_request_ctx(request_ctx)

    if not caminho_ficheiro:
        if modo_nao_interativo:
            raise ValueError("Faltou o parâmetro caminho_ficheiro em modo não-interativo.")
        try:
            import tkinter as tk
            from tkinter import filedialog

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            caminho_ficheiro = filedialog.askopenfilename(
                title="Selecione o ficheiro Excel",
                filetypes=(("Ficheiros Excel", "*.xlsx"), ("Todos os ficheiros", "*.*")),
            )
            root.destroy()
        except Exception as exc:
            raise RuntimeError(f"Falha ao selecionar ficheiro Excel: {exc}") from exc
        if not caminho_ficheiro:
            raise ValueError("Operação cancelada: ficheiro não selecionado.")

    caminho_ficheiro = os.path.abspath(caminho_ficheiro)
    if not os.path.exists(caminho_ficheiro):
        raise FileNotFoundError(f"Ficheiro não encontrado: {caminho_ficheiro}")

    wb = load_workbook(caminho_ficheiro)
    try:
        ws, header_row, header_map = localizar_folha_cabecalho(wb, NOME_SHEET)
        registros = ler_registros(ws, header_row, header_map)
        grupos = agrupar_registros(registros)
        if not grupos:
            print("Nenhuma linha elegível para processamento.")
            return False

        total_linhas = len(registros)
        total_ob40 = 0
        for grupo in grupos:
            validacao = validar_grupo(grupo)
            if validacao.ok:
                total_ob40 += len(validacao.dados.get("ob40_pairs", []))

        if pedir_confirmacao and not modo_nao_interativo:
            request_para_confirmar = request_ctx_normalizado if request_ctx_normalizado else {"request_number": "", "request_option": "", "request_desc": ""}
            if not _confirmar_execucao(ambiente, request_para_confirmar, grupos, total_ob40):
                print("Execução cancelada pelo utilizador.")
                return False

        print(f"Total de linhas lidas: {total_linhas}")
        print(f"Total de códigos agrupados: {len(grupos)}")
        print(f"Quantidade de operações OB40: {total_ob40}")

        resultados: dict[str, dict[str, str]] = {}
        total_concluido = 0
        total_erro = 0
        total_existente = 0
        total_ignorado = 0

        try:
            session = obter_sessao_sap(ambiente)
        except Exception as exc:
            session = None
            erro_sessao = str(exc)
            for grupo in grupos:
                resultados.update(_resultado_para_linhas(grupo, "ERRO", erro_sessao))
            gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, registros, resultados)
            print(f"Total concluído: 0")
            print(f"Total já existente: 0")
            print(f"Total com erro: {len(grupos)}")
            print(f"Total ignorado: 0")
            print(f"Tempo total: {formatar_tempo(time.time() - tempo_inicio_total)}")
            return False

        cliente = CodigoIvaSapGuiClient(session=session, request_ctx=request_ctx_normalizado)

        for idx, grupo in enumerate(grupos, start=1):
            tempo_inicio_grupo = time.time()
            chave = f"{grupo['COUNTRY']} | {grupo['VAT_CODE']}"
            print(f"[{idx}/{len(grupos)}] {chave}")

            try:
                validacao = validar_grupo(grupo)
                if not validacao.ok:
                    _aplicar_resultado_grupo(grupo, resultados, "ERRO", validacao.msg)
                    total_erro += 1
                    print(f"  [VALIDAÇÃO] ERRO: {validacao.msg}")
                    print("  [RESULTADO] ERRO")
                    continue

                print("  [VALIDAÇÃO] OK")
                dados = validacao.dados
                action = dados["action"]

                if action == "VALIDAR":
                    try:
                        status, msg, consulta = _processar_consulta(cliente, grupo, dados)
                        _aplicar_resultado_grupo(grupo, resultados, status, msg)
                        if status == "CONCLUIDO":
                            total_concluido += 1
                            total_existente += 1
                        elif status == "IGNORADO":
                            total_ignorado += 1
                        else:
                            total_erro += 1
                        print(f"  [SAP] {msg}")
                        print(f"  [RESULTADO] {status}")
                    except CodigoIvaSapGuiUnavailable as exc:
                        msg = str(exc)
                        _aplicar_resultado_grupo(grupo, resultados, "ERRO", msg)
                        total_erro += 1
                        print(f"  [ERRO] {msg}")
                        print("  [RESULTADO] ERRO")
                    continue

                try:
                    request_para_criar = _exigir_request_para_criar(request_ctx_normalizado, ambiente)
                except Exception as exc:
                    msg = str(exc)
                    _aplicar_resultado_grupo(grupo, resultados, "ERRO", msg)
                    total_erro += 1
                    print(f"  [REQUEST] ERRO: {msg}")
                    print("  [RESULTADO] ERRO")
                    continue

                try:
                    status, msg = _processar_criacao(cliente, grupo, dados, request_para_criar)
                    _aplicar_resultado_grupo(grupo, resultados, status, msg)
                    if status == "CONCLUIDO":
                        total_concluido += 1
                    elif status == "IGNORADO":
                        total_ignorado += 1
                    else:
                        total_erro += 1
                    if "já existia" in msg.lower():
                        total_existente += 1
                    print(f"  [SAP] {msg}")
                    print(f"  [RESULTADO] {status}")
                except CodigoIvaSapGuiUnavailable as exc:
                    msg = str(exc)
                    _aplicar_resultado_grupo(grupo, resultados, "ERRO", msg)
                    total_erro += 1
                    print(f"  [ERRO] {msg}")
                    print("  [RESULTADO] ERRO")
                except Exception as exc:
                    msg = str(exc)
                    _aplicar_resultado_grupo(grupo, resultados, "ERRO", msg)
                    total_erro += 1
                    print(f"  [ERRO] {msg}")
                    print("  [RESULTADO] ERRO")
            finally:
                try:
                    gravar_resultados_excel(caminho_ficheiro, NOME_SHEET, header_map, registros, resultados)
                except Exception:
                    pass
                print(f"  [TEMPO] {formatar_tempo(time.time() - tempo_inicio_grupo)}")

        resumo = _resumo_totais(resultados)
        print(f"Total concluído: {total_concluido}")
        print(f"Total já existente: {total_existente}")
        print(f"Total com erro: {total_erro}")
        print(f"Total ignorado: {total_ignorado}")
        print(f"Tempo total: {formatar_tempo(time.time() - tempo_inicio_total)}")

        if nome_pasta:
            print(f"Evidência funcional desativada neste subprocesso: {nome_pasta}")

        return resumo["ERRO"] == 0
    finally:
        try:
            wb.close()
        except Exception:
            pass


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Processo de códigos IVA.")
    parser.add_argument("ambiente", nargs="?", default="DEV")
    parser.add_argument("xlsx", nargs="?")
    args = parser.parse_args()
    executar(args.ambiente, request_ctx={}, caminho_ficheiro=args.xlsx)
