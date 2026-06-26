"""
Centralized JIRA attachment service for the SapScript project.

Handles: listing, streaming download, cache, sanitization, extension/size
validation, manifest creation, text extraction, and normalized Excel copy.

Used by:
  - sap_agent/jira_client.py  (agent analysis — Windows worker)
  - web_api/jira_client.py    (fetch_ticket_details — Docker/Linux)
  - web_api/main.py           (chat context enrichment)

Design rules:
  - Heavy deps (PIL, winocr, PyPDF2, openpyxl, extract_msg) are imported
    lazily so the module is importable on Docker/Linux without crashing.
  - The original downloaded file is NEVER modified.
  - A separate normalized copy is created for Excel files when requested.
  - Per-attachment errors are recorded in the manifest without interrupting
    the processing of other attachments.
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Environment-driven configuration
# ---------------------------------------------------------------------------

def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except ValueError:
        return default


def _env_bool(name: str, default: bool) -> bool:
    return os.getenv(name, "true" if default else "false").strip().lower() in (
        "1", "true", "yes", "sim"
    )


def _env_extensions(name: str, default: str) -> set[str]:
    raw = os.getenv(name, default).strip()
    return {ext.strip().lower() for ext in raw.split(",") if ext.strip()}


MAX_SIZE_BYTES: int = _env_int("JIRA_ATTACHMENT_MAX_SIZE_MB", 10) * 1024 * 1024
MAX_COUNT: int = _env_int("JIRA_ATTACHMENT_MAX_COUNT", 10)
ALLOWED_EXTENSIONS: set[str] = _env_extensions(
    "JIRA_ATTACHMENT_ALLOWED_EXTENSIONS",
    ".xlsx,.xlsm,.pdf,.png,.jpg,.jpeg,.gif,.msg,.eml,.txt,.csv,.json,.xml,.log",
)
CACHE_BASE_DIR: str = os.getenv("JIRA_ATTACHMENT_CACHE_DIR", "/data/jira").strip()
OCR_PROVIDER: str = os.getenv("JIRA_ATTACHMENT_OCR_PROVIDER", "windows_ocr").strip().lower()
EXTRACT_XLSX_TEXT: bool = _env_bool("JIRA_ATTACHMENT_EXTRACT_XLSX_TEXT", True)

# Maximum characters sent to the LLM per attachment
ATTACHMENT_TEXT_MAX_CHARS: int = 4000


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class AttachmentMeta:
    attachment_id: str
    filename: str
    size: int
    created: str
    mime_type: str
    content_url: str


@dataclass
class AttachmentResult:
    attachment_id: str
    filename: str
    size: int
    created: str
    extension: str
    original_path: str
    normalized_path: str = ""
    text: str = ""
    text_truncated: bool = False
    error: str = ""
    skipped: bool = False
    skip_reason: str = ""


# ---------------------------------------------------------------------------
# Filename sanitization
# ---------------------------------------------------------------------------

def sanitize_filename(filename: str) -> str:
    """
    Returns a filesystem-safe filename.

    - Normalizes Unicode (é→e, ã→a) via NFKD + ASCII drop
    - Removes characters forbidden in Windows/Linux filenames
    - Separates stem (max 80 chars) and suffix (max 20 chars)
    - Never returns an empty string or path-traversal sequence
    """
    if not filename or not filename.strip():
        return "anexo_sem_nome"

    try:
        normalized = unicodedata.normalize("NFKD", filename.strip())
        ascii_name = normalized.encode("ascii", "ignore").decode("ascii")
    except Exception:
        ascii_name = filename.strip()

    # Forbidden on Windows and ambiguous on Linux: < > : " / \ | ? * and control chars
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", ascii_name)
    # Collapse ".." sequences to prevent path traversal
    safe = re.sub(r'\.{2,}', ".", safe)
    safe = safe.strip(". ")

    if not safe:
        return "anexo_sem_nome"

    p = Path(safe)
    stem = p.stem[:80].strip(". ") or "anexo"
    suffix = (p.suffix or "")[:20]
    return stem + suffix


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_extension(filename: str, allowed: set[str] | None = None) -> bool:
    """Returns True if the file extension is in the allowed set (case-insensitive)."""
    ext = Path(filename.lower()).suffix
    return ext in (allowed if allowed is not None else ALLOWED_EXTENSIONS)


def validate_size(size_bytes: int, max_bytes: int | None = None) -> bool:
    """Returns True if size_bytes does not exceed the limit."""
    return size_bytes <= (max_bytes if max_bytes is not None else MAX_SIZE_BYTES)


# ---------------------------------------------------------------------------
# Cache & manifest
# ---------------------------------------------------------------------------

def _manifest_path(ticket_dir: Path) -> Path:
    return ticket_dir / "manifest.json"


def load_manifest(ticket_dir: Path) -> dict[str, Any]:
    """Loads the manifest JSON for a ticket directory. Returns {} on miss."""
    path = _manifest_path(ticket_dir)
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_manifest(ticket_dir: Path, manifest: dict[str, Any]) -> None:
    """Atomically persists the manifest JSON."""
    _manifest_path(ticket_dir).write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _cache_key(att: AttachmentMeta) -> str:
    return f"{att.attachment_id}|{att.filename}|{att.size}|{att.created}"


# ---------------------------------------------------------------------------
# Text extraction  (all imports are lazy — safe on Docker/Linux)
# ---------------------------------------------------------------------------

def extract_text_from_content(filename: str, content: bytes) -> str:
    """
    Extracts readable text from raw file bytes based on the file extension.
    Returns an empty string if extraction is not supported or dependencies are missing.
    """
    ext = Path(filename.lower()).suffix

    if ext in (".png", ".jpg", ".jpeg", ".gif"):
        return _ocr_image(filename, content)
    if ext == ".pdf":
        return _extract_pdf(content)
    if ext == ".msg":
        return _extract_msg(filename, content)
    if ext == ".eml":
        return _extract_eml(content)
    if ext in (".xlsx", ".xlsm", ".xls"):
        return _extract_xlsx(content) if EXTRACT_XLSX_TEXT else ""
    if ext in (".txt", ".csv", ".json", ".xml", ".log"):
        return _decode_text(content)
    return ""


def _ocr_image(filename: str, content: bytes) -> str:
    if OCR_PROVIDER != "windows_ocr":
        return ""
    try:
        from PIL import Image  # noqa: PLC0415
        import winocr  # noqa: PLC0415

        image = Image.open(io.BytesIO(content))
        result = winocr.recognize_pil_sync(image)
        return str(result.get("text") or "").strip()
    except ImportError:
        logger.debug(
            "[ATTACH] winocr/PIL indisponível (ambiente Docker/Linux) — OCR ignorado para %s",
            filename,
        )
        return ""
    except Exception as exc:
        raise RuntimeError(f"Erro OCR em {filename}: {exc}") from exc


def _extract_pdf(content: bytes) -> str:
    try:
        import PyPDF2  # noqa: PLC0415

        pdf = PyPDF2.PdfReader(io.BytesIO(content))
        pages = [page.extract_text() for page in pdf.pages if page.extract_text()]
        return "\n".join(pages).strip()
    except ImportError:
        logger.warning("[ATTACH] PyPDF2 não instalado — extração de PDF ignorada")
        return ""
    except Exception as exc:
        raise RuntimeError(f"Erro ao extrair PDF: {exc}") from exc


def _extract_msg(filename: str, content: bytes) -> str:
    try:
        import extract_msg  # noqa: PLC0415

        msg = extract_msg.Message(io.BytesIO(content))
        return f"Assunto: {msg.subject}\n\n{msg.body}".strip()
    except ImportError:
        logger.warning("[ATTACH] extract_msg não instalado — extração de MSG ignorada")
        return ""
    except Exception as exc:
        raise RuntimeError(f"Erro ao extrair MSG {filename}: {exc}") from exc


def _extract_eml(content: bytes) -> str:
    try:
        import email  # noqa: PLC0415
        from email import policy  # noqa: PLC0415

        msg = email.message_from_bytes(content, policy=policy.default)
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body += part.get_payload(decode=True).decode(errors="ignore") + "\n"
        else:
            body = msg.get_payload(decode=True).decode(errors="ignore")
        return f"Assunto: {msg['subject']}\n\n{body}".strip()
    except Exception as exc:
        raise RuntimeError(f"Erro ao extrair EML: {exc}") from exc


def _extract_xlsx(content: bytes) -> str:
    """
    Reads up to 200 rows × 30 columns per sheet for text summarization.
    Does not break on large workbooks — truncates instead.
    """
    try:
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet_texts: list[str] = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_text: list[str] = []
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if r_idx > 200:
                    rows_text.append("[... linhas truncadas]")
                    break
                vals = [str(c).strip() if c is not None else "" for c in row[:30]]
                line = " | ".join(v for v in vals if v)
                if line:
                    rows_text.append(line)
            if rows_text:
                sheet_texts.append(f"Folha: {sheet_name}\n" + "\n".join(rows_text))
        wb.close()
        return "\n\n".join(sheet_texts).strip()
    except ImportError:
        logger.warning("[ATTACH] openpyxl não instalado — extração de XLSX ignorada")
        return ""
    except Exception as exc:
        raise RuntimeError(f"Erro ao extrair XLSX: {exc}") from exc


def _decode_text(content: bytes) -> str:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            return content.decode(enc).strip()
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace").strip()


# ---------------------------------------------------------------------------
# Normalized Excel copy (original is NEVER modified)
# ---------------------------------------------------------------------------

def create_normalized_excel(original_path: Path, normalized_path: Path) -> None:
    """
    Creates a copy of an Excel file with leading spaces stripped from string cells.
    The original file is never touched.
    Falls back to a plain file copy when openpyxl is unavailable.
    """
    try:
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(str(original_path))
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith(" "):
                        cell.value = cell.value[1:]
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(normalized_path))
        wb.close()
        logger.info("[ATTACH] Cópia normalizada criada: %s", normalized_path)
    except ImportError:
        import shutil

        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(original_path), str(normalized_path))
        logger.warning(
            "[ATTACH] openpyxl indisponível — cópia normalizada idêntica ao original"
        )


# ---------------------------------------------------------------------------
# Best-XLSX selector by Jira metadata (not local mtime)
# ---------------------------------------------------------------------------

def get_best_xlsx(attachments_meta: list[dict[str, Any]]) -> dict[str, Any] | None:
    """
    Returns the most recent XLSX/XLSM attachment according to the Jira `created`
    and `id` fields — NOT the local filesystem mtime.

    Both fields are compared as strings; ISO-8601 dates sort correctly lexicographically.
    """
    xlsx_exts = {".xlsx", ".xlsm"}
    candidates = [
        att
        for att in attachments_meta
        if Path(
            sanitize_filename(str(att.get("filename") or "")).lower()
        ).suffix
        in xlsx_exts
    ]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda a: (str(a.get("created") or ""), str(a.get("id") or "")),
    )


# ---------------------------------------------------------------------------
# Streaming download
# ---------------------------------------------------------------------------

def _download_streaming(
    url: str, auth: tuple[str, str], target: Path, max_bytes: int
) -> None:
    with requests.get(url, auth=auth, stream=True, timeout=60) as r:
        r.raise_for_status()
        total = 0
        with open(target, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    total += len(chunk)
                    if total > max_bytes:
                        raise ValueError(
                            f"Anexo excede tamanho máximo "
                            f"({max_bytes // (1024 * 1024)} MB)"
                        )
                    f.write(chunk)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def process_ticket_attachments(
    ticket_key: str,
    attachments_meta: list[dict[str, Any]],
    auth: tuple[str, str],
    cache_base_dir: str | None = None,
    only_extensions: set[str] | None = None,
    make_normalized_excel: bool = False,
    overwrite: bool = False,
) -> list[AttachmentResult]:
    """
    Downloads and processes JIRA ticket attachments.

    For each attachment (up to MAX_COUNT):
      1. Skips if extension not in allowed set or size exceeds MAX_SIZE_BYTES.
      2. Uses file-based cache keyed by attachment_id|filename|size|created.
         If the cache entry contains previously extracted text, returns it
         without re-reading the file.
      3. Downloads via streaming into  <cache_dir>/<KEY>/original/<filename>.
         The original file is NEVER modified.
      4. Optionally creates a normalized copy (leading spaces stripped) in
         <cache_dir>/<KEY>/normalized/<filename>.
      5. Extracts text and caps it at ATTACHMENT_TEXT_MAX_CHARS, marking
         truncated results.
      6. Records per-attachment errors in the manifest without raising, so
         other attachments are still processed.

    Returns a list of AttachmentResult (one per entry in attachments_meta,
    including skipped ones).
    """
    base = cache_base_dir or CACHE_BASE_DIR
    allowed = only_extensions if only_extensions is not None else ALLOWED_EXTENSIONS

    key_upper = ticket_key.strip().upper()
    ticket_dir = Path(base) / key_upper
    original_dir = ticket_dir / "original"
    normalized_dir = ticket_dir / "normalized"
    original_dir.mkdir(parents=True, exist_ok=True)

    manifest = load_manifest(ticket_dir)
    manifest.setdefault("ticket_key", key_upper)
    manifest.setdefault("attachments", {})

    results: list[AttachmentResult] = []

    for att_raw in attachments_meta[:MAX_COUNT]:
        att = AttachmentMeta(
            attachment_id=str(att_raw.get("id") or ""),
            filename=sanitize_filename(str(att_raw.get("filename") or "anexo")),
            size=int(att_raw.get("size") or 0),
            created=str(att_raw.get("created") or ""),
            mime_type=str(att_raw.get("mimeType") or ""),
            content_url=str(att_raw.get("content") or ""),
        )

        ext = Path(att.filename.lower()).suffix
        result = AttachmentResult(
            attachment_id=att.attachment_id,
            filename=att.filename,
            size=att.size,
            created=att.created,
            extension=ext,
            original_path="",
        )

        if not validate_extension(att.filename, allowed):
            result.skipped = True
            result.skip_reason = f"extensão não permitida: {ext}"
            results.append(result)
            continue

        if not validate_size(att.size):
            result.skipped = True
            result.skip_reason = (
                f"tamanho excede o máximo "
                f"({att.size} bytes > {MAX_SIZE_BYTES} bytes)"
            )
            results.append(result)
            continue

        ck = _cache_key(att)
        original_path = original_dir / att.filename

        try:
            cached_entry = manifest["attachments"].get(ck, {})
            is_downloaded = original_path.exists()

            if is_downloaded and not overwrite and cached_entry and "text" in cached_entry:
                # Full cache hit — no I/O needed
                result.original_path = str(original_path)
                result.normalized_path = cached_entry.get("normalized_path", "")
                result.text = cached_entry.get("text", "")
                result.text_truncated = cached_entry.get("text_truncated", False)
                logger.debug("[ATTACH] Cache hit: %s", att.filename)
                results.append(result)
                continue

            # Download if missing
            if not is_downloaded or overwrite:
                if not att.content_url:
                    result.error = "URL de conteúdo ausente no Jira"
                    results.append(result)
                    continue
                _download_streaming(att.content_url, auth, original_path, MAX_SIZE_BYTES)
                logger.info("[ATTACH] Descarregado: %s (%d bytes)", att.filename, att.size)

            result.original_path = str(original_path)

            # Optional normalized Excel copy
            if make_normalized_excel and ext in (".xlsx", ".xlsm", ".xls"):
                norm_path = normalized_dir / att.filename
                create_normalized_excel(original_path, norm_path)
                result.normalized_path = str(norm_path)

            # Text extraction
            raw_text = extract_text_from_content(att.filename, original_path.read_bytes())

            if len(raw_text) > ATTACHMENT_TEXT_MAX_CHARS:
                result.text = raw_text[:ATTACHMENT_TEXT_MAX_CHARS]
                result.text_truncated = True
            else:
                result.text = raw_text

            # Persist to manifest (including text for future cache hits)
            manifest["attachments"][ck] = {
                "attachment_id": att.attachment_id,
                "filename": att.filename,
                "size": att.size,
                "created": att.created,
                "extension": ext,
                "original_path": result.original_path,
                "normalized_path": result.normalized_path,
                "text": result.text,
                "text_truncated": result.text_truncated,
                "error": "",
            }

        except Exception as exc:
            error_msg = str(exc)
            logger.error("[ATTACH] Erro ao processar %s: %s", att.filename, error_msg)
            result.error = error_msg
            manifest["attachments"][ck] = {
                "attachment_id": att.attachment_id,
                "filename": att.filename,
                "error": error_msg,
            }

        results.append(result)

    save_manifest(ticket_dir, manifest)
    return results
