import os
import re
from pathlib import Path

import requests
from dotenv import load_dotenv

DEFAULT_SYNC_PROJECTS = ("IT - Salsa Jeans", "SAP - Desenvolvimento")


def _load_project_env() -> None:
    """
    Carrega o .env do projeto a partir de caminhos previsíveis no host e no container.
    Isto evita depender do working directory do Docker.
    """
    module_dir = Path(__file__).resolve().parent
    project_root = module_dir.parent
    candidate_paths = [
        project_root / ".env",
        project_root.parent / ".env",
        Path("/sap-script/.env"),
        Path("/srv/sap-script-web/.env"),
    ]

    for env_path in candidate_paths:
        try:
            if env_path.is_file():
                # O runtime do worker/web API pode arrancar com variáveis vazias.
                # Forçamos o .env do projeto para garantir que o JIRA lê as credenciais corretas.
                load_dotenv(dotenv_path=env_path, override=True)
                return
        except Exception:
            continue

    load_dotenv(override=True)


_load_project_env()


def _split_env_csv(value: str) -> list[str]:
    return [item.strip() for item in (value or "").split(",") if item.strip()]


def get_jira_sync_projects() -> list[str]:
    projects = _split_env_csv(os.getenv("JIRA_SYNC_PROJECTS", ""))
    return projects or list(DEFAULT_SYNC_PROJECTS)


def build_project_scope_jql(status_clause: str) -> str:
    projects = get_jira_sync_projects()
    if not projects:
        return status_clause
    project_clause = " OR ".join(f'project = "{project}"' for project in projects)
    return f"({project_clause}) AND {status_clause}"


def build_jql_or_clause(field: str, values: list[str]) -> str:
    clean_values = [v.strip() for v in values if v and v.strip()]
    if not clean_values:
        return ""
    if len(clean_values) == 1:
        return f'{field} = "{clean_values[0]}"'
    joined = " OR ".join(f'{field} = "{value}"' for value in clean_values)
    return f"({joined})"
def _safe_filename(filename: str) -> str:
    """Sanitiza o nome do ficheiro removendo caracteres inválidos."""
    sanitized = re.sub(r'[<>:"/\\|?*]+', "_", str(filename or "").strip())
    return sanitized or "anexo_sem_nome"


def download_ticket_attachments_to_dir(
    ticket_key: str,
    output_dir_container: str,
    output_dir_windows: str,
    only_xlsx: bool = True,
    overwrite: bool = False,
) -> list[str]:
    """
    Descarrega os anexos de um ticket JIRA para uma pasta local.

    Args:
        ticket_key:           Chave do ticket (ex: 'IZ-56680')
        output_dir_container: Caminho base dentro do container (ex: '/data/jira')
        output_dir_windows:   Caminho base no Windows para o worker (ex: r'C:\\Jira')
        only_xlsx:            Se True, descarrega apenas ficheiros .xlsx
        overwrite:            Se True, sobrescreve ficheiros já existentes

    Returns:
        Lista de paths Windows dos ficheiros descarregados (.xlsx mais recente primeiro).
    """
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        print(f"[DOWNLOAD] Credenciais JIRA não configuradas para {ticket_key}.")
        return []

    auth = (jira_email, jira_token)
    headers = {"Accept": "application/json"}

    ticket_key_upper = ticket_key.strip().upper()

    # 1. Buscar anexos e sumário do ticket
    url = f"{jira_base}/{jira_api_path}/issue/{ticket_key_upper}"
    try:
        res = requests.get(url, params={"fields": "attachment,summary"}, auth=auth,
                           headers=headers, timeout=30)
        res.raise_for_status()
    except Exception as e:
        print(f"[DOWNLOAD] Erro ao obter detalhes de {ticket_key_upper}: {e}")
        return []

    data = res.json()
    summary = data.get("fields", {}).get("summary", "")
    attachments = data.get("fields", {}).get("attachment", []) or []

    if not attachments:
        print(f"[DOWNLOAD] {ticket_key_upper}: sem anexos.")
        return []

    # Sanitizar o sumário e limitar a 30 caracteres
    clean_summary = ""
    if summary:
        clean_summary = _safe_filename(summary).strip()
        clean_summary = clean_summary[:30].strip()

    # Nome da pasta: CHAVE ou CHAVE_SUMARIO
    folder_name = ticket_key_upper
    if clean_summary:
        folder_name = f"{ticket_key_upper}_{clean_summary}"

    # 2. Criar pasta no container: /data/jira/{folder_name}/
    issue_folder = Path(output_dir_container) / folder_name
    issue_folder.mkdir(parents=True, exist_ok=True)

    downloaded_windows = []

    for att in attachments:
        filename = _safe_filename(att.get("filename", "anexo"))
        if only_xlsx and not filename.lower().endswith(".xlsx"):
            continue

        content_url = att.get("content")
        if not content_url:
            continue

        target = issue_folder / filename
        if target.exists() and not overwrite:
            print(f"[DOWNLOAD] {ticket_key_upper}: já existe -> {filename}")
            # Mesmo que já exista, inclui na lista
            win_path = str(Path(output_dir_windows) / folder_name / filename)
            downloaded_windows.append(win_path)
            continue

        try:
            with requests.get(content_url, auth=auth, stream=True, timeout=60) as r:
                r.raise_for_status()
                with open(target, "wb") as f:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
            
            # Clean excel file leading spaces right after download!
            if filename.lower().endswith(".xlsx"):
                try:
                    clean_excel_leading_spaces(str(target))
                except Exception as exc:
                    print(f"[DOWNLOAD] Erro ao limpar espaços do excel: {exc}")

            print(f"[DOWNLOAD] {ticket_key_upper}: descarregado -> {filename}")
            win_path = str(Path(output_dir_windows) / folder_name / filename)
            downloaded_windows.append(win_path)
        except Exception as e:
            print(f"[DOWNLOAD] {ticket_key_upper}: erro ao descarregar {filename}: {e}")

    # Ordena por data de modificação (mais recente primeiro) dentro do container
    def _mtime(win_path: str) -> float:
        container_path = issue_folder / Path(win_path).name
        try:
            return container_path.stat().st_mtime
        except Exception:
            return 0.0

    downloaded_windows.sort(key=_mtime, reverse=True)
    return downloaded_windows


def _parse_issue(issue: dict) -> dict:
    """Converte um objeto issue da API JIRA num dict normalizado."""
    fields = issue.get("fields", {})
    assignee_data = fields.get("assignee") or {}
    raw_assignee = (
        assignee_data.get("displayName")
        or assignee_data.get("emailAddress")
        or ""
    )
    assignee_parts = raw_assignee.strip().split()
    if len(assignee_parts) > 2:
        assignee_name = f"{assignee_parts[0]} {assignee_parts[-1]}"
    else:
        assignee_name = raw_assignee

    status_data = fields.get("status") or {}
    status_name = status_data.get("name") or "Unknown"

    reporter_data = fields.get("reporter") or {}
    raw_creator = reporter_data.get("displayName") or ""
    creator_parts = raw_creator.strip().split()
    if len(creator_parts) > 2:
        creator_name = f"{creator_parts[0]} {creator_parts[-1]}"
    else:
        creator_name = raw_creator

    # Project field
    project_data = fields.get("project") or {}
    project_name = project_data.get("name") or ""

    # Priority option field
    priority_data = fields.get("customfield_15815")
    priority_name = ""
    if isinstance(priority_data, dict):
        priority_name = priority_data.get("value") or ""
    elif isinstance(priority_data, str):
        priority_name = priority_data

    # Ticket type option field
    type_data = fields.get("customfield_15810")
    type_name = ""
    if isinstance(type_data, dict):
        type_name = type_data.get("value") or ""
    elif isinstance(type_data, str):
        type_name = type_data

    # Team field (customfield_15839)
    team_data = fields.get("customfield_15839")
    team_name = ""
    if isinstance(team_data, dict):
        team_name = team_data.get("value") or ""
    elif isinstance(team_data, str):
        team_name = team_data

    # Stream field (customfield_15260)
    stream_data = fields.get("customfield_15260")
    stream_name = ""
    if isinstance(stream_data, dict):
        stream_name = stream_data.get("value") or ""
    elif isinstance(stream_data, str):
        stream_name = stream_data

    # Process field (customfield_15845)
    process_data = fields.get("customfield_15845")
    process_name = ""
    if isinstance(process_data, dict):
        process_name = process_data.get("value") or ""
    elif isinstance(process_data, str):
        process_name = process_data

    # Time to resolution field (customfield_14560)
    sla_data = fields.get("customfield_14560")
    time_to_resolution = ""
    if isinstance(sla_data, dict):
        ongoing = sla_data.get("ongoingCycle")
        if isinstance(ongoing, dict):
            breached = ongoing.get("breached", False)
            rem = ongoing.get("remainingTime")
            friendly = ""
            if isinstance(rem, dict):
                friendly = rem.get("friendly", "")
            if breached:
                if friendly:
                    time_to_resolution = friendly if friendly.startswith("-") else f"-{friendly}"
                else:
                    time_to_resolution = "Excedido"
            else:
                time_to_resolution = friendly or "Pendente"
        else:
            completed = sla_data.get("completedCycles")
            if isinstance(completed, list) and len(completed) > 0:
                last_cycle = completed[-1]
                if isinstance(last_cycle, dict):
                    breached = last_cycle.get("breached", False)
                    elapsed = last_cycle.get("elapsedTime")
                    friendly_elapsed = ""
                    if isinstance(elapsed, dict):
                        friendly_elapsed = elapsed.get("friendly", "")
                    status = "Resolvido"
                    if breached:
                        status = "Resolvido com atraso"
                    if friendly_elapsed:
                        time_to_resolution = f"{status} ({friendly_elapsed})"
                    else:
                        time_to_resolution = status

    # Supplier option field (customfield_14595)
    supplier_data = fields.get("customfield_14595")
    supplier_name = ""
    if isinstance(supplier_data, dict):
        supplier_name = supplier_data.get("value") or ""
    elif isinstance(supplier_data, str):
        supplier_name = supplier_data

    linked_keys: list[str] = []
    done_statuses = ["DONE", "CONCLU", "RESOLV", "FECHADO", "FECHADA", "CLOSED"]
    
    for link in fields.get("issuelinks", []) or []:
        # "inwardIssue" = ticket que linka ESTE; "outwardIssue" = ticket que ESTE linka
        for direction in ("inwardIssue", "outwardIssue"):
            linked_issue = link.get(direction)
            if linked_issue and linked_issue.get("key"):
                issue_fields = linked_issue.get("fields") or {}
                status_obj = issue_fields.get("status") or {}
                link_status = status_obj.get("name", "").upper()
                
                # Check if the linked issue status is in the done_statuses
                if not any(ds in link_status for ds in done_statuses):
                    linked_keys.append(linked_issue["key"])

    return {
        "key": issue.get("key"),
        "summary": fields.get("summary") or "",
        "status": status_name,
        "assignee": assignee_name,
        "created_at": fields.get("created") or "",
        "updated_at": fields.get("updated") or "",
        "priority": priority_name,
        "ticket_type": type_name,
        "creator": creator_name,
        "project": project_name,
        "team": team_name,
        "stream": stream_name,
        "process": process_name,
        "time_to_resolution": time_to_resolution,
        "supplier": supplier_name,
        "linked_keys": linked_keys,
        "resolved_at": fields.get("resolutiondate") or "",
    }


def _fetch_single_issue(
    key: str,
    jira_base: str,
    jira_api_path: str,
    auth: tuple,
    headers: dict,
) -> dict | None:
    """Busca e parseia um único issue JIRA pelo seu key."""
    url = f"{jira_base}/{jira_api_path}/issue/{key.strip().upper()}"
    params = {
        "fields": ",".join([
            "summary", "status", "assignee", "created", "updated",
            "reporter", "project", "customfield_15815", "customfield_15810",
            "customfield_15839", "customfield_15260", "customfield_15845",
            "customfield_14560", "customfield_14595", "issuelinks",
            "resolutiondate",
        ])
    }
    try:
        res = requests.get(url, auth=auth, headers=headers, params=params, timeout=15)
        res.raise_for_status()
        return _parse_issue(res.json())
    except Exception as e:
        print(f"[JIRA SYNC] Erro ao buscar ticket linkado {key}: {e}")
        return None


def fetch_jira_tickets_from_api(jql: str = None, on_page_fetched = None) -> list[dict]:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        print(
            "Jira integration credentials are not configured in environment variables."
        )
        return []

    # Query padrão: tickets relevantes dos projetos acompanhados pelo cockpit.
    # O filtro por responsável foi removido para que o browser faça o recorte final.
    if not jql:
        jql = os.getenv("JIRA_SYNC_JQL") or build_project_scope_jql("statusCategory != Done")

    url = f"{jira_base}/{jira_api_path}/search/jql"
    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    payload = {
        "jql": jql,
        "fields": [
            "summary",
            "status",
            "assignee",
            "created",
            "updated",
            "reporter",
            "project",
            "customfield_15815",
            "customfield_15810",
            "customfield_15839",
            "customfield_15260",
            "customfield_15845",
            "customfield_14560",
            "customfield_14595",
            "issuelinks",
            "resolutiondate",
        ],
        "maxResults": 100,
    }

    tickets = []
    next_page_token = None

    while True:
        if next_page_token:
            payload["nextPageToken"] = next_page_token
        else:
            payload.pop("nextPageToken", None)

        try:
            response = requests.post(
                url, auth=auth, headers=headers, json=payload, timeout=15
            )
            response.raise_for_status()
        except Exception as e:
            print(f"Jira API connection error: {e}")
            if tickets:
                break
            raise

        data = response.json()
        issues = data.get("issues", [])
        page_tickets = []

        for issue in issues:
            parsed_ticket = _parse_issue(issue)
            tickets.append(parsed_ticket)
            page_tickets.append(parsed_ticket)

        if on_page_fetched and page_tickets:
            try:
                on_page_fetched(page_tickets)
            except Exception as cb_err:
                print(f"[JIRA SYNC] Error in on_page_fetched callback: {cb_err}")

        next_page_token = data.get("nextPageToken")
        is_last = data.get("isLast", True)
        if not next_page_token or is_last:
            break

    # ─────────────────────────────────────────────────────────────────────────

    return tickets


def assign_jira_ticket(key: str, assignee_name: str) -> bool:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        print("Jira integration credentials are not configured in environment variables.")
        return False

    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    try:
        # If it's a special string like 'Sem responsável', we want to unassign (set assignee to None)
        if not assignee_name or assignee_name.lower() in ("sem responsável", "unassigned", "none", ""):
            assign_url = f"{jira_base}/{jira_api_path}/issue/{key}/assignee"
            res = requests.put(assign_url, auth=auth, headers=headers, json={"accountId": None}, timeout=15)
            return res.ok

        # Search for assignee_name to find their accountId
        search_url = f"{jira_base}/{jira_api_path}/user/search"
        params = {"query": assignee_name, "maxResults": 1}
        res = requests.get(search_url, auth=auth, headers=headers, params=params, timeout=15)
        res.raise_for_status()
        users = res.json()
        if not users:
            print(f"User {assignee_name} not found in search.")
            return False

        account_id = users[0].get("accountId")
        if not account_id:
            print("No accountId found for user.")
            return False

        # Assign issue
        assign_url = f"{jira_base}/{jira_api_path}/issue/{key}/assignee"
        res = requests.put(assign_url, auth=auth, headers=headers, json={"accountId": account_id}, timeout=15)
        res.raise_for_status()
        return True
    except Exception as e:
        print(f"Error assigning ticket {key} to {assignee_name}: {e}")
        return False


def update_jira_ticket_type(key: str, ticket_type: str) -> bool:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        print("Jira integration credentials are not configured in environment variables.")
        return False

    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    url = f"{jira_base}/{jira_api_path}/issue/{key}"
    payload = {
        "fields": {
            "customfield_15810": {
                "value": ticket_type
            } if ticket_type else None
        }
    }

    try:
        res = requests.put(url, auth=auth, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        return True
    except Exception as e:
        print(f"Error updating ticket type for {key} to {ticket_type}: {e}")
        return False


def get_jira_issue_transitions(key: str) -> list[dict]:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        return []

    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    url = f"{jira_base}/{jira_api_path}/issue/{key}/transitions"
    try:
        res = requests.get(url, auth=auth, headers=headers, timeout=15)
        res.raise_for_status()
        data = res.json()
        transitions = data.get("transitions", [])
        return [{"id": t.get("id"), "name": t.get("name")} for t in transitions]
    except Exception as e:
        print(f"Error fetching transitions for ticket {key}: {e}")
        return []


def transition_jira_issue(key: str, transition_id: str) -> bool:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        return False

    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    url = f"{jira_base}/{jira_api_path}/issue/{key}/transitions"
    payload = {
        "transition": {
            "id": transition_id
        }
    }

    try:
        res = requests.post(url, auth=auth, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        return True
    except Exception as e:
        print(f"Error transitioning ticket {key} with transition {transition_id}: {e}")
        return False


def update_jira_ticket_supplier(key: str, supplier: str) -> bool:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        print("Jira integration credentials are not configured in environment variables.")
        return False

    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    url = f"{jira_base}/{jira_api_path}/issue/{key}"
    payload = {
        "fields": {
            "customfield_14595": {
                "value": supplier
            } if supplier else None
        }
    }

    try:
        res = requests.put(url, auth=auth, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        return True
    except Exception as e:
        print(f"Error updating supplier for {key} to {supplier}: {e}")
        return False


def _parse_jira_adf(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        if "text" in value:
            return str(value["text"])
        return "\n".join(_parse_jira_adf(item) for item in value.get("content", []))
    if isinstance(value, list):
        return "\n".join(_parse_jira_adf(item) for item in value)
    return str(value)


def add_jira_comment(key: str, comment_text: str) -> bool:
    """
    Adiciona um comentário público (Reply to customer) a um ticket JIRA.

    Args:
        key:          Chave do ticket (ex: 'IZ-56680')
        comment_text: Texto do comentário em formato plano

    Returns:
        True se bem-sucedido, False caso contrário.
    """
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        print(f"[COMMENT] Credenciais JIRA não configuradas para {key}.")
        return False

    auth = (jira_email, jira_token)
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    url = f"{jira_base}/{jira_api_path}/issue/{key.upper().strip()}/comment"

    # Jira Cloud REST API v3 usa Atlassian Document Format (ADF)
    payload = {
        "body": {
            "type": "doc",
            "version": 1,
            "content": [
                {
                    "type": "paragraph",
                    "content": [
                        {
                            "type": "text",
                            "text": comment_text,
                        }
                    ],
                }
            ],
        }
    }

    try:
        res = requests.post(url, auth=auth, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        print(f"[COMMENT] Comentário adicionado ao ticket {key}.")
        return True
    except Exception as e:
        print(f"[COMMENT] Erro ao adicionar comentário ao ticket {key}: {e}")
        return False


def fetch_ticket_details(ticket_key: str) -> dict:
    """
    Busca os detalhes de um único ticket (Summary, Description, Comments) do JIRA,
    convertendo campos no formato ADF (Atlassian Document Format) para texto simples.
    """
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        raise RuntimeError(
            "Credenciais do Jira não configuradas. Preencha JIRA_DADOS_COMP_HASH, "
            "JIRA_EMAIL e JIRA_TOKEN no ficheiro .env e reinicie a aplicação."
        )

    auth = (jira_email, jira_token)
    headers = {"Accept": "application/json"}
    url = f"{jira_base}/{jira_api_path}/issue/{ticket_key.upper().strip()}"

    try:
        res = requests.get(
            url,
            auth=auth,
            headers=headers,
            params={"fields": "summary,description,comment,customfield_15845"},
            timeout=15,
        )
        res.raise_for_status()
        data = res.json()
        fields = data.get("fields", {})

        # Converter descrição ADF para texto simples
        desc_raw = fields.get("description")
        description = _parse_jira_adf(desc_raw)

        # Converter comentários ADF para lista de textos simples
        comments_raw = fields.get("comment", {}).get("comments", [])
        comments = []
        for c in comments_raw:
            author = c.get("author", {}).get("displayName", "User")
            body = _parse_jira_adf(c.get("body"))
            comments.append(f"{author}: {body}")

        # IT SALSA - Categoria SAP (customfield_15845)
        categoria_raw = fields.get("customfield_15845")
        if isinstance(categoria_raw, dict):
            categoria_sap = categoria_raw.get("value") or ""
        elif isinstance(categoria_raw, str):
            categoria_sap = categoria_raw
        else:
            categoria_sap = ""

        return {
            "summary": str(fields.get("summary") or ""),
            "description": description,
            "comments": comments,
            "categoria_sap": categoria_sap,
        }
    except requests.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else 0
        if status in (401, 403):
            raise RuntimeError(
                "O Jira recusou as credenciais (token inválido/expirado ou utilizador sem acesso). "
                "Crie um novo API token Atlassian, atualize JIRA_EMAIL e JIRA_TOKEN no .env "
                "e reinicie a aplicação."
            ) from exc
        if status == 404:
            raise RuntimeError(
                f"O ticket {ticket_key.upper().strip()} não existe ou o utilizador não tem acesso."
            ) from exc
        raise RuntimeError(f"O Jira devolveu o erro HTTP {status or 'desconhecido'}.") from exc
    except requests.RequestException as exc:
        raise RuntimeError(
            "Não foi possível ligar ao Jira. Verifique a internet, VPN/proxy e tente novamente."
        ) from exc
    except (TypeError, ValueError) as exc:
        raise RuntimeError("O Jira devolveu uma resposta inválida.") from exc


def fetch_single_ticket_for_trigger(key: str) -> dict | None:
    jira_base = os.getenv("JIRA_DADOS_COMP_HASH", "").strip().rstrip("/")
    jira_email = os.getenv("JIRA_EMAIL", "").strip()
    jira_token = os.getenv("JIRA_TOKEN", "").strip()
    jira_api_path = os.getenv("JIRA_DADOS_HASH", "rest/api/3").strip().strip("/")

    if not jira_base or not jira_email or not jira_token:
        return None

    auth = (jira_email, jira_token)
    headers = {"Accept": "application/json"}
    return _fetch_single_issue(key, jira_base, jira_api_path, auth, headers)


def clean_excel_leading_spaces(file_path: str) -> None:
    """
    Abre o ficheiro Excel com openpyxl e, para cada célula de texto em todas as folhas,
    se o valor começar com um espaço (" "), remove apenas esse primeiro caractere.
    Grava as alterações mantendo a formatação e as fórmulas intactas.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        modified = False
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith(" "):
                        cell.value = val[1:]
                        modified = True
        if modified:
            wb.save(file_path)
            print(f"[EXCEL CLEAN] Ficheiro {file_path} limpo e gravado com sucesso.")
        wb.close()
    except Exception as e:
        print(f"[EXCEL CLEAN ERROR] Erro ao processar/limpar espaços de {file_path}: {e}")


