# -*- coding: utf-8 -*-
"""
Ler ficheiros de favoritos SAP e escrever apenas transações válidas no Excel selecionado por popup.

Regras:
- O Excel é sempre escolhido por popup em primeiro plano
- A sheet a usar é: "Users Ativos"
- O script procura dinamicamente a linha do cabeçalho
- A coluna obrigatória é: "Usuário"
- Para cada utilizador:
    - procura o ficheiro com o nome exato do utilizador em C:\Favoritos
    - lê esse ficheiro
    - extrai a transação real removendo o prefixo técnico SAP, por exemplo:
        TR0000100002BNK_APP    Autorizar pagamentos
      resultado:
        BNK_APP
    - ignora valores técnicos/numéricos como:
        0000102387
    - escreve as transações válidas a partir da coluna G para a frente
- Também escreve STATUS / MSG / TIMESTEMP no próprio Excel

Requisitos:
    pip install openpyxl
"""

import os
import re
import sys
import traceback
from datetime import datetime
from tkinter import Tk, filedialog, messagebox

from openpyxl import load_workbook


# =========================================================
# CONFIGURAÇÕES
# =========================================================
SHEET_NAME = "Users Ativos"
FAVORITOS_DIR = r"C:\Favoritos"
HEADER_SEARCH_MAX_ROWS = 20

HEADER_USUARIO = "Usuário"

COLUNA_INICIAL_TRANSACOES = 7  # G
HEADER_STATUS = "STATUS"
HEADER_MSG = "MSG"
HEADER_TIMESTEMP = "TIMESTEMP"


# =========================================================
# UTILITÁRIOS
# =========================================================
def log(msg: str) -> None:
    print(msg, flush=True)


def agora_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalizar_texto(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_cabecalho(valor) -> str:
    return normalizar_texto(valor).upper()


def selecionar_excel_popup() -> str:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update()

    caminho = filedialog.askopenfilename(
        parent=root,
        title="Selecionar ficheiro Excel",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("All files", "*.*"),
        ],
    )

    root.update()
    root.destroy()

    if not caminho:
        raise RuntimeError("Nenhum ficheiro Excel foi selecionado no popup.")

    return caminho


# =========================================================
# EXCEL
# =========================================================
def localizar_linha_cabecalho(ws):
    """
    Procura a linha do cabeçalho nas primeiras HEADER_SEARCH_MAX_ROWS linhas.
    """
    limite_linhas = min(ws.max_row, HEADER_SEARCH_MAX_ROWS)
    limite_colunas = ws.max_column

    for row_idx in range(1, limite_linhas + 1):
        mapa = {}

        for col_idx in range(1, limite_colunas + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            chave = normalizar_cabecalho(valor)
            if chave:
                mapa[chave] = col_idx

        if normalizar_cabecalho(HEADER_USUARIO) in mapa:
            return row_idx, mapa

    raise RuntimeError(
        f"Não foi possível localizar a linha do cabeçalho com a coluna '{HEADER_USUARIO}'."
    )


def garantir_coluna(ws, header_row: int, headers_map: dict, nome_coluna: str) -> int:
    chave = normalizar_cabecalho(nome_coluna)

    if chave in headers_map:
        return headers_map[chave]

    nova_coluna = ws.max_column + 1
    ws.cell(row=header_row, column=nova_coluna).value = nome_coluna
    headers_map[chave] = nova_coluna
    return nova_coluna


def obter_linhas_utilizadores(ws, header_row: int, headers_map: dict):
    col_usuario = headers_map[normalizar_cabecalho(HEADER_USUARIO)]
    col_status = garantir_coluna(ws, header_row, headers_map, HEADER_STATUS)
    col_msg = garantir_coluna(ws, header_row, headers_map, HEADER_MSG)
    col_ts = garantir_coluna(ws, header_row, headers_map, HEADER_TIMESTEMP)

    linhas = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        usuario = normalizar_texto(ws.cell(row=row_idx, column=col_usuario).value)
        if not usuario:
            continue

        linhas.append(
            {
                "row": row_idx,
                "usuario": usuario,
                "status_col": col_status,
                "msg_col": col_msg,
                "ts_col": col_ts,
            }
        )

    return linhas


def limpar_transacoes_da_linha(ws, row_idx: int, start_col: int = COLUNA_INICIAL_TRANSACOES):
    """
    Limpa as transações antigas da coluna G para a frente.
    """
    for col_idx in range(start_col, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).value = None


def escrever_transacoes_na_linha(ws, row_idx: int, transacoes: list[str], start_col: int = COLUNA_INICIAL_TRANSACOES):
    """
    Escreve uma transação por célula: G, H, I, J...
    """
    for i, transacao in enumerate(transacoes):
        ws.cell(row=row_idx, column=start_col + i).value = transacao


def escrever_resultado(ws, row_idx: int, status_col: int, msg_col: int, ts_col: int, status: str, msg: str):
    ws.cell(row=row_idx, column=status_col).value = status
    ws.cell(row=row_idx, column=msg_col).value = msg
    ws.cell(row=row_idx, column=ts_col).value = agora_str()


# =========================================================
# FAVORITOS SAP
# =========================================================
def caminho_favorito_do_usuario(usuario: str) -> str:
    """
    Procura o ficheiro com nome exato do utilizador em C:\Favoritos
    Exemplo:
        utilizador = 168801
        ficheiro esperado = C:\Favoritos\168801
    """
    caminho = os.path.join(FAVORITOS_DIR, usuario)

    if not os.path.isfile(caminho):
        raise FileNotFoundError(f"Ficheiro não encontrado: {caminho}")

    return caminho


def ler_linhas_ficheiro(caminho: str) -> list[str]:
    """
    Lê o ficheiro tentando os encodings mais comuns.
    """
    encodings = ["utf-8", "latin-1", "cp1252"]
    ultimo_erro = None

    for enc in encodings:
        try:
            with open(caminho, "r", encoding=enc) as f:
                return f.readlines()
        except Exception as e:
            ultimo_erro = e

    raise RuntimeError(f"Não foi possível ler o ficheiro: {caminho}. Erro: {ultimo_erro}")


def extrair_transacao_da_linha(linha: str) -> str:
    """
    Extrai o código real da transação removendo o prefixo técnico do SAP.

    Exemplo de linha:
        TR0000100002BNK_APP                                         Autorizar pagamentos

    Resultado:
        BNK_APP
    """
    linha = linha.rstrip("\r\n")

    if not linha.strip():
        return ""

    texto = linha.strip()

    # Caso padrão do export SAP:
    # TR0000100002BNK_APP    Autorizar pagamentos
    match = re.match(r"^TR\d{10}([A-Z0-9_/\-]+)", texto)
    if match:
        return match.group(1).strip()

    # Fallback: separa pelo bloco grande de espaços/tabs e usa a primeira parte
    partes = re.split(r"\s{2,}|\t+", texto, maxsplit=1)
    if partes and partes[0]:
        primeira_parte = partes[0].strip()

        # Se ainda vier com o prefixo técnico colado, remove
        match_fallback = re.match(r"^TR\d{10}(.+)$", primeira_parte)
        if match_fallback:
            return match_fallback.group(1).strip()

        return primeira_parte

    return ""


def transacao_valida(valor: str) -> bool:
    """
    Mantém apenas valores que parecem realmente uma transação SAP.

    Regras:
    - ignora vazio
    - ignora valores apenas numéricos, ex: 0000102387
    - aceita apenas códigos que começam por letra
    - depois podem ter letras, números, underscore, slash ou hífen
    """
    valor = normalizar_texto(valor).upper()

    if not valor:
        return False

    if re.fullmatch(r"\d+", valor):
        return False

    if re.fullmatch(r"[A-Z][A-Z0-9_/\-]*", valor):
        return True

    return False


def extrair_transacoes_do_ficheiro(caminho: str) -> list[str]:
    """
    Extrai as transações válidas do ficheiro removendo duplicados,
    mas preservando a ordem original.
    """
    linhas = ler_linhas_ficheiro(caminho)

    transacoes = []
    vistos = set()

    for linha in linhas:
        transacao = extrair_transacao_da_linha(linha)

        if not transacao:
            continue

        if not transacao_valida(transacao):
            continue

        if transacao not in vistos:
            vistos.add(transacao)
            transacoes.append(transacao)

    return transacoes


# =========================================================
# MAIN
# =========================================================
def main():
    log("###################################################################################")
    log("🚀 INÍCIO | LEITURA DOS FAVORITOS SAP")
    log("###################################################################################")

    # Excel sempre vem do popup
    excel_path = selecionar_excel_popup()
    log(f"[POPUP] Excel selecionado: {excel_path}")

    if not os.path.isdir(FAVORITOS_DIR):
        raise RuntimeError(f"A diretoria de favoritos não existe: {FAVORITOS_DIR}")

    wb = load_workbook(excel_path)

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f'A sheet "{SHEET_NAME}" não existe no Excel selecionado.')

    ws = wb[SHEET_NAME]

    header_row, headers_map = localizar_linha_cabecalho(ws)
    log(f"[EXCEL] Linha do cabeçalho localizada: {header_row}")

    linhas = obter_linhas_utilizadores(ws, header_row, headers_map)

    if not linhas:
        raise RuntimeError("Não foram encontrados utilizadores abaixo da coluna 'Usuário'.")

    log(f"[EXCEL] Total de utilizadores encontrados: {len(linhas)}")
    log(f"[DIR] Diretoria de favoritos: {FAVORITOS_DIR}")

    total_ok = 0
    total_erro = 0

    for idx, item in enumerate(linhas, start=1):
        row_idx = item["row"]
        usuario = item["usuario"]

        log("------------------------------------------------------------")
        log(f"[{idx}/{len(linhas)}] PROCESSANDO | LINHA={row_idx} | USUÁRIO={usuario}")

        try:
            caminho_ficheiro = caminho_favorito_do_usuario(usuario)
            transacoes = extrair_transacoes_do_ficheiro(caminho_ficheiro)

            limpar_transacoes_da_linha(ws, row_idx, COLUNA_INICIAL_TRANSACOES)
            escrever_transacoes_na_linha(ws, row_idx, transacoes, COLUNA_INICIAL_TRANSACOES)

            mensagem = f"{len(transacoes)} transação(ões) válida(s) encontrada(s)"
            escrever_resultado(
                ws=ws,
                row_idx=row_idx,
                status_col=item["status_col"],
                msg_col=item["msg_col"],
                ts_col=item["ts_col"],
                status="OK",
                msg=mensagem,
            )

            wb.save(excel_path)
            total_ok += 1

            log(f"[OK] FICHEIRO={caminho_ficheiro} | TOTAL_TRANSAÇÕES_VÁLIDAS={len(transacoes)}")

        except Exception as e:
            limpar_transacoes_da_linha(ws, row_idx, COLUNA_INICIAL_TRANSACOES)

            escrever_resultado(
                ws=ws,
                row_idx=row_idx,
                status_col=item["status_col"],
                msg_col=item["msg_col"],
                ts_col=item["ts_col"],
                status="ERRO",
                msg=normalizar_texto(str(e)),
            )

            wb.save(excel_path)
            total_erro += 1

            log(f"[ERRO] LINHA={row_idx} | USUÁRIO={usuario} | MSG={normalizar_texto(str(e))}")

    wb.save(excel_path)

    log("###################################################################################")
    log(f"[FIM] SUCESSO={total_ok} | ERROS={total_erro} | EXCEL={excel_path}")
    log("###################################################################################")

    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showinfo(
        "Concluído",
        f"Processamento concluído.\n\nSucesso: {total_ok}\nErros: {total_erro}\n\nExcel: {excel_path}",
        parent=root,
    )
    root.destroy()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\n[ERRO FATAL]")
        print("".join(traceback.format_exception_only(type(exc), exc)).strip())
        sys.exit(1)